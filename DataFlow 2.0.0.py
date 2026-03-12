# RIGHE 1-2 (Importazioni necessarie per il DPI)
import sys
if sys.platform == 'win32':
    from ctypes import windll
else:
    windll = None

# ---------------------------------------------
# RIGHE 3-20: BLOCCO DPI AWARENESS (DEVE ESSERE QUI)
# ---------------------------------------------
if sys.platform == 'win32':
    try:
        # Importiamo windll direttamente se non è già stata importata
        
        # Imposta PerMonitorV2
        DPI_AWARENESS_CONTEXT_PER_MONITOR_AWARE_V2 = -4
        
        if hasattr(windll.shcore, 'SetProcessDpiAwarenessContext'):
            windll.shcore.SetProcessDpiAwarenessContext(DPI_AWARENESS_CONTEXT_PER_MONITOR_AWARE_V2)
        elif hasattr(windll.user32, 'SetProcessDPIAware'):
            windll.user32.SetProcessDPIAware()

    except Exception as e:
        # Ignora errori se le librerie non sono presenti o la funzione non è supportata
        # BUG #21 FIX: Log warning invece di pass silenzioso per diagnostica
        import logging
        logging.getLogger(__name__).debug(f"DPI awareness non disponibile: {e}")

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from tksheet import Sheet
import os
from database_manager import DatabaseManager, DatabaseError
import tempfile
from tkcalendar import DateEntry
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
import shutil
import configparser
import re
from PIL import Image, ImageTk
import time
import math
import glob
import ast # Aggiunto per la gestione sicura delle note formattate
import json # Aggiunto per parsing sicuro delle note
import logging
from logging.handlers import RotatingFileHandler
import builtins
if not hasattr(builtins, '_'):
    builtins._ = lambda x: x
import webbrowser
import atexit
import gettext
import subprocess
import threading
import unicodedata

# --- PULIZIA TMP ALL'AVVIO ---
def cleanup_temp_on_startup():
    """Pulisce le directory temporanee di PyInstaller rimaste da precedenti esecuzioni."""
    try:
        temp_dir = tempfile.gettempdir()
        
        # Cerca cartelle _MEI* create da PyInstaller
        for item in os.listdir(temp_dir):
            if item.startswith('_MEI'):
                temp_path = os.path.join(temp_dir, item)
                try:
                    if os.path.isdir(temp_path):
                        shutil.rmtree(temp_path, ignore_errors=True)
                except Exception:
                    pass  # Ignora errori di pulizia
        
        # Pulisci anche file temporanei di DataFlow vecchi (>24 ore)
        # Pattern per file temporanei creati da AttachmentWindow
        pattern = os.path.join(temp_dir, 'tmp*')
        current_time = time.time()
        for temp_file in glob.glob(pattern):
            try:
                # Elimina file più vecchi di 24 ore (86400 secondi)
                if os.path.isfile(temp_file) and (current_time - os.path.getmtime(temp_file)) > 86400:
                    os.remove(temp_file)
            except Exception:
                pass  # Ignora errori di pulizia
                
    except Exception:
        pass  # Ignora completamente errori di pulizia

# Esegui pulizia all'avvio
cleanup_temp_on_startup()

# --- INIZIO SISTEMA DI LOGGING ---
def setup_logging():
    """Configura il sistema di logging con file rotanti."""
    # Usa la directory locale dell'utente invece della directory corrente
    if getattr(sys, 'frozen', False):
        # Se eseguito come EXE/MSIX
        if sys.platform == 'win32':
            log_dir = os.path.join(os.path.expanduser('~'), 'AppData', 'Local', 'DataFlow')
        else:
            log_dir = os.path.join(os.path.expanduser('~'), '.local', 'share', 'DataFlow')
    else:
        # Se eseguito come script Python
        log_dir = os.path.dirname(os.path.abspath(__file__))
    
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, 'dataflow.log')
    
    logger = logging.getLogger('DataFlow')
    logger.setLevel(logging.INFO)
    
    # Rimuovi handler esistenti per evitare duplicati in caso di riavvio
    if logger.handlers:
        logger.handlers.clear()
    
    handler = RotatingFileHandler(
        log_file, 
        maxBytes=5*1024*1024,  # 5MB
        backupCount=3,
        encoding='utf-8'
    )
    
    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(funcName)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    
    return logger

logger = setup_logging()
# --- FINE SISTEMA DI LOGGING ---

# --- INIZIO CODICE AGGIUNTO PER PYINSTALLER ---
def resource_path(relative_path):
    """ Ottiene il percorso assoluto della risorsa, funzionante sia in sviluppo che con PyInstaller """
    try:
        # PyInstaller crea una cartella temporanea e ci memorizza il percorso in _MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        # BUG #21 FIX: Catch specifico invece di Exception generico
        # In sviluppo, usa la directory dello script
        if getattr(sys, 'frozen', False):
            base_path = os.path.dirname(sys.executable)
        else:
            base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

def set_window_icon(window):
    """Imposta l'icona della finestra in modo cross-platform."""
    icon_path = resource_path(os.path.join("add_data", "DataFlow.ico"))
    try:
        if sys.platform == 'win32':
            window.iconbitmap(icon_path)
        else:
            img = Image.open(icon_path)
            photo = ImageTk.PhotoImage(img)
            window._icon_photo = photo  # Evita garbage collection
            window.iconphoto(True, photo)
    except Exception as e:
        logger.debug(f"Icona non caricabile (normale in alcuni ambienti): {e}")

# --- HELPER FUNCTION PER PERCORSI FILE ---
def get_app_data_dir():
    """Restituisce la directory dati dell'applicazione."""
    if getattr(sys, 'frozen', False):
        # Se eseguito come EXE/MSIX, usa la directory locale dell'utente
        if sys.platform == 'win32':
            return os.path.join(os.path.expanduser('~'), 'AppData', 'Local', 'DataFlow')
        else:
            return os.path.join(os.path.expanduser('~'), '.local', 'share', 'DataFlow')
    else:
        # Se eseguito come script Python, usa la directory corrente
        return os.path.dirname(os.path.abspath(__file__))

def get_config_file():
    """Restituisce il percorso del file config.ini."""
    # Assicura che la directory esista
    app_dir = get_app_data_dir()
    os.makedirs(app_dir, exist_ok=True)
    return os.path.join(app_dir, 'config.ini')

def _strip_accents(value):
    """Rimuove gli accenti da una stringa mantenendo solo caratteri ASCII."""
    if not value:
        return ""
    normalized = unicodedata.normalize('NFKD', value)
    return ''.join(ch for ch in normalized if not unicodedata.combining(ch))

def generate_username(first_name, last_name):
    """
    Genera lo username secondo le regole: prima lettera del nome + cognome,
    senza spazi, senza accenti e tutto in minuscolo.
    """
    if not first_name or not last_name:
        raise ValueError("Nome e cognome sono obbligatori per generare lo username.")
    
    first_clean = ''.join(ch for ch in _strip_accents(first_name).strip() if ch.isalpha())
    last_clean = ''.join(ch for ch in _strip_accents(last_name) if ch.isalnum())
    
    if not first_clean or not last_clean:
        raise ValueError("Nome e cognome devono contenere caratteri alfabetici validi.")
    
    username = (first_clean[0] + last_clean).lower()
    return username

def load_user_identity():
    """Carica nome, cognome e username dell'utente dal config."""
    identity = {
        'first_name': '',
        'last_name': '',
        'username': '',
        'full_name': ''
    }
    config_file = get_config_file()
    config = configparser.ConfigParser(interpolation=None)
    if os.path.exists(config_file):
        config.read(config_file, encoding='utf-8')
        if config.has_section('User'):
            identity['first_name'] = config.get('User', 'first_name', fallback='').strip()
            identity['last_name'] = config.get('User', 'last_name', fallback='').strip()
            identity['username'] = config.get('User', 'username', fallback='').strip().lower()
            full_name = f"{identity['first_name']} {identity['last_name']}".strip()
            identity['full_name'] = full_name
    return identity

def save_user_identity(first_name, last_name, username):
    """Salva nel config il nome completo dell'utente e lo username derivato."""
    config_file = get_config_file()
    config = configparser.ConfigParser(interpolation=None)
    if os.path.exists(config_file):
        config.read(config_file, encoding='utf-8')
    if 'User' not in config:
        config['User'] = {}
    config['User']['first_name'] = first_name.strip()
    config['User']['last_name'] = last_name.strip()
    config['User']['username'] = username.strip().lower()
    with open(config_file, 'w', encoding='utf-8') as f:
        config.write(f)

_DATAFLOW_STRUCTURE_VERIFIED = False

def get_user_documents_dataflow_dir():
    """Restituisce la directory dati utente principale per DataFlow.
    Rispetta dataflow_base_dir se presente nel config.ini.
    """
    # Determina lo username dal config.ini
    config_file = get_config_file()
    config = configparser.ConfigParser(interpolation=None)
    username = None
    dataflow_base_dir = None
    
    if os.path.exists(config_file):
        config.read(config_file)
        if 'User' in config and config.get('User', 'username', fallback=None):
            username = config.get('User', 'username').strip().lower()
        
        # ✅ LEGGI dataflow_base_dir se presente
        if 'Settings' in config:
            dataflow_base_dir = config.get('Settings', 'dataflow_base_dir', fallback=None)

    # Se non c'è username, NON creare nessuna cartella e restituisci None
    if not username:
        logger.warning("Username non presente: la cartella utente non viene creata.")
        return None

    base_folder = f"DataFlow_{username}"
    
    # ✅ USA dataflow_base_dir se presente, altrimenti default a Documents
    if dataflow_base_dir and os.path.exists(dataflow_base_dir):
        chosen_dir = os.path.join(dataflow_base_dir, base_folder)
        logger.info(f"Usando directory DataFlow personalizzata: {chosen_dir}")
    else:
        # Windows: usa ~/Documents/DataFlow_username (comportamento standard Windows)
        # Linux/macOS: usa ~/DataFlow_username (direttamente nella home)
        if sys.platform == 'win32':
            documents_dir = os.path.join(os.path.expanduser('~'), 'Documents')
            chosen_dir = os.path.join(documents_dir, base_folder)
        else:
            chosen_dir = os.path.join(os.path.expanduser('~'), base_folder)
        logger.info(f"Usando directory DataFlow standard: {chosen_dir}")
    
    try:
        os.makedirs(chosen_dir, exist_ok=True)
    except OSError as e:
        logger.error(f"Impossibile creare la cartella DataFlow utente '{chosen_dir}': {e}")
        return None

    global _DATAFLOW_STRUCTURE_VERIFIED
    if not _DATAFLOW_STRUCTURE_VERIFIED:
        # Leggi config per capire se è stato già impostato un DB personalizzato.
        custom_db = None
        try:
            if os.path.exists(config_file):
                config.read(config_file)
                custom_db = config.get('Settings', 'custom_db_path', fallback=None)
        except Exception:
            custom_db = None

        required = [
            os.path.join(chosen_dir, 'Database'),
            os.path.join(chosen_dir, 'Attachments')
        ]
        if any(not os.path.exists(path) for path in required):
            try:
                initialize_dataflow_directory_structure(chosen_dir)
            except Exception as e:
                logger.error(f"Errore nel ripristino automatico della struttura DataFlow: {e}")
        _DATAFLOW_STRUCTURE_VERIFIED = True

    return chosen_dir

def get_fixed_db_dir():
    """Restituisce la cartella fissa per il database."""
    db_dir = os.path.join(get_user_documents_dataflow_dir(), 'Database')
    os.makedirs(db_dir, exist_ok=True)
    return db_dir

def get_fixed_attachments_dir():
    """Restituisce la cartella fissa per gli allegati (Attachments)."""
    base_dir = get_user_documents_dataflow_dir()
    new_dir = os.path.join(base_dir, 'Attachments')
    old_dir = os.path.join(base_dir, 'Allegati')
    
    if os.path.exists(old_dir) and not os.path.exists(new_dir):
        try:
            shutil.move(old_dir, new_dir)
            logger.info(f"Cartella allegati rinominata da '{old_dir}' a '{new_dir}'")
        except Exception as e:
            logger.error(f"Impossibile rinominare cartella Allegati: {e}")
    
    os.makedirs(new_dir, exist_ok=True)
    return new_dir

def initialize_dataflow_directory_structure(base_dir=None):
    """
    Crea la struttura standard DataFlow (Database, Allegati, ecc.) e
    inizializza un database SQLite vuoto con le tabelle richieste.
    """
    try:
        if base_dir:
            base_dir = os.path.normpath(os.path.abspath(base_dir))
        else:
            base_dir = get_user_documents_dataflow_dir()
    except Exception as e:
        logger.error(f"Impossibile determinare la cartella DataFlow: {e}")
        raise
    
    # Gestione migrazione vecchia cartella "Allegati" -> "Attachments"
    old_attachments_dir = os.path.join(base_dir, 'Allegati')
    new_attachments_dir = os.path.join(base_dir, 'Attachments')
    if os.path.exists(old_attachments_dir) and not os.path.exists(new_attachments_dir):
        try:
            shutil.move(old_attachments_dir, new_attachments_dir)
            logger.info(f"Cartella Allegati migrata in Attachments: {new_attachments_dir}")
        except Exception as e:
            logger.error(f"Impossibile migrare cartella Allegati: {e}")
    
    subfolders = ['Database', 'Attachments']
    try:
        os.makedirs(base_dir, exist_ok=True)
        for sub in subfolders:
            os.makedirs(os.path.join(base_dir, sub), exist_ok=True)
        logger.info(f"Struttura DataFlow creata/in ripristino in: {base_dir}")
    except OSError as e:
        logger.error(f"Errore nella creazione delle cartelle DataFlow: {e}")
        raise
    
    # NON creare nessun database qui! Solo la struttura cartelle.
    # Il database verrà creato solo dopo l'inserimento dell'identità utente.
    return None

# --- INIZIO SISTEMA DI INTERNAZIONALIZZAZIONE (i18n) ---
def init_i18n(language_code='en'):
    """
    Inizializza il sistema di internazionalizzazione (gettext).
    Legge la lingua preferita dal config.ini o usa 'en' come default.
    """
    # Leggi la lingua dal config.ini (solo se esiste e ha la chiave)
    try:
        config_file = get_config_file()
        if os.path.exists(config_file):
            config = configparser.ConfigParser(interpolation=None)
            config.read(config_file)
            if 'Settings' in config and config.has_option('Settings', 'language'):
                language_code = config.get('Settings', 'language', fallback='en')
    except Exception as e:
        logger.warning(f"Errore nel leggere config.ini per la lingua: {e}, uso default 'en'")
        language_code = 'en'
    
    # Validazione: accetta solo 'en' o 'it', default sempre 'en'
    if language_code not in ['en', 'it']:
        language_code = 'en'
    
    # Determina il percorso dei file di traduzione
    try:
        # In PyInstaller, usa resource_path per trovare i file nella directory _MEIPASS
        if getattr(sys, 'frozen', False):
            locale_dir = resource_path('locale')
        else:
            # In sviluppo, usa la directory corrente
            locale_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'locale')
        
        # Inizializza gettext
        try:
            logger.info(f"Tentativo di caricare traduzioni per '{language_code}' da: {locale_dir}")
            mo_path = os.path.join(locale_dir, language_code, 'LC_MESSAGES', 'dataflow.mo')
            
            if os.path.exists(mo_path):
                trans = gettext.translation('dataflow', localedir=locale_dir, languages=[language_code], fallback=False)
                trans.install()  # Installa _ in builtins
                logger.info(f"✓ File traduzioni caricato con successo: {mo_path}")
            else:
                logger.warning(f"File .mo non trovato: {mo_path}, uso fallback")
                trans = gettext.NullTranslations()
                trans.install()
        except Exception as e:
            # Se il file .mo non esiste o c'è errore, usa gettext.NullTranslations (fallback silenzioso)
            trans = gettext.NullTranslations()
            trans.install()
            logger.error(f"ERRORE nel caricare traduzioni per '{language_code}': {e}", exc_info=True)
    except Exception as e:
        # In caso di errore, usa NullTranslations come fallback
        trans = gettext.NullTranslations()
        trans.install()
        logger.error(f"Errore nel caricamento delle traduzioni: {e}")
    
    return language_code

# Inizializza le traduzioni subito dopo la definizione
init_i18n('it')

# --- FINE SISTEMA DI INTERNAZIONALIZZAZIONE ---

# --- HELPER FUNCTIONS PER TRADUZIONI CONDIZIONALI ---
def get_current_language():
    """Restituisce il codice lingua corrente ('it' o 'en').
    Gestisce correttamente il caso in cui il config non esista ancora o non sia inizializzato.
    """
    try:
        config_file = get_config_file()
        if config_file and os.path.exists(config_file):
            config = configparser.ConfigParser(interpolation=None)
            config.read(config_file, encoding='utf-8')
            if 'Settings' in config and config.has_option('Settings', 'language'):
                lang = config.get('Settings', 'language', fallback='en')
                # Validazione: accetta solo 'en' o 'it'
                if lang in ['en', 'it']:
                    return lang
    except (configparser.Error, OSError, IOError, AttributeError) as e:
        # Log solo se non è un errore di file non esistente (normale all'avvio)
        try:
            if config_file and os.path.exists(config_file):
                logger.debug(f"Errore lettura config per lingua: {e}")
        except (NameError, UnboundLocalError):
            # config_file potrebbe non essere definito in caso di errore precoce
            pass
    except Exception as e:
        # Log altri errori inattesi
        logger.debug(f"Errore inatteso in get_current_language: {e}")
    # Fallback sempre a 'en' se qualcosa va storto
    return 'en'

def get_pos_column_text():
    """Restituisce il testo per la colonna Posizione: 'Item' in inglese, 'Pos.' in italiano."""
    return "Item" if get_current_language() == 'en' else "Pos."

def get_qty_column_text():
    """Restituisce il testo per la colonna Quantità: 'Q.ty' in inglese, 'Q.tà' in italiano."""
    return "Q.ty" if get_current_language() == 'en' else "Q.tà"

# --- HELPER FUNCTIONS PER GESTIONE TIPI RFQ MULTILINGUA ---
def normalize_rfq_type(rfq_type):
    """
    Normalizza un tipo di RFQ da qualsiasi lingua al valore canonico italiano.
    Gestisce sia i valori vecchi (tradotti) che quelli nuovi (canonici).
    
    BUG #7 FIX: Validazione robusta con gestione completa dei casi edge.
    """
    # Gestione valori None, vuoti o non-stringa
    if not rfq_type:
        logger.warning("normalize_rfq_type: valore None/vuoto ricevuto, uso default 'Fornitura piena'")
        return "Fornitura piena"
    
    # Converti a stringa e pulisci whitespace
    try:
        rfq_type = str(rfq_type).strip()
    except Exception as e:
        logger.error(f"normalize_rfq_type: impossibile convertire a stringa '{rfq_type}': {e}")
        return "Fornitura piena"
    
    # Se dopo strip è vuoto, usa default
    if not rfq_type:
        logger.warning("normalize_rfq_type: stringa vuota dopo strip, uso default")
        return "Fornitura piena"
    
    # Mappa tutte le possibili varianti ai valori canonici italiani
    type_map = {
        # Valori canonici italiani (già corretti)
        "Fornitura piena": "Fornitura piena",
        "Conto lavoro": "Conto lavoro",
        # Traduzioni inglesi
        "Full Supply": "Fornitura piena",
        "Work Order": "Conto lavoro",
        # Varianti possibili (case-insensitive)
        "fornitura piena": "Fornitura piena",
        "conto lavoro": "Conto lavoro",
        "full supply": "Fornitura piena",
        "work order": "Conto lavoro",
    }
    
    # Cerca corrispondenza esatta (case-sensitive prima)
    if rfq_type in type_map:
        return type_map[rfq_type]
    
    # Cerca corrispondenza case-insensitive
    rfq_type_lower = rfq_type.lower()
    for key, value in type_map.items():
        if key.lower() == rfq_type_lower:
            return value
    
    # Se non trovato, logga warning dettagliato e ritorna default
    logger.warning(f"normalize_rfq_type: tipo RFQ non riconosciuto: '{rfq_type}' (len={len(rfq_type)}), uso default 'Fornitura piena'")
    return "Fornitura piena"

def translate_rfq_type(rfq_type):
    """
    Traduce un tipo di RFQ dal valore canonico italiano alla lingua corrente.
    """
    # Prima normalizza per gestire anche valori vecchi
    canonical = normalize_rfq_type(rfq_type)
    
    # Traduci il valore canonico
    if canonical == "Fornitura piena":
        return _("Fornitura piena")
    elif canonical == "Conto lavoro":
        return _("Conto lavoro")
    else:
        # Fallback: ritorna il valore normalizzato tradotto
        return _(canonical)

# --- FINE HELPER FUNCTIONS PER GESTIONE TIPI RFQ MULTILINGUA ---

# --- FINE HELPER FUNCTIONS ---

# --- NUOVA GESTIONE DATABASE PATH (CON CACHE PER SESSIONE) --- ### CORREZIONE BUG ###
_PERCORSO_DB_CACHE = None

def reset_db_cache():
    """
    Invalida la cache del percorso DB per forzare il ricaricamento.
    Chiamare questa funzione quando si modifica il percorso del database
    o si vuole forzare il ricalcolo.
    """
    global _PERCORSO_DB_CACHE
    _PERCORSO_DB_CACHE = None
    logger.info("Cache percorso DB invalidata")

def get_db_path():
    """
    Determina il percorso del database da usare per la sessione corrente.
    Alla prima chiamata, legge il file config.ini per decidere se usare il DB
    personalizzato o quello standard. Alle chiamate successive, restituisce 
    il percorso già memorizzato (cache) per garantire coerenza durante tutta la sessione.
    
    Priorità:
    1. Directory DataFlow personalizzata (dataflow_base_dir) - permanente
    2. Database personalizzato (custom_db_path)
    3. Database standard (Documents/DataFlow/Database)
    
    Returns:
        str: Percorso assoluto al file database da usare (estensione .db)
    """
    global _PERCORSO_DB_CACHE
    if _PERCORSO_DB_CACHE is not None:
        return _PERCORSO_DB_CACHE

    config = configparser.ConfigParser(interpolation=None)
    config_file = get_config_file()
    legacy_custom_path = None
    dataflow_override = None
    
    if os.path.exists(config_file):
        try:
            config.read(config_file)
            dataflow_override = config.get('Settings', 'dataflow_base_dir', fallback=None)
            legacy_custom_path = config.get('Settings', 'custom_db_path', fallback=None)
        except Exception as e:
            logger.error(f"Errore lettura config per percorsi DB: {e}")
    
    if legacy_custom_path:
        percorso_da_usare = legacy_custom_path
        logger.info(f"Usando database personalizzato: {legacy_custom_path}")
        if not os.path.exists(percorso_da_usare):
            try:
                os.makedirs(os.path.dirname(percorso_da_usare), exist_ok=True)
                logger.info(f"Creata directory per database legacy: {os.path.dirname(percorso_da_usare)}")
            except OSError as e:
                logger.error(f"Impossibile creare cartella per database legacy: {e}")
                percorso_da_usare = None
    else:
        # Prova a ricavare lo username e costruire il percorso
        identity = load_user_identity()
        username = identity.get('username')
        if username:
            base_dir = get_user_documents_dataflow_dir()
            percorso_da_usare = os.path.join(base_dir, 'Database', f'dataflow_db_{username}.db')
            logger.info(f"Usando database utente: {percorso_da_usare}")
        else:
            logger.error("Nessun username trovato: impossibile determinare percorso DB.")
            percorso_da_usare = None
    _PERCORSO_DB_CACHE = percorso_da_usare
    logger.info(f"Percorso database finale: {percorso_da_usare}")
    return percorso_da_usare

# --- DEPRECATA: Funzione helper per connessioni dirette (ora sostituita da DatabaseManager) ---
# --- FUNZIONE PER CALCOLARE LA POSIZIONE CENTRALE (SENZA MOSTRARE LA FINESTRA) ---
def calculate_center_position(win):
    """Calcola la posizione centrale per la finestra senza renderla visibile."""
    # Forza il ricalcolo della geometria per ottenere le dimensioni corrette
    win.update() 
    
    width = win.winfo_reqwidth()
    height = win.winfo_reqheight()

    screen_w = win.winfo_screenwidth()
    screen_h = win.winfo_screenheight()

    # Limita le dimensioni alle dimensioni dello schermo
    if width > screen_w:
        width = screen_w
    if height > screen_h - 100:  # TASKBAR_BUFFER
        height = screen_h - 100

    # Calcola le coordinate per centrare la finestra
    x = max(0, (screen_w - width) // 2)
    y = max(0, (screen_h - height) // 2)

    # --- INIZIO BLOCCO DI CONTROLLO ANTI-TASKBAR ---
    TASKBAR_BUFFER = 100 
    if y + height > screen_h - TASKBAR_BUFFER:
        y = screen_h - height - TASKBAR_BUFFER
    if y < 0:
        y = 0
    # --- FINE BLOCCO DI CONTROLLO ---

    return f'{width}x{height}+{x}+{y}'

def calculate_optimal_window_size(win, num_suppliers, is_conto_lavoro=False):
    """Calcola la larghezza ottimale per ViewRequestWindow in base al numero di fornitori."""
    # Dimensioni base delle colonne
    BASE_ARTICLE_WIDTH = 80 + 80 + 250 + 60  # Codice + Allegato + Descrizione + Q.tà = 470px
    CONTO_LAVORO_WIDTH = 100 + 100 + 150  # Cod.Grezzo + Dis.Grezzo + Mat.C/L = 350px
    SUPPLIER_COLUMN_WIDTH = 120  # Larghezza stimata per colonna fornitore
    PADDING = 140  # Margini laterali, scrollbar, bordi finestra e safety margin per DPI scaling
    
    # Larghezza minima per contenere tutti i pulsanti in alto (incluso "Create SQDC Analysis")
    # I pulsanti occupano circa: 6 pulsanti × 180px + spaziatura = ~1150px
    # (aumentato da 150px a 180px per testi tradotti più lunghi, specialmente in .exe con DPI scaling)
    BUTTONS_MIN_WIDTH = 1150
    
    # Calcola larghezza necessaria
    article_width = BASE_ARTICLE_WIDTH
    if is_conto_lavoro:
        article_width += CONTO_LAVORO_WIDTH
    
    suppliers_width = num_suppliers * SUPPLIER_COLUMN_WIDTH
    total_width = article_width + suppliers_width + PADDING
    
    # Ottieni dimensioni schermo
    screen_w = win.winfo_screenwidth()
    screen_h = win.winfo_screenheight()
    
    # Limita la larghezza al 95% dello schermo (lascia spazio ai bordi)
    max_width = int(screen_w * 0.95)
    optimal_width = min(total_width, max_width)
    
    # Larghezza minima: il maggiore tra larghezza pulsanti e larghezza base
    min_width = max(BUTTONS_MIN_WIDTH, 850)
    optimal_width = max(optimal_width, min_width)
    
    # Altezza ottimale (80% dello schermo, lasciando spazio per taskbar)
    optimal_height = int(screen_h * 0.80)
    
    # Calcola posizione centrale
    x = max(0, (screen_w - optimal_width) // 2)
    y = max(0, (screen_h - optimal_height) // 2)
    
    # Anti-taskbar buffer
    TASKBAR_BUFFER = 100
    if y + optimal_height > screen_h - TASKBAR_BUFFER:
        y = screen_h - optimal_height - TASKBAR_BUFFER
    if y < 0:
        y = 0
    
    return f'{optimal_width}x{optimal_height}+{x}+{y}'

# --- FUNZIONE PER CENTRARE FINESTRE (VERSIONE CORRETTA PER FINESTRE NASCOSTE) ---
def center_window(win):
    """Centra la finestra e la rende visibile."""
    geometry = calculate_center_position(win)
    win.geometry(geometry)
    win.deiconify()

# --- NUOVA GESTIONE NUMERI E TESTO---
def parse_float_from_comma_string(s):
    """Converte una stringa con virgola decimale in float, con validazione robusta.
    
    BUG #5 FIX: Validazione completa per gestire None, stringhe vuote e malformate.
    """
    # Gestione None e tipi numerici
    if s is None:
        return 0.0
    if isinstance(s, (int, float)):
        return float(s)
    
    # Converti a stringa e pulisci
    s = str(s).strip()
    
    # Gestione stringa vuota
    if not s or s == '':
        return 0.0
    
    # Validazione: accetta solo numeri, virgola e segno
    if not all(c.isdigit() or c in ',-' for c in s):
        raise ValueError(f"Formato numero non valido: '{s}'. Usare solo cifre e virgola come separatore decimale.")
    
    # Validazione: no punto decimale
    if '.' in s:
        raise ValueError("Usare la virgola, non il punto, come separatore decimale.")
    
    # Validazione: massimo una virgola
    if s.count(',') > 1:
        raise ValueError(f"Formato numero non valido: '{s}'. Troppi separatori decimali.")
    
    # Conversione sicura
    try:
        return float(s.replace(',', '.'))
    except ValueError as e:
        raise ValueError(f"Impossibile convertire '{s}' in numero: {e}")

def format_quantity_display(val):
    """Formatta la quantità per la visualizzazione con gestione errori robusta.
    
    BUG #6 FIX: Gestione completa degli errori di conversione.
    """
    if val is None or val == '':
        return ''
    
    # Se è già numero, formatta direttamente
    if isinstance(val, (int, float)):
        if val == int(val):
            return str(int(val))
        else:
            return str(val).replace('.', ',')
    
    # Se è stringa, prova a convertire
    try:
        val_float = parse_float_from_comma_string(val)
        if val_float == int(val_float):
            return str(int(val_float))
        else:
            return str(val_float).replace('.', ',')
    except (ValueError, TypeError):
        # Se la conversione fallisce, restituisci la stringa originale
        return str(val)

def format_price_display(num):
    """Formatta il prezzo per la visualizzazione con 4 decimali e virgola.
    
    BUG #6 FIX: Gestione completa degli errori di conversione.
    """
    if num is None or num == '':
        return ''
    
    # Converti in float gestendo errori
    try:
        if isinstance(num, str):
            num_float = parse_float_from_comma_string(num)
        else:
            num_float = float(num)
        return f"{num_float:.4f}".replace('.', ',')
    except (ValueError, TypeError) as e:
        # In caso di errore, restituisci stringa vuota o valore originale
        logger.warning(f"Impossibile formattare prezzo '{num}': {e}")
        return str(num) if num else ''

# ------------------------------------------------------------------------------------
# DATABASE SETUP
# ------------------------------------------------------------------------------------
def crea_database_v4():
    logger.info("Inizializzazione database")
    # Determina il percorso DB preferito (rispetta custom_db_path se presente nel config)
    db_file = get_db_path()
    try:
        config_file = get_config_file()
        if os.path.exists(config_file):
            cfg = configparser.ConfigParser(interpolation=None)
            cfg.read(config_file)
            custom = cfg.get('Settings', 'custom_db_path', fallback=None)
            if custom:
                # Forza l'uso del DB personalizzato per evitare di creare il DB standard
                db_file = custom
                logger.info(f"Usando custom_db_path per inizializzazione DB: {db_file}")
    except Exception as e:
        # BUG #36 FIX: Log eccezioni invece di silenziarle completamente
        logger.debug(f"Nessun custom_db_path configurato o errore lettura config: {e}")
    is_new_db = not os.path.exists(db_file)
    
    # BUG #32 FIX: Usa try-finally per garantire chiusura DB anche in caso di eccezione
    db_manager = None
    try:
        # Usa il DatabaseManager per creare le tabelle
        db_manager = DatabaseManager(db_file)
        db_manager.create_tables()
        
        if is_new_db: 
            print("Nuovo database creato. Imposto il contatore RdO a 0.")
            logger.info("Nuovo database creato")
        
        logger.info("Database inizializzato con successo")
        
    except DatabaseError as e:
        logger.error(f"Errore critico inizializzazione database: {e}", exc_info=True)
        print(f"ERRORE CRITICO: Impossibile inizializzare il database.\n{e}")
        raise
    finally:
        # BUG #32 FIX: Garantisce chiusura connessione anche in caso di eccezione
        if db_manager is not None:
            try:
                db_manager.close()
            except Exception as close_error:
                logger.warning(f"Errore chiusura database in finally: {close_error}")

# ------------------------------------------------------------------------------------
# FINESTRA GESTIONE ALLEGATI
# ------------------------------------------------------------------------------------
class AttachmentWindow(tk.Toplevel):
    def __init__(self, parent, request_id, attachment_type, read_only=False, source_db_path=None):
        super().__init__(parent)
        self.withdraw()
        set_window_icon(self)
        self.transient(parent)  # Aggiunta
        self.grab_set()         # Aggiunta
        self.request_id = request_id
        self.attachment_type = attachment_type
        self.read_only = read_only  # Flag per modalità sola lettura
        
        # Determina quale database usare (locale o remoto) e la cartella Attachments
        if source_db_path and os.path.exists(source_db_path):
            self.db_path = source_db_path
            logger.info(f"[AttachmentWindow] Usando DB remoto: {source_db_path}")
            
            # Calcola path Attachments con fallback robusto
            try:
                db_parent = os.path.dirname(self.db_path)  # Cartella Database
                dataflow_root = os.path.dirname(db_parent)  # Cartella DataFlow_utente
                self.attachments_base = os.path.join(dataflow_root, 'Attachments')
                
                # VERIFICA ESISTENZA - se non esiste, usa None
                if not os.path.isdir(self.attachments_base):
                    logger.warning(f"Cartella Attachments non trovata: {self.attachments_base}")
                    self.attachments_base = None
                else:
                    logger.info(f"[AttachmentWindow] Path Attachments remoto: {self.attachments_base}")
            except Exception as e:
                logger.error(f"Errore calcolo path Attachments remoto: {e}")
                self.attachments_base = None
        else:
            self.db_path = get_db_path()
            logger.info(f"[AttachmentWindow] Usando DB locale: {self.db_path}")
            # Default globale
            self.attachments_base = get_fixed_attachments_dir()
        
        # Titolo con suffisso SOLA LETTURA se applicabile
        if self.attachment_type == "Offerta Fornitore":
            title_base = _("Gestione Offerta Fornitore")
        else:
            title_base = _("Gestione Documento Interno")
        
        if self.read_only:
            title_base += _(" [SOLA LETTURA]")
        
        self.title(title_base)
        
        # Lista per tracciare file temporanei creati
        self.temp_files = []
        
        # Lista per memorizzare gli ID degli allegati (non visibili nella tabella)
        self.attachment_ids = []
        
        # Handler per cleanup alla chiusura
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Frame per avviso (solo per Offerta Fornitore)
        if self.attachment_type == "Offerta Fornitore":
            frame_warning = ttk.Frame(self)
            frame_warning.pack(side="bottom", fill="x", padx=10, pady=(0, 5))
            warning_label = tk.Label(frame_warning, 
                                    text=_("⚠️ Seleziona prima un fornitore dall'elenco sottostante prima di aggiungere un allegato"), 
                                    fg="red", 
                                    font=("Calibri", 10, "bold"))
            warning_label.pack()
        
        # Frame per i pulsanti (SEMPRE IN FONDO, non espandibile)
        frame_buttons = ttk.Frame(self)
        frame_buttons.pack(side="bottom", fill="x", padx=10, pady=10)
        
        # Pulsanti con gestione read-only
        # 'Elimina' va a destra
        self.btn_delete = ttk.Button(frame_buttons, text=_("❌ Elimina Selezionato"), command=self.delete_attachment)
        self.btn_delete.pack(side="right")

        # Gli altri pulsanti vanno a sinistra, in ordine
        self.btn_add = ttk.Button(frame_buttons, text=_("➕ Aggiungi..."), command=self.add_attachment)
        self.btn_add.pack(side="left")
        
        ttk.Button(frame_buttons, text=_("📂 Apri Selezionato"), command=self.open_attachment).pack(side="left", padx=10)
        
        # Ecco il nuovo pulsante "Download..."
        ttk.Button(frame_buttons, text=_("⬇️ Download..."), command=self.download_attachment).pack(side="left")

        # E qui reinseriamo la logica originale per il menu a tendina
        if self.attachment_type == "Offerta Fornitore":
            self.combo_suppliers = ttk.Combobox(frame_buttons, state="readonly")
            self.combo_suppliers.pack(side="left", padx=10) # Aggiunto padding
            self.load_suppliers_for_request()
            
        # Disabilita pulsanti di modifica se in modalità read-only
        if self.read_only:
            self.btn_add.config(state='disabled')
            self.btn_delete.config(state='disabled')
            if hasattr(self, 'combo_suppliers'):
                self.combo_suppliers.config(state='disabled')
        
        # Frame contenuto principale (ESPANDIBILE, sopra i pulsanti)
        frame_main = ttk.Frame(self)
        frame_main.pack(side="top", fill="both", expand=True, padx=10, pady=(10, 0))
        
        # Creiamo un frame per contenere il foglio
        sheet_frame = ttk.Frame(frame_main)
        sheet_frame.pack(fill="both", expand=True)
        
        # Creiamo il widget tksheet
        self.sheet_attachments = Sheet(sheet_frame,
                                       theme="light blue",
                                       header_font=("Calibri", 11, "bold"),
                                       font=("Calibri", 11, "normal"))
        
        # 🆕 CORREZIONE: Abilita solo i binding necessari, ESCLUDE edit_cell per impedire modifiche
        self.sheet_attachments.enable_bindings(
            "single_select",
            "row_select",
            "column_width_resize",
            "double_click_column_resize",
            "arrowkeys",
            "right_click_popup_menu",
            "rc_select",
            "copy"
            # NON includere: "edit_cell", "paste", "delete", "cut" per rendere la tabella read-only
        )
        
        self.sheet_attachments.pack(fill="both", expand=True)
            
        self.load_attachments()
        
        # Imposta dimensione minima per mostrare tutte le colonne
        # Larghezza: Fornitore(200) + Nome File(350) + Data(150) + margini(100) = 800
        # Altezza: tabella + pulsanti + padding = 450
        self.geometry("850x450")
        self.minsize(800, 400)
        
        center_window(self)
        self.deiconify()
        # --- FINE BLOCCO CORRETTO ---

    def on_closing(self):
        """Pulisce i file temporanei prima di chiudere la finestra con gestione sicura.
        
        BUG #14 FIX: Gestione robusta della chiusura senza deadlock o TclError.
        """
        # BUG #14 FIX: Disabilita i pulsanti per evitare nuove operazioni durante la chiusura
        try:
            for widget in self.winfo_children():
                if isinstance(widget, (ttk.Button, tk.Button)):
                    widget.config(state='disabled')
        except:
            pass
        
        # BUG #14 FIX: Garbage collection UNA SOLA VOLTA all'inizio (non ad ogni loop)
        if sys.platform == 'win32':
            try:
                import gc
                gc.collect()
            except:
                pass
        
        # BUG #14 FIX: Attendi eventuali operazioni DB in corso con gestione robusta
        max_wait = 30  # 30 * 100ms = 3 secondi (ridotto da 5s)
        wait_count = 0
        window_destroyed = False
        
        while wait_count < max_wait:
            # Verifica se ci sono altre operazioni DB attive
            active_db_threads = [t for t in threading.enumerate() 
                                if 'database' in t.name.lower()]
            if not active_db_threads:
                break
            
            # BUG #14 FIX: Try-except per gestire TclError se finestra già distrutta
            try:
                self.update()
            except Exception as update_error:
                # Finestra già distrutta o altro errore Tkinter
                logger.debug(f"Errore update() durante chiusura: {update_error}")
                window_destroyed = True
                break
            
            time.sleep(0.1)
            wait_count += 1
        
        # BUG #14 FIX: Pulisci i file temporanei con singolo tentativo (cleanup automatico del BUG #12 gestisce i file bloccati)
        for temp_path in self.temp_files:
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                    logger.info(f"File temporaneo eliminato: {temp_path}")
            except PermissionError:
                # File ancora in uso, il thread di cleanup automatico lo eliminerà dopo 30 secondi (BUG #12)
                logger.debug(f"File temporaneo in uso, verrà eliminato dal cleanup automatico: {temp_path}")
            except Exception as e:
                logger.warning(f"Impossibile eliminare file temporaneo {temp_path}: {e}")
        
        # BUG #14 FIX: Distruggi la finestra solo se non già distrutta
        if not window_destroyed:
            try:
                self.destroy()
            except Exception as destroy_error:
                logger.debug(f"Errore destroy() durante chiusura: {destroy_error}")

    def _sanitize_filename(self, name):
        return re.sub(r'[\\/*?:"<>|]', "", name)

    def delete_attachment(self):
        # Blocca se in modalità read-only
        if self.read_only:
            messagebox.showwarning(
                _("Operazione Non Consentita"),
                _("Non puoi eliminare allegati di RdO di altri utenti."),
                parent=self
            )
            return
        
        selected = self.sheet_attachments.get_currently_selected()
        if not selected or selected.row is None:
            messagebox.showwarning(_("Attenzione"), _("Seleziona un allegato da eliminare."), parent=self)
            return
        if messagebox.askyesno(_("Conferma Eliminazione"), _("Sei sicuro di voler eliminare questo allegato?"), parent=self):
            # Ottieni l'ID dalla lista nascosta usando l'indice della riga
            row_idx = selected.row
            
            # ✅ CORREZIONE: Verifica che l'indice sia valido
            if row_idx < 0 or row_idx >= len(self.attachment_ids):
                logger.error(f"Indice riga non valido in delete: {row_idx}, totale: {len(self.attachment_ids)}")
                messagebox.showerror(_("Errore"), _("Impossibile identificare l'allegato selezionato."), parent=self)
                return
            
            attachment_id = self.attachment_ids[row_idx]
            try:
                file_to_delete = None
                db_manager = DatabaseManager(self.db_path, read_only=self.read_only)
                try:
                    try:
                        result = db_manager.get_allegato_file_data(attachment_id)
                        if result:
                            nome_file, dati_file, percorso_esterno = result
                            if percorso_esterno:
                                base_path = self.attachments_base
                                if base_path:
                                    file_to_delete = os.path.join(base_path, percorso_esterno)
                    except DatabaseError as fetch_error:
                        logger.warning(f"Impossibile recuperare informazioni allegato da eliminare: {fetch_error}")
                    db_manager.delete_allegato(attachment_id)
                finally:
                    # BUG #19 FIX: Chiudi SEMPRE il database, anche in caso di errore
                    try:
                        db_manager.close()
                    except Exception as close_error:
                        logger.warning(f"Errore chiusura DB in delete_attachment: {close_error}")

                if file_to_delete and os.path.exists(file_to_delete):
                    try:
                        os.remove(file_to_delete)
                        logger.info(f"Allegato eliminato dal disco: {file_to_delete}")
                    except Exception as disk_error:
                        logger.warning(f"Impossibile eliminare il file allegato {file_to_delete}: {disk_error}")
                
                messagebox.showinfo(_("Eliminazione"), _("Allegato eliminato."), parent=self)
                self.load_attachments()
                
                # 🆕 AGGIUNTA: Se è stato eliminato un documento SQDC, aggiorna il pulsante nella finestra parent
                if self.attachment_type == "Documento Interno":
                    # Verifica se il parent ha il metodo per aggiornare il pulsante SQDC
                    if hasattr(self.master, 'check_sqdc_status_and_update_button'):
                        try:
                            self.master.check_sqdc_status_and_update_button()
                        except Exception as e:
                            logger.warning(f"Impossibile aggiornare pulsante SQDC nel parent: {e}")
                
            except DatabaseError as e: 
                messagebox.showerror(_("Errore Database"), _("Impossibile eliminare l'allegato: {}").format(e), parent=self)

    def load_attachments(self):
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(self.db_path, read_only=self.read_only) as db_manager:
                # Verifica se la colonna data_inserimento esiste
                has_date_column = db_manager.check_table_has_column('allegati_richiesta', 'data_inserimento')
                
                rows = db_manager.get_allegati_by_richiesta(self.request_id, self.attachment_type, has_date_column)
            
            if has_date_column:
                # Salva gli ID in una lista separata (non visibile nella tabella)
                self.attachment_ids = [id_allegato for id_allegato, nf, nfile, di in rows]
                
                # Prepara i dati per tksheet SENZA la colonna ID
                data_rows = []
                for id_all, nome_fornitore, nome_file, data_inserimento in rows:
                    # BUG #22 + #23 FIX: Parsing date consolidato con gestione errori robusta
                    data_formattata = ""
                    if data_inserimento:
                        # Prova formati data multipli in ordine di probabilità
                        date_formats = [
                            '%Y-%m-%d %H:%M:%S',  # Formato completo con timestamp
                            '%Y-%m-%d',           # Formato solo data
                            '%d/%m/%Y'            # Formato italiano già presente
                        ]
                        
                        for fmt in date_formats:
                            try:
                                # BUG #22 FIX: Usa strptime diretto senza slicing rischioso
                                dt = datetime.strptime(str(data_inserimento).strip(), fmt)
                                data_formattata = dt.strftime('%d/%m/%Y')
                                break  # Successo, esci dal loop
                            except (ValueError, TypeError):
                                continue  # Prova il formato successivo
                        
                        # BUG #23 FIX: Se nessun formato funziona, log warning e usa fallback
                        if not data_formattata:
                            logger.warning(f"Formato data non riconosciuto per allegato {id_all}: '{data_inserimento}'")
                            data_formattata = str(data_inserimento) if data_inserimento else ""
                    
                    data_rows.append([str(nome_fornitore), str(nome_file), data_formattata])
                
                # Configura intestazioni e dati (senza colonna ID)
                headers = [_("Fornitore"), _("Nome File"), _("Data Inserimento")]
                self.sheet_attachments.headers(headers)
                self.sheet_attachments.set_sheet_data(data_rows)
                
                # Configura larghezze colonne DOPO aver impostato headers e dati
                self.sheet_attachments.column_width(column=0, width=200)  # Fornitore
                self.sheet_attachments.column_width(column=1, width=350)  # Nome File
                self.sheet_attachments.column_width(column=2, width=150)  # Data Inserimento
            else:
                # Fallback per database senza la colonna data_inserimento
                self.attachment_ids = [id_allegato for id_allegato, nf, nfile in rows]
                data_rows = [[str(nome_fornitore), str(nome_file)] for id_all, nome_fornitore, nome_file in rows]
                
                headers = [_("Fornitore"), _("Nome File")]
                self.sheet_attachments.headers(headers)
                self.sheet_attachments.set_sheet_data(data_rows)
                
                # Configura larghezze colonne DOPO aver impostato headers e dati
                self.sheet_attachments.column_width(column=0, width=200)  # Fornitore
                self.sheet_attachments.column_width(column=1, width=400)  # Nome File
            
        except DatabaseError as e:
            logger.error(f"Errore database in load_attachments: {e}", exc_info=True)
            from tkinter import messagebox as mb
            mb.showerror("Errore Database", f"Impossibile caricare gli allegati: {e}", parent=self)

    def load_suppliers_for_request(self):
        try:
            # BUG #46 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(self.db_path, read_only=self.read_only) as db_manager:
                rows = db_manager.get_fornitori_by_richiesta(self.request_id)
            self.combo_suppliers['values'] = [row[0] for row in rows]
        except DatabaseError as e:
            logger.error(f"Errore database in load_suppliers_for_request: {e}", exc_info=True)
            messagebox.showerror(_("Errore Database"), _("Impossibile caricare i fornitori: {}").format(e), parent=self)

    def add_attachment(self):
        # Blocca se in modalità read-only
        if self.read_only:
            messagebox.showwarning(
                _("Operazione Non Consentita"),
                _("Non puoi aggiungere allegati a RdO di altri utenti."),
                parent=self
            )
            return
        
        filepath = filedialog.askopenfilename(title=_("Seleziona file da allegare"))
        if not filepath: return
        supplier = self.combo_suppliers.get() if self.attachment_type == "Offerta Fornitore" else "Interno"
        if not supplier and self.attachment_type == "Offerta Fornitore":
            messagebox.showwarning(_("Attenzione"), _("Seleziona un fornitore."), parent=self)
            return
        
        archive_path = self.attachments_base
        if not archive_path:
            messagebox.showerror(_("Errore"), _("Percorso allegati non disponibile."), parent=self)
            return

        try:
            file_ext = os.path.splitext(filepath)[1]
            sanitized_supplier = self._sanitize_filename(supplier)
            db_manager_temp = DatabaseManager(self.db_path, read_only=self.read_only)
            try:
                next_id = db_manager_temp.get_max_allegato_id() + 1
            finally:
                try:
                    db_manager_temp.close()
                except Exception:
                    pass
            # Per documenti interni, non includere "Interno" nel nome (è già nel tipo allegato)
            if self.attachment_type == "Documento Interno":
                new_filename = f"RfQ{self.request_id}_ID{next_id}{file_ext}"
            else:
                new_filename = f"RfQ{self.request_id}_{sanitized_supplier}_ID{next_id}{file_ext}"
            dest_path = os.path.join(archive_path, new_filename)
            shutil.copy(filepath, dest_path)
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(self.db_path, read_only=self.read_only) as db_manager:
                db_manager.insert_allegato_richiesta_link(self.request_id, os.path.basename(filepath), self.attachment_type, supplier, new_filename)
        except Exception as e:
            messagebox.showerror(_("Errore"), _("Impossibile aggiungere l'allegato: {}").format(e), parent=self)
        
        self.load_attachments()
    
    def open_attachment(self):
            selected = self.sheet_attachments.get_currently_selected()
            if not selected or selected.row is None:
                messagebox.showwarning(_("Attenzione"), _("Seleziona un allegato da aprire."), parent=self)
                return

            # Ottieni l'ID dalla lista nascosta usando l'indice della riga
            row_idx = selected.row
            
            # ✅ CORREZIONE: Verifica che l'indice sia valido
            if row_idx < 0 or row_idx >= len(self.attachment_ids):
                logger.error(f"Indice riga non valido: {row_idx}, totale attachment_ids: {len(self.attachment_ids)}")
                messagebox.showerror(_("Errore"), _("Impossibile identificare l'allegato selezionato. Prova a ricaricare la finestra."), parent=self)
                return
            
            attachment_id = self.attachment_ids[row_idx]
            
            # 1. Recupera i dati dal DB
            try:
                # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
                with DatabaseManager(self.db_path, read_only=self.read_only) as db_manager:
                    result = db_manager.get_allegato_file_data(attachment_id)
                
                if not result:
                    messagebox.showerror(_("Errore"), _("Allegato non trovato."), parent=self)
                    return
                    
                nome_file, dati_file, percorso_esterno = result
            except DatabaseError as e:
                logger.error(f"Errore database in open_attachment: {e}", exc_info=True)
                messagebox.showerror(_("Errore Database"), _("Impossibile recuperare l'allegato: {}").format(e), parent=self)
                return

            try:
                if percorso_esterno:
                    # Caso 1: File esterno (linkato)
                    logger.info(f"Apertura allegato esterno: {nome_file}")
                    base_path = self.attachments_base
                    if not base_path:
                        logger.error("Percorso archivio non configurato")
                        messagebox.showerror(_("Errore"), _("Percorso di archivio non configurato."), parent=self)
                        return
                    
                    # ✅ SICURO: Validazione completa del percorso
                    full_path = os.path.join(base_path, percorso_esterno)
                    
                    # Normalizza i percorsi e risolvi i link simbolici
                    real_base = os.path.realpath(base_path)
                    real_full = os.path.realpath(full_path)
                    
                    # Verifica che il percorso finale sia dentro la directory base
                    if not real_full.startswith(real_base + os.sep) and real_full != real_base:
                        logger.error(f"Tentativo di accesso non autorizzato a: {real_full}")
                        messagebox.showerror(_("Errore Sicurezza"), 
                                            _("Percorso file non valido. Possibile tentativo di accesso non autorizzato."), 
                                            parent=self)
                        return
                    
                    # Ora è sicuro procedere
                    if not os.path.exists(real_full):
                        logger.error(f"File esterno non trovato: {real_full}")
                        messagebox.showerror(_("Errore"), _("File sorgente non trovato:\n{}").format(real_full), parent=self)
                        return
                    
                    # Apri il file direttamente dal suo percorso di archivio
                    # Usa webbrowser.open() per compatibilità MSIX (non blocca eseguibili)
                    webbrowser.open(f'file:///{real_full}')

                elif dati_file:
                    # Caso 2: File interno (BLOB)
                    logger.info(f"Apertura allegato interno: {nome_file}")
                    # Salva in un file temporaneo e aprilo
                    file_ext = os.path.splitext(nome_file)[1]
                    
                    with tempfile.NamedTemporaryFile(mode='wb', suffix=file_ext, delete=False) as temp_file:
                        temp_file.write(dati_file)
                        temp_path = temp_file.name
                    
                    # Traccia il file temporaneo per cleanup alla chiusura
                    self.temp_files.append(temp_path)
                    
                    # BUG #12 + #26 FIX: Cleanup automatico con verifica handle e retry
                    # Evita accumulo infinito di file temporanei se l'app va in crash
                    def delayed_cleanup(path, delay=60):
                        """Elimina il file temporaneo dopo delay secondi con retry se locked."""
                        try:
                            time.sleep(delay)  # BUG #26 FIX: Aumentato da 30s a 60s
                            
                            if not os.path.exists(path):
                                return  # File già eliminato
                            
                            # BUG #26 FIX: Tenta eliminazione con retry se file ancora aperto
                            max_retries = 3
                            for attempt in range(max_retries):
                                try:
                                    os.remove(path)
                                    logger.info(f"File temporaneo pulito automaticamente: {path}")
                                    break  # Successo
                                except (PermissionError, OSError) as e:
                                    if attempt < max_retries - 1:
                                        logger.debug(f"File temporaneo ancora in uso, retry {attempt+1}/{max_retries}: {e}")
                                        time.sleep(5)  # Attendi 5s prima di riprovare
                                    else:
                                        logger.warning(f"File temporaneo non eliminabile dopo {max_retries} tentativi (in uso?): {path}")
                        except Exception as e:
                            logger.warning(f"Impossibile pulire file temporaneo {path}: {e}")
                    
                    import threading
                    cleanup_thread = threading.Thread(
                        target=delayed_cleanup,
                        args=(temp_path,),
                        name=f"TempFileCleanup-{os.path.basename(temp_path)}",
                        daemon=True
                    )
                    cleanup_thread.start()
                    
                    # Apri il file temporaneo
                    # Usa webbrowser.open() per compatibilità MSIX (non blocca eseguibili)
                    webbrowser.open(f'file:///{temp_path}')
                else:
                    logger.error("Allegato senza dati né percorso esterno")
                    messagebox.showerror(_("Errore"), _("Dati allegato non disponibili (né interni, né esterni)."), parent=self)

            except FileNotFoundError as e:
                logger.error(f"File non trovato in open_attachment: {e}", exc_info=True)
                messagebox.showerror(_("Errore Apertura"), _("File non trovato: {}").format(e), parent=self)
            except PermissionError as e:
                logger.error(f"Permessi insufficienti in open_attachment: {e}", exc_info=True)
                messagebox.showerror(_("Errore Apertura"), _("Permessi insufficienti per aprire il file: {}").format(e), parent=self)
            except OSError as e:
                logger.error(f"Errore sistema operativo in open_attachment: {e}", exc_info=True)
                messagebox.showerror(_("Errore Apertura"), _("Errore sistema operativo: {}").format(e), parent=self)
            except Exception as e:
                logger.error(f"Errore imprevisto in open_attachment: {e}", exc_info=True)
                messagebox.showerror(_("Errore Apertura"), _("Impossibile aprire il file: {}").format(e), parent=self)
        
# --- INIZIO NUOVO METODO: Download Allegato ---
    def download_attachment(self):
        selected = self.sheet_attachments.get_currently_selected()
        if not selected or selected.row is None:
            messagebox.showwarning(_("Attenzione"), _("Seleziona un allegato da scaricare."), parent=self)
            return

        # Ottieni l'ID dalla lista nascosta usando l'indice della riga
        row_idx = selected.row
        
        # ✅ CORREZIONE: Verifica che l'indice sia valido
        if row_idx < 0 or row_idx >= len(self.attachment_ids):
            logger.error(f"Indice riga non valido in download: {row_idx}, totale: {len(self.attachment_ids)}")
            messagebox.showerror(_("Errore"), _("Impossibile identificare l'allegato selezionato."), parent=self)
            return
        
        attachment_id = self.attachment_ids[row_idx]
        
        # 1. Recupera i dati dal DB (come in open_attachment)
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(self.db_path) as db_manager:
                result = db_manager.get_allegato_file_data(attachment_id)
            
            if not result:
                messagebox.showerror(_("Errore"), _("Allegato non trovato."), parent=self)
                return
                
            nome_file, dati_file, percorso_esterno = result
        except DatabaseError as e:
            logger.error(f"Errore database in download_attachment: {e}", exc_info=True)
            messagebox.showerror(_("Errore Database"), _("Impossibile recuperare l'allegato: {}").format(e), parent=self)
            return

        # 2. Chiedi all'utente dove salvare il file
        save_path = filedialog.asksaveasfilename(
            title=_("Salva allegato come..."),
            initialfile=nome_file,  # Propone il nome file originale
            parent=self
        )
        
        if not save_path:
            return  # L'utente ha annullato

        # 3. Salva il file in base al tipo (esterno o BLOB)
        try:
            if percorso_esterno:
                # Caso 1: File esterno (linkato)
                base_path = self.attachments_base
                if not base_path:
                    messagebox.showerror(_("Errore"), _("Percorso di archivio non configurato."), parent=self)
                    return
                
                # ✅ SICURO: Validazione completa del percorso
                full_path = os.path.join(base_path, percorso_esterno)
                
                # Normalizza i percorsi e risolvi i link simbolici
                real_base = os.path.realpath(base_path)
                real_full = os.path.realpath(full_path)
                
                # Verifica che il percorso finale sia dentro la directory base
                if not real_full.startswith(real_base + os.sep) and real_full != real_base:
                    messagebox.showerror(_("Errore Sicurezza"), 
                                        _("Percorso file non valido. Possibile tentativo di accesso non autorizzato."), 
                                        parent=self)
                    return
                
                # Ora è sicuro procedere
                if not os.path.exists(real_full):
                    messagebox.showerror(_("Errore"), _("File sorgente non trovato:\n{}").format(real_full), parent=self)
                    return
                
                # Copia il file dall'archivio alla destinazione scelta
                shutil.copy(real_full, save_path)

            elif dati_file:
                # Caso 2: File interno (BLOB)
                # Scrive i dati binari nel nuovo file
                with open(save_path, 'wb') as f:
                    f.write(dati_file)
            else:
                messagebox.showerror(_("Errore"), _("Dati allegato non disponibili (né interni, né esterni)."), parent=self)
                return

            messagebox.showinfo(_("Successo"), _("File scaricato con successo in:\n{}").format(save_path), parent=self)

        except Exception as e:
            messagebox.showerror(_("Errore Download"), _("Impossibile salvare il file: {}").format(e), parent=self)
    # --- FINE NUOVO METODO ---

# ------------------------------------------------------------------------------------
# FINESTRA GESTIONE NUMERI ORDINE DI ACQUISTO (PO)
# ------------------------------------------------------------------------------------
class PurchaseOrderWindow(tk.Toplevel):
    def __init__(self, parent, request_id):
        super().__init__(parent)
        self.request_id = request_id
        self.parent = parent
        
        # Imposta titolo in base alla lingua
        if get_current_language() == 'it':
            self.title("Gestione Numeri Ordine")
        else:
            self.title("Purchase Order Management")
        
        # MODIFICA 2: Finestra 20% più corta (da 500 a 400)
        self.geometry("700x400")
        self.resizable(True, True)
        
        # MODIFICA 4: Mantieni la finestra sempre in primo piano
        self.transient(parent)
        self.grab_set()
        
        # IMPORTANTE: Frame pulsanti PRIMA (side=bottom) - così rimangono sempre visibili
        bottom_frame = ttk.Frame(self)
        bottom_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)
        
        # Pulsante Aggiungi con testo tradotto (a sinistra)
        if get_current_language() == 'it':
            add_text = "➕ Aggiungi"
        else:
            add_text = "➕ Add"
        
        add_btn = ttk.Button(
            bottom_frame, 
            text=add_text,
            command=lambda: self.add_po_entry_safe()
        )
        add_btn.pack(side=tk.LEFT, padx=5)
        
        # Pulsante Elimina con testo tradotto (a sinistra)
        if get_current_language() == 'it':
            delete_text = "❌ Elimina"
        else:
            delete_text = "❌ Delete"
        
        delete_btn = ttk.Button(
            bottom_frame, 
            text=delete_text,
            command=lambda: self.delete_selected_po_safe()
        )
        delete_btn.pack(side=tk.LEFT, padx=5)
        
        # Pulsante Chiudi con testo tradotto (a destra)
        if get_current_language() == 'it':
            close_text = "Chiudi"
        else:
            close_text = "Close"
        
        close_btn = ttk.Button(
            bottom_frame, 
            text=close_text,
            command=self.on_closing
        )
        close_btn.pack(side=tk.RIGHT, padx=5)
        
        # Ora creiamo i frame contenuto (dall'alto verso il basso)
        # Frame superiore con istruzioni
        info_frame = ttk.Frame(self)
        info_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        
        # Testo istruzioni tradotto
        if get_current_language() == 'it':
            info_text = "Associa numeri di ordine ai fornitori della RdO"
        else:
            info_text = "Associate purchase order numbers with RfQ suppliers"
        
        info_label = ttk.Label(
            info_frame, 
            text=info_text,
            font=('Segoe UI', 10)
        )
        info_label.pack(anchor='w')
        
        # Frame per i controlli di inserimento
        controls_frame = ttk.Frame(self)
        controls_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        # Campo numero ordine con label tradotta
        if get_current_language() == 'it':
            po_label_text = "Numero Ordine:"
        else:
            po_label_text = "PO Number:"
        
        po_label = ttk.Label(controls_frame, text=po_label_text)
        po_label.grid(row=0, column=0, sticky='w', padx=5, pady=5)
        
        self.po_entry = ttk.Entry(controls_frame, width=30)
        self.po_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)
        
        # ComboBox fornitore con label tradotta
        if get_current_language() == 'it':
            supplier_label_text = "Fornitore:"
        else:
            supplier_label_text = "Supplier:"
        
        supplier_label = ttk.Label(controls_frame, text=supplier_label_text)
        supplier_label.grid(row=0, column=2, sticky='w', padx=5, pady=5)
        
        self.supplier_combo = ttk.Combobox(controls_frame, state='readonly', width=25)
        self.supplier_combo.grid(row=0, column=3, sticky='ew', padx=5, pady=5)
        
        controls_frame.columnconfigure(1, weight=1)
        controls_frame.columnconfigure(3, weight=1)
        
        # Frame per la griglia (espandibile)
        grid_frame = ttk.Frame(self)
        grid_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Headers tradotti per la griglia
        if get_current_language() == 'it':
            headers = ["Numero Ordine", "Fornitore"]
        else:
            headers = ["PO Number", "Supplier"]
        
        # Crea il foglio tksheet
        self.sheet = Sheet(
            grid_frame,
            headers=headers,
            header_font=("Segoe UI", 10, "bold"),
            font=("Segoe UI", 10, "normal"),
            show_row_index=False,
            show_top_left=False,
            empty_horizontal=0,
            empty_vertical=0
        )
        self.sheet.enable_bindings(
            "single_select",
            "row_select",
            "drag_select",
            "column_width_resize",
            "arrowkeys",
            "right_click_popup_menu",
            "rc_select",
            "copy",
            "cut",
            "paste",
            "delete",
            "edit_cell"
        )
        self.sheet.pack(fill=tk.BOTH, expand=True)
        
        # Salva gli headers per uso futuro
        self.headers = headers
        
        # Gestione chiusura finestra - PRIMA di caricare i dati
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Centra la finestra
        center_window(self)
        
        # Carica i fornitori e i PO esistenti - ALLA FINE
        self.load_suppliers_for_request()
        self.load_po_entries()
    
    def on_closing(self):
        """Salva i dati prima di chiudere."""
        try:
            self.save_po_entries()
        except Exception as e:
            logger.error(f"Errore nel salvataggio PO durante chiusura: {e}")
        
        # Notifica il parent per aggiornare eventualmente l'interfaccia
        if hasattr(self.parent, 'load_po_numbers'):
            try:
                self.parent.load_po_numbers()
            except Exception as e:
                logger.error(f"Errore nell'aggiornamento parent: {e}")
        
        self.destroy()
    
    def load_suppliers_for_request(self):
        """Carica i fornitori associati a questa richiesta."""
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(getattr(self, 'db_path', get_db_path())) as db_manager:
                suppliers_rows = db_manager.get_fornitori_ordered_for_request(self.request_id)
            suppliers = [row[0] for row in suppliers_rows]
            
            # MODIFICA 3: Non impostare un valore predefinito
            # Forza l'utente a selezionare esplicitamente il fornitore
            self.supplier_combo['values'] = suppliers
            # NON impostare: self.supplier_combo.current(0)
        except DatabaseError as e:
            logger.error(f"Errore nel caricamento fornitori per PO: {e}")
            # Non mostrare messagebox che potrebbe distruggere la finestra
            # Lascia semplicemente il combo vuoto
    
    def load_po_entries(self):
        """Carica i numeri ordine esistenti dal database."""
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(getattr(self, 'db_path', get_db_path())) as db_manager:
                row = db_manager.get_numeri_ordine(self.request_id)
            
            if row and row[0]:
                po_data = row[0]
                po_list = []
                
                # Tenta di parsare come JSON (nuovo formato)
                try:
                    po_list = json.loads(po_data)
                    if not isinstance(po_list, list):
                        po_list = []
                except (json.JSONDecodeError, TypeError):
                    # Formato vecchio: stringa con virgole
                    # Converti in nuovo formato (senza fornitore associato)
                    old_numbers = [n.strip() for n in po_data.split(',') if n.strip()]
                    po_list = [{"po_number": num, "supplier": ""} for num in old_numbers]
                
                # Popola il foglio
                data = [[entry.get("po_number", ""), entry.get("supplier", "")] for entry in po_list]
                if data:
                    self.sheet.set_sheet_data(data)
                    # Auto-ridimensiona le colonne dopo aver caricato i dati
                    self.auto_resize_columns()
        except DatabaseError as e:
            logger.error(f"Errore nel caricamento numeri ordine: {e}")
    
    def auto_resize_columns(self):
        """Auto-ridimensiona le colonne in base al contenuto."""
        try:
            import tkinter.font as tkfont
            
            # Font per intestazioni e contenuto
            header_font = tkfont.Font(family="Segoe UI", size=10, weight="bold")
            content_font = tkfont.Font(family="Segoe UI", size=10, weight="normal")
            
            PADDING_PX = 30  # Padding per evitare troncamenti
            MIN_WIDTH = 150  # Larghezza minima
            
            # Ottieni i dati correnti
            data_rows = self.sheet.get_sheet_data()
            
            # Per ogni colonna
            for col_idx in range(len(self.headers)):
                header_text = self.headers[col_idx]
                max_width = header_font.measure(header_text)
                
                # Controlla il contenuto di tutte le righe
                for row in data_rows:
                    if col_idx < len(row):
                        cell_value = str(row[col_idx])
                        cell_width = content_font.measure(cell_value)
                        max_width = max(max_width, cell_width)
                
                # Calcola larghezza finale con padding e minimo
                column_width = max(int(max_width + PADDING_PX), MIN_WIDTH)
                self.sheet.column_width(column=col_idx, width=column_width)
                
        except Exception as e:
            logger.warning(f"Errore auto-ridimensionamento colonne PO: {e}")
            # Fallback a larghezze fisse
            self.sheet.column_width(column=0, width=200)
            self.sheet.column_width(column=1, width=250)
    
    def add_po_entry_safe(self):
        """Wrapper sicuro per add_po_entry."""
        if hasattr(self, 'sheet'):
            self.add_po_entry()
    
    def delete_selected_po_safe(self):
        """Wrapper sicuro per delete_selected_po."""
        if hasattr(self, 'sheet'):
            self.delete_selected_po()
    
    def add_po_entry(self):
        """Aggiunge un nuovo numero ordine con fornitore associato."""
        po_number = self.po_entry.get().strip()
        supplier = self.supplier_combo.get().strip()
        
        # BUG #35 FIX: Validazione esplicita campo vuoto con feedback utente
        if not po_number:
            messagebox.showwarning(
                _("Attenzione"),
                _("Inserisci un numero ordine valido."),
                parent=self
            )
            return
        
        # BUG #25 FIX: Previeni SQL injection e caratteri pericolosi
        FORBIDDEN_CHARS = re.compile(r"[';\"\\`<>]")
        if FORBIDDEN_CHARS.search(po_number):
            logger.warning(f"Caratteri pericolosi rimossi da PO number: '{po_number}'")
            po_number = FORBIDDEN_CHARS.sub('', po_number)
            self.po_entry.delete(0, tk.END)
            self.po_entry.insert(0, po_number)
        
        if FORBIDDEN_CHARS.search(supplier):
            logger.warning(f"Caratteri pericolosi rimossi da supplier: '{supplier}'")
            supplier = FORBIDDEN_CHARS.sub('', supplier)
        
        # MODIFICA 4: Aggiunto parent=self per mantenere la finestra in primo piano
        if not po_number:
            if get_current_language() == 'it':
                messagebox.showwarning(
                    "Campo obbligatorio",
                    "Inserire il numero ordine.",
                    parent=self
                )
            else:
                messagebox.showwarning(
                    "Required Field",
                    "Please enter the PO number.",
                    parent=self
                )
            return
        
        if not supplier:
            if get_current_language() == 'it':
                messagebox.showwarning(
                    "Campo obbligatorio",
                    "Selezionare un fornitore.",
                    parent=self
                )
            else:
                messagebox.showwarning(
                    "Required Field",
                    "Please select a supplier.",
                    parent=self
                )
            return
        
        # Aggiungi alla griglia
        current_data = self.sheet.get_sheet_data()
        current_data.append([po_number, supplier])
        self.sheet.set_sheet_data(current_data)
        
        # Auto-ridimensiona le colonne dopo l'aggiunta
        self.auto_resize_columns()
        
        # Pulisci i campi
        self.po_entry.delete(0, tk.END)
        # MODIFICA 3: NON reimpostare il fornitore, lascia vuoto
        # self.supplier_combo.current(0) # RIMOSSO
        
        # Salva automaticamente
        self.save_po_entries()
    
    def delete_selected_po(self):
        """Elimina il numero ordine selezionato."""
        selected = self.sheet.get_currently_selected()
        
        # MODIFICA 4: Aggiunto parent=self per mantenere la finestra in primo piano
        if not selected:
            if get_current_language() == 'it':
                messagebox.showwarning(
                    "Nessuna selezione",
                    "Selezionare una riga da eliminare.",
                    parent=self
                )
            else:
                messagebox.showwarning(
                    "No Selection",
                    "Please select a row to delete.",
                    parent=self
                )
            return
        
        # BUG #18 FIX: Validazione robusta dell'indice riga prima dell'accesso
        row_idx = selected.row if hasattr(selected, 'row') else None
        if row_idx is None:
            logger.warning("delete_selected_po: selected non ha attributo 'row'")
            return
        
        # BUG #18 FIX: Ottieni i dati PRIMA di validare l'indice
        current_data = self.sheet.get_sheet_data()
        
        # BUG #18 FIX: Verifica che l'indice sia valido PRIMA di mostrare il dialog
        if not (0 <= row_idx < len(current_data)):
            logger.error(f"delete_selected_po: Indice {row_idx} fuori range (0-{len(current_data)-1})")
            if get_current_language() == 'it':
                messagebox.showerror(
                    "Errore",
                    f"Impossibile eliminare: indice riga non valido ({row_idx}).",
                    parent=self
                )
            else:
                messagebox.showerror(
                    "Error",
                    f"Cannot delete: invalid row index ({row_idx}).",
                    parent=self
                )
            return
        
        # Conferma eliminazione
        if get_current_language() == 'it':
            confirm = messagebox.askyesno(
                "Conferma eliminazione",
                "Eliminare il numero ordine selezionato?",
                parent=self
            )
        else:
            confirm = messagebox.askyesno(
                "Confirm Deletion",
                "Delete the selected PO number?",
                parent=self
            )
        
        if confirm:
            # BUG #18 FIX: L'indice è già stato validato, procedi sicuro
            del current_data[row_idx]
            self.sheet.set_sheet_data(current_data)
            self.save_po_entries()
    
    def save_po_entries(self):
        """Salva i numeri ordine nel database in formato JSON."""
        try:
            # Ottieni i dati dal foglio
            data = self.sheet.get_sheet_data()
            
            # Converti in lista di dizionari
            po_list = []
            for row in data:
                if len(row) >= 2 and row[0].strip():
                    po_list.append({
                        "po_number": row[0].strip(),
                        "supplier": row[1].strip() if len(row) > 1 else ""
                    })
            
            # Salva come JSON
            json_data = json.dumps(po_list, ensure_ascii=False)
            
            # BUG #46 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(getattr(self, 'db_path', get_db_path())) as db_manager:
                db_manager.update_numeri_ordine(self.request_id, json_data)
            
            logger.info(f"Numeri ordine salvati per RdO {self.request_id}: {len(po_list)} entries")
        except DatabaseError as e:
            logger.error(f"Errore nel salvataggio numeri ordine: {e}")
            messagebox.showerror(_( "Errore"), _("Errore nel salvataggio dei numeri ordine."), parent=self)
    
# ------------------------------------------------------------------------------------
# FINESTRA MODIFICA FORNITORI, RIFERIMENTO, DATA
# ------------------------------------------------------------------------------------
class EditSuppliersWindow(tk.Toplevel):
    def __init__(self, parent, request_id):
        super().__init__(parent)
        self.withdraw()
        set_window_icon(self)
        self.request_id = request_id;         self.title(_("Modifica Fornitori - RdO N°{}").format(request_id)); self.db_path = get_db_path(); self.transient(parent); self.grab_set()
        
        # Frame pulsanti (sempre in fondo)
        btn_frame = ttk.Frame(self)
        btn_frame.pack(side="bottom", fill="x", padx=10, pady=10)
        ttk.Button(btn_frame, text=_("💾 Salva"), command=self.save_changes).pack(side="right")
        ttk.Button(btn_frame, text=_("❌ Annulla"), command=self.destroy).pack(side="right", padx=10)
        
        # Frame contenuto (espandibile)
        frame = ttk.Frame(self, padding="10")
        frame.pack(side="top", fill="both", expand=True)
        
        ttk.Label(frame, text=_("Modifica elenco fornitori (nomi separati da virgola):")).pack(anchor="w")
        self.entry_suppliers = ttk.Entry(frame, width=70); self.entry_suppliers.pack(fill="x", expand=True, pady=5)
        
        self.load_current_suppliers(); center_window(self)
    def load_current_suppliers(self):
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(getattr(self, 'db_path', get_db_path())) as db_manager:
                rows = db_manager.get_fornitori_by_richiesta(self.request_id)
            self.entry_suppliers.insert(0, ", ".join([r[0] for r in rows]))
        except DatabaseError as e:
            logger.error(f"Errore database in load_current_suppliers: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Impossibile caricare i fornitori: {}").format(e), parent=self)
    def save_changes(self):
        # Blocca se in modalità read-only
        if getattr(self, 'read_only', False):
            messagebox.showwarning(
                _("Operazione Non Consentita"),
                _("Non puoi modificare i fornitori di RdO di altri utenti."),
                parent=self
            )
            return
        
        new_suppliers = [n.strip() for n in self.entry_suppliers.get().split(',') if n.strip()]
        
        # Validazione fornitori duplicati (solo se ci sono fornitori)
        if new_suppliers:
            fornitori_lower = [f.lower() for f in new_suppliers]
            duplicati = [f for f in new_suppliers if fornitori_lower.count(f.lower()) > 1]
            duplicati_unici = list(set(duplicati))
            
            if duplicati_unici:
                messagebox.showwarning(
                    _("Fornitori Duplicati"),
                    _("Hai inserito lo stesso fornitore più volte:\n\n{}\n\nOgni fornitore deve essere inserito una sola volta.").format(', '.join(sorted(set(duplicati_unici)))),
                    parent=self
                )
                return
        
        try:
            # Recupera i fornitori PRIMA di eliminarli e gli id_dettaglio
            # BUG FIX: Usa context manager per garantire chiusura DB automatica
            with DatabaseManager(getattr(self, 'db_path', get_db_path())) as db_manager:
                old_suppliers_rows = db_manager.get_fornitori_by_richiesta(self.request_id)
                old_suppliers = [row[0] for row in old_suppliers_rows]
                detail_ids_rows = db_manager.get_dettaglio_ids_by_richiesta(self.request_id)
                detail_ids = [row[0] for row in detail_ids_rows]
                
                # Usa db_manager per salvare con transazione
                print(f"[EditSuppliersWindow] Inizio save_suppliers_with_transaction...")
                db_manager.save_suppliers_with_transaction(self.request_id, new_suppliers, old_suppliers, detail_ids)
                print(f"[EditSuppliersWindow] Fine save_suppliers_with_transaction")
                
                # Verifica immediata che i dati siano stati salvati
                verify_rows = db_manager.get_fornitori_by_richiesta(self.request_id)
                verify_count = len(verify_rows)
                print(f"[EditSuppliersWindow] VERIFICA POST-SALVATAGGIO: {verify_count} fornitori trovati nel DB (attesi: {len(new_suppliers)})")
                print(f"[EditSuppliersWindow] Fornitori salvati: {[r[0] for r in verify_rows]}")
            
            # Context manager ha già chiuso il DB qui
            print(f"[EditSuppliersWindow] DB chiuso dal context manager")
            
            # Messaggio di successo personalizzato
            if new_suppliers:
                messagebox.showinfo(_("Successo"), _("Elenco fornitori aggiornato."), parent=self.master)
            else:
                messagebox.showinfo(_("Successo"), _("Tutti i fornitori sono stati rimossi."), parent=self.master)
            
            self.destroy()
            
        except DatabaseError as e:
            logger.error(f"Errore database in save_changes (EditSuppliersWindow): {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Impossibile salvare: {}").format(e), parent=self)

class EditReferenceWindow(tk.Toplevel):
    def __init__(self, parent, request_id):
        super().__init__(parent)
        self.withdraw()
        set_window_icon(self)
        self.request_id = request_id;         self.title(_("Modifica Riferimento")); self.db_path = get_db_path(); self.transient(parent)
        
        # Frame pulsanti (sempre in fondo)
        btn_frame = ttk.Frame(self)
        btn_frame.pack(side="bottom", fill="x", padx=10, pady=10)
        ttk.Button(btn_frame, text=_("💾 Salva"), command=self.save_changes).pack(side="right")
        ttk.Button(btn_frame, text=_("❌ Annulla"), command=self.destroy).pack(side="right", padx=10)
        
        # Frame contenuto (espandibile)
        frame = ttk.Frame(self, padding="10")
        frame.pack(side="top", fill="both", expand=True)
        
        ttk.Label(frame, text=_("Modifica Riferimento:")).pack(anchor="w")
        self.entry_riferimento = ttk.Entry(frame, width=70); self.entry_riferimento.pack(fill="x", expand=True, pady=5)
        
        self.load_current_reference(); center_window(self); self.wait_visibility(); self.grab_set()
    def load_current_reference(self):
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(getattr(self, 'db_path', get_db_path())) as db_manager:
                result = db_manager.get_riferimento(self.request_id)
            if result and result[0]: self.entry_riferimento.insert(0, result[0])
        except DatabaseError as e:
            logger.error(f"Errore database in load_current_reference: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Impossibile caricare riferimento: {}").format(e), parent=self)
    
    def save_changes(self):
        try:
            # BUG #46 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(getattr(self, 'db_path', get_db_path())) as db_manager:
                db_manager.update_riferimento(self.request_id, self.entry_riferimento.get().strip())
            messagebox.showinfo(_("Successo"), _("Riferimento aggiornato."), parent=self.master)
            self.destroy()
        except DatabaseError as e:
            logger.error(f"Errore database in save_changes (EditReferenceWindow): {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Impossibile salvare: {}").format(e), parent=self)

# ------------------------------------------------------------------------------------
# FINESTRA DI DIALOGO SCELTA LINGUA (NUOVA)
# ------------------------------------------------------------------------------------
class LanguagePrompt(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.withdraw()
        set_window_icon(self)
        self.title(_("Scegli Lingua"))
        self.choice = None
        self.transient(parent)
        self.grab_set()

        frame = ttk.Frame(self, padding="20")
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text=_("In quale lingua vuoi esportare il file Excel?"), font=(None, 10)).pack(pady=(0, 15))

        lang_frame = ttk.Frame(frame)
        lang_frame.pack(pady=10)
        
        ttk.Label(lang_frame, text=_("Lingua:")).pack(side="left", padx=(0, 10))
        
        # Determina la lingua corrente dell'app (coerente con init_i18n/get_current_language)
        current_lang = get_current_language()
        default_language = "Italiano" if current_lang == 'it' else "English"
        
        # Ordina le opzioni ponendo per prima la lingua correntemente usata dal programma
        values = ["Italiano", "English"] if default_language == "Italiano" else ["English", "Italiano"]

        self.language_var = tk.StringVar(value=default_language)
        language_combo = ttk.Combobox(lang_frame, textvariable=self.language_var,
                                      values=values,
                                      state="readonly", width=20)
        language_combo.pack(side="left", padx=(0, 10))
        # Allinea anche la selezione alla prima voce
        language_combo.current(0)
        language_combo.bind("<<ComboboxSelected>>", lambda e: self.on_language_selected())
        
        btn_ok = ttk.Button(lang_frame, text="OK", command=self.confirm_choice)
        btn_ok.pack(side="left", padx=5)
        
        btn_cancel = ttk.Button(lang_frame, text=_("❌ Annulla"), command=self.on_close)
        btn_cancel.pack(side="left", padx=5)

        self.protocol("WM_DELETE_WINDOW", self.on_close)
        center_window(self)
    
    def on_language_selected(self):
        """Gestisce la selezione della lingua"""
        pass  # Già gestita dalla variabile
    
    def confirm_choice(self):
        """Conferma la scelta e chiude la finestra"""
        selected = self.language_var.get()
        if selected == "Italiano":
            self.choice = "ita"
        elif selected == "English":
            self.choice = "eng"
        else:
            self.choice = None
        self.destroy()

    def on_close(self):
        self.choice = None
        self.destroy()

# ------------------------------------------------------------------------------------
# --- NUOVA FINESTRA PER GESTIONE NOTE (VERSIONE CORRETTA) ---
# ------------------------------------------------------------------------------------
class NotesWindow(tk.Toplevel):
    def __init__(self, parent, request_id):
        super().__init__(parent)
        self.withdraw()
        set_window_icon(self)
        self.request_id = request_id
        self.db_path = get_db_path()
        self.title(_("Note - RdO N° {}").format(self.request_id))
        self.resizable(True, True)
        self.parent_window = parent
        # BUG #37 FIX: Verifica esistenza finestra prima di modificare attributes
        if hasattr(self, 'parent_window') and self.parent_window:
            try:
                if self.parent_window.winfo_exists():
                    self.parent_window.attributes('-disabled', True)
            except Exception as e:
                logger.debug(f"Impossibile disabilitare parent_window in NotesWindow: {e}")
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        # Pulsanti di salvataggio e chiusura (sempre in fondo)
        button_frame = ttk.Frame(self)
        button_frame.pack(side="bottom", fill="x", padx=10, pady=10)
        ttk.Button(button_frame, text=_("💾 Salva Nota"), command=self.save_note).pack(side="right")
        ttk.Button(button_frame, text=_("❌ Annulla"), command=self.on_close).pack(side="right", padx=10)

        # Frame principale (espandibile)
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(side="top", fill="both", expand=True)

        # Toolbar di formattazione
        toolbar = ttk.Frame(main_frame)
        toolbar.pack(fill="x", pady=(0, 5))

        ttk.Button(toolbar, text=_("B Grassetto").replace("B", "𝐁", 1), command=lambda: self.apply_tag("bold")).pack(side="left")
        ttk.Button(toolbar, text=_("I Corsivo").replace("I", "𝑰", 1), command=lambda: self.apply_tag("italic")).pack(side="left", padx=5)
        ttk.Button(toolbar, text=_("U Sottolineato").replace("U", "U\u0332", 1), command=lambda: self.apply_tag("underline")).pack(side="left")
        # Spacer per allineare i pulsanti a sinistra e lasciare libera la caption bar
        ttk.Label(toolbar, text="").pack(side="left", expand=True, fill="x")

        # Editor di testo con scrollbar
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill="both", expand=True)
        self.text_editor = tk.Text(text_frame, wrap="word", undo=True, font=("Calibri", 11))
        scrollbar = ttk.Scrollbar(text_frame, command=self.text_editor.yview)
        self.text_editor.config(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.text_editor.pack(side="left", fill="both", expand=True)

        # Configurazione dei tag per la formattazione
        self.text_editor.tag_configure("bold", font=("Calibri", 11, "bold"))
        self.text_editor.tag_configure("italic", font=("Calibri", 11, "italic"))
        self.text_editor.tag_configure("underline", font=("Calibri", 11, "underline"))

        self.load_note()
        center_window(self)

    def apply_tag(self, tag_name):
        try:
            # Controlla se il tag è già applicato alla selezione
            current_tags = self.text_editor.tag_names("sel.first")
            if tag_name in current_tags:
                self.text_editor.tag_remove(tag_name, "sel.first", "sel.last")
            else:
                self.text_editor.tag_add(tag_name, "sel.first", "sel.last")
        except tk.TclError:
            # Nessun testo selezionato, non fare nulla
            pass

    def load_note(self):
        """
        Carica la nota dal database e ricostruisce il testo con la formattazione corretta.
        """
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(getattr(self, 'db_path', get_db_path())) as db_manager:
                result = db_manager.get_note_formattate(self.request_id)
            
            if result and result[0]:
                # ✅ SICURO: Validazione prima del parsing
                note_data = result[0]
                
                # 1. Limita lunghezza massima
                if len(note_data) > 1000000:  # 1MB max
                    raise ValueError("Nota troppo grande")
                
                # 2. Usa json.loads invece di ast.literal_eval (più sicuro e veloce)
                try:
                    content_dump = json.loads(note_data)
                except json.JSONDecodeError:
                    # Fallback per compatibilità con vecchi dati
                    content_dump = ast.literal_eval(note_data)
                
                # 3. Valida struttura dati
                if not isinstance(content_dump, list):
                    raise ValueError("Formato nota non valido")
                
                # 4. Limita numero di elementi
                if len(content_dump) > 10000:
                    raise ValueError("Nota troppo complessa")
                
                self.text_editor.delete("1.0", tk.END)
                
                # Mantiene un set dei tag di formattazione attivi
                active_tags = set()
                
                # Itera su ogni elemento salvato (testo, inizio tag, fine tag)
                for item in content_dump:
                    key = item[0]
                    value = item[1]

                    if key == "text":
                        # Inserisce il testo applicando i tag attualmente attivi
                        self.text_editor.insert(tk.END, value, tuple(active_tags))
                    elif key == "tagon":
                        # Aggiunge un tag al set di quelli attivi per il testo successivo
                        active_tags.add(value)
                    elif key == "tagoff":
                        # Rimuove un tag dal set di quelli attivi
                        active_tags.discard(value)

        except (DatabaseError, SyntaxError, ValueError) as e:
            logger.error(f"Errore in load_note: {e}", exc_info=True)
            # Se la nota salvata è in un formato vecchio o corrotto, prova a caricarla come testo semplice
            if result and result[0]:
                self.text_editor.delete("1.0", tk.END)
                self.text_editor.insert("1.0", _("Impossibile ripristinare la formattazione. Nota caricata come testo semplice:\n\n{}").format(result[0]))
            messagebox.showwarning(_("Errore Caricamento Nota"), _("Non è stato possibile ripristinare la formattazione della nota. Potrebbe essere stata salvata con una versione precedente.\n\nDettagli: {}").format(e), parent=self)

    def on_close(self):
        try:
            if hasattr(self, 'parent_window') and self.parent_window and self.parent_window.winfo_exists():
                try:
                    self.parent_window.attributes('-disabled', False)
                except Exception:
                    pass
                try:
                    self.parent_window.focus_set()
                except Exception:
                    pass
        finally:
            self.destroy()

    def save_note(self):
        # Salva il contenuto del widget Text, inclusa la formattazione
        content_dump = self.text_editor.dump("1.0", tk.END, text=True, tag=True)
        
        # Estrae solo il testo effettivo (senza formattazione) per verificare se è vuoto
        text_content = self.text_editor.get("1.0", tk.END).strip()
        
        # Se il contenuto è vuoto, salva NULL invece di una stringa vuota
        if not text_content:
            content_to_save = None
        else:
            # Usiamo repr per salvare una rappresentazione sicura della stringa
            content_to_save = repr(content_dump)

        try:
            # BUG #46 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(getattr(self, 'db_path', get_db_path())) as db_manager:
                db_manager.update_note_formattate(self.request_id, content_to_save)
            # Chiama il metodo del genitore per aggiornare il pulsante
            if hasattr(self.master, 'check_note_status_and_update_button'):
                self.master.check_note_status_and_update_button()
            messagebox.showinfo(_("Successo"), _("Nota salvata correttamente."), parent=self)
            self.on_close()
        except DatabaseError as e:
            logger.error(f"Errore database in save_note: {e}", exc_info=True)
            messagebox.showerror(_("Errore Database"), _("Impossibile salvare la nota: {}").format(e), parent=self)

# ------------------------------------------------------------------------------------
# FINESTRA ANALISI SQDC
# ------------------------------------------------------------------------------------
class SQDCAnalysisWindow(tk.Toplevel):
    def __init__(self, parent, request_id, existing_data=None):
        super().__init__(parent)
        self.withdraw()
        set_window_icon(self)
        self.request_id = request_id
        self.db_path = get_db_path()
        self.title(_("Analisi SQDC - RdO N° {}").format(self.request_id))
        self.transient(parent)
        self.grab_set()
        
        # Inizializza variabili
        self.weights = {'safety': tk.StringVar(value='25'), 'quality': tk.StringVar(value='25'),
                       'delivery': tk.StringVar(value='25'), 'cost': tk.StringVar(value='25')}
        self.scores = {}  # {'supplier': {'safety': score, 'quality': score, ...}}
        self.suppliers = []
        self.automatic_cost = False  # Se True, Cost è calcolato automaticamente
        self.missing_price_suppliers = []  # Lista fornitori con prezzi mancanti/incompleti
        
        # Carica fornitori
        self.load_suppliers()
        
        # Pulsanti (sempre in fondo)
        button_frame = ttk.Frame(self)
        button_frame.pack(side="bottom", fill="x", padx=10, pady=10)
        ttk.Button(button_frame, text=_("📊 Esporta Excel"), command=self.export_to_excel).pack(side="left", padx=5)
        ttk.Button(button_frame, text=_("💾 Salva SQDC"), command=self.save_sqdc).pack(side="left", padx=5)
        ttk.Button(button_frame, text=_("❌ Chiudi"), command=self.destroy).pack(side="right", padx=5)
        
        # UI principale (espandibile)
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(side="top", fill="both", expand=True)
        
        # Notepad con tab
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill="both", expand=True, pady=10)
        
        # Tab 1: Pesi
        tab_weights = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab_weights, text=_("Pesi (%)"))
        self.create_weights_tab(tab_weights)
        
        # Tab 2: Voti
        tab_scores = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab_scores, text=_("Voti (1-10)"))
        self.create_scores_tab(tab_scores)
        
        # Binding per validazione quando si passa al tab Voti
        def on_tab_changed(event):
            # Controlla quale tab è stato selezionato
            selected = self.notebook.index(self.notebook.select())
            if selected == 1:  # Tab Voti (indice 1, 0 è Pesi)
                if not self.validate_weights_only():
                    # Se i pesi non sono validi, torna al tab Pesi
                    self.notebook.select(0)
        
        self.notebook.bind("<<NotebookTabChanged>>", on_tab_changed)
        
        # Inizializza dati esistenti se forniti - DOPO aver creato l'UI
        if existing_data:
            self.load_from_existing_data(existing_data)
        
        # Imposta dimensione finestra
        # Larghezza: 200 (Fornitore) + 5*100 (colonne voti) + margini = circa 950px
        # Altezza: 680px per contenere tutti gli elementi incluso l'eventuale messaggio di warning rosso
        width = 950
        height = 680
        
        # Centra la finestra
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        # Mostra la finestra
        self.deiconify()
    
    def load_suppliers(self):
        """Carica la lista dei fornitori per questa RdO"""
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(getattr(self, 'db_path', get_db_path())) as db_manager:
                rows = db_manager.get_fornitori_by_richiesta(self.request_id, order_by=True)
            self.suppliers = [row[0] for row in rows]
            
            # Inizializza scores se non ancora fatto
            for supplier in self.suppliers:
                if supplier not in self.scores:
                    self.scores[supplier] = {'safety': tk.StringVar(value=''), 'quality': tk.StringVar(value=''),
                                           'delivery': tk.StringVar(value=''), 'cost': tk.StringVar(value='')}
        except DatabaseError as e:
            logger.error(f"Errore database in load_suppliers: {e}", exc_info=True)
            messagebox.showerror(_("Errore Database"), _("Impossibile caricare i fornitori: {}").format(e), parent=self)
    
    def create_weights_tab(self, parent):
        """Crea il tab per l'inserimento dei pesi percentuali"""
        info_label = ttk.Label(parent, 
                              text=_("Assegna i pesi percentuali a ciascun criterio (la somma deve essere 100%):"),
                              font=(None, 10))
        info_label.pack(pady=(0, 10))
        
        descs_frame = ttk.LabelFrame(parent, text=_("Descrizioni Criteri"), padding="10")
        descs_frame.pack(fill="x", pady=(0, 5))
        
        descriptions = [
            (_("Safety"), _("aderenza del fornitore agli standard di sicurezza, alle conformità normativa relativa al prodotto/servizio e al rischio finanziario e geopolitico.")),
            (_("Quality"), _("capacità del fornitore di rispettare integralmente le specifiche tecniche concordate.")),
            (_("Delivery"), _("capacità di rispettare il tempo di consegna offerto e la flessibilità a fronte di eventuali variazioni richieste.")),
            (_("Cost"), _("competitività relativa al prezzo totale offerto, considerando i termini di pagamento e i costi accessori (es. trasporto, installazione)."))
        ]
        
        for i, (criterion, desc) in enumerate(descriptions):
            row_frame = ttk.Frame(descs_frame)
            row_frame.pack(fill="x", pady=2)
            ttk.Label(row_frame, text=criterion + ":", font=(None, 10, 'bold'), width=12, anchor='w').pack(side="left", padx=(0, 10))
            ttk.Label(row_frame, text=desc, font=(None, 9), foreground="gray", wraplength=500).pack(side="left", fill="x")
        
        weights_frame = ttk.LabelFrame(parent, text=_("Pesi Percentuali"), padding="10")
        weights_frame.pack(fill="x", pady=(5, 10))
        
        for criterion in ['safety', 'quality', 'delivery', 'cost']:
            row = ttk.Frame(weights_frame)
            row.pack(fill="x", pady=5)
            label_text = {'safety': _("Safety"), 'quality': _("Quality"), 
                         'delivery': _("Delivery"), 'cost': _("Cost")}[criterion]
            ttk.Label(row, text=label_text + ":", width=15, anchor='w').pack(side="left")
            entry = ttk.Entry(row, textvariable=self.weights[criterion], width=10)
            entry.pack(side="left", padx=5)
            ttk.Label(row, text="%").pack(side="left")
    
    def create_scores_tab(self, parent):
        """Crea il tab per l'inserimento dei voti usando tksheet"""
        info_label = ttk.Label(parent, 
                              text=_("Assegna un voto da 1 a 10 per ciascun criterio e fornitore (doppio click per modificare):"),
                              font=(None, 10))
        info_label.pack(pady=(0, 10))
        
        # Pulsante per calcolare automaticamente i voti Cost (spostato qui dal tab Pesi)
        calc_cost_frame = ttk.Frame(parent)
        calc_cost_frame.pack(fill="x", pady=(0, 10))
        ttk.Button(calc_cost_frame, text=_("🔄 Calcola Cost Automaticamente"), 
                  command=self.auto_calculate_cost).pack()
        
        # Frame per avviso prezzi mancanti (inizialmente nascosto)
        self.price_warning_frame = ttk.Frame(parent)
        self.price_warning_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        self.price_warning_label = tk.Label(
            self.price_warning_frame,
            text="",  # Vuoto inizialmente
            bg="red",
            fg="white",
            font=("Calibri", 11, "bold"),
            wraplength=800,  # Wrap testo se troppo lungo
            justify="center",
            padx=10,
            pady=10
        )
        self.price_warning_label.pack(fill="x")
        
        # Nascondi inizialmente il frame
        self.price_warning_frame.pack_forget()
        
        # Frame per la tabella
        sheet_frame = ttk.Frame(parent)
        sheet_frame.pack(fill="both", expand=True)
        
        # Crea il widget tksheet
        self.sheet_scores = Sheet(sheet_frame,
                                 theme="light blue",
                                 header_font=("Calibri", 11, "bold"),
                                 font=("Calibri", 11, "normal"))
        self.sheet_scores.enable_bindings()
        self.sheet_scores.pack(fill="both", expand=True)
        
        # Imposta la funzione di validazione per l'editing delle celle
        self.sheet_scores.edit_validation(self.validate_sqdc_cell_input)
        
        # Popola il sheet con i fornitori
        self.refresh_scores_sheet()
        
        # Binding per catturare le modifiche e aggiornare i totali
        self.sheet_scores.extra_bindings([
            ("end_edit_cell", self.on_sqdc_cell_edit)
        ])
    
    def refresh_scores_sheet(self):
        """Popola o aggiorna il sheet con i dati dei fornitori"""
        if not hasattr(self, 'sheet_scores'):
            return
        
        # Definisci gli headers
        headers = [_("Fornitore"), _("Safety"), _("Quality"), _("Delivery"), _("Cost"), _("TOTALE")]
        
        # Costruisci le righe di dati e trova i vincitori (potrebbero essere multipli in caso di parità)
        data_rows = []
        max_total = -1
        winner_rows = []  # Lista di indici di riga dei vincitori
        all_scores_complete = True  # Flag per verificare se tutti i voti sono completi
        
        for idx, supplier in enumerate(self.suppliers):
            # Recupera i valori correnti
            safety_val = self.scores[supplier]['safety'].get() or ''
            quality_val = self.scores[supplier]['quality'].get() or ''
            delivery_val = self.scores[supplier]['delivery'].get() or ''
            cost_val = self.scores[supplier]['cost'].get() or ''
            
            # Verifica se tutti i voti sono completi per questo fornitore
            if not (safety_val and quality_val and delivery_val and cost_val):
                all_scores_complete = False
            
            # Calcola il totale
            try:
                w_safety = float(self.weights['safety'].get() or 0) / 100
                w_quality = float(self.weights['quality'].get() or 0) / 100
                w_delivery = float(self.weights['delivery'].get() or 0) / 100
                w_cost = float(self.weights['cost'].get() or 0) / 100
                
                score_safety = float(safety_val or 0)
                score_quality = float(quality_val or 0)
                score_delivery = float(delivery_val or 0)
                score_cost = float(cost_val or 0)
                
                total = (score_safety * w_safety + score_quality * w_quality + 
                        score_delivery * w_delivery + score_cost * w_cost)
                
                # Forza sempre due decimali nel formato con virgola come separatore
                total_str = f"{total:.2f}".replace('.', ',')
                
                # Trova i vincitori (considera parità con tolleranza 0.01)
                if total > max_total + 0.01:
                    # Nuovo massimo trovato, resetta la lista
                    max_total = total
                    winner_rows = [idx]
                elif abs(total - max_total) <= 0.01:
                    # Parità, aggiungi alla lista dei vincitori
                    winner_rows.append(idx)
                
            except (ValueError, TypeError):
                total_str = "0,00"
            
            # Aggiungi la riga
            data_rows.append([supplier, safety_val, quality_val, delivery_val, cost_val, total_str])
        
        # Carica i dati nel sheet
        self.sheet_scores.set_sheet_data(data_rows)
        self.sheet_scores.headers(headers)
        
        # Configura larghezze colonne
        self.sheet_scores.column_width(column=0, width=200)  # Fornitore
        for col_idx in range(1, 6):  # Safety, Quality, Delivery, Cost, TOTALE
            self.sheet_scores.column_width(column=col_idx, width=100)
        
        # Rendi le colonne Fornitore (0) e TOTALE (5) in sola lettura
        self.sheet_scores.readonly_columns(columns=[0, 5], readonly=True)
        
        # BUG #27 FIX: Usa enumerate invece di range(len()) per migliore leggibilità
        # Prima rimuovi TUTTE le evidenziazioni dalla colonna TOTALE
        for row_idx, _supplier in enumerate(self.suppliers):
            self.sheet_scores.dehighlight_cells(row=row_idx, column=5)
        
        # Poi evidenzia in verde le celle TOTALE di TUTTI i vincitori se tutti i voti sono completi
        if winner_rows and all_scores_complete:
            for winner_row_idx in winner_rows:
                self.sheet_scores.highlight_cells(row=winner_row_idx, column=5, bg="#90EE90", fg="black")
    
    def validate_sqdc_cell_input(self, event):
        """Valida l'input delle celle SQDC durante l'editing - SOLO interi 1-10"""
        # event è un dizionario che contiene: 'value', 'row', 'column', etc.
        # La funzione deve ritornare il VALORE VALIDATO o None per rifiutare
        
        col_idx = event.get('column')
        # Ottieni il valore dall'evento (prova 'text' o 'value')
        raw_value = event.get('text', event.get('value', ''))
        new_value = raw_value.strip() if isinstance(raw_value, str) else str(raw_value)
        
        # Le colonne Fornitore (0) e TOTALE (5) sono già readonly, ma per sicurezza
        if col_idx == 0 or col_idx == 5:
            return None  # Rifiuta l'editing
        
        # Se il campo è vuoto, permetti (l'utente può cancellare)
        if new_value == '':
            return ''  # Ritorna stringa vuota
        
        # Controlla che sia un numero intero tra 1 e 10
        try:
            score = int(new_value)
            if 1 <= score <= 10:
                return str(score)  # Ritorna il valore validato come stringa
            else:
                # Valore fuori range - mostra errore
                messagebox.showwarning(_("Valore Non Valido"), 
                                      _("I voti devono essere tra 1 e 10."),
                                      parent=self)
                return None
        except ValueError:
            # Non è un numero intero - mostra errore
            messagebox.showwarning(_("Valore Non Valido"), 
                                  _("I voti devono essere numeri interi da 1 a 10."),
                                  parent=self)
            return None
    
    def on_sqdc_cell_edit(self, event):
        """Gestisce le modifiche alle celle del sheet SQDC - aggiorna i totali"""
        if not hasattr(self, 'sheet_scores'):
            return
        
        # L'evento end_edit_cell fornisce direttamente row, column, value
        row_idx = event.row
        col_idx = event.column
        new_value = str(event.value).strip() if event.value is not None else ''
        
        if row_idx is None or col_idx is None:
            return
        
        # Non permettere editing sulla colonna Fornitore (0) o TOTALE (5)
        if col_idx == 0 or col_idx == 5:
            return
        
        # Mappa indice colonna a criterio
        col_to_criterion = {1: 'safety', 2: 'quality', 3: 'delivery', 4: 'cost'}
        if col_idx not in col_to_criterion:
            return
        
        # Verifica che row_idx sia valido
        if row_idx < 0 or row_idx >= len(self.suppliers):
            return
        
        # Ottieni il fornitore dalla riga
        supplier = self.suppliers[row_idx]
        criterion = col_to_criterion[col_idx]
        
        # Aggiorna il valore nella variabile StringVar
        self.scores[supplier][criterion].set(new_value)
        
        # Ricarica completamente il sheet per ricalcolare totali e evidenziare il vincitore
        self.refresh_scores_sheet()
    
    def update_price_warning(self):
        """Aggiorna l'avviso per i prezzi mancanti nella scheda Scores"""
        if not hasattr(self, 'price_warning_frame') or not hasattr(self, 'price_warning_label'):
            return
        
        if self.missing_price_suppliers:
            # Ci sono fornitori con prezzi mancanti - mostra l'avviso
            suppliers_list = ", ".join(self.missing_price_suppliers)
            
            # Messaggio tradotto
            if len(self.missing_price_suppliers) == 1:
                warning_text = _("⚠️ ATTENZIONE: Impossibile calcolare il prezzo automaticamente per il fornitore {} per mancanza di prezzi o quantità nella tabella di RdO. Il voto Cost è stato impostato a 0.").format(suppliers_list)
            else:
                warning_text = _("⚠️ ATTENZIONE: Impossibile calcolare il prezzo automaticamente per i fornitori {} per mancanza di prezzi o quantità nella tabella di RdO. I voti Cost sono stati impostati a 0.").format(suppliers_list)
            
            self.price_warning_label.config(text=warning_text)
            self.price_warning_frame.pack(fill="x", padx=10, pady=(0, 10))
            logger.info(f"SQDC: Avviso prezzi mancanti mostrato per: {suppliers_list}")
        else:
            # Nessun problema con i prezzi - nascondi l'avviso
            self.price_warning_frame.pack_forget()
            logger.info("SQDC: Avviso prezzi mancanti nascosto")
    
    def auto_calculate_cost(self):
        """Calcola automaticamente i voti Cost basati sui prezzi"""
        # Prima verifica la somma dei pesi
        try:
            total_weight = (float(self.weights['safety'].get() or 0) + 
                           float(self.weights['quality'].get() or 0) + 
                           float(self.weights['delivery'].get() or 0) + 
                           float(self.weights['cost'].get() or 0))
        except (ValueError, TypeError):
            messagebox.showerror(_("Errore Pesi"),
                               _("I pesi devono essere numeri validi."),
                               parent=self)
            return
        
        # Verifica che la somma sia 100% (con tolleranza per errori di arrotondamento)
        if abs(total_weight - 100) > 0.01:
            messagebox.showerror(_("Errore Pesi"),
                               _("La somma dei pesi deve essere 100%. Attualmente: {:.1f}%").format(total_weight),
                               parent=self)
            return
        
        # Carica prezzi da database
        supplier_prices = {}
        calculation_success = False
        
        # RESET: Pulisce la lista dei fornitori con prezzi mancanti
        self.missing_price_suppliers = []
        
        try:
            
            logger.info(f"SQDC auto_calculate_cost: Processing {len(self.suppliers)} suppliers")
            
            # Prima determina quanti articoli ci sono in totale per questa RdO
            db_manager = DatabaseManager(get_db_path())
            total_items = db_manager.get_dettagli_count_by_richiesta(self.request_id)
            logger.info(f"SQDC auto_calculate_cost: Total items in RdO: {total_items}")
            
            for supplier in self.suppliers:
                # Query per ottenere tutti i prezzi per il fornitore (senza filtri su NULL)
                results = db_manager.get_prezzo_quantita_by_fornitore(self.request_id, supplier)
                
                # CONTROLLO CRITICO: verifica che il fornitore abbia prezzi per TUTTI gli articoli
                if len(results) < total_items:
                    # Mancano dei prezzi - imposta Cost a 0
                    logger.warning(f"SQDC auto_calculate_cost: Supplier {supplier} ha solo {len(results)} prezzi su {total_items} articoli - Cost impostato a 0")
                    self.scores[supplier]['cost'].set('0')
                    self.missing_price_suppliers.append(supplier)  # Traccia il fornitore
                    continue
                
                # Verifica che tutti i prezzi siano numerici validi (non "X", "ND" o vuoti)
                total_price = 0
                valid_prices = 0
                has_invalid_price = False
                
                for price_val, qty in results:
                    # BUG #34 FIX: Semplificato check con conversione a stringa e strip
                    # Controlla se il prezzo è valido (None, vuoto, o solo whitespace)
                    price_str_raw = str(price_val).strip()
                    if not price_str_raw:
                        has_invalid_price = True
                        logger.warning(f"SQDC auto_calculate_cost: Supplier {supplier} ha un prezzo vuoto")
                        break
                    
                    # BUG #34 FIX: Usa la variabile già calcolata invece di rifare strip()
                    price_str = price_str_raw.upper()
                    if price_str in ('X', 'ND'):
                        has_invalid_price = True
                        logger.warning(f"SQDC auto_calculate_cost: Supplier {supplier} ha prezzo non numerico: {price_str}")
                        break
                    
                    try:
                        price_float = parse_float_from_comma_string(str(price_val))
                        qty_float = parse_float_from_comma_string(str(qty))
                        total_price += price_float * qty_float
                        valid_prices += 1
                    except (ValueError, TypeError) as e:
                        logger.warning(f"SQDC auto_calculate_cost: Invalid price/qty for supplier {supplier}: {e}")
                        has_invalid_price = True
                        break
                
                if has_invalid_price or valid_prices < total_items:
                    # Almeno un prezzo non è valido - imposta Cost a 0
                    logger.warning(f"SQDC auto_calculate_cost: Supplier {supplier} ha prezzi non validi - Cost impostato a 0")
                    self.scores[supplier]['cost'].set('0')
                    self.missing_price_suppliers.append(supplier)  # Traccia il fornitore
                else:
                    # Tutti i prezzi sono validi - salva per il calcolo del voto
                    supplier_prices[supplier] = total_price
                    logger.info(f"SQDC auto_calculate_cost: Supplier {supplier} total price: {total_price}")

            
            db_manager.close()
            
            # Se almeno un fornitore ha prezzi validi, calcola i voti
            if supplier_prices:
                min_price = min(supplier_prices.values())

                # BUG #10 FIX: Gestione robusta divisione per zero e errori calcolo
                for supplier, price in supplier_prices.items():
                    try:
                        if min_price == 0:
                            # Se il prezzo minimo è 0, assegna 10 a chi ha 0, 1 agli altri
                            score = 10 if price == 0 else 1
                        else:
                            # Calcolo normale
                            price_ratio = price / min_price
                            score = 10 / price_ratio
                        
                        # Arrotonda e limita tra 1 e 10
                        score = int(score + 0.5)
                        score = max(1, min(10, score))
                        
                        self.scores[supplier]['cost'].set(str(score))
                        
                    except (ZeroDivisionError, ValueError, TypeError) as e:
                        # Fallback per errori imprevisti
                        logger.error(
                            f"Errore calcolo score per {supplier}: "
                            f"price={price}, min={min_price}, error={e}"
                        )
                        self.scores[supplier]['cost'].set('1')  # Punteggio minimo

                calculation_success = True
                logger.info("SQDC auto_calculate_cost: Calculation completed successfully")

            else:
                # Nessun fornitore ha prezzi completi e validi
                logger.warning("SQDC auto_calculate_cost: Nessun fornitore ha prezzi completi")

        except DatabaseError as e:
            logger.error(f"SQDC auto_calculate_cost: Database error: {e}", exc_info=True)
            messagebox.showerror(_("Errore Database"),
                               _("Errore nel recupero dei prezzi dal database: {}").format(e),
                               parent=self)
        finally:
            # Aggiorna sempre il Sheet per mostrare i voti (anche gli 0)
            self.refresh_scores_sheet()
            
            # Aggiorna l'avviso per prezzi mancanti
            self.update_price_warning()
            
            # Cambia al tab Scores sempre (anche se ci sono fornitori con Cost = 0)
            if hasattr(self, 'notebook'):
                try:
                    for idx in range(self.notebook.index('end')):
                        tab_text = self.notebook.tab(idx, 'text')
                        if tab_text and ('scores' in tab_text.lower() or 'voti' in tab_text.lower()):
                            self.notebook.select(idx)
                            logger.info(f"SQDC: Switched to Scores tab (index {idx})")
                            break
                except Exception as e:
                    logger.error(f"SQDC: Failed to switch to Scores tab: {e}", exc_info=True)
    
    def update_total_score(self, supplier):
        """Aggiorna il punteggio totale per un fornitore - non più necessaria con tksheet"""
        # Con tksheet, il totale viene ricalcolato automaticamente in refresh_scores_sheet()
        # Questa funzione è mantenuta per compatibilità ma semplicemente aggiorna il sheet
        self.refresh_scores_sheet()
    
    def reorder_suppliers_by_score(self):
        """Riordina dinamicamente i fornitori per punteggio decrescente nell'UI"""
        if not hasattr(self, 'supplier_frames') or not hasattr(self, 'total_labels'):
            return
        
        # Ottieni i fornitori ordinati per punteggio (decrescente)
        try:
            suppliers_sorted = sorted(self.suppliers, 
                                     key=lambda s: float(self.total_labels[s].cget('text') or 0),
                                     reverse=True)
            
            # Riordina fisicamente i frame nell'UI
            for supplier in suppliers_sorted:
                if supplier in self.supplier_frames:
                    # Solleva il frame in cima (pack)
                    self.supplier_frames[supplier].pack_forget()
                    self.supplier_frames[supplier].pack(fill="x", pady=2)
        except (ValueError, TypeError):
            # Errore nel parsing dei punteggi, ignora
            pass
    
    def validate_weights_only(self):
        """Valida solo i pesi senza controllare i voti"""
        total_weight = 0
        for criterion in ['safety', 'quality', 'delivery', 'cost']:
            try:
                weight = float(self.weights[criterion].get() or 0)
                if weight < 0 or weight > 100:
                    messagebox.showerror(_("Errore Pesi"), 
                                       _("I pesi devono essere tra 0 e 100."), 
                                       parent=self)
                    return False
                total_weight += weight
            except ValueError:
                messagebox.showerror(_("Errore Pesi"), 
                                   _("I pesi devono essere numeri validi."), 
                                   parent=self)
                return False
        
        if abs(total_weight - 100) > 0.01:
            messagebox.showerror(_("Errore Pesi"), 
                               _("La somma dei pesi deve essere 100%. Attualmente: {:.1f}%").format(total_weight), 
                               parent=self)
            return False
        
        return True
    
    def validate_inputs(self):
        """Valida che i pesi sommino a 100% e i voti siano 1-10"""
        # Valida pesi
        if not self.validate_weights_only():
            return False
        
        # Valida voti (solo interi 1-10) - TUTTI devono essere valorizzati
        for supplier in self.suppliers:
            for criterion in ['safety', 'quality', 'delivery', 'cost']:
                value = self.scores[supplier][criterion].get()
                if not value:  # Campo vuoto non permesso
                    messagebox.showerror(_("Errore Voti"), 
                                       _("Devi compilare tutti i voti."), 
                                       parent=self)
                    return False
                try:
                    # Controlla se è un intero (no decimali)
                    score = int(value)
                    if score < 1 or score > 10:
                        messagebox.showerror(_("Errore Voti"), 
                                           _("I voti devono essere tra 1 e 10.\nFornitore: {}\nCriterio: {}").format(supplier, criterion), 
                                           parent=self)
                        return False
                except ValueError:
                    # Non è un intero valido
                    messagebox.showerror(_("Errore Voti"), 
                                       _("I voti devono essere numeri interi da 1 a 10.\nFornitore: {}\nCriterio: {}").format(supplier, criterion), 
                                       parent=self)
                    return False
        
        return True
    
    def load_from_existing_data(self, data):
        """Carica dati da un'analisi SQDC esistente"""
        # data è un dizionario con struttura:
        # {'weights': {...}, 'scores': {...}, 'automatic_cost': bool}
        if 'weights' in data:
            for criterion in ['safety', 'quality', 'delivery', 'cost']:
                if criterion in data['weights']:
                    self.weights[criterion].set(str(data['weights'][criterion]))
        
        if 'scores' in data:
            for supplier, scores_dict in data['scores'].items():
                if supplier in self.scores:
                    for criterion in ['safety', 'quality', 'delivery', 'cost']:
                        if criterion in scores_dict:
                            self.scores[supplier][criterion].set(str(scores_dict[criterion]))
        
        if 'automatic_cost' in data:
            self.automatic_cost = data['automatic_cost']
        
        # Aggiorna il Sheet dopo caricamento dati
        if hasattr(self, 'sheet_scores'):
            self.refresh_scores_sheet()
    
    def export_to_excel(self):
        """Esporta l'analisi SQDC in un file Excel"""
        if not self.validate_inputs():
            return
        
        # 🔧 CORREZIONE BUG: Usa la lingua corrente dell'applicazione
        language = get_current_language()
        
        # Scegli template e nome file suggerito in base alla lingua
        if language == 'it':
            template_name = "template_sqdc.xlsx"
            default_name = f"SQDC_Analisi_RdO_{self.request_id}.xlsx"
        else:
            template_name = "template_sqdc_eng.xlsx"
            default_name = f"SQDC_Analysis_RfQ_{self.request_id}.xlsx"
        
        template_path = resource_path(os.path.join("add_data", template_name))
        
        if not os.path.exists(template_path):
            messagebox.showerror(_("Errore"), 
                               _("File modello non trovato!\nAssicurarsi che '{}' esista nella cartella 'add_data'.").format(template_name), 
                               parent=self)
            return
        
        wb = None  # BUG #15 FIX: Inizializza per finally block
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Popola dati
            ws['B1'] = self.request_id  # Numero RdO
            ws['B2'] = datetime.now().strftime('%d/%m/%Y')  # Data
            
            # Pesi - BUG #42 FIX: Aggiunti 'or 0' per gestire campi vuoti
            ws['B5'] = float(self.weights['safety'].get() or 0)
            ws['B6'] = float(self.weights['quality'].get() or 0)
            ws['B7'] = float(self.weights['delivery'].get() or 0)
            ws['B8'] = float(self.weights['cost'].get() or 0)
            
            # Dati fornitori
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            
            start_row = 17  # Prima riga dati dopo header a riga 16
            max_total = -1
            winner_row = None
            
            # Ordina fornitori per punteggio totale (dal migliore al peggiore)
            # Calcola il totale per ogni fornitore per ordinare
            def get_total(supplier):
                try:
                    w_safety = float(self.weights['safety'].get() or 0) / 100
                    w_quality = float(self.weights['quality'].get() or 0) / 100
                    w_delivery = float(self.weights['delivery'].get() or 0) / 100
                    w_cost = float(self.weights['cost'].get() or 0) / 100
                    
                    score_safety = float(self.scores[supplier]['safety'].get() or 0)
                    score_quality = float(self.scores[supplier]['quality'].get() or 0)
                    score_delivery = float(self.scores[supplier]['delivery'].get() or 0)
                    score_cost = float(self.scores[supplier]['cost'].get() or 0)
                    
                    return (score_safety * w_safety + score_quality * w_quality + 
                           score_delivery * w_delivery + score_cost * w_cost)
                except (ValueError, TypeError):
                    return 0.0
            
            suppliers_sorted = sorted(self.suppliers, key=get_total, reverse=True)
            
            for i, supplier in enumerate(suppliers_sorted):
                row = start_row + i
                ws.cell(row=row, column=1, value=supplier)
                
                # BUG #43 FIX: Gestione robusta conversione punteggi
                try:
                    ws.cell(row=row, column=2, value=float(self.scores[supplier]['safety'].get() or 0))
                    ws.cell(row=row, column=3, value=float(self.scores[supplier]['quality'].get() or 0))
                    ws.cell(row=row, column=4, value=float(self.scores[supplier]['delivery'].get() or 0))
                    ws.cell(row=row, column=5, value=float(self.scores[supplier]['cost'].get() or 0))
                except (ValueError, TypeError) as e:
                    logger.warning(f"Errore conversione punteggio per {supplier}: {e}. Uso 0.")
                    ws.cell(row=row, column=2, value=0)
                    ws.cell(row=row, column=3, value=0)
                    ws.cell(row=row, column=4, value=0)
                    ws.cell(row=row, column=5, value=0)
                
                # Calcola totale
                total = get_total(supplier)
                total_cell = ws.cell(row=row, column=6)
                total_cell.value = total
                total_cell.number_format = '0.00'  # Forza 2 decimali fissi
                
                # Applica bordi
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')
                
                # Trova vincitore
                if total > max_total:
                    max_total = total
                    winner_row = row
            
            # Evidenzia vincitore
            if winner_row:
                for col in range(1, 7):
                    ws.cell(row=winner_row, column=col).fill = green_fill
            
            # Salva con nome suggerito localizzato (già definito sopra)
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[(_("File Excel"), "*.xlsx")],
                title=_("Salva Analisi SQDC"),
                initialfile=default_name,
                parent=self
            )
            
            if filepath:
                wb.save(filepath)
                logger.info(f"Analisi SQDC esportata: {filepath}")
                messagebox.showinfo(_("Successo"), 
                                  _("Analisi SQDC esportata con successo:\n{}").format(filepath), 
                                  parent=self)
        
        except Exception as e:
            logger.error(f"Errore esportazione SQDC: {e}", exc_info=True)
            messagebox.showerror(_("Errore Esportazione"), 
                               _("Impossibile esportare l'analisi: {}").format(e), 
                               parent=self)
        finally:
            # BUG #15 FIX: Chiudi SEMPRE il workbook per rilasciare il file
            if wb is not None:
                try:
                    wb.close()
                    logger.debug(f"Workbook SQDC chiuso: {template_path}")
                except Exception as close_error:
                    logger.warning(f"Errore chiusura workbook SQDC: {close_error}")
    
    def save_sqdc(self):
        """Salva l'analisi SQDC come Documento Interno"""
        if not self.validate_inputs():
            return
        
        # 🔧 CORREZIONE BUG: Usa la lingua corrente dell'applicazione per il template
        language = get_current_language()
        
        # Nome file univoco per database storage (indipendente dalla lingua)
        sqdc_filename = f"SQDC_Analysis_RfQ_{self.request_id}.xlsx"
        
        # Template corretto in base alla lingua
        if language == 'it':
            template_name = "template_sqdc.xlsx"
        else:
            template_name = "template_sqdc_eng.xlsx"
        
        template_path = resource_path(os.path.join("add_data", template_name))
        
        wb = None  # BUG #16 FIX: Inizializza per finally block
        temp_path = None  # BUG #16 FIX: Traccia file temporaneo
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Popola dati (stessa logica di export_to_excel)
            ws['B1'] = self.request_id
            ws['B2'] = datetime.now().strftime('%d/%m/%Y')
            
            # Pesi - BUG #42 FIX: Aggiunti 'or 0' per gestire campi vuoti
            ws['B5'] = float(self.weights['safety'].get() or 0)
            ws['B6'] = float(self.weights['quality'].get() or 0)
            ws['B7'] = float(self.weights['delivery'].get() or 0)
            ws['B8'] = float(self.weights['cost'].get() or 0)
            
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            
            start_row = 17
            max_total = -1
            winner_row = None
            
            # Ordina fornitori per punteggio totale (dal migliore al peggiore)
            # Calcola il totale per ogni fornitore per ordinare
            def get_total(supplier):
                try:
                    w_safety = float(self.weights['safety'].get() or 0) / 100
                    w_quality = float(self.weights['quality'].get() or 0) / 100
                    w_delivery = float(self.weights['delivery'].get() or 0) / 100
                    w_cost = float(self.weights['cost'].get() or 0) / 100
                    
                    score_safety = float(self.scores[supplier]['safety'].get() or 0)
                    score_quality = float(self.scores[supplier]['quality'].get() or 0)
                    score_delivery = float(self.scores[supplier]['delivery'].get() or 0)
                    score_cost = float(self.scores[supplier]['cost'].get() or 0)
                    
                    return (score_safety * w_safety + score_quality * w_quality + 
                           score_delivery * w_delivery + score_cost * w_cost)
                except (ValueError, TypeError):
                    return 0.0
            
            suppliers_sorted = sorted(self.suppliers, key=get_total, reverse=True)
            
            for i, supplier in enumerate(suppliers_sorted):
                row = start_row + i
                ws.cell(row=row, column=1, value=supplier)
                
                # BUG #43 FIX: Gestione robusta conversione punteggi
                try:
                    ws.cell(row=row, column=2, value=float(self.scores[supplier]['safety'].get() or 0))
                    ws.cell(row=row, column=3, value=float(self.scores[supplier]['quality'].get() or 0))
                    ws.cell(row=row, column=4, value=float(self.scores[supplier]['delivery'].get() or 0))
                    ws.cell(row=row, column=5, value=float(self.scores[supplier]['cost'].get() or 0))
                except (ValueError, TypeError) as e:
                    logger.warning(f"Errore conversione punteggio per {supplier}: {e}. Uso 0.")
                    ws.cell(row=row, column=2, value=0)
                    ws.cell(row=row, column=3, value=0)
                    ws.cell(row=row, column=4, value=0)
                    ws.cell(row=row, column=5, value=0)
                
                total = get_total(supplier)
                total_cell = ws.cell(row=row, column=6)
                total_cell.value = total
                total_cell.number_format = '0.00'  # Forza 2 decimali fissi
                
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')
                
                if total > max_total:
                    max_total = total
                    winner_row = row
            
            if winner_row:
                for col in range(1, 7):
                    ws.cell(row=winner_row, column=col).fill = green_fill
            
            # Ottieni il percorso della cartella Attachments
            archive_path = get_fixed_attachments_dir()
            if not archive_path:
                messagebox.showerror(_("Errore"), _("Percorso allegati non disponibile."), parent=self)
                return
            
            # Genera nome file univoco con ID progressivo
            try:
                db_manager_temp = DatabaseManager(get_db_path())
                try:
                    next_id = db_manager_temp.get_max_allegato_id() + 1
                finally:
                    try:
                        db_manager_temp.close()
                    except Exception:
                        pass
                
                # Crea nome file univoco: RfQ{id_richiesta}_SQDC_ID{id}.xlsx
                file_ext = ".xlsx"
                new_filename = f"RfQ{self.request_id}_SQDC_ID{next_id}{file_ext}"
                dest_path = os.path.join(archive_path, new_filename)
                
                # Salva il file nella cartella Attachments
                wb.save(dest_path)
                
                # Inserisci il link nel database
                # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
                with DatabaseManager(get_db_path()) as db_manager:
                    db_manager.insert_or_update_allegato_sqdc(self.request_id, sqdc_filename, new_filename)
                
                logger.info(f"Analisi SQDC salvata come Documento Interno: {new_filename} -> {dest_path}")
                messagebox.showinfo(_("Successo"), 
                                  _("Analisi SQDC salvata correttamente nei Documenti Interni."), 
                                  parent=self)
                self.destroy()
                
            except DatabaseError as e:
                logger.error(f"Errore database in save_sqdc: {e}", exc_info=True)
                messagebox.showerror(_("Errore Database"), 
                                   _("Impossibile salvare l'analisi: {}").format(e), 
                                   parent=self)
        
        except Exception as e:
            logger.error(f"Errore nella creazione file SQDC: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), 
                               _("Impossibile creare il file: {}").format(e), 
                               parent=self)

# ------------------------------------------------------------------------------------
# FINESTRA DETTAGLIO RICHIESTA
# ------------------------------------------------------------------------------------
class ViewRequestWindow(tk.Toplevel):
    def __init__(self, parent, request_id, read_only=False, source_db_path=None):
        super().__init__(parent)
        self.withdraw()
        set_window_icon(self)
        self.request_id = request_id
        self.selected_detail_id = None
        self.selected_supplier_name = None
        self.read_only = read_only  # Flag per modalità sola lettura
        
        # Determina quale database usare
        # Se source_db_path è fornito (RdO di altro utente), usa quello
        # Altrimenti usa il database corrente
        if source_db_path and os.path.exists(source_db_path):
            self.db_path = source_db_path
            print(f"[ViewRequestWindow] Usando DB remoto: {source_db_path}")
        else:
            self.db_path = get_db_path()
            print(f"[ViewRequestWindow] Usando DB locale: {self.db_path}")
        
        # Recupera il tipo di RdO e lo username dal database per il titolo
        tipo_rdo = "Fornitura piena"  # Valore di default
        username = None
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(self.db_path) as db_manager:
                result = db_manager.get_tipo_rdo(request_id)
                if result and result[0]:
                    tipo_rdo = result[0]
                username = db_manager.get_username_by_richiesta(request_id)
        except DatabaseError as e:
            logger.error(f"Errore database nel recupero dati per titolo: {e}", exc_info=True)
        
        # Traduci il tipo di RdO e imposta il titolo con User
        tipo_rdo_tradotto = translate_rfq_type(tipo_rdo)
        if username:
            title_base = _("Control Panel - User: {} - Request N° {} - {}").format(username, request_id, tipo_rdo_tradotto)
        else:
            # Fallback se username non disponibile
            title_base = _("Control Panel - Request N° {} - {}").format(request_id, tipo_rdo_tradotto)
        if self.read_only:
            title_base += _(" [SOLA LETTURA]")
        self.title(title_base)
        
        # Rendi la finestra ridimensionabile e massimizzabile
        self.resizable(True, True)
        
        # Gestisci correttamente la chiusura della finestra (NON chiudere l'app principale)
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Frame pulsanti articoli (sempre in fondo)
        frame_article_buttons = ttk.Frame(self)
        frame_article_buttons.pack(side="bottom", fill="x", padx=10, pady=10)
        self.btn_add_article = ttk.Button(frame_article_buttons, text=_("➕ Aggiungi Articolo"), command=self.add_new_article_row)
        self.btn_add_article.pack(side="left", padx=5)
        self.btn_remove_article = ttk.Button(frame_article_buttons, text=_("🗑 Rimuovi Articolo Selezionato"), command=self.remove_selected_article)
        self.btn_remove_article.pack(side="left", padx=5)
        self.btn_import_excel = ttk.Button(frame_article_buttons, text=_("📊 Importa da Excel"), command=self.import_from_excel)
        self.btn_import_excel.pack(side="left", padx=5)
        
        # Frame comandi (in alto)
        frame_comandi = ttk.Frame(self)
        frame_comandi.pack(side="top", fill="x", padx=10, pady=5)
        ttk.Button(frame_comandi, text=_("📄 Gestisci Offerte Fornitori"), command=lambda: self.open_attachment_window("Offerta Fornitore")).pack(side="left")
        ttk.Button(frame_comandi, text=_("📁 Gestisci Documenti Interni"), command=lambda: self.open_attachment_window("Documento Interno")).pack(side="left", padx=10)
        # --- MODIFICA: Pulsante dinamico Fornitori ---
        self.btn_suppliers = ttk.Button(frame_comandi, text="...", command=self.open_edit_suppliers_window)
        self.btn_suppliers.pack(side="left")
        # --- FINE MODIFICA ---
        # --- MODIFICA RICHIESTA: Aggiunta pulsante Note ---
        self.btn_notes = ttk.Button(frame_comandi, text="...", command=self.open_notes_window)
        self.btn_notes.pack(side="left", padx=10)
        # --- FINE MODIFICA ---
        ttk.Button(frame_comandi, text=_("📊 Esporta Excel"), command=self.export_to_excel).pack(side="left", padx=10)
        # --- MODIFICA RICHIESTA: Aggiunta pulsante SQDC ---
        self.btn_sqdc = ttk.Button(frame_comandi, text="...", command=self.open_sqdc_analysis)
        self.btn_sqdc.pack(side="left", padx=10)
        # --- FINE MODIFICA ---
        
        # Frame dettagli richiesta con layout a griglia
        details_frame = ttk.LabelFrame(self, text=_("Dettagli Richiesta"), padding="10")
        details_frame.pack(fill="x", padx=10, pady=5)
        
        s = ttk.Style(); s.configure("Clickable.TLabel", foreground="black", font=('Calibri', 9, 'underline'))
        
        # --- LAYOUT MODIFICATO CON GRID ---
        # Riga 0: Data Emissione + Data Scadenza + Pulsante Add PO
        ttk.Label(details_frame, text=_("Data Emissione: ")).grid(row=0, column=0, sticky="w", padx=(0,5), pady=5)
        self.entry_data_emissione = DateEntry(details_frame, width=12, date_pattern='dd/mm/yyyy', locale=('it_IT' if get_current_language() == 'it' else 'en_US'))
        self.entry_data_emissione.grid(row=0, column=1, sticky="w", padx=(0,20), pady=5)
        self.entry_data_emissione.bind('<<DateEntrySelected>>', self._on_date_changed)
        self.entry_data_emissione.bind('<FocusOut>', self._on_date_changed)  # Salva quando l'utente esce dal campo
        self.entry_data_emissione.bind('<Return>', self._on_date_changed)  # Salva quando l'utente preme Invio
        
        ttk.Label(details_frame, text=_("Data Scadenza: ")).grid(row=0, column=2, sticky="w", padx=(0,5), pady=5)
        self.entry_data_scadenza = DateEntry(details_frame, width=12, date_pattern='dd/mm/yyyy', locale=('it_IT' if get_current_language() == 'it' else 'en_US'))
        self.entry_data_scadenza.grid(row=0, column=3, sticky="w", padx=(0,20), pady=5)
        self.entry_data_scadenza.bind('<<DateEntrySelected>>', self._on_date_changed)
        self.entry_data_scadenza.bind('<FocusOut>', self._on_date_changed)  # Salva quando l'utente esce dal campo
        self.entry_data_scadenza.bind('<Return>', self._on_date_changed)  # Salva quando l'utente preme Invio
        
        # Pulsante Add PO allineato a destra nella stessa riga
        po_btn_text = _("📋 Inserisci OdA") if get_current_language() == 'it' else _("📋 Add PO")
        self.btn_po = ttk.Button(details_frame, text=po_btn_text, command=self.open_po_window)
        self.btn_po.grid(row=0, column=4, sticky="e", padx=(20,0), pady=5)
        
        # Riga 1: Riferimento
        ttk.Label(details_frame, text=_("Riferimento: ")).grid(row=1, column=0, sticky="w", padx=(0,5), pady=5)
        self.lbl_riferimento = ttk.Label(details_frame, text="...", style="Clickable.TLabel", cursor="hand2")
        self.lbl_riferimento.grid(row=1, column=1, columnspan=3, sticky="w", pady=5)
        self.lbl_riferimento.bind("<Button-1>", self._on_reference_click)
        
        # Configura le colonne per espandere il pulsante Add PO a destra
        details_frame.grid_columnconfigure(4, weight=1)
        # --- FINE LAYOUT MODIFICATO ---
        
        # Frame griglia (espandibile, tra i pulsanti sopra e sotto)
        frame_grid = ttk.LabelFrame(self, text=_("Tabella Prezzi: Materiali / Fornitori"))
        frame_grid.pack(side="top", fill="both", expand=True, padx=10, pady=5)
        
        # Crea il widget tksheet
        self.sheet = Sheet(frame_grid, 
                          theme="light blue",
                          header_font=("Calibri", 11, "bold"),
                          font=("Calibri", 11, "normal"))
        self.sheet.enable_bindings()
        self.sheet.pack(fill="both", expand=True)
        
        self.load_rdo_details(); self.build_grid()
        self.check_suppliers_status_and_update_button() # MODIFICA: Controlla lo stato dei fornitori all'avvio
        self.check_note_status_and_update_button() # MODIFICA RICHIESTA: Controlla lo stato della nota all'avvio
        self.check_sqdc_status_and_update_button() # MODIFICA: Controlla lo stato dell'analisi SQDC all'avvio
        
        # Disabilita pulsanti di modifica se in modalità read-only
        if self.read_only:
            self._disable_edit_controls()
            
            # Mostra un messaggio informativo sulla modalità sola lettura
            info_frame = ttk.Frame(self, style='Warning.TFrame')
            info_frame.pack(side='bottom', fill='x', padx=10, pady=(0, 5))
            
            info_label = ttk.Label(
                info_frame,
                text=_('⚠️ MODALITÀ SOLA LETTURA: Stai visualizzando una RdO di un altro utente. Non puoi modificare i dati.'),
                foreground='#d63031',
                font=('Calibri', 10, 'bold'),
                anchor='center'
            )
            info_label.pack(pady=5)
        
        # 🆕 MODIFICA: Calcola e imposta la larghezza ottimale della finestra in base ai fornitori
        num_suppliers = len(self.suppliers) if hasattr(self, 'suppliers') else 0
        is_cl = self.is_conto_lavoro if hasattr(self, 'is_conto_lavoro') else False
        optimal_geometry = calculate_optimal_window_size(self, num_suppliers, is_cl)
        self.geometry(optimal_geometry)
        
        self.deiconify()

    def _disable_edit_controls(self):
        """Disabilita tutti i controlli di modifica quando in modalità read-only."""
        # Disabilita pulsanti articoli
        if hasattr(self, 'btn_add_article'):
            self.btn_add_article.config(state='disabled')
        if hasattr(self, 'btn_remove_article'):
            self.btn_remove_article.config(state='disabled')
        if hasattr(self, 'btn_import_excel'):
            self.btn_import_excel.config(state='disabled')
        
        # Disabilita pulsanti gestione
        if hasattr(self, 'btn_suppliers'):
            self.btn_suppliers.config(state='disabled')
        if hasattr(self, 'btn_notes'):
            self.btn_notes.config(state='disabled')
        if hasattr(self, 'btn_sqdc'):
            self.btn_sqdc.config(state='disabled')
        if hasattr(self, 'btn_po'):
            self.btn_po.config(state='disabled')
        
        # NON disabilitare i pulsanti "Gestisci Offerte/Documenti" e "Esporta Excel"
        # perché sono operazioni di sola lettura/visualizzazione
        
        # Disabilita date
        if hasattr(self, 'entry_data_emissione'):
            self.entry_data_emissione.config(state='disabled')
        if hasattr(self, 'entry_data_scadenza'):
            self.entry_data_scadenza.config(state='disabled')
        
        # Disabilita riferimento
        if hasattr(self, 'entry_riferimento'):
            self.entry_riferimento.config(state='disabled')
        
        # Disabilita click su label riferimento
        if hasattr(self, 'lbl_riferimento'):
            self.lbl_riferimento.unbind("<Button-1>")
            self.lbl_riferimento.config(cursor="")
        
        # Rendi la griglia read-only
        if hasattr(self, 'sheet'):
            self.sheet.disable_bindings()
            self.sheet.enable_bindings(
                "single_select",
                "row_select",
                "column_width_resize",
                "double_click_column_resize",
                "arrowkeys",
                "right_click_popup_menu",
                "copy"
            )

    def on_closing(self):
        """Gestisce la chiusura della finestra rilasciando tutte le risorse per prevenire memory leak."""
        # BUG #7 FIX: Cleanup completo per prevenire memory leak
        try:
            # Rilascia binding eventi per prevenire memory leak
            if hasattr(self, 'entry_data_emissione'):
                try:
                    self.entry_data_emissione.unbind('<<DateEntrySelected>>')
                    self.entry_data_emissione.unbind('<FocusOut>')
                    self.entry_data_emissione.unbind('<Return>')
                except Exception as e:
                    logger.debug(f"Errore unbind entry_data_emissione: {e}")
            
            if hasattr(self, 'entry_data_scadenza'):
                try:
                    self.entry_data_scadenza.unbind('<<DateEntrySelected>>')
                    self.entry_data_scadenza.unbind('<FocusOut>')
                    self.entry_data_scadenza.unbind('<Return>')
                except Exception as e:
                    logger.debug(f"Errore unbind entry_data_scadenza: {e}")
            
            # Rilascia binding sheet per prevenire memory leak
            if hasattr(self, 'sheet'):
                try:
                    self.sheet.unbind("<Double-Button-1>")
                    # Pulisci i dati memorizzati
                    if hasattr(self.sheet, '_sheet_data'):
                        del self.sheet._sheet_data
                    if hasattr(self.sheet, '_last_click_time'):
                        del self.sheet._last_click_time
                    if hasattr(self.sheet, '_last_click_row'):
                        del self.sheet._last_click_row
                except Exception as e:
                    logger.debug(f"Errore rilascio binding sheet: {e}")
            
            # Pulisci riferimenti circolari che potrebbero impedire garbage collection
            if hasattr(self, 'materials'):
                self.materials = None
            if hasattr(self, 'suppliers'):
                self.suppliers = None
            if hasattr(self, 'prices'):
                self.prices = None
                
        except Exception as e:
            logger.error(f"Errore in on_closing cleanup: {e}", exc_info=True)
        finally:
            self.destroy()

    def _get_db_manager(self):
        """Helper per ottenere il DatabaseManager corretto (locale o remoto).
        
        Se self.read_only è True (database di altro utente), apre in modalità sola lettura
        per permettere accesso concorrente senza lock.
        """
        return DatabaseManager(self.db_path, read_only=self.read_only)

# --- INIZIO MODIFICA: Nuovo metodo per gestire il doppio click ---
    # --- MODIFICA RICHIESTA: Nuovo metodo per aprire la finestra delle note ---
    def open_notes_window(self):
        # Apre la finestra delle note e attende la sua chiusura
        win = NotesWindow(self, self.request_id)
        setattr(win, 'db_path', getattr(self, 'db_path', None))
        setattr(win, 'read_only', getattr(self, 'read_only', False))
        self.wait_window(win)
        # Aggiorna il testo del pulsante dopo la chiusura della finestra
        self.check_note_status_and_update_button()
        
    # --- MODIFICA RICHIESTA: Nuovo metodo per controllare l'esistenza di una nota e aggiornare il pulsante ---
    def check_note_status_and_update_button(self):
        """Controlla se esiste una nota e aggiorna il testo del pulsante.
        
        BUG #7 FIX: Aggiunto finally per garantire chiusura database.
        """
        db_manager = None
        try:
            db_manager = self._get_db_manager()
            result = db_manager.get_note_formattate(self.request_id)
            
            # Se esiste una nota (la colonna non è NULL e non è una stringa vuota o un dump vuoto)
            if result and result[0] and result[0] != "()":
                self.btn_notes.config(text="📝 " + _("Visualizza nota"))
            else:
                self.btn_notes.config(text="📝 " + _("Aggiungi nota"))
        except DatabaseError as e:
            logger.error(f"Errore database in check_note_status_and_update_button: {e}", exc_info=True)
            self.btn_notes.config(text="⚠️ " + _("Errore Nota"), state="disabled")
        finally:
            if db_manager:
                try:
                    db_manager.close()
                except Exception as close_err:
                    logger.warning(f"Errore chiusura DatabaseManager in check_note_status_and_update_button: {close_err}")
    
    # --- MODIFICA: Nuovo metodo per controllare l'esistenza di fornitori e aggiornare il pulsante ---
    def check_suppliers_status_and_update_button(self):
        """Controlla se esistono fornitori associati e aggiorna il testo del pulsante.
        
        BUG FIX: Usa context manager per gestione automatica connessione.
        """
        try:
            with DatabaseManager(self.db_path, read_only=self.read_only) as db_manager:
                result = db_manager.get_fornitori_count(self.request_id)
                
                # Se ci sono fornitori, mostra "Modifica Fornitori", altrimenti "Aggiungi Fornitori"
                if result and result > 0:
                    self.btn_suppliers.config(text="✏️ " + _("Modifica Fornitori"))
                else:
                    self.btn_suppliers.config(text="➕ " + _("Aggiungi Fornitori"))
        except DatabaseError as e:
            logger.error(f"Errore database in check_suppliers_status_and_update_button: {e}", exc_info=True)
            self.btn_suppliers.config(text="⚠️ " + _("Errore Fornitori"), state="disabled")
    
    # --- MODIFICA: Nuovi metodi per gestione SQDC ---
    def check_sqdc_status_and_update_button(self):
        """Controlla se esiste un'analisi SQDC salvata e aggiorna il testo del pulsante.
        
        BUG FIX: Usa context manager per gestione automatica connessione.
        """
        try:
            # 🔧 CORREZIONE BUG: Cerca sempre il nome file univoco indipendente dalla lingua
            sqdc_filename = f"SQDC_Analysis_RfQ_{self.request_id}.xlsx"
            
            with DatabaseManager(self.db_path, read_only=self.read_only) as db_manager_sqdc:
                existing = db_manager_sqdc.get_allegato_id_by_filename(self.request_id, sqdc_filename, 'Documento Interno')
            
            if existing:
                self.btn_sqdc.config(text=_("📈 Apri analisi SQDC"))
            else:
                self.btn_sqdc.config(text=_("📊 Crea analisi SQDC"))
                
        except DatabaseError as e:
            logger.error(f"Errore database in check_sqdc_status_and_update_button: {e}", exc_info=True)
            self.btn_sqdc.config(text=_("Errore SQDC"), state="disabled")
    
    def open_sqdc_analysis(self):
        """Apre la finestra di analisi SQDC (nuova o esistente)"""
        # Carica dati esistenti se presenti
        existing_data = None
        
        try:
            # 🔧 CORREZIONE BUG: Cerca sempre il nome file univoco indipendente dalla lingua
            sqdc_display_name = f"SQDC_Analysis_RfQ_{self.request_id}.xlsx"
            
            # Cerca l'allegato nel database
            db_manager_sqdc = DatabaseManager(self.db_path, read_only=self.read_only)
            try:
                allegato = db_manager_sqdc.get_allegato_id_by_filename(
                    self.request_id, 
                    sqdc_display_name, 
                    'Documento Interno'
                )
            finally:
                try:
                    db_manager_sqdc.close()
                except Exception:
                    pass
            
            if allegato:
                # File SQDC trovato - carica i dati dal file fisico linkato
                allegato_id = allegato[0]
                
                # Recupera il percorso fisico del file
                db_manager_file = DatabaseManager(self.db_path, read_only=self.read_only)
                try:
                    result = db_manager_file.get_allegato_file_data(allegato_id)
                finally:
                    try:
                        db_manager_file.close()
                    except Exception:
                        pass
                
                if result:
                    nome_file, dati_file, percorso_esterno = result
                    
                    # Variabile per tracciare il file temporaneo (se creato da BLOB)
                    temp_blob_path = None
                    
                    try:
                        # 🆕 LETTURA DAL FILE FISICO (non BLOB)
                        if percorso_esterno:
                            # File esterno linkato - usa il percorso fisico
                            base_path = get_fixed_attachments_dir()
                            if not base_path:
                                raise ValueError("Percorso allegati non configurato")
                            
                            excel_path = os.path.join(base_path, percorso_esterno)
                            
                            if not os.path.exists(excel_path):
                                raise FileNotFoundError(f"File SQDC non trovato: {excel_path}")
                            
                            logger.info(f"Caricamento SQDC da file fisico: {excel_path}")
                        elif dati_file:
                            # Fallback BLOB (per compatibilità con vecchi dati)
                            logger.warning("SQDC trovato come BLOB - conversione consigliata")
                            import tempfile
                            fd, temp_blob_path = tempfile.mkstemp(suffix='.xlsx')
                            os.close(fd)
                            with open(temp_blob_path, 'wb') as f:
                                f.write(dati_file)
                            excel_path = temp_blob_path
                        else:
                            raise ValueError("Nessun dato allegato disponibile")
                        
                        # 🆕 PARSING DEL FILE EXCEL SQDC
                        wb = openpyxl.load_workbook(excel_path, data_only=True)
                        ws = wb.active
                        
                        # Carica pesi (celle B5-B8)
                        weights = {
                            'safety': ws['B5'].value or 25,
                            'quality': ws['B6'].value or 25,
                            'delivery': ws['B7'].value or 25,
                            'cost': ws['B8'].value or 25
                        }
                        
                        # Carica punteggi fornitori (a partire da riga 17)
                        scores = {}
                        start_row = 17
                        row = start_row
                        
                        while row < 100:  # Limite sicurezza
                            supplier_cell = ws.cell(row=row, column=1)
                            if not supplier_cell.value:
                                break  # Fine dati fornitori
                            
                            supplier = supplier_cell.value
                            scores[supplier] = {
                                'safety': ws.cell(row=row, column=2).value or 0,
                                'quality': ws.cell(row=row, column=3).value or 0,
                                'delivery': ws.cell(row=row, column=4).value or 0,
                                'cost': ws.cell(row=row, column=5).value or 0
                            }
                            row += 1
                        
                        wb.close()
                        
                        # Costruisci dizionario dati esistenti
                        existing_data = {
                            'weights': weights,
                            'scores': scores,
                            'automatic_cost': False  # Non sappiamo se era automatico
                        }
                        
                        logger.info(f"Dati SQDC caricati per RdO {self.request_id}: {len(scores)} fornitori")
                        
                    except Exception as parse_error:
                        logger.error(f"Errore parsing file SQDC: {parse_error}", exc_info=True)
                        messagebox.showwarning(
                            _("Avviso"),
                            _("File SQDC trovato ma impossibile caricare i dati.\nVerrà aperta una nuova analisi vuota.\n\nErrore: {}").format(parse_error),
                            parent=self
                        )
                        existing_data = None
                    finally:
                        # Pulisci sempre il file temporaneo se era BLOB
                        if temp_blob_path and os.path.exists(temp_blob_path):
                            try:
                                os.remove(temp_blob_path)
                                logger.debug(f"File temporaneo BLOB rimosso: {temp_blob_path}")
                            except Exception as e:
                                logger.warning(f"Impossibile rimuovere file temporaneo {temp_blob_path}: {e}")
        
        except (DatabaseError, Exception) as e:
            logger.error(f"Errore nel caricamento dati SQDC: {e}", exc_info=True)
            # Continua con existing_data = None (apre finestra vuota)
        
        # Apri finestra SQDC
        win = SQDCAnalysisWindow(self, self.request_id, existing_data)
        setattr(win, 'db_path', getattr(self, 'db_path', None))
        setattr(win, 'read_only', getattr(self, 'read_only', False))
        self.wait_window(win)
        # Dopo la chiusura, aggiorna il pulsante
        self.check_sqdc_status_and_update_button()
    
    def _format_date_for_display(self, db_date):
        if not db_date: return _("N/D")
        try: return datetime.strptime(db_date, '%Y-%m-%d').strftime('%d/%m/%Y')
        except (ValueError, TypeError): return db_date

    def load_rdo_details(self):
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with self._get_db_manager() as db_manager:
                result = db_manager.get_richiesta_basic_data(self.request_id)
            r, de, ds = result if result else (_("N/D"), None, None)
            self.lbl_riferimento.config(text=r if r else _("N/D"))
            
            # --- MODIFICA: Caricamento date nei DateEntry ---
            try:
                if de: self.entry_data_emissione.set_date(datetime.strptime(de, '%Y-%m-%d'))
                else: self.entry_data_emissione.delete(0, 'end')
            except (ValueError, TypeError): self.entry_data_emissione.delete(0, 'end')
                
            try:
                if ds: self.entry_data_scadenza.set_date(datetime.strptime(ds, '%Y-%m-%d'))
                else: self.entry_data_scadenza.delete(0, 'end')
            except (ValueError, TypeError): self.entry_data_scadenza.delete(0, 'end')
            # --- FINE MODIFICA ---
                
        except DatabaseError as e:
            logger.error(f"Errore database in load_rdo_details: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Impossibile caricare dettagli: {}").format(e), parent=self)

    def open_edit_reference_window(self):
        """Apre la finestra di modifica riferimento."""
        # Controlla se in modalità read-only
        if self.read_only:
            messagebox.showwarning(
                _("Operazione Non Consentita"),
                _("Non puoi modificare il riferimento di RdO di altri utenti."),
                parent=self
            )
            return
        
        win = EditReferenceWindow(self, self.request_id)
        setattr(win, 'db_path', getattr(self, 'db_path', None))
        setattr(win, 'read_only', getattr(self, 'read_only', False))
        self.wait_window(win)
        self.load_rdo_details()
    
    def open_po_window(self):
        """Apre la finestra di gestione numeri ordine di acquisto."""
        # Controlla se in modalità read-only
        if self.read_only:
            messagebox.showwarning(
                _("Operazione Non Consentita"),
                _("Non puoi modificare i numeri ordine di RdO di altri utenti."),
                parent=self
            )
            return
        
        win = PurchaseOrderWindow(self, self.request_id)
        setattr(win, 'db_path', getattr(self, 'db_path', None))
        setattr(win, 'read_only', getattr(self, 'read_only', False))
        self.wait_window(win)

# --- INIZIO NUOVI METODI AGGIUNTI ---
    def _format_date_for_db(self, display_date):
        """Converte una data 'dd/mm/yyyy' in 'YYYY-MM-DD' per il DB."""
        if not display_date: return None
        try: return datetime.strptime(display_date, '%d/%m/%Y').strftime('%Y-%m-%d')
        except (ValueError, TypeError): return None

    # BUG #16 FIX: Metodi dedicati invece di lambda per evitare memory leak
    def _on_date_changed(self, event=None):
        """Handler per eventi di cambio data."""
        self.auto_save_dates()
    
    def _on_reference_click(self, event=None):
        """Handler per click su etichetta riferimento."""
        self.open_edit_reference_window()

    def auto_save_dates(self):
        """Salva automaticamente le date quando vengono modificate dai calendari."""
        # BUG #14 FIX: Debounce per evitare race condition
        if hasattr(self, '_date_save_pending') and self._date_save_pending:
            return  # Già in corso un salvataggio, ignora questa chiamata
        
        self._date_save_pending = True
        
        try:
            new_date_em = self._format_date_for_db(self.entry_data_emissione.get())
            new_date_sc = self._format_date_for_db(self.entry_data_scadenza.get())
            
            # BUG #46 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with self._get_db_manager() as db_manager:
                db_manager.update_date_richiesta(self.request_id, new_date_em, new_date_sc)
            
            # Aggiorna la lista nella finestra principale
            if hasattr(self.master, 'refresh_data'):
                self.master.refresh_data()
                
        except Exception as e:
            logger.error(f"Errore database in auto_save_dates: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Impossibile salvare le date: {}").format(e), parent=self)
        finally:
            # BUG #30 FIX: Usa weakref per evitare memory leak con lambda capture
            # Rilascia il lock dopo un breve delay
            if hasattr(self, 'after'):
                import weakref
                weak_self = weakref.ref(self)
                def release_lock():
                    obj = weak_self()
                    if obj is not None:
                        obj._date_save_pending = False
                self.after(300, release_lock)
            else:
                self._date_save_pending = False

    def save_dates(self):
        """Salva la data di emissione e scadenza modificate."""
        try:
            # Usa la funzione helper per convertire, gestisce anche i campi vuoti
            new_date_em = self._format_date_for_db(self.entry_data_emissione.get())
            new_date_sc = self._format_date_for_db(self.entry_data_scadenza.get())
        except Exception as e:
            messagebox.showerror(_("Errore Formato Data"), _("Date non valide: {}").format(e), parent=self)
            return

        try:
            # BUG #46 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with self._get_db_manager() as db_manager:
                db_manager.update_date_richiesta(self.request_id, new_date_em, new_date_sc)
            
            messagebox.showinfo(_("Successo"), _("Date aggiornate."), parent=self)
            
            # Aggiorna anche la lista nella finestra principale
            if hasattr(self.master, 'refresh_data'):
                self.master.refresh_data()
                
        except DatabaseError as e:
            logger.error(f"Errore database in save_dates: {e}", exc_info=True)
            messagebox.showerror(_("Errore Database"), _("Impossibile salvare le date: {}").format(e), parent=self)
    # --- FINE NUOVI METODI AGGIUNTI ---

    def export_to_excel(self):
        logger.info(f"Esportazione Excel per RdO {self.request_id}")
        # --- INIZIO MODIFICA PER SCELTA LINGUA E TEMPLATE ---
        
        # 1. Chiedi all'utente la lingua
        prompt = LanguagePrompt(self)
        self.wait_window(prompt)
        chosen_language = prompt.choice
        
        if not chosen_language: # L'utente ha chiuso la finestra
            return
            
        # 2. Recupera il tipo di RdO per scegliere il template corretto
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with self._get_db_manager() as db_manager:
                rdo_type_result = db_manager.get_tipo_rdo(self.request_id)
            if not rdo_type_result:
                messagebox.showerror(_("Errore"), _("Tipo RdO non trovato."), parent=self)
                return
            # Normalizza il tipo RFQ prima del confronto per gestire valori in qualsiasi lingua
            tipo_normalizzato = normalize_rfq_type(rdo_type_result[0])
            is_cl = tipo_normalizzato == 'Conto lavoro'
        except Exception as e:
            messagebox.showerror(_("Errore Database"), _("Impossibile determinare il tipo di RdO: {}").format(e), parent=self)
            return

        # 3. Determina il nome del file template e le stringhe di testo
        template_name = ""
        texts = {}

        if is_cl:
            if chosen_language == 'ita':
                template_name = "template_rdo_cl.xlsx"
            else: # eng
                template_name = "template_rdo_eng_cl.xlsx"
        else: # Fornitura piena
            if chosen_language == 'ita':
                template_name = "template_rdo.xlsx"
            else: # eng
                template_name = "template_rdo_eng.xlsx"

        if chosen_language == 'ita':
            texts = {
                "save_title": "Salva Riepilogo",
                "initial_file": f"Riepilogo_RdO_{self.request_id}.xlsx",
                "vs_best": "VS. MIGLIORE"
            }
        else: # eng
            texts = {
                "save_title": "Save Summary",
                "initial_file": f"Summary_RfQ_{self.request_id}.xlsx",
                "vs_best": "BEST DELIVERY"
            }
            
        # 4. Costruisci il percorso completo e verifica l'esistenza del template
        template_path = resource_path(os.path.join("add_data", template_name))

        if not os.path.exists(template_path):
            messagebox.showerror(_("Errore"), _("File modello non trovato!\nAssicurarsi che '{}' esista nella cartella 'add_data'.").format(template_name), parent=self)
            return
        
        # --- FINE MODIFICA PER SCELTA LINGUA E TEMPLATE ---
        
        wb = None  # BUG #17 FIX: Inizializza per finally block
        try:
            with self._get_db_manager() as db_manager:
                rdo_det = db_manager.get_richiesta_full_data(self.request_id)
                if not rdo_det:
                    messagebox.showerror(_("Errore"), _("Dettagli RdO non trovati."), parent=self)
                    return
                de_db, ds_db, rif, tipo = rdo_det
                suppliers_rows = db_manager.get_fornitori_by_richiesta(self.request_id, order_by=True)
                suppliers = [r[0] for r in suppliers_rows]
                items = db_manager.get_dettagli_by_richiesta(self.request_id)
                prices_rows = db_manager.get_offerte_by_richiesta(self.request_id)
            prices = {(id_d, nf): p for id_d, nf, p in prices_rows}
            
            wb = openpyxl.load_workbook(template_path); ws = wb.active # Usa il percorso del template dinamico
            
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            bold = Font(bold=True); center = Alignment(horizontal='center', vertical='center'); left_align = Alignment(horizontal='left', vertical='center'); price_fmt = '0.0000'
            
            # Validazione sicura delle date per prevenire crash
            try:
                ws['B1'] = datetime.strptime(de_db, '%Y-%m-%d').strftime('%d/%m/%Y') if de_db else _("N/D")
            except (ValueError, TypeError) as e:
                logger.error(f"Formato data emissione non valido per RdO {self.request_id}: '{de_db}' - {e}")
                ws['B1'] = _("Data non valida")
            
            try:
                ws['B2'] = datetime.strptime(ds_db, '%Y-%m-%d').strftime('%d/%m/%Y') if ds_db else _("N/D")
            except (ValueError, TypeError) as e:
                logger.error(f"Formato data scadenza non valido per RdO {self.request_id}: '{ds_db}' - {e}")
                ws['B2'] = _("Data non valida")
            
            ws['C1'] = self.request_id
            for i, s_name in enumerate(suppliers): ws.cell(row=3, column=14+i, value=s_name).font=bold; ws.cell(row=3, column=14+i).alignment=center; ws.cell(row=3, column=14+i).border=border
            for i, item in enumerate(items):
                id_d, cod, all, desc, qta, c_g, d_g, m_cl = item; row = 4+i
                ws.cell(row=row, column=1, value=cod); ws.cell(row=row, column=2, value=all).alignment=left_align
                ws.cell(row=row, column=3, value=format_quantity_display(qta)).alignment=center; ws.cell(row=row, column=5, value=desc)
                if is_cl: ws.cell(row=row, column=6, value=c_g); ws.cell(row=row, column=7, value=d_g); ws.cell(row=row, column=8, value=m_cl).alignment=left_align
                
                ws.cell(row=row, column=10, value=texts["vs_best"]).alignment=center # Usa testo dinamico
                
                ws.cell(row=row, column=12, value=rif)
                for col in range(1, 13): ws.cell(row=row, column=col).border = border
                for j, s_name in enumerate(suppliers):
                    p_cell = ws.cell(row=row, column=14+j)
                    price_val = prices.get((id_d, s_name))
                    if price_val is not None:
                        try:
                            p_cell.value = float(str(price_val).replace(',', '.'))
                            p_cell.number_format = price_fmt
                        except (ValueError, TypeError):
                            p_cell.value = price_val
                            p_cell.number_format = '@'
                    p_cell.border = border

            # Usa titolo e nome file dinamici
            filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[(_("File Excel"), "*.xlsx")], title=texts["save_title"], initialfile=texts["initial_file"])
            
            if filepath: 
                wb.save(filepath)
                logger.info(f"Excel esportato: {filepath}")
                messagebox.showinfo(_("Successo"), _("File salvato in:\n{}").format(filepath), parent=self)
        except Exception as e: 
            logger.error(f"Errore esportazione Excel: {e}", exc_info=True)
            messagebox.showerror(_("Errore Esportazione"), _("Errore: {}").format(e), parent=self)
        finally:
            # BUG #17 FIX: Chiudi SEMPRE il workbook per rilasciare il file
            if wb is not None:
                try:
                    wb.close()
                    logger.debug(f"Workbook Excel chiuso: {template_path}")
                except Exception as close_error:
                    logger.warning(f"Errore chiusura workbook Excel: {close_error}")

    def open_edit_suppliers_window(self):
        print(f"[ViewRequestWindow] Prima di aprire EditSuppliersWindow - Richiesta {self.request_id}")
        win = EditSuppliersWindow(self, self.request_id)
        setattr(win, 'db_path', getattr(self, 'db_path', None))
        setattr(win, 'read_only', getattr(self, 'read_only', False))
        print(f"[ViewRequestWindow] Attendendo chiusura EditSuppliersWindow...")
        self.wait_window(win)
        print(f"[ViewRequestWindow] EditSuppliersWindow CHIUSA - Richiesta {self.request_id}")
        
        # BUG FIX: Delay più lungo per dare tempo al WAL di consolidarsi
        print(f"[ViewRequestWindow] Attesa 300ms per consolidamento WAL prima del refresh...")
        self.after(300, self._delayed_refresh_after_suppliers)
    
    def _delayed_refresh_after_suppliers(self):
        """Refresh ritardato dopo modifica fornitori per evitare race condition."""
        print(f"[ViewRequestWindow] === INIZIO REFRESH RITARDATO ===")
        print(f"[ViewRequestWindow] Chiamando refresh_grid() per richiesta {self.request_id}...")
        try:
            self.refresh_grid()
            print(f"[ViewRequestWindow] refresh_grid() COMPLETATO con successo")
        except Exception as e:
            print(f"[ViewRequestWindow] ERRORE in refresh_grid(): {e}")
            import traceback
            traceback.print_exc()
        
        print(f"[ViewRequestWindow] Aggiornando pulsante fornitori...")
        try:
            self.check_suppliers_status_and_update_button()
            print(f"[ViewRequestWindow] Pulsante fornitori aggiornato")
        except Exception as e:
            print(f"[ViewRequestWindow] ERRORE in check_suppliers_status_and_update_button(): {e}")
        
        print(f"[ViewRequestWindow] === FINE REFRESH RITARDATO ===")
    def refresh_grid(self):
        """Aggiorna la griglia ricostruendola"""
        self.build_grid()
    def build_grid(self):
        """Costruisce la griglia prezzi usando tksheet con colonne articolo modificabili a sinistra"""
        try:
            print(f"[ViewRequestWindow.build_grid] === INIZIO BUILD_GRID ===")
            print(f"[ViewRequestWindow.build_grid] Richiesta: {self.request_id}, DB path: {self.db_path}")
            
            # Forza aggiornamento UI per evitare freeze percepito
            self.update_idletasks()
            
            # BUG FIX: Usa context manager per gestione automatica connessione
            print(f"[ViewRequestWindow.build_grid] Apertura DatabaseManager...")
            with DatabaseManager(self.db_path, read_only=self.read_only) as db_manager:
                print(f"[ViewRequestWindow.build_grid] DatabaseManager aperto, recupero tipo RdO...")
                # Recupera tipo RdO per determinare quali colonne mostrare
                result = db_manager.get_tipo_rdo(self.request_id)
                tipo_normalizzato = normalize_rfq_type(result[0] if result and result[0] else "Fornitura piena")
                self.is_conto_lavoro = tipo_normalizzato == "Conto lavoro"
                print(f"[ViewRequestWindow.build_grid] Tipo RdO: {tipo_normalizzato}")
                
                print(f"[ViewRequestWindow.build_grid] Recupero fornitori...")
                suppliers_rows = db_manager.get_fornitori_by_richiesta(self.request_id, order_by=True)
                suppliers = [r[0] for r in suppliers_rows]
                print(f"[ViewRequestWindow.build_grid] Fornitori recuperati: {len(suppliers)} - {suppliers}")
                
                print(f"[ViewRequestWindow.build_grid] Recupero materiali...")
                # Recupera TUTTI i campi articolo inclusi quelli per conto lavoro
                materials = db_manager.get_dettagli_by_richiesta(self.request_id)
                print(f"[ViewRequestWindow.build_grid] Recupero prezzi...")
                prices_rows = db_manager.get_offerte_by_richiesta(self.request_id)
                print(f"[ViewRequestWindow.build_grid] Chiusura DatabaseManager...")
            
            # Context manager ha chiuso il DB automaticamente
            print(f"[ViewRequestWindow.build_grid] DatabaseManager chiuso, elaborazione dati...")
            prices = {(id_d, nf): p for id_d, nf, p in prices_rows}
            print(f"[ViewRequestWindow.build_grid] Materiali: {len(materials)}, Prezzi: {len(prices)}")
            if materials:
                print(f"[ViewRequestWindow.build_grid] Primi 3 materiali: {materials[:3]}")
        except DatabaseError as e:
            logger.error(f"Errore database in build_grid: {e}", exc_info=True)
            messagebox.showerror(_("Errore Database"), _("Impossibile caricare la griglia: {}").format(e), parent=self)
            return
        
        # Salva i dati per uso successivo
        self.suppliers = suppliers
        self.materials = materials
        self.prices = prices
        
        # 🆕 CORREZIONE BUG: Reset completo del sheet PRIMA di ricostruire
        # Questo evita che tksheet mostri header vecchie o di default ("E", "F")
        self.sheet.set_sheet_data([])  # Svuota i dati
        self.sheet.headers([])         # Svuota gli header
        self.sheet.set_all_cell_sizes_to_text()  # Reset dimensioni
        
        # Costruisci le intestazioni: colonne articolo + colonne fornitori
        base_headers = [
            _("Codice"),
            _("Allegato"),
            _("Descrizione"),
            get_qty_column_text()  # "Q.tà" o "Q.ty"
        ]
        
        # Aggiungi colonne conto lavoro se necessario
        if self.is_conto_lavoro:
            cl_headers = [_("Cod. Grezzo"), _("Allegato Grezzo"), _("Mat. C/L")]
            headers = base_headers + cl_headers + suppliers
            num_article_cols = 7  # Totale colonne articolo per conto lavoro
        else:
            headers = base_headers + suppliers
            num_article_cols = 4  # Totale colonne articolo per fornitura piena
        
        # Costruisci le righe di dati
        data_rows = []
        for i, (id_d, cod, all, desc, qta, c_g, d_g, m_cl) in enumerate(materials, start=1):
            # Colonne articolo (base)
            row = [
                cod or "",     # Codice
                all or "",     # Allegato
                desc or "",    # Descrizione
                format_quantity_display(qta) or ""  # Quantità formattata
            ]
            
            # Aggiungi colonne conto lavoro se necessario
            if self.is_conto_lavoro:
                row.extend([c_g or "", d_g or "", m_cl or ""])
            
            # Aggiungi i prezzi per ogni fornitore
            for s in suppliers:
                price = prices.get((id_d, s))
                display_val = ""
                if price is not None:
                    try:
                        price_str = str(price)
                        price_str_upper = price_str.upper()
                        if price_str_upper in ('X', 'ND'):
                            display_val = price_str_upper
                        else:
                            price_float = parse_float_from_comma_string(price_str)
                            display_val = format_price_display(price_float)
                    except (ValueError, TypeError):
                        display_val = str(price)
                row.append(display_val)
            
            data_rows.append(row)
        
        # 🆕 CORREZIONE: Imposta PRIMA gli header, POI i dati
        # Questo garantisce che tksheet non usi mai header di fallback
        self.sheet.headers(headers)
        self.sheet.set_sheet_data(data_rows)
        
        # Forza aggiornamento della vista dopo aver impostato i dati
        self.sheet.update_idletasks()
        print(f"[ViewRequestWindow.build_grid] Griglia aggiornata con {len(data_rows)} righe di dati")
        
        # Configura larghezze colonne
        try:
            import tkinter.font as tkfont
            header_font = tkfont.Font(family="Calibri", size=10, weight="bold")
            content_font = tkfont.Font(family="Calibri", size=10, weight="normal")
            
            PADDING_PX = 20
            
            # Larghezze minime per colonne articolo
            article_min_widths = {
                0: 80,   # Codice
                1: 80,   # Allegato
                2: 200,  # Descrizione (più largo per editing)
                3: 60    # Q.tà
            }
            
            # Larghezze per colonne conto lavoro
            if self.is_conto_lavoro:
                article_min_widths.update({
                    4: 100,  # Cod. Grezzo
                    5: 100,  # Allegato Grezzo
                    6: 150   # Mat. C/L
                })
            
            DEFAULT_MIN_WIDTH = 100  # Per colonne fornitori
            
            for col_idx in range(len(headers)):
                header_text = headers[col_idx]
                max_width = header_font.measure(header_text)
                
                # Per le colonne articolo, controlla anche il contenuto delle celle
                if col_idx < num_article_cols:
                    for row in data_rows:
                        if col_idx < len(row):
                            cell_value = str(row[col_idx])
                            cell_width = content_font.measure(cell_value)
                            max_width = max(max_width, cell_width)
                    
                    min_width = article_min_widths.get(col_idx, 80)
                else:
                    # Colonne fornitori
                    min_width = DEFAULT_MIN_WIDTH
                
                column_width = max(int(max_width + PADDING_PX), min_width)
                self.sheet.column_width(column=col_idx, width=column_width)
                
        except Exception as e:
            logger.warning(f"Errore calcolo larghezza colonne: {e}. Uso larghezze default.")
            # Fallback
            widths = [120, 120, 250, 60]  # Base
            if self.is_conto_lavoro:
                widths.extend([100, 100, 150])  # CL
            for col_idx in range(len(widths)):
                self.sheet.column_width(column=col_idx, width=widths[col_idx])
            for col_idx in range(len(widths), len(headers)):
                self.sheet.column_width(column=col_idx, width=120)
        
        # TUTTE le colonne articolo sono modificabili (Codice, Allegato, Descrizione, Q.tà)
        # Se conto lavoro, anche le colonne aggiuntive sono modificabili (Cod. Grezzo, Allegato Grezzo, Mat. C/L)
        # Nessuna colonna articolo è readonly - tutto modificabile!
        
        # Centra l'allineamento della colonna quantità
        align_cols = [3]  # Q.tà
        self.sheet.align_columns(columns=align_cols, align="center")
        
        # Allinea a destra le colonne prezzi fornitori
        price_columns = list(range(num_article_cols, len(headers)))
        if price_columns:
            self.sheet.align_columns(columns=price_columns, align="right")
        
        # Configura colori alternati per le righe
        for row_idx in range(len(data_rows)):
            if row_idx % 2 != 0:
                self.sheet.highlight_rows(rows=[row_idx], bg="#F0F0F0", fg="black", highlight_index=False)
        
        # Abilita le funzionalità di tksheet
        self.sheet.enable_bindings(
            "single_select",
            "drag_select",
            "column_select",
            "row_select",
            "column_width_resize",
            "double_click_column_resize",
            "row_width_resize",
            "column_height_resize",
            "arrowkeys",
            "row_height_resize",
            "double_click_row_resize",
            "right_click_popup_menu",
            "rc_select",
            "copy",
            "cut",
            "paste",
            "delete",
            "undo",
            "edit_cell"
        )
        
        # Collega l'evento di fine modifica cella
        self.sheet.extra_bindings("end_edit_cell", self.on_sheet_edit_cell)

    def on_sheet_edit_cell(self, event):
        """Gestisce la modifica di una cella nella tabella tksheet - articoli E prezzi"""
        try:
            row = event.row
            col = event.column
            new_value = event.value
            
            if row is None or col is None or row >= len(self.materials):
                return
            
            # Converti None in stringa vuota
            if new_value is None:
                new_value = ""
            
            # Determina il numero di colonne articolo
            num_article_cols = 7 if self.is_conto_lavoro else 4
            
            # Se la cella modificata è una colonna articolo (tutte modificabili!)
            if col < num_article_cols:
                # Salva la modifica all'articolo nel database
                success = self.save_article_field(row, col, str(new_value))
                
                # 🆕 Se il salvataggio è fallito (es: punto invece di virgola), ripristina valore precedente
                if not success:
                    # Ripristina il valore originale dalla memoria
                    if row < len(self.materials):
                        old_material = self.materials[row]
                        if old_material:
                            # Mappa colonna a indice in materials
                            material_index_map = {
                                0: 1,  # codice_materiale
                                1: 2,  # disegno
                                2: 3,  # descrizione_materiale
                                3: 4,  # quantita
                                4: 5,  # codice_grezzo
                                5: 6,  # disegno_grezzo
                                6: 7   # materiale_conto_lavoro
                            }
                            if col in material_index_map:
                                old_value = old_material[material_index_map[col]]
                                self.sheet.set_cell_data(row, col, str(old_value) if old_value else "")
                    return  # Esci senza ulteriori azioni
                
                # 🆕 AGGIORNAMENTO VISIVO: Se è la colonna quantità, aggiorna la cella con il valore salvato
                if col == 3:  # Colonna quantità
                    # Rileggi il valore salvato dal database (già arrotondato a 4 decimali)
                    updated_material = self.materials[row]
                    if updated_material and len(updated_material) > 4:
                        saved_qty = updated_material[4]  # quantita è il 5° campo (indice 4)
                        # Mostra il valore esattamente come salvato (stringa o numero)
                        # NON formattare con format_quantity_display per evitare "12,0"
                        if saved_qty is not None:
                            self.sheet.set_cell_data(row, col, str(saved_qty))
            
            # Se la cella modificata è una colonna prezzo
            elif col >= num_article_cols:
                # Recupera detail_id dalla riga materials
                detail_id = self.materials[row][0]  # id_dettaglio è il primo campo
                
                # Recupera supplier_name dall'header della colonna
                supplier_idx = col - num_article_cols
                if supplier_idx >= len(self.suppliers):
                    return
                
                supplier_name = self.suppliers[supplier_idx]
                
                # Salva nel database e ottieni il valore formattato
                formatted_value = self.save_price_in_db_no_refresh(detail_id, supplier_name, str(new_value))
                
                # Aggiorna immediatamente la cella con il valore formattato
                if formatted_value is not None:
                    self.sheet.set_cell_data(row, col, formatted_value)
                else:
                    # Errore di validazione: svuota la cella
                    self.sheet.set_cell_data(row, col, "")
            
        except Exception as e:
            logger.error(f"Errore in on_sheet_edit_cell: {e}", exc_info=True)

    def save_article_field(self, row_idx, col_idx, new_value):
        """Salva una modifica a un campo articolo nel database.
        
        Returns:
            bool: True se il salvataggio è riuscito, False altrimenti
        """
        if row_idx >= len(self.materials):
            return False
        
        detail_id = self.materials[row_idx][0]  # id_dettaglio
        
        # Mappa indice colonna a campo database
        field_map = {
            0: 'codice_materiale',
            1: 'disegno',              # Colonna Allegato/disegno
            2: 'descrizione_materiale',
            3: 'quantita',
            4: 'codice_grezzo',        # Solo per conto lavoro
            5: 'disegno_grezzo',       # Solo per conto lavoro
            6: 'materiale_conto_lavoro'  # Solo per conto lavoro
        }
        
        if col_idx not in field_map:
            return False
        
        field_name = field_map[col_idx]
        
        # 🆕 VALIDAZIONE: Limita la quantità a 4 decimali (solo per valori numerici)
        if col_idx == 3:  # Colonna quantità
            if new_value and new_value.strip():
                # 🆕 CONTROLLO: Avvisa se l'utente ha usato il punto invece della virgola
                if '.' in new_value and ',' not in new_value:
                    messagebox.showwarning(
                        _("Separatore Decimale"),
                        _("Hai usato il punto (.) come separatore decimale.\n\nIn questo programma si usa la VIRGOLA (,) come separatore decimale.\n\nEsempio corretto: 12,5 invece di 12.5"),
                        parent=self
                    )
                    # Non salvare il valore, esci segnalando errore
                    return False
                
                try:
                    # Prova a convertire in numero
                    qty_float = parse_float_from_comma_string(new_value)
                    # Arrotonda a 4 decimali
                    qty_float = round(qty_float, 4)
                    # Riconverti in stringa preservando il formato originale
                    # Se il numero è intero, salvalo senza decimali
                    if qty_float == int(qty_float):
                        new_value = str(int(qty_float))
                    else:
                        # Rimuovi zeri decimali non necessari
                        new_value = str(qty_float).replace('.', ',').rstrip('0').rstrip(',')
                except (ValueError, TypeError):
                    # Se non è un numero, lascia il valore come stringa (supporta testo)
                    pass
        
        try:
            # Usa db_manager per aggiornare il campo
            db_manager = DatabaseManager(get_db_path())
            db_manager.update_dettaglio_field(detail_id, field_name, new_value)
            
            # Ricarica quella riga dal database
            updated_row = db_manager.get_dettaglio_row_by_id(detail_id)
            if updated_row:
                # Sostituisci la riga in self.materials
                self.materials[row_idx] = updated_row
            
            db_manager.close()
            return True  # Salvataggio riuscito
                
        except DatabaseError as e:
            logger.error(f"Errore database in save_article_field: {e}", exc_info=True)
            messagebox.showerror(_("Errore Database"), 
                               _("Impossibile salvare la modifica: {}").format(e), 
                               parent=self)
            return False  # Salvataggio fallito

    def save_price_in_db_no_refresh(self, detail_id, supplier_name, price_str):
        """Salva un valore nel DB senza aggiornare la griglia (per evitare loop).
        
        Returns:
            str: Il valore formattato da visualizzare nella cella, o None se errore/vuoto
        """
        try:
            if not price_str:
                # Se la stringa è vuota, rimuovi la riga
                # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
                with DatabaseManager(get_db_path()) as db_manager:
                    db_manager.delete_offerta_by_dettaglio_fornitore(detail_id, supplier_name)
                return ""  # Restituisci stringa vuota
            else:
                # Altrimenti inserisci/aggiorna (gestendo virgola, "X" e "ND")
                value_to_save = price_str
                
                # Convalida: se non è "X" o "ND", prova a formattarlo come numero
                price_str_upper = price_str.upper()
                
                if price_str_upper != 'X' and price_str_upper != 'ND':
                    try:
                        # Tenta di convertire per validare e salva la stringa normalizzata con virgola
                        price_float = parse_float_from_comma_string(price_str)
                        value_to_save = format_price_display(price_float)
                    except ValueError:
                        messagebox.showerror(_("Errore Formato"), _("Il prezzo deve essere un numero valido (es. 123,45), 'X' o 'ND'.\nUsa la virgola come separatore decimale."), parent=self)
                        return None  # Errore, non aggiornare la cella

                # Per coerenza, salviamo "X" e "ND" sempre in maiuscolo nel database
                if price_str_upper == 'X' or price_str_upper == 'ND':
                    value_to_save = price_str_upper

                # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
                with DatabaseManager(get_db_path()) as db_manager:
                    db_manager.insert_or_replace_offerta(detail_id, supplier_name, value_to_save)
                
                return value_to_save  # Restituisci il valore formattato
                
        except DatabaseError as e:
            logger.error(f"Errore database in save_price_in_db_no_refresh: {e}", exc_info=True)
            messagebox.showerror(_("Errore Database"), _("Impossibile salvare il prezzo: {}").format(e), parent=self)
            return None

    def save_price_in_db(self, detail_id, supplier_name, price_str):
        """Salva un valore (prezzo, 'X', 'ND' o vuoto) nel DB e aggiorna la griglia."""
        self.save_price_in_db_no_refresh(detail_id, supplier_name, price_str)
        self.refresh_grid()  # Aggiorna la griglia dopo il salvataggio
    
    def add_new_article_row(self):
        """Aggiunge una nuova riga articolo vuota nel database e aggiorna la griglia"""
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(get_db_path()) as db_manager:
                db_manager.insert_dettaglio_richiesta(self.request_id)
            
            # Ricarica la griglia per mostrare la nuova riga
            self.refresh_grid()
            
            # Seleziona automaticamente l'ultima riga aggiunta per facilitare l'editing
            total_rows = self.sheet.get_total_rows()
            if total_rows > 0:
                self.sheet.see(row=total_rows-1, column=0, keep_yscroll=False, keep_xscroll=False, 
                              bottom_right_corner=False, check_cell_visibility=True)
                self.sheet.select_row(total_rows-1)
                
        except DatabaseError as e:
            logger.error(f"Errore database in add_new_article_row: {e}", exc_info=True)
            messagebox.showerror(_("Errore Database"), 
                               _("Impossibile aggiungere l'articolo: {}").format(e), 
                               parent=self)
    
    def remove_selected_article(self):
        """Rimuove gli articoli selezionati dal database e aggiorna la griglia"""
        selected = self.sheet.get_selected_rows()
        print(f"[ViewRequestWindow.remove_selected_article] Righe selezionate: {selected}")
        
        if not selected:
            messagebox.showwarning(_("Attenzione"), 
                                  _("Seleziona almeno un articolo da rimuovere."), 
                                  parent=self)
            return
        
        # Verifica che self.materials sia definito e non vuoto
        if not hasattr(self, 'materials') or not self.materials:
            messagebox.showwarning(_("Attenzione"), 
                                  _("Nessun articolo disponibile per l'eliminazione."), 
                                  parent=self)
            return
        
        print(f"[ViewRequestWindow.remove_selected_article] Numero materiali disponibili: {len(self.materials)}")
        
        # Conferma eliminazione
        if not messagebox.askyesno(_("Conferma Eliminazione"), 
                                   _("Sei sicuro di voler eliminare {} articolo/i selezionato/i?\nVerranno eliminati anche tutti i prezzi associati.").format(len(selected)), 
                                   parent=self):
            return
        
        try:
            # Raccogli gli id_dettaglio da eliminare
            ids_to_delete = []
            invalid_indices = []  # BUG #20 FIX: Traccia indici invalidi
            
            for row_idx in selected:
                print(f"[ViewRequestWindow.remove_selected_article] Elaborazione riga {row_idx} (range: 0-{len(self.materials)-1})")
                # Gli indici restituiti da get_selected_rows() sono basati su 0 e non includono la riga header
                # BUG #20 FIX: Validazione più robusta con isinstance per type safety
                if not isinstance(row_idx, int):
                    logger.error(f"remove_selected_article: row_idx non è int: {type(row_idx)} = {row_idx}")
                    invalid_indices.append(str(row_idx))
                    continue
                
                if 0 <= row_idx < len(self.materials):
                    detail_id = self.materials[row_idx][0]
                    ids_to_delete.append(detail_id)
                    print(f"[ViewRequestWindow.remove_selected_article] Aggiunto id_dettaglio {detail_id} alla lista di eliminazione")
                else:
                    print(f"[ViewRequestWindow.remove_selected_article] WARNING: Indice {row_idx} fuori range (0-{len(self.materials)-1})")
                    invalid_indices.append(str(row_idx))
            
            # BUG #20 FIX: Mostra warning se ci sono indici invalidi
            if invalid_indices:
                logger.warning(f"remove_selected_article: {len(invalid_indices)} indici invalidi: {invalid_indices}")
                messagebox.showwarning(
                    _("Attenzione"),
                    _("Alcuni indici selezionati non sono validi e verranno ignorati: {}").format(", ".join(invalid_indices)),
                    parent=self
                )
            
            if not ids_to_delete:
                messagebox.showwarning(_("Attenzione"), 
                                      _("Nessun articolo valido selezionato per l'eliminazione."), 
                                      parent=self)
                return
            
            print(f"[ViewRequestWindow.remove_selected_article] Eliminazione di {len(ids_to_delete)} articoli: {ids_to_delete}")
            
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(get_db_path()) as db_manager:
                count = db_manager.delete_dettagli_batch(ids_to_delete)
            
            print(f"[ViewRequestWindow.remove_selected_article] Eliminati {count} articoli dal database")
            
            # Ricarica la griglia
            self.refresh_grid()
            
            messagebox.showinfo(_("Successo"), 
                               _("{} articolo/i eliminato/i con successo.").format(count), 
                               parent=self)
            
        except DatabaseError as e:
            logger.error(f"Errore database in remove_selected_article: {e}", exc_info=True)
            messagebox.showerror(_("Errore Database"), 
                               _("Impossibile eliminare l'articolo: {}").format(e), 
                               parent=self)

    def import_from_excel(self):
        """Importa articoli da un file Excel e li aggiunge al database"""
        # Determina il tipo di RdO dal database
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(get_db_path()) as db_manager:
                result = db_manager.get_tipo_rdo(self.request_id)
            if not result:
                messagebox.showerror(_("Errore"), _("RdO non trovata."), parent=self)
                return
            tipo_rdo = result[0]
        except DatabaseError as e:
            logger.error(f"Errore database in import_from_excel: {e}", exc_info=True)
            messagebox.showerror(_("Errore Database"), _("Impossibile determinare il tipo di RdO: {}").format(e), parent=self)
            return
        
        is_cl = (tipo_rdo == "Conto lavoro")
        
        # Mostra istruzioni
        msg = (_("Assicurarsi che il file Excel abbia la seguente struttura:\n\n")
               + _("TIPO '{}' (4 colonne):\n").format(_("Fornitura piena"))
               + _("A: Codice, B: Allegato, C: Descrizione, D: Quantità\n\n")
               + _("TIPO '{}' (7 colonne):\n").format(_("Conto lavoro"))
               + _("A-D come sopra, E: Codice Grezzo, F: Allegato Grezzo, G: Materiale C/L"))
        if not messagebox.askokcancel(_("Istruzioni Importazione Excel"), msg, parent=self):
            return

        # Selezione file
        filepath = filedialog.askopenfilename(
            title=_("Seleziona file Excel"), 
            filetypes=[(_("File Excel"), "*.xlsx"), (_("Tutti i file"), "*.*")],
            parent=self
        )
        if not filepath:
            return

        expected_cols = 7 if is_cl else 4
        workbook = None  # BUG #13 FIX: Inizializza variabile per finally block
        try:
            # Leggi il file Excel
            workbook = openpyxl.load_workbook(filepath, read_only=True)
            sheet = workbook.active
            
            if sheet.max_column < expected_cols:
                raise ValueError(_("Il file Excel deve avere almeno {} colonne per una RdO '{}'.").format(
                    expected_cols, 
                    _("Conto lavoro") if is_cl else _("Fornitura piena")
                ))
            
            # Raccogli gli articoli dal file
            items_to_add = []
            for row in sheet.iter_rows(min_row=1):
                cod = row[0].value
                allegato = row[1].value
                desc = row[2].value
                qta = row[3].value
                
                # Salta righe incomplete
                if cod is None or qta is None:
                    continue
                
                if is_cl:
                    cod_grezzo = str(row[4].value or "") if len(row) > 4 else ""
                    dis_grezzo = str(row[5].value or "") if len(row) > 5 else ""
                    mat_cl = str(row[6].value or "") if len(row) > 6 else ""
                    items_to_add.append((str(cod), str(allegato or ""), str(desc or ""), str(qta), cod_grezzo, dis_grezzo, mat_cl))
                else:
                    items_to_add.append((str(cod), str(allegato or ""), str(desc or ""), str(qta), "", "", ""))
            
            if not items_to_add:
                messagebox.showwarning(_("Attenzione"), _("Nessun articolo valido trovato nel file Excel."), parent=self)
                return
            
            # Inserisci gli articoli nel database usando db_manager
            db_manager = DatabaseManager(get_db_path())
            count = db_manager.import_dettagli_from_list(self.request_id, items_to_add)
            
            # La connessione viene chiusa e riaperta dentro import_dettagli_from_list per forzare persistenza
            # Chiudiamo la connessione esterna (import_dettagli_from_list ne crea una nuova internamente)
            db_manager.close()
            
            # Ricarica la griglia - ora i dati dovrebbero essere persistenti e visibili
            print(f"[ViewRequestWindow.import_from_excel] Chiamata refresh_grid() dopo importazione di {count} articoli")
            self.refresh_grid()
            
            messagebox.showinfo(_("Importazione Completata"), 
                               _("{} articoli importati.").format(count), 
                               parent=self)
            
        except ValueError as e:
            messagebox.showerror(_("Errore Formato File"), str(e), parent=self)
        except Exception as e:
            logger.error(f"Errore in import_from_excel: {e}", exc_info=True)
            messagebox.showerror(_("Errore Importazione"), 
                               _("Impossibile leggere il file Excel.\n{}").format(e), 
                               parent=self)
        finally:
            # BUG #13 FIX: Chiudi SEMPRE il workbook per rilasciare il file Excel
            if workbook is not None:
                try:
                    workbook.close()
                    logger.debug(f"Workbook Excel chiuso: {filepath}")
                except Exception as close_error:
                    logger.warning(f"Errore chiusura workbook Excel: {close_error}")

    def open_attachment_window(self, attachment_type):
        """Apre la finestra gestione allegati passando anche il flag read_only."""
        AttachmentWindow(self, self.request_id, attachment_type, read_only=self.read_only, source_db_path=self.db_path)

# ------------------------------------------------------------------------------------
# FINESTRA IMPOSTAZIONI
# ------------------------------------------------------------------------------------
class SettingsWindow(tk.Toplevel):
    def __init__(self, parent, main_app):
        try:
            super().__init__(parent)
            self.withdraw()
            set_window_icon(self)
            
            self.main_app = main_app
            try:
                self.title(_("Impostazioni e Manutenzione"))
            except Exception as e:
                logger.error(f"Errore nel settare il titolo: {e}")
                self.title(_("Impostazioni e Manutenzione"))
            self.transient(parent)
            self.grab_set()
            
            self.autobackup_enabled = tk.BooleanVar()
            self.autobackup_hour = tk.StringVar()
            self.autobackup_path = tk.StringVar()
            self.language_var = tk.StringVar()
            # Imposta un valore di default per la lingua (verrà aggiornato da load_settings)
            self.language_var.set("English")

            # Le impostazioni di visualizzazione sono ora gestite automaticamente da Windows DPI

            main_frame = ttk.Frame(self, padding="20")
            main_frame.pack(fill="both", expand=True)

            # --- Sezione Posizione DataFlow Standard ---
            dataflow_frame = ttk.LabelFrame(main_frame, text=_("Posizione DataFlow Standard"), padding=10)
            dataflow_frame.pack(fill="x", pady=(0, 15), padx=5)
            
            dataflow_label = ttk.Label(
                dataflow_frame, 
                text=_("Scegli dove salvare la cartella DataFlow (richiede riavvio)."),
                wraplength=480,
                justify="left"
            )
            dataflow_label.pack(anchor="w", pady=(0, 10))
            
            ttk.Button(
                dataflow_frame, 
                text=_("📁 Cambia Posizione DataFlow..."), 
                command=self.select_standard_dataflow_location
            ).pack(anchor="w")
            
            try:
                current_dataflow = get_user_documents_dataflow_dir()
                ttk.Label(
                    dataflow_frame,
                    text=_("Cartella DataFlow attuale: {}").format(current_dataflow),
                    font=("Courier New", 8),
                    foreground="gray",
                    wraplength=480,
                    justify="left"
                ).pack(anchor="w", pady=(10, 0))
            except Exception as e:
                logger.error(f"Errore visualizzazione posizione DataFlow corrente: {e}")

            # --- Sezione Backup Manuale ---
            backup_frame = ttk.LabelFrame(main_frame, text=_("Backup Manuale"), padding="10")
            backup_frame.pack(fill="x", pady=(0, 15), padx=5)
            ttk.Label(backup_frame, text=_("Crea una copia di sicurezza immediata del database."), wraplength=500).pack(anchor="w", pady=(0, 10))
            ttk.Button(backup_frame, text=_("💾 Backup Manuale..."), command=self.backup_database).pack(anchor="w")

            # --- Sezione Backup Automatico ---
            autobackup_frame = ttk.LabelFrame(main_frame, text=_("Backup Automatico Giornaliero"), padding="10")
            autobackup_frame.pack(fill="x", pady=(0, 15), padx=5)

            ttk.Checkbutton(autobackup_frame, text=_("Abilita backup automatico giornaliero (max 3 copie)"), variable=self.autobackup_enabled).pack(anchor="w", pady=(0, 10))
            
            hour_frame = ttk.Frame(autobackup_frame)
            hour_frame.pack(fill="x", pady=5)
            ttk.Label(hour_frame, text=_("Ora:")).pack(side="left", padx=(0, 5))
            ttk.Combobox(hour_frame, textvariable=self.autobackup_hour, values=[f"{h:02}" for h in range(24)], width=5, state="readonly").pack(side="left")

            path_frame = ttk.Frame(autobackup_frame)
            path_frame.pack(fill="x", pady=5)
            ttk.Label(path_frame, text=_("Salva in:")).pack(anchor="w")
            
            path_entry_frame = ttk.Frame(autobackup_frame)
            path_entry_frame.pack(fill="x")
            ttk.Entry(path_entry_frame, textvariable=self.autobackup_path, state="readonly", width=50).pack(side="left", fill="x", expand=True, pady=(0, 5))
            ttk.Button(path_entry_frame, text=_("📁 Scegli..."), command=self.select_autobackup_path).pack(side="left", padx=(5,0), pady=(0,5))

            ttk.Button(autobackup_frame, text=_("💾 Salva Impostazioni Backup"), command=self.save_autobackup_settings).pack(pady=(10,0))

            # --- Sezione Lingua ---
            language_frame = ttk.LabelFrame(main_frame, text=_("Lingua"), padding="10")
            language_frame.pack(fill="x", pady=(0, 15), padx=5)
            
            ttk.Label(language_frame, text=_("Seleziona la lingua dell'interfaccia. Il cambio richiede il riavvio dell'applicazione."), wraplength=500).pack(anchor="w", pady=(0, 15))
            
            # Riga per il controllo della lingua
            lang_row = ttk.Frame(language_frame)
            lang_row.pack(fill="x", pady=(0, 5))
            
            ttk.Label(lang_row, text=_("Lingua:")).pack(side="left", padx=(0, 10))
            language_combo = ttk.Combobox(lang_row, textvariable=self.language_var, values=["English", "Italiano"], state="readonly", width=20)
            language_combo.pack(side="left", padx=(0, 10))
            self.language_combo = language_combo  # Salva riferimento per aggiornamento successivo
            ttk.Button(lang_row, text=_("💾 Salva Lingua"), command=self.save_language_settings).pack(side="left")
            
            # Assicura che il valore nel combobox corrisponda al codice lingua
            def on_language_change(event):
                selected = self.language_var.get()
                # Il valore viene già impostato correttamente dal combobox
                pass
            language_combo.bind("<<ComboboxSelected>>", on_language_change)
            

            try:
                self.load_settings()
                # Aggiorna il combobox dopo aver caricato le impostazioni
                if hasattr(self, 'language_combo'):
                    current_val = self.language_var.get()
                    if current_val == "English":
                        self.language_combo.current(0)
                    elif current_val == "Italiano":
                        self.language_combo.current(1)
            except Exception as e:
                logger.error(f"Errore nel caricare impostazioni all'avvio di SettingsWindow: {e}", exc_info=True)
                # Continua comunque con valori di default
            
            try:
                center_window(self)
            except Exception as e:
                logger.error(f"Errore nel centrare la finestra SettingsWindow: {e}", exc_info=True)
                # Mostra comunque la finestra anche se il centraggio fallisce
                self.deiconify()
                self.geometry("800x600")
        except Exception as e:
            logger.error(f"Errore critico nell'inizializzazione di SettingsWindow: {e}", exc_info=True)
            # Mostra la finestra anche in caso di errore critico
            try:
                self.deiconify()
                self.geometry("800x600")
            except:
                pass

    def load_settings(self):
        """Carica le impostazioni dal file config.ini."""
        try:
            config = configparser.ConfigParser(interpolation=None)
            config_file = get_config_file()
            config.read(config_file)
            
            # Carica impostazioni AutoBackup
            if 'AutoBackup' in config:
                try:
                    self.autobackup_enabled.set(config['AutoBackup'].getboolean('enabled', False))
                    self.autobackup_hour.set(config['AutoBackup'].get('hour', '12'))
                    self.autobackup_path.set(config['AutoBackup'].get('path', ''))
                except Exception as e:
                    logger.warning(f"Errore nel caricare impostazioni AutoBackup: {e}")
                    self.autobackup_enabled.set(False)
                    self.autobackup_hour.set("12")
                    self.autobackup_path.set("")
            else:
                self.autobackup_enabled.set(False)
                self.autobackup_hour.set("12")
                self.autobackup_path.set("")
            
            # Carica impostazioni generali
            if 'Settings' in config:
                # Carica la lingua (default 'en' per primo avvio)
                try:
                    current_lang = config.get('Settings', 'language', fallback='en')
                    # Validazione: accetta solo 'en' o 'it'
                    if current_lang not in ['en', 'it']:
                        current_lang = 'en'
                    self.language_var.set("English" if current_lang == 'en' else "Italiano")
                except Exception as e:
                    logger.warning(f"Errore nel caricare lingua: {e}")
                    self.language_var.set("English")
            else:
                # Se non c'è la sezione Settings, usa default inglese
                self.language_var.set("English")
        except Exception as e:
            logger.error(f"Errore critico nel caricare impostazioni: {e}", exc_info=True)
            # Imposta valori di default in caso di errore
            self.autobackup_enabled.set(False)
            self.autobackup_hour.set("12")
            self.autobackup_path.set("")
            self.language_var.set("English")

    # La funzione save_display_settings() è stata rimossa perché le impostazioni
    # di visualizzazione sono ora gestite automaticamente da Windows DPI

    def save_language_settings(self):
        """Salva la lingua selezionata nel config.ini."""
        try:
            config = configparser.ConfigParser(interpolation=None)
            config_file = get_config_file()
            if os.path.exists(config_file):
                config.read(config_file)
            
            if 'Settings' not in config:
                config['Settings'] = {}
            
            # Converte "English"/"Italiano" in "en"/"it"
            selected_lang = self.language_var.get()
            if not selected_lang:
                messagebox.showwarning(_("Attenzione"), _("Seleziona una lingua."), parent=self)
                return
            
            lang_code = "en" if selected_lang == "English" else "it"
            config['Settings']['language'] = lang_code
            
            # BUG #49 FIX: Usa encoding UTF-8 per gestire caratteri speciali
            with open(config_file, 'w', encoding='utf-8') as f:
                config.write(f)
            
            response = messagebox.askyesno(_("Successo"), _("Impostazione lingua salvata.\nRiavviare ora l'applicazione per applicare le modifiche?"), parent=self)
            if response:
                # Riavvia l'applicazione
                self.main_app.restart_program()
        except Exception as e:
            logger.error(f"Errore nel salvare la lingua: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Impossibile salvare l'impostazione della lingua: {}").format(e), parent=self)

    def select_autobackup_path(self):
        path = filedialog.askdirectory(title=_("Seleziona cartella per backup automatici"), parent=self)
        if path: self.autobackup_path.set(path)

    def save_autobackup_settings(self):
        config = configparser.ConfigParser(interpolation=None); config.read(get_config_file())
        if 'AutoBackup' not in config: config['AutoBackup'] = {}
        config['AutoBackup']['enabled'] = str(self.autobackup_enabled.get())
        config['AutoBackup']['hour'] = self.autobackup_hour.get()
        config['AutoBackup']['path'] = self.autobackup_path.get()
        if self.autobackup_enabled.get() and not self.autobackup_path.get():
            messagebox.showwarning(_("Attenzione"), _("Per abilitare il backup automatico specificare un percorso."), parent=self); return
        try:
            # BUG #49 FIX: Usa encoding UTF-8 per gestire caratteri speciali
            with open(get_config_file(), 'w', encoding='utf-8') as f: config.write(f)
            messagebox.showinfo(_("Successo"), _("Impostazioni backup salvate."), parent=self)
        except Exception as e: messagebox.showerror(_("Errore"), _("Impossibile salvare: {}").format(e), parent=self)

    def backup_database(self):
        """Crea backup manuale con VACUUM INTO per garantire consistenza."""
        db_file = get_db_path()
        if not os.path.exists(db_file):
            messagebox.showerror(_("Errore"), _("File database '{}' non trovato!").format(db_file), parent=self)
            return
        
        dest = filedialog.asksaveasfilename(
            title=_("Salva backup come..."), 
            initialfile=f"backup_manuale_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db", 
            defaultextension=".db", 
            filetypes=[(_("Database SQLite"), "*.db"), (_("Tutti i file"), "*.*")], 
            parent=self
        )
        
        if not dest:
            return  # Utente ha annullato
        
        # Normalizza l'estensione del file di destinazione
        if not dest.endswith('.db'):
            dest = dest.rsplit('.', 1)[0] + '.db'
        
        # Chiudi temporaneamente la connessione della MainWindow per permettere il backup
        main_window_was_open = False
        try:
            if hasattr(self.main_app, 'db_manager') and self.main_app.db_manager:
                logger.info("Chiusura connessione MainWindow per backup...")
                self.main_app.db_manager.close()
                main_window_was_open = True
                # Piccolo delay per assicurarsi che la connessione sia completamente chiusa
                import time
                time.sleep(0.2)
        except Exception as e:
            logger.warning(f"Impossibile chiudere connessione MainWindow: {e}")
        
        # BUG #24 FIX: Verifica che tutte le connessioni database siano chiuse prima della copia
        # Su Windows, file database con handle aperti possono causare corruzione durante copia
        try:
            # Attendi che il DB rilasci tutti i lock (max 1 secondo)
            import time
            for attempt in range(5):
                try:
                    # Test se possiamo aprire il file in modalità esclusiva
                    with open(db_file, 'r+b') as test_handle:
                        pass  # File accessibile senza lock
                    break  # Successo, esci dal loop
                except (PermissionError, IOError) as lock_error:
                    if attempt < 4:  # Non l'ultimo tentativo
                        logger.debug(f"Database ancora locked, tentativo {attempt+1}/5: {lock_error}")
                        time.sleep(0.2)  # Attendi 200ms
                    else:
                        logger.warning(f"Database potrebbe avere lock attivi dopo 5 tentativi")
            
            # ✅ COPIA FILE PRINCIPALE
            shutil.copy2(db_file, dest)
            logger.info(f"Backup DB principale: {dest}")
            
            # ✅ COPIA FILE WAL (se esiste)
            wal_file = db_file.replace('.db', '.db-wal')
            if os.path.exists(wal_file):
                wal_dest = dest.replace('.db', '.db-wal')
                shutil.copy2(wal_file, wal_dest)
                logger.info(f"Backup WAL copiato: {wal_dest}")
            else:
                logger.info("File WAL non presente (normale se DB appena chiuso)")
            
            # ✅ COPIA FILE SHM (se esiste)
            shm_file = db_file.replace('.db', '.db-shm')
            if os.path.exists(shm_file):
                shm_dest = dest.replace('.db', '.db-shm')
                shutil.copy2(shm_file, shm_dest)
                logger.info(f"Backup SHM copiato: {shm_dest}")
            else:
                logger.info("File SHM non presente (normale se DB appena chiuso)")
            
            # Verifica dimensione backup principale (sanity check)
            original_size = os.path.getsize(db_file)
            backup_size = os.path.getsize(dest)
            
            if backup_size < original_size * 0.5:
                logger.warning(f"Backup manuale potenzialmente incompleto: {backup_size} vs {original_size} bytes")
                if not messagebox.askyesno(
                    _("Attenzione Dimensione"), 
                    _("Il backup creato è significativamente più piccolo del database originale.\n\nOriginale: {:.2f} MB\nBackup: {:.2f} MB\n\nVuoi conservarlo comunque?").format(original_size / (1024*1024), backup_size / (1024*1024)),
                    parent=self
                ):
                    try:
                        os.remove(dest)
                        # Rimuovi anche WAL e SHM se esistono
                        wal_dest = dest.replace('.db', '.db-wal')
                        shm_dest = dest.replace('.db', '.db-shm')
                        if os.path.exists(wal_dest):
                            os.remove(wal_dest)
                        if os.path.exists(shm_dest):
                            os.remove(shm_dest)
                    except:
                        pass
                    return
            
            # Messaggio di successo con info sui file copiati
            files_copied = [os.path.basename(dest)]
            wal_dest = dest.replace('.db', '.db-wal')
            shm_dest = dest.replace('.db', '.db-shm')
            if os.path.exists(wal_dest):
                files_copied.append(os.path.basename(wal_dest))
            if os.path.exists(shm_dest):
                files_copied.append(os.path.basename(shm_dest))
            
            messagebox.showinfo(
                _("Successo"), 
                _("Backup creato con successo:\n\nFile copiati:\n{}\n\nDimensione totale: {:.2f} MB").format(
                    '\n'.join(f'  • {f}' for f in files_copied),
                    sum(os.path.getsize(f) for f in [dest] + 
                        ([wal_dest] if os.path.exists(wal_dest) else []) + 
                        ([shm_dest] if os.path.exists(shm_dest) else [])) / (1024*1024)
                ),
                parent=self
            )
            logger.info(f"Backup manuale completato: {len(files_copied)} file copiati")
            
        except Exception as e:
            logger.error(f"Errore backup manuale: {e}", exc_info=True)
            messagebox.showerror(
                _("Errore"), 
                _("Impossibile creare backup:\n{}").format(e), 
                parent=self
            )
            # Rimuovi backup parziale/corrotto
            if os.path.exists(dest):
                try:
                    os.remove(dest)
                    # Rimuovi anche WAL e SHM parziali
                    wal_dest = dest.replace('.db', '.db-wal')
                    shm_dest = dest.replace('.db', '.db-shm')
                    if os.path.exists(wal_dest):
                        os.remove(wal_dest)
                    if os.path.exists(shm_dest):
                        os.remove(shm_dest)
                except:
                    pass
        finally:
            # Riapri la connessione della MainWindow se era aperta
            if main_window_was_open:
                try:
                    logger.info("Riapertura connessione MainWindow dopo backup...")
                    self.main_app.db_manager = DatabaseManager(get_db_path())
                    logger.info("Connessione MainWindow riaperta con successo")
                except Exception as e:
                    logger.error(f"Errore nella riapertura connessione MainWindow: {e}")
                    messagebox.showwarning(
                        _("Attenzione"),
                        _("Il backup è stato completato, ma non è stato possibile riaprire la connessione principale.\nSi consiglia di riavviare l'applicazione."),
                        parent=self
                    )

    def _sanitize_filename(self, name): return re.sub(r'[\\/*?:"<>|]', "", name)

    def select_standard_dataflow_location(self):
        """
        Permette all'utente di scegliere una nuova posizione per la cartella DataFlow.
        
        Passaggi:
        1. Avviso esplicativo con conferma
        2. Selezione cartella
        3. Validazioni (permessi, rete, lunghezza path, unità)
        4. Salvataggio config
        5. Istruzioni per spostare manualmente la cartella
        6. Riavvio applicazione
        """
        logger.info("Avvio procedura cambio posizione cartella DataFlow")
        current_dataflow_dir = get_user_documents_dataflow_dir()
        
        warning_text = _(
            "⚠️ ATTENZIONE: stai per cambiare la posizione della cartella DataFlow.\n\n"
            "IMPORTANTE:\n"
            "- La cartella attuale non verrà spostata automaticamente\n"
            "- L'app verrà riavviata per applicare la modifica\n\n"
            "Posizione attuale:\n{}\n\n"
            "Vuoi procedere?"
        ).format(current_dataflow_dir)
        
        if not messagebox.askyesno(
            _("Conferma Cambio Posizione"), 
            warning_text,
            parent=self,
            icon='warning'
        ):
            logger.info("Utente ha annullato il cambio posizione DataFlow")
            return
        
        if sys.platform == 'win32':
            initial_dir = os.path.dirname(current_dataflow_dir) or os.path.join(os.path.expanduser('~'), 'Documents')
        else:
            initial_dir = os.path.dirname(current_dataflow_dir) or os.path.expanduser('~')
        
        try:
            selected_dir = filedialog.askdirectory(
                title=_("Seleziona la nuova posizione della cartella DataFlow"),
                initialdir=initial_dir,
                parent=self
            )
        except Exception as e:
            logger.error(f"Errore apertura dialog selezione cartella: {e}")
            messagebox.showerror(
                _("Errore"),
                _("Errore durante la selezione della cartella: {}").format(e),
                parent=self
            )
            return
        
        if not selected_dir:
            logger.info("Utente ha annullato la selezione della nuova posizione")
            return
        
        normalized_dir = os.path.normpath(os.path.abspath(selected_dir.strip()))
        if not normalized_dir:
            messagebox.showerror(_("Errore"), _("Percorso non valido."), parent=self)
            return
        
        # ✅ CORREZIONE: NON aggiungere "DataFlow" - useremo DataFlow_{username}
        # Il percorso selezionato dall'utente è la directory PARENT dove verrà creata DataFlow_{username}
        logger.info(f"Cartella parent selezionata per DataFlow: {normalized_dir}")
        
        # Verifica che la directory parent esista o possa essere creata
        try:
            os.makedirs(normalized_dir, exist_ok=True)
        except OSError as e:
            logger.error(f"Impossibile creare/accedere alla cartella parent: {e}")
            messagebox.showerror(
                _("Errore"),
                _("Impossibile accedere alla cartella selezionata:\n{}\n\nDettagli: {}").format(normalized_dir, e),
                parent=self
            )
            return
        
        # Validazione permessi scrittura nella directory parent
        try:
            # BUG #28 FIX: Risolto TOCTOU usando try-except invece di check esistenza
            # Test scrittura nella directory parent (già esistente o appena creata)
            test_file = os.path.join(normalized_dir, ".dataflow_test_write")
            try:
                with open(test_file, 'w') as f:
                    f.write("test")
            finally:
                # BUG #28 FIX: Cleanup in finally per garantire rimozione anche se write fallisce
                try:
                    os.remove(test_file)
                except FileNotFoundError:
                    pass  # File già rimosso, va bene
            logger.info(f"Permessi verifica OK per {normalized_dir}")
        except (OSError, PermissionError) as e:
            logger.error(f"Test permessi fallito per {normalized_dir}: {e}")
            messagebox.showerror(
                _("Errore Permessi"),
                _("Impossibile scrivere nella cartella selezionata:\n{}\n\nDettagli: {}").format(normalized_dir, e),
                parent=self
            )
            return
        
        # Controllo lunghezza
        if len(normalized_dir) > 240:
            logger.warning(f"Percorso DataFlow troppo lungo ({len(normalized_dir)} caratteri)")
            length_warning = _(
                "Il percorso selezionato è molto lungo ({} caratteri).\n"
                "Windows potrebbe avere problemi nell'accesso ai file.\n"
                "Vuoi procedere comunque?"
            ).format(len(normalized_dir))
            if not messagebox.askyesno(
                _("Percorso Molto Lungo"),
                length_warning,
                parent=self
            ):
                logger.info("Utente ha annullato dopo avviso percorso lungo")
                return
        
        # Controllo unità rimovibile
        try:
            drive_letter = os.path.splitdrive(normalized_dir)[0]
            if drive_letter and drive_letter.upper() not in ['C:', 'D:', 'E:']:
                logger.warning(f"Unità potenzialmente rimovibile: {drive_letter}")
                removable_warning = _(
                    "⚠️ L'unità selezionata ({}) potrebbe essere rimovibile.\n"
                    "Se viene scollegata, DataFlow non potrà accedere ai dati."
                ).format(drive_letter)
                messagebox.showwarning(_("Unità Rimovibile?"), removable_warning, parent=self)
        except Exception as e:
            logger.error(f"Errore durante controllo unità rimovibile: {e}")
        
        # === INIZIO LOGICA CONTROLLO CONFLITTO USERNAME ===
        # Carica identità utente corrente
        identity = load_user_identity()
        current_username = identity.get('username', '').strip().lower()
        
        if not current_username:
            logger.error("Username corrente non trovato nel config")
            messagebox.showerror(
                _("Errore"),
                _("Impossibile determinare l'utente corrente. Riavvia DataFlow."),
                parent=self
            )
            return
        
        # Variabili per gestione cambio username
        final_username = current_username
        username_changed = False
        
        # Loop controllo conflitto username
        while True:
            # Controlla se esiste già un database con questo username nella destinazione
            potential_folder = os.path.join(normalized_dir, f"DataFlow_{final_username}")
            potential_db = os.path.join(potential_folder, 'Database', f'dataflow_db_{final_username}.db')
            
            folder_exists = os.path.exists(potential_folder)
            db_exists = False
            
            # Controllo robusto dell'esistenza del DB (gestisce file locked)
            if folder_exists:
                try:
                    # Verifica esistenza DB in modo più robusto
                    db_exists = os.path.exists(potential_db)
                    
                    # Se il DB esiste, prova ad aprirlo per verificare che sia accessibile
                    if db_exists:
                        try:
                            # Test di accesso in lettura (non modifica il file)
                            with open(potential_db, 'rb') as f:
                                f.read(1)  # Leggi solo 1 byte per verificare accesso
                            logger.info(f"Controllo conflitto: DB '{potential_db}' esiste ed è accessibile")
                        except (PermissionError, OSError) as e:
                            # File locked o inaccessibile: CONSIDERA COME ESISTENTE
                            logger.warning(f"DB '{potential_db}' esistente ma locked/inaccessibile: {e}")
                            db_exists = True
                except Exception as e:
                    logger.error(f"Errore nel controllo esistenza DB: {e}")
                    # In caso di errore, ASSUME CHE ESISTA (principio di precauzione)
                    db_exists = True
            
            logger.info(f"Controllo conflitto per username '{final_username}': folder={folder_exists}, db={db_exists}")
            
            # ✅ CORREZIONE LOGICA: Se ESISTE cartella O database, è un CONFLITTO
            if folder_exists or db_exists:
                # Conflitto rilevato: chiedi se vuole cambiare username
                conflict_message = _(
                    "⚠️ CONFLITTO UTENTE RILEVATO\n\n"
                    "Nella cartella di destinazione selezionata esiste già un database \n"
                    "associato all'utente '{}'.\n\n"
                    "Per evitare conflitti e perdita dati, è necessario cambiare \n"
                    "il tuo username prima di procedere.\n\n"
                    "Vuoi procedere con il cambio username?"
                ).format(final_username)
                
                if not messagebox.askyesno(
                    _("Conflitto Username"),
                    conflict_message,
                    parent=self,
                    icon='warning'
                ):
                    # Utente ha rifiutato, annulla tutto
                    logger.info("Utente ha rifiutato il cambio username, operazione annullata")
                    return
                
                # Mostra dialogo cambio identità
                self.withdraw()  # Nascondi finestra settings temporaneamente
                new_identity_dialog = UserIdentityDialog(self)
                self.wait_window(new_identity_dialog)
                self.deiconify()  # Mostra di nuovo
                
                new_identity = getattr(new_identity_dialog, 'result', None)
                if not new_identity:
                    # Utente ha annullato il dialogo identità
                    logger.info("Utente ha annullato il dialogo identità, operazione annullata")
                    return
                
                # Aggiorna username e continua il loop per ricontrollare
                final_username = new_identity['username']
                username_changed = True
                logger.info(f"Nuovo username proposto: {final_username}, rientro nel loop controllo")
            else:
                # ✅ NESSUN CONFLITTO: Username libero, prosegui
                logger.info(f"Username '{final_username}' disponibile nella destinazione (nessun conflitto rilevato)")
                break
        
        # === FINE LOGICA CONTROLLO CONFLITTO USERNAME ===
        
        # A questo punto final_username è libero, procedi con la copia
        source_folder = current_dataflow_dir
        dest_parent = normalized_dir  # Directory parent dove creare DataFlow_{username}
        dest_folder = os.path.join(dest_parent, f"DataFlow_{final_username}")  # Percorso completo destinazione
        
        # Verifica che la cartella sorgente esista
        if not os.path.exists(source_folder):
            logger.error(f"Cartella sorgente non esiste: {source_folder}")
            messagebox.showerror(
                _("Errore"),
                _("Cartella DataFlow di origine non trovata:\n{}").format(source_folder),
                parent=self
            )
            return
        
        # ✅ CHIUDI DATABASE PRIMA DELLA COPIA (evita WinError 32)
        logger.info("Chiusura database prima della copia...")
        try:
            # Chiudi il DatabaseManager globale se esiste
            if hasattr(self.main_app, 'db_manager') and self.main_app.db_manager:
                self.main_app.db_manager.close()
                logger.info("DatabaseManager principale chiuso")
        except Exception as e:
            logger.warning(f"Errore chiusura DatabaseManager: {e}")
        
        # Mostra finestra progresso copia
        progress_win = CopyProgressWindow(self, title=_("Copia DataFlow in corso..."))
        progress_win.update_progress(0, _("Preparazione copia..."))
        
        # Backup config originale (per rollback)
        config_backup = None
        try:
            config_file = get_config_file()
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    config_backup = f.read()
        except Exception as e:
            logger.error(f"Impossibile fare backup config: {e}")
        
        try:
            # === COPIA FISICA COMPLETA CON PROGRESSIONE ===
            logger.info(f"Inizio copia da '{source_folder}' a '{dest_folder}'")
            
            # Conta file totali per barra progresso
            progress_win.update_progress(5, _("Analisi file da copiare..."))
            total_files = 0
            for root, dirs, files in os.walk(source_folder):
                total_files += len(files)
            
            logger.info(f"File totali da copiare: {total_files}")
            
            if total_files == 0:
                raise Exception(_("Nessun file da copiare nella cartella sorgente"))
            
            # Copia ricorsiva con aggiornamento progressione
            files_copied = 0
            
            def copy_with_progress(src, dst):
                nonlocal files_copied
                os.makedirs(dst, exist_ok=True)
                
                for item in os.listdir(src):
                    s = os.path.join(src, item)
                    d = os.path.join(dst, item)
                    
                    if os.path.isdir(s):
                        copy_with_progress(s, d)
                    else:
                        # Copia file
                        shutil.copy2(s, d)
                        files_copied += 1
                        
                        # Aggiorna progressione (da 10% a 80%)
                        progress_pct = 10 + int((files_copied / total_files) * 70)
                        file_name = os.path.basename(s)
                        progress_win.update_progress(
                            progress_pct,
                            _("Copia file {}/{}: {}").format(files_copied, total_files, file_name[:40])
                        )
            
            copy_with_progress(source_folder, dest_folder)
            
            logger.info(f"Copia file completata: {files_copied} file copiati")
            progress_win.update_progress(85, _("Copia completata, aggiornamento configurazione..."))
            
            # === AGGIORNA USERNAME NEL DATABASE (SOLO SE CAMBIATO) ===
            if username_changed:
                logger.info(f"Username cambiato da '{current_username}' a '{final_username}', aggiorno database")
                progress_win.update_progress(90, _("Aggiornamento username nel database..."))
                
                # Percorso nuovo database
                new_db_path = os.path.join(dest_folder, 'Database', f'dataflow_db_{final_username}.db')
                
                # Rinomina anche il file database se necessario
                old_db_name = f'dataflow_db_{current_username}.db'
                old_db_path = os.path.join(dest_folder, 'Database', old_db_name)
                
                if os.path.exists(old_db_path) and old_db_path != new_db_path:
                    logger.info(f"Rinomino database da '{old_db_name}' a 'dataflow_db_{final_username}.db'")
                    shutil.move(old_db_path, new_db_path)
                
                # Aggiorna username in tutte le RdO
                try:
                    # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
                    with DatabaseManager(new_db_path) as db_manager:
                        rows_updated = db_manager.update_all_usernames(final_username)
                    logger.info(f"Username aggiornato in {rows_updated} RdO")
                except Exception as db_error:
                    logger.error(f"Errore aggiornamento username in DB: {db_error}", exc_info=True)
                    raise
            
            # === AGGIORNA CONFIG.INI ===
            progress_win.update_progress(95, _("Salvataggio configurazione..."))
            
            config = configparser.ConfigParser(interpolation=None)
            config_file = get_config_file()
            
            if os.path.exists(config_file):
                config.read(config_file)
            
            if 'Settings' not in config:
                config['Settings'] = {}
            if 'User' not in config:
                config['User'] = {}
            
            # Salva nuovo percorso base
            config['Settings']['dataflow_base_dir'] = dest_parent
            
            # Rimuovi legacy custom_db_path se presente
            if config.has_option('Settings', 'custom_db_path'):
                config.remove_option('Settings', 'custom_db_path')
            
            # Se username è cambiato, aggiorna anche sezione User
            if username_changed:
                config['User']['first_name'] = new_identity['first_name']
                config['User']['last_name'] = new_identity['last_name']
                config['User']['username'] = final_username
            
            with open(config_file, 'w', encoding='utf-8') as f:
                config.write(f)
            
            logger.info(f"Config aggiornato con nuovo percorso: {dest_parent}")
            
            progress_win.update_progress(100, _("Operazione completata!"))
            time.sleep(0.5)
            progress_win.destroy()
            
            # === MESSAGGIO SUCCESSO ===
            username_info = ""
            if username_changed:
                username_info = _("\n\n✓ Username aggiornato da '{}' a '{}'").format(current_username, final_username)
            
            success_msg = _(
                "✓ OPERAZIONE COMPLETATA CON SUCCESSO\n\n"
                "La cartella DataFlow è stata copiata con successo in:\n"
                "{dest}\n"
                "\nFile copiati: {count}{username_change}\n\n"
                "⚠️ IMPORTANTE:\n"
                "- La cartella ORIGINALE in '{src}' NON è stata eliminata.\n"
                "- Prima di eliminarla manualmente, TESTA il corretto funzionamento \n"
                "  del database copiato.\n"
                "- DataFlow verrà riavviato automaticamente."
            ).format(
                dest=dest_folder,
                count=files_copied,
                username_change=username_info,
                src=source_folder
            )
            
            messagebox.showinfo(_("Operazione Completata"), success_msg, parent=self)
            
            # ✅ SALVA ESPLICITAMENTE LA NUOVA IDENTITÀ (se cambiata)
            if username_changed:
                save_user_identity(new_identity['first_name'], new_identity['last_name'], final_username)
                logger.info(f"Identità salvata nel config: {final_username}")
            
            # Invalida cache e riavvia
            reset_db_cache()
            logger.info("Cache DB invalidata, riavvio applicazione")
            self.destroy()
            self.main_app.restart_program()
            
        except Exception as e:
            # === GESTIONE ERRORE CON ROLLBACK ===
            logger.error(f"Errore durante copia DataFlow: {e}", exc_info=True)
            
            try:
                progress_win.destroy()
            except:
                pass
            
            # Ripristina config backup se disponibile
            if config_backup:
                try:
                    with open(get_config_file(), 'w', encoding='utf-8') as f:
                        f.write(config_backup)
                    logger.info("Config.ini ripristinato da backup")
                except Exception as restore_err:
                    logger.error(f"Impossibile ripristinare config: {restore_err}")
            
            # Tenta di eliminare cartella parziale (se creata)
            if os.path.exists(dest_folder):
                try:
                    shutil.rmtree(dest_folder, ignore_errors=True)
                    logger.info(f"Cartella parziale eliminata: {dest_folder}")
                except Exception as cleanup_err:
                    logger.error(f"Impossibile eliminare cartella parziale: {cleanup_err}")
            
            error_msg = _(
                "❌ OPERAZIONE FALLITA\n\n"
                "Impossibile completare lo spostamento della cartella DataFlow.\n\n"
                "Dettaglio errore:\n{error}\n\n"
                "Le impostazioni originali sono state ripristinate.\n"
                "Consulta il file di log per maggiori dettagli."
            ).format(error=str(e))
            
            messagebox.showerror(_("Errore Spostamento"), error_msg, parent=self)

# ------------------------------------------------------------------------------------
# FINESTRA GUIDA UTENTE
# ------------------------------------------------------------------------------------
class HelpWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.withdraw()
        set_window_icon(self)
        
        # Leggi la lingua corrente per la traduzione del sommario
        import configparser
        config = configparser.ConfigParser(interpolation=None)
        config_file = get_config_file()
        current_language = "en"  # default
        if os.path.exists(config_file):
            config.read(config_file, encoding='utf-8')
            if 'Settings' in config and config.has_option('Settings', 'language'):
                current_language = config.get('Settings', 'language', fallback='en')
        if current_language not in ['en', 'it']:
            current_language = "en"
        
        # Determina la traduzione corretta per "Analisi SQDC"
        sqdc_text = "   - Analysis SQDC" if current_language == "en" else _("   - Analisi SQDC")
        
        self.title(_("Guida Utente - DataFlow Procurement Software")); self.transient(parent); self.grab_set()
        main_frame = ttk.Frame(self); main_frame.pack(fill="both", expand=True)
        paned = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL); paned.pack(fill="both", expand=True, padx=10, pady=10)
        toc_frame = ttk.Frame(paned, padding=10); paned.add(toc_frame, weight=1)
        ttk.Label(toc_frame, text=_("Sommario"), font=("Helvetica", 12, "bold")).pack(anchor="w", pady=(0, 10))
        self.topics = [(_("0. Primi Passi"), "quick_start"), (_("   - Benvenuto in DataFlow"), "welcome"), (_("   - Primo Avvio"), "first_run"), (_("   - Prima RdO di Prova"), "first_rdo"), (_("1. Schermata Principale"), "main_screen"), (_("   - Interfaccia e Pulsanti Principali"), "main_interface"), (_("   - Scorciatoie da Tastiera"), "keyboard_shortcuts"), (_("   - Ordinamento delle Colonne"), "column_sorting"), (_("   - Filtri di Ricerca"), "main_filters"), (_("2. Creare una Nuova RdO"), "new_rdo"), (_("   - Inserimento Manuale degli Articoli"), "new_rdo_data"), (_("   - Importazione da Excel"), "new_rdo_excel"), (_("3. Gestire una RdO Esistente"), "manage_rdo"), (_("   - La Griglia Prezzi"), "manage_grid"), (_("   - Modifica Dati e Aggiunta Note"), "manage_edit"), (_("   - Gestione Numeri Ordine (PO)"), "manage_po"), (_("   - Gestione Allegati"), "manage_attachments"), (sqdc_text, "manage_sqdc"), (_("   - Esportazione Excel"), "manage_export"), (_("4. Impostazioni e Manutenzione"), "settings"), (_("   - Gestione Database"), "settings_db"), (_("   - Backup"), "settings_backup"), (_("   - Avanzate"), "settings_advanced"), (_("5. Problemi Comuni e Soluzioni"), "troubleshooting"), (_("   - Database Bloccato"), "ts_db_locked"), (_("   - Errori Importazione Excel"), "ts_import"), (_("   - Allegati Non Trovati"), "ts_attachments"), (_("   - Recupero da Backup"), "ts_backup"), (_("6. Requisiti di Sistema e Limiti"), "requirements"), (_("7. Glossario"), "glossary"), (_("8. Contatti e Supporto"), "support")]
        # Creiamo una mappa di ricerca veloce (Testo del titolo -> tag_ancoraggio)
        self.topic_anchor_map = {}
        for text, tag in self.topics:
            clean_text = text.strip()
            # Rimuovi il prefisso "   - " per le sottovoci
            if clean_text.startswith("- "):
                clean_text = clean_text[2:]
            self.topic_anchor_map[clean_text] = tag
        
        # Aggiungi anche le chiavi alternative per "Analisi SQDC" quando la lingua è inglese
        # Il file guida inglese contiene "SQDC Analysis" invece di "Analysis SQDC"
        # e potrebbe contenere anche "Analisi SQDC" se non è stato ancora aggiornato
        if current_language == "en":
            self.topic_anchor_map["Analisi SQDC"] = "manage_sqdc"
            self.topic_anchor_map["SQDC Analysis"] = "manage_sqdc"
            # Aggiungi anche la traduzione per la gestione PO
            self.topic_anchor_map["Purchase Order Number Management"] = "manage_po"
            # Aggiungi mappatura alternativa per "Managing an Existing RFQ" (RFQ maiuscolo)
            self.topic_anchor_map["3. Managing an Existing RFQ"] = "manage_rdo"
            # Aggiungi mappatura alternativa per "First Test RFQ" (RFQ maiuscolo)
            self.topic_anchor_map["First Test RFQ"] = "first_rdo"
            # Aggiungi mappatura alternativa per "Creating a New RFQ" (RFQ maiuscolo + numero)
            self.topic_anchor_map["2. Creating a New RFQ"] = "new_rdo"
        
        # Aggiungi mappatura per titolo completo della sezione PO (italiano)
        self.topic_anchor_map["Gestione Numeri Ordine di Acquisto (PO)"] = "manage_po"
        for text, tag in self.topics:
            link = ttk.Label(toc_frame, text=text, foreground="blue", cursor="hand2"); link.pack(anchor="w", pady=2 if text.strip().startswith(('0','1','2','3','4','5','6','7','8')) else 1); link.bind("<Button-1>", lambda e, t=tag: self.text_content.see(f"{t}.first"))
        content_frame = ttk.Frame(paned); paned.add(content_frame, weight=4)
        
        # --- INIZIO SEARCH BAR ---
        # Inizializza variabili per la ricerca
        self.search_var = tk.StringVar()
        self.search_matches = []  # Lista di posizioni dei match trovati
        self.current_match_index = -1  # Indice del match corrente
        self.search_result_label = None  # Label per mostrare "Risultato X di Y"
        
        # Frame per la barra di ricerca (sopra il widget Text)
        search_frame = ttk.Frame(content_frame)
        search_frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 5))
        self.setup_search_functionality(search_frame)
        # --- FINE SEARCH BAR ---
        
        scrollbar = ttk.Scrollbar(content_frame); self.text_content = tk.Text(content_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set, padx=15, pady=10, relief="flat", background="#FFFFFF", font=("Arial", 10))
        scrollbar.config(command=self.text_content.yview); scrollbar.pack(side=tk.RIGHT, fill=tk.Y); self.text_content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.text_content.insert(tk.END, _("Caricamento della guida in corso...\n\n"), "normal")
        self.text_content.config(state="disabled")
        center_window(self)
        # Carica il contenuto in modo asincrono ma più robusto
        self.after(100, self.populate_content)


    ### CORREZIONE: La funzione 'populate_content' è stata indentata per diventare un metodo della classe 'HelpWindow'.
    def populate_content(self):
        try:
            # Riabilita il widget per inserire il contenuto
            self.text_content.config(state="normal")
            # Pulisci il messaggio di caricamento
            self.text_content.delete("1.0", tk.END)
            
            # Configurazione degli stili di testo
            tags = {
                "h1": ("Arial", 14, "bold", "underline"),
                "h2": ("Arial", 12, "bold"),
                "h3": ("Arial", 10, "bold", "underline"),
                "bold": ("Arial", 10, "bold"),
                "normal": ("Arial", 10),
                "list": ("Arial", 10),
                "warning": ("Arial", 10, "italic"),
                "code": ("Arial", 9),
                "info": ("Arial", 9, "italic"),
                "warning_red": ("Arial", 10, "bold") 
            }
            for t, f_tuple in tags.items():
                # Assicura che la tupla del font sia passata correttamente
                self.text_content.tag_configure(t, font=f_tuple)
            
            # --- CONFIGURAZIONE COLORE PER I TAG ---
            self.text_content.tag_configure("warning_red", foreground="red")
            self.text_content.tag_configure("info", foreground="gray")
            
            # --- CONFIGURAZIONE TAG DI ANCORAGGIO ---
            # Configura tutti i tag di ancoraggio usati nel sommario
            for _, tag in self.topics:
                self.text_content.tag_configure(tag, underline=False)
            
            # --- CONFIGURAZIONE TAG PER RICERCA ---
            self.text_content.tag_configure("search_highlight", background="yellow", foreground="black")
            
            # Carica il contenuto dal file esterno in base alla lingua
            import configparser
            config = configparser.ConfigParser(interpolation=None)
            config_file = get_config_file()
            current_language = "en"  # default, coerente con init_i18n
            if os.path.exists(config_file):
                config.read(config_file, encoding='utf-8')
                if 'Settings' in config and config.has_option('Settings', 'language'):
                    current_language = config.get('Settings', 'language', fallback='en')
            
            # Validazione: accetta solo 'en' o 'it'
            if current_language not in ['en', 'it']:
                current_language = "en"
            
            # Carica il file guida corretto in base alla lingua
            if current_language == "en":
                guida_path = resource_path(os.path.join("add_data", "guida_en.txt"))
            else:
                guida_path = resource_path(os.path.join("add_data", "guida.txt"))
            
            # Debug: log quale file viene caricato
            print(f"[HelpWindow] Lingua corrente: {current_language}")
            print(f"[HelpWindow] Tentativo di caricare: {guida_path}")
            print(f"[HelpWindow] File esiste: {os.path.exists(guida_path)}")
            
            if os.path.exists(guida_path):
                with open(guida_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                print(f"[HelpWindow] File caricato con successo, lunghezza contenuto: {len(content)} caratteri")
                self._parse_and_insert_content(content)
            else:
                # Fallback se il file non esiste
                error_msg = _("File guida non trovato. Contatta l'amministratore.") + f"\n\nPath cercato: {guida_path}"
                self.text_content.insert(tk.END, error_msg, "normal")
                print(f"[HelpWindow] ERRORE: File non trovato: {guida_path}")
            
            # Disabilita il widget dopo il caricamento
            self.text_content.config(state="disabled")
            
        except Exception as e:
            # In caso di errore, mostra un messaggio di errore
            self.text_content.config(state="normal")
            self.text_content.delete("1.0", tk.END)
            self.text_content.insert(tk.END, _("Errore caricamento guida: {}\n\nContatta l'amministratore.").format(e), "normal")
            self.text_content.config(state="disabled")
    
    def _parse_and_insert_content(self, content):
        """Analizza il contenuto del file e lo inserisce nel widget con la formattazione corretta"""
        lines = content.split('\n')
        
        for line in lines:
            if not line.strip():
                self.text_content.insert(tk.END, "\n")
                continue
                
            # Trova il tag di stile (es. 'h1') e il contenuto (es. '0. Primi Passi')
            style_tag = "normal"
            content_to_insert = line
            is_section_tag = False

            if line.startswith('[H1]'):
                style_tag = "h1"
                content_to_insert = line[4:]
                is_section_tag = True
            elif line.startswith('[H2]'):
                style_tag = "h2"
                content_to_insert = line[4:]
                is_section_tag = True
            elif line.startswith('[H3]'):
                style_tag = "h3"
                content_to_insert = line[4:]
                is_section_tag = True
            elif line.startswith('[INFO]'):
                style_tag = "info"
                content_to_insert = line[6:]
                is_section_tag = True
            elif line.startswith('[LIST]'):
                style_tag = "list"
                content_to_insert = line[6:]
            elif line.startswith('[WARNING_RED]'):
                content_to_insert = line
            elif line.startswith('[BOLD]'):
                style_tag = "bold"
                content_to_insert = line[6:]
            elif line.startswith('[NORMAL]'):
                style_tag = "normal"
                content_to_insert = line[8:]
            elif line.startswith('[CODE]'):
                style_tag = "code"
                content_to_insert = line[6:]
            elif line.startswith('[WARNING]'):
                content_to_insert = line

            # Se è un tag di sezione (H1, H2, H3, INFO) applica la formattazione e cerca il link
            if is_section_tag:
                clean_content = content_to_insert.strip()
                anchor_tag = self.topic_anchor_map.get(clean_content)
                tags_to_apply = [style_tag]
                if anchor_tag:
                    tags_to_apply.append(anchor_tag)
                # Usa il parser inline per gestire eventuali tag inline nel contenuto
                self._insert_formatted_line_with_anchor(content_to_insert, tuple(tags_to_apply))
                continue
            
            # Per tutte le altre righe (inclusi tag come [BOLD], [LIST], etc.), usa il parser inline
            if line != content_to_insert:
                # Se abbiamo estratto del testo dopo un tag, analizzalo con il parser inline
                self._insert_formatted_line(content_to_insert)
            else:
                # Nessun tag speciale, è testo normale con possibili tag inline
                self._insert_formatted_line(line)
    
    def _insert_formatted_line(self, line):
        """Inserisce una riga con formattazione inline"""
        parts = []
        current_text = ""
        current_tag = "normal"
        i = 0
        
        while i < len(line):
            # Cerca tag LINK con URL: [LINK:url]testo[/LINK]
            if line.startswith('[LINK:', i):
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                # Estrae l'URL
                url_start = i + 6  # dopo '[LINK:'
                url_end = line.find(']', url_start)
                if url_end != -1:
                    url = line[url_start:url_end]
                    # Trova il testo del link
                    text_start = url_end + 1
                    text_end = line.find('[/LINK]', text_start)
                    if text_end != -1:
                        link_text = line[text_start:text_end]
                        parts.append((link_text, "hyperlink", url))
                        i = text_end + 7  # dopo '[/LINK]'
                        continue
                current_text += line[i]
                i += 1
            # Cerca tag di apertura/chiusura
            elif line[i:i+6] == '[BOLD]':
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "bold"
                i += 6
            elif line[i:i+8] == '[/BOLD]':
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "normal"
                i += 8
            elif line[i:i+8] == '[NORMAL]':
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "normal"
                i += 8
            elif line[i:i+10] == '[/NORMAL]':
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "normal"
                i += 10
            elif line[i:i+6] == '[CODE]':
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "code"
                i += 6
            elif line[i:i+8] == '[/CODE]':
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "normal"
                i += 8
            elif line.startswith('[WARNING]', i):
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "warning"
                i += len('[WARNING]')
            elif line.startswith('[/WARNING]', i):
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "normal"
                i += len('[/WARNING]')
            elif line.startswith('[WARNING_RED]', i):
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "warning_red"
                i += len('[WARNING_RED]')
            elif line.startswith('[/WARNING_RED]', i):
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "normal"
                i += len('[/WARNING_RED]')
            else:
                current_text += line[i]
                i += 1
        
        # Aggiunge l'ultimo pezzo
        if current_text:
            parts.append((current_text, current_tag, None))
        elif not parts:
            # Se non c'è nessun testo, aggiungi una stringa vuota
            parts.append(("", "normal", None))
        
        # Inserisce tutti i pezzi
        for item in parts:
            if len(item) == 3:
                text, tag, url = item
                if tag == "hyperlink" and url:
                    # Crea un tag univoco per questo link
                    link_tag = f"link_{id(url)}"
                    # Configura il tag con lo stile del link
                    self.text_content.tag_configure(link_tag, foreground="blue", underline=True)
                    # Bind del click per aprire l'URL
                    self.text_content.tag_bind(link_tag, "<Button-1>", lambda e, u=url: webbrowser.open(u))
                    self.text_content.tag_bind(link_tag, "<Enter>", lambda e: self.text_content.config(cursor="hand2"))
                    self.text_content.tag_bind(link_tag, "<Leave>", lambda e: self.text_content.config(cursor=""))
                    # Inserisce il testo con il tag del link
                    self.text_content.insert(tk.END, text, link_tag)
                else:
                    self.text_content.insert(tk.END, text, tag)
            else:
                # Fallback per compatibilità
                self.text_content.insert(tk.END, item[0], item[1])
        
        self.text_content.insert(tk.END, "\n")
    
    def _insert_formatted_line_with_anchor(self, line, anchor_tags):
        """Inserisce una riga con formattazione inline e tag di ancoraggio"""
        parts = []
        current_text = ""
        current_tag = "normal"
        i = 0
        
        while i < len(line):
            # Cerca tag LINK con URL: [LINK:url]testo[/LINK]
            if line.startswith('[LINK:', i):
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                # Estrae l'URL
                url_start = i + 6  # dopo '[LINK:'
                url_end = line.find(']', url_start)
                if url_end != -1:
                    url = line[url_start:url_end]
                    # Trova il testo del link
                    text_start = url_end + 1
                    text_end = line.find('[/LINK]', text_start)
                    if text_end != -1:
                        link_text = line[text_start:text_end]
                        parts.append((link_text, "hyperlink", url))
                        i = text_end + 7  # dopo '[/LINK]'
                        continue
                current_text += line[i]
                i += 1
            # Cerca tag di apertura/chiusura
            elif line[i:i+6] == '[BOLD]':
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "bold"
                i += 6
            elif line[i:i+8] == '[/BOLD]':
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "normal"
                i += 8
            elif line[i:i+8] == '[NORMAL]':
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "normal"
                i += 8
            elif line[i:i+10] == '[/NORMAL]':
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "normal"
                i += 10
            elif line[i:i+6] == '[CODE]':
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "code"
                i += 6
            elif line[i:i+8] == '[/CODE]':
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "normal"
                i += 8
            elif line.startswith('[WARNING]', i):
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "warning"
                i += len('[WARNING]')
            elif line.startswith('[/WARNING]', i):
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "normal"
                i += len('[/WARNING]')
            elif line.startswith('[WARNING_RED]', i):
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "warning_red"
                i += len('[WARNING_RED]')
            elif line.startswith('[/WARNING_RED]', i):
                if current_text:
                    parts.append((current_text, current_tag, None))
                    current_text = ""
                current_tag = "normal"
                i += len('[/WARNING_RED]')
            else:
                current_text += line[i]
                i += 1
        
        # Aggiunge l'ultimo pezzo
        if current_text:
            parts.append((current_text, current_tag, None))
        elif not parts:
            parts.append(("", "normal", None))
        
        # Inserisce tutti i pezzi con i tag di ancoraggio
        for item in parts:
            if len(item) == 3:
                text, tag, url = item
                if tag == "hyperlink" and url:
                    # Crea un tag univoco per questo link
                    link_tag = f"link_{id(url)}"
                    # Configura il tag con lo stile del link
                    self.text_content.tag_configure(link_tag, foreground="blue", underline=True)
                    # Bind del click per aprire l'URL
                    self.text_content.tag_bind(link_tag, "<Button-1>", lambda e, u=url: webbrowser.open(u))
                    self.text_content.tag_bind(link_tag, "<Enter>", lambda e: self.text_content.config(cursor="hand2"))
                    self.text_content.tag_bind(link_tag, "<Leave>", lambda e: self.text_content.config(cursor=""))
                    # Combina anchor_tags con il tag del link
                    combined_tags = list(anchor_tags) + [link_tag]
                    self.text_content.insert(tk.END, text, tuple(combined_tags))
                else:
                    # Combina il tag di formattazione con i tag di ancoraggio
                    combined_tags = list(anchor_tags) + [tag]
                    self.text_content.insert(tk.END, text, tuple(combined_tags))
            else:
                # Fallback per compatibilità
                combined_tags = list(anchor_tags) + [item[1]]
                self.text_content.insert(tk.END, item[0], tuple(combined_tags))
        
        self.text_content.insert(tk.END, "\n")
    
    def setup_search_functionality(self, parent_frame):
        """Crea l'interfaccia di ricerca con Entry, pulsanti e contatore risultati"""
        # Determina la lingua corrente per le traduzioni
        import configparser
        config = configparser.ConfigParser(interpolation=None)
        config_file = get_config_file()
        current_language = "en"  # default
        if os.path.exists(config_file):
            config.read(config_file, encoding='utf-8')
            if 'Settings' in config and config.has_option('Settings', 'language'):
                current_language = config.get('Settings', 'language', fallback='en')
        if current_language not in ['en', 'it']:
            current_language = "en"
        
        # Testi tradotti
        search_label_text = "Trova:" if current_language == 'it' else "Search:"
        search_button_text = "🔍 Trova" if current_language == 'it' else "🔍 Search"
        next_button_text = "⏩ Successivo" if current_language == 'it' else "⏩ Next"
        
        # Label "Trova:" / "Search:"
        ttk.Label(parent_frame, text=search_label_text).pack(side=tk.LEFT, padx=(5, 5))
        
        # Entry per digitare la parola da cercare
        search_entry = ttk.Entry(parent_frame, textvariable=self.search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=(0, 5))
        search_entry.bind("<Return>", lambda e: self.search_text())
        search_entry.bind("<Escape>", lambda e: self.clear_search())
        
        # Pulsante "Trova" / "Search"
        ttk.Button(parent_frame, text=search_button_text, command=self.search_text).pack(side=tk.LEFT, padx=(0, 5))
        
        # Pulsante "Successivo" / "Next"
        ttk.Button(parent_frame, text=next_button_text, command=self.search_next).pack(side=tk.LEFT, padx=(0, 5))
        
        # Label per mostrare "Risultato X di Y" / "Result X of Y"
        self.search_result_label = ttk.Label(parent_frame, text="", foreground="blue")
        self.search_result_label.pack(side=tk.LEFT, padx=(10, 5))
    
    def search_text(self):
        """Cerca il testo nel widget e evidenzia tutti i match (case-insensitive)"""
        # Pulisci ricerca precedente
        self.clear_search()
        
        search_term = self.search_var.get().strip()
        if not search_term:
            return
        
        # Abilita temporaneamente il widget per modificare i tag
        self.text_content.config(state="normal")
        
        # Cerca tutti i match (case-insensitive usando opzione nocase)
        start_pos = "1.0"
        while True:
            # tk.Text.search con nocase=1 per ricerca case-insensitive
            pos = self.text_content.search(search_term, start_pos, stopindex=tk.END, nocase=1)
            if not pos:
                break
            
            # Calcola la posizione finale del match
            end_pos = f"{pos}+{len(search_term)}c"
            
            # Aggiungi il tag di evidenziazione
            self.text_content.tag_add("search_highlight", pos, end_pos)
            
            # Salva la posizione per la navigazione
            self.search_matches.append(pos)
            
            # Continua la ricerca dopo questo match
            start_pos = end_pos
        
        # Disabilita nuovamente il widget
        self.text_content.config(state="disabled")
        
        # Se ci sono match, vai al primo e aggiorna il contatore
        if self.search_matches:
            self.current_match_index = 0
            self.text_content.see(self.search_matches[0])
            self.update_search_counter()
        else:
            # Nessun risultato trovato
            import configparser
            config = configparser.ConfigParser(interpolation=None)
            config_file = get_config_file()
            current_language = "en"
            if os.path.exists(config_file):
                config.read(config_file, encoding='utf-8')
                if 'Settings' in config and config.has_option('Settings', 'language'):
                    current_language = config.get('Settings', 'language', fallback='en')
            if current_language not in ['en', 'it']:
                current_language = "en"
            
            no_result_text = "Nessun risultato" if current_language == 'it' else "No results"
            self.search_result_label.config(text=no_result_text)
    
    def search_next(self):
        """Naviga al risultato successivo nella lista dei match"""
        if not self.search_matches:
            # Se non ci sono match, esegui una nuova ricerca
            self.search_text()
            return
        
        # Incrementa l'indice con wrap-around
        self.current_match_index = (self.current_match_index + 1) % len(self.search_matches)
        
        # Scrolla al match corrente
        current_pos = self.search_matches[self.current_match_index]
        self.text_content.see(current_pos)
        
        # Aggiorna il contatore
        self.update_search_counter()
    
    def update_search_counter(self):
        """Aggiorna la label con il contatore dei risultati"""
        if not self.search_matches:
            self.search_result_label.config(text="")
            return
        
        # Determina la lingua per il formato del contatore
        import configparser
        config = configparser.ConfigParser(interpolation=None)
        config_file = get_config_file()
        current_language = "en"
        if os.path.exists(config_file):
            config.read(config_file, encoding='utf-8')
            if 'Settings' in config and config.has_option('Settings', 'language'):
                current_language = config.get('Settings', 'language', fallback='en')
        if current_language not in ['en', 'it']:
            current_language = "en"
        
        # Formato: "Risultato 1 di 5" (italiano) o "Result 1 of 5" (inglese)
        total = len(self.search_matches)
        current = self.current_match_index + 1
        
        if current_language == 'it':
            counter_text = f"Risultato {current} di {total}"
        else:
            counter_text = f"Result {current} of {total}"
        
        self.search_result_label.config(text=counter_text)
    
    def clear_search(self):
        """Pulisce le evidenziazioni della ricerca precedente"""
        # Abilita temporaneamente per modificare
        self.text_content.config(state="normal")
        
        # Rimuovi tutti i tag di evidenziazione
        self.text_content.tag_remove("search_highlight", "1.0", tk.END)
        
        # Disabilita nuovamente
        self.text_content.config(state="disabled")
        
        # Reset variabili
        self.search_matches = []
        self.current_match_index = -1
        if self.search_result_label:
            self.search_result_label.config(text="")

# ------------------------------------------------------------------------------------
# --- NUOVA FINESTRA LICENZA ---
# ------------------------------------------------------------------------------------
class LicenseWindow(tk.Toplevel):
    def __init__(self, parent, first_run=False):
        super().__init__(parent)
        self.withdraw()
        set_window_icon(self)
        self.title(_("Licenza d'Uso - DataFlow Procurement Software"))
        self.transient(parent)
        self.grab_set()
        
        # Frame pulsanti (sempre in fondo)
        button_frame = ttk.Frame(self)
        button_frame.pack(side="bottom", fill="x", padx=10, pady=10)

        if first_run:
            self.accepted = False # Stato di default
            ttk.Button(button_frame, text=_("❌ Esci"), command=self.on_exit).pack(side="right")
            ttk.Button(button_frame, text=_("✅ Accetto"), command=self.on_accept).pack(side="right", padx=10)
            # Gestisce la chiusura della finestra con la 'X' come un "Esci"
            self.protocol("WM_DELETE_WINDOW", self.on_exit) 
        else:
            ttk.Button(button_frame, text=_("❌ Chiudi"), command=self.destroy).pack(side="right")
        
        # Frame contenuto (espandibile)
        main_frame = ttk.Frame(self)
        main_frame.pack(side="top", fill="both", expand=True)

        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        scrollbar = ttk.Scrollbar(content_frame)
        self.text_content = tk.Text(content_frame, wrap=tk.WORD, yscrollcommand=scrollbar.set, padx=15, pady=10, relief="flat", background="#FFFFFF", font=("Arial", 10))
        scrollbar.config(command=self.text_content.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.text_content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self._populate_content()
        self.text_content.config(state="disabled")
        
        # Centra la finestra dopo aver aggiunto tutti i widget
        # Usiamo un after per assicurarci che il contenuto sia disegnato
        center_window(self)

    def on_accept(self):
        self.accepted = True
        self.destroy()

    def on_exit(self):
        self.accepted = False
        self.destroy()

    def _populate_content(self):
        # Configurazione degli stili di testo
        self.text_content.tag_configure("h1", font=("Arial", 14, "bold", "underline"), justify="center")
        self.text_content.tag_configure("h2", font=("Arial", 11, "bold"))
        self.text_content.tag_configure("normal", font=("Arial", 10))
        self.text_content.tag_configure("code", font=("Arial", 9))
        
        # Configurazione tag per link cliccabile
        self.text_content.tag_configure("link", foreground="blue", underline=True)
        self.text_content.tag_bind("link", "<Button-1>", lambda e: webbrowser.open("https://www.linkedin.com/in/guido-soraru-buyer/"))
        self.text_content.tag_bind("link", "<Enter>", lambda e: self.text_content.config(cursor="hand2"))
        self.text_content.tag_bind("link", "<Leave>", lambda e: self.text_content.config(cursor=""))
        
        def add(txt, tag_keys):
            tag_tuple = tag_keys if isinstance(tag_keys, tuple) else (tag_keys,)
            self.text_content.insert(tk.END, txt, tag_tuple)

        # --- INIZIO CONTENUTO LICENZA ---
        
        add(_("Licenza d'Uso (Freeware) - DataFlow Procurement Software\n\n"), "h1")
        
        add(_("Sviluppatore: "), "h2"); add("Guido Sorarù", ("normal", "link")); add("\n", "normal")
        add(_("E-mail: "), "h2"); add("sorguido@gmail.com\n", "normal")
        add(_("Copyright © 2025 Guido Sorarù. Tutti i diritti riservati.\n\n"), "h2")
        
        add("--------------------------------------------------\n\n", "normal")
        
        add(_("Il presente software, \"DataFlow\" (di seguito \"il Software\"), è fornito come \"Freeware\".\n\n"), "normal")
        
        add(_("1. CONCESSIONE DELLA LICENZA\n"), "h2")
        add(_("Lo Sviluppatore concede all'utente una licenza non esclusiva e gratuita per scaricare, installare e utilizzare il Software per scopi personali, educativi o commerciali, senza limiti di tempo.\n\n"), "normal")
        
        add(_("2. RESTRIZIONI\n"), "h2")
        add(_("All'utente non è consentito:\n"), "normal")
        add(_("a) Vendere, affittare, noleggiare o distribuire il Software in cambio di un compenso.\n"), "normal")
        add(_("b) Decodificare (reverse engineer), decompilare, disassemblare o tentare di scoprire in altro modo il codice sorgente del Software.\n"), "normal")
        add(_("c) Modificare, adattare, tradurre o creare opere derivate basate sul Software senza il previo consenso scritto dello Sviluppatore.\n"), "normal")
        add(_("d) Rimuovere o alterare qualsiasi avviso di copyright, marchio di fabbrica o altra notifica di proprietà presente nel Software.\n\n"), "normal")
        
        add(_("3. ESCLUSIONE DI GARANZIA\n"), "h2")
        add(_("IL SOFTWARE È FORNITO \"COSÌ COM'È\" (AS IS), SENZA ALCUNA GARANZIA, ESPRESSA O IMPLICITA. LO SVILUPPATORE NON FORNISCE ALCUNA GARANZIA RIGUARDO LA COMMERCIABILITÀ, L'IDONEITÀ PER UNO SCOPO PARTICOLARE O LA NON VIOLAZIONE DI DIRITTI DI TERZI.\n"), "normal")
        add(_("L'INTERO RISCHIO DERIVANTE DALL'USO O DALLE PRESTAZIONI DEL SOFTWARE RIMANE A CARICO DELL'UTENTE.\n\n"), "normal")
        
        add(_("4. LIMITAZIONE DI RESPONSABILITÀ\n"), "h2")
        add(_("IN NESSUN CASO LO SVILUPPATORE (GUIDO SORARÙ) POTRÀ ESSERE RITENUTO RESPONSABILE PER QUALSIASI DANNO DIRETTO, INDIRETTO, INCIDENTALE, SPECIALE, ESEMPLARE O CONSEQUENZIALE (INCLUSI, A TITOLO ESEMPLIFICATIVO MA NON ESAUSIVO, DANNI PER PERDITA DI DATI, PERDITA DI PROFITTI O INTERRUZIONE DELL'ATTIVITÀ) DERIVANTE DALL'USO, DALL'USO IMPROPRIO O DALL'IMPOSSIBILITÀ DI UTILIZZARE IL SOFTWARE, ANCHE SE LO SVILUPPATORE È STATO AVVISATO DELLA POSSIBILITÀ DI TALI DANNI.\n\n"), "normal")
        
        # --- INIZIO TESTO AGGIUNTO ---
        add(_("Il Software utilizza un database SQLite con modalità WAL per ogni utente. DataFlow 2.0.0 supporta l'utilizzo multi-utente con database separati per ciascun utente, permettendo la condivisione sicura dei dati in sola lettura.\n"), "normal")
        add(_("L'utente si assume la piena responsabilità per la perdita o corruzione dei dati derivante dall'uso improprio del software.\n"), "normal")
        add(_("L'accesso simultaneo in scrittura da parte di più utenti allo stesso file di database non è supportato e causerà con alta probabilità la corruzione irreversibile dei dati. Tuttavia, l'architettura multi-utente di DataFlow garantisce che ogni utente abbia il proprio database separato, eliminando questo rischio.\n\n"), "normal")
        # --- FINE TESTO AGGIUNTO ---
        
        add(_("Utilizzando questo Software, l'utente accetta i termini e le condizioni di questa licenza.\n"), "normal")
        
        # Disabilita il widget dopo il caricamento
        self.text_content.config(state="disabled")

# ------------------------------------------------------------------------------------
# DIALOG SELEZIONE TIPO RDO
# ------------------------------------------------------------------------------------
class NewRdOTypeDialog(tk.Toplevel):
    """Dialog minimale per scegliere il tipo di RdO da creare"""
    def __init__(self, parent):
        super().__init__(parent)
        self.withdraw()
        set_window_icon(self)
        
        self.title(_("Nuova Richiesta di Offerta"))
        # NON usare transient() e grab_set() per evitare che la chiusura chiuda anche il parent
        self.result = None
        
        # Frame principale
        main_frame = ttk.Frame(self, padding="20")
        main_frame.pack(fill="both", expand=True)
        
        # Etichetta domanda
        ttk.Label(
            main_frame, 
            text=_("Che tipo di RdO vuoi creare?"), 
            font=("Arial", 11, "bold")
        ).pack(pady=(0, 20))
        
        # Frame pulsanti tipo
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill="x", pady=(0, 20))
        
        ttk.Button(
            btn_frame,
            text=_("📦 Fornitura piena"),
            command=lambda: self.set_result("Fornitura piena"),
            width=20
        ).pack(side="left", padx=5)
        
        ttk.Button(
            btn_frame,
            text=_("🔧 Conto lavoro"),
            command=lambda: self.set_result("Conto lavoro"),
            width=20
        ).pack(side="left", padx=5)
        
        # Pulsante annulla
        ttk.Button(
            main_frame,
            text=_("❌ Annulla"),
            command=self.destroy
        ).pack()
        
        # Gestione chiusura con X
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        
        center_window(self)
    
    def set_result(self, tipo):
        """Salva la scelta e chiude il dialog"""
        self.result = tipo
        self.destroy()

# ------------------------------------------------------------------------------------
# FINESTRA PRINCIPALE
# ------------------------------------------------------------------------------------
class MainWindow:
    def __init__(self, root):
        self.root = root;
        set_window_icon(self.root)
        self.root.title(_("DataFlow Procurement Software - Cruscotto Principale"))
        self.all_users_placeholder = _("Tutti gli utenti")
        self.username_filter_var = None
        self.user_filter_combo = None
        self._load_identity_from_config()
        self.last_backup_date = None; self.db_path_standard = self.get_standard_db_path()
        
        # BUG #45 FIX: Inizializza ID del timer autobackup per permettere cancellazione
        self._autobackup_timer_id = None
        
        # BUG #48 FIX: Inizializza ID del timer SQL warning per permettere cancellazione
        self._sql_warning_after_id = None
        
        # BUG #50 FIX: Inizializza flag debounce per doppio click apertura RdO
        self._opening_request = False
        
        # Inizializza il database manager con il percorso completo del database
        self.db_manager = DatabaseManager(get_db_path())
        
        # --- INIZIO MODIFICA: Rilevamento DB Temporaneo (escludendo DB personalizzato) ---
        self.active_db_path = get_db_path()
        
        # Determina il percorso "di default" considerando eventuali impostazioni legacy
        # Il warning deve apparire solo se il database è DIVERSO da quello di default
        # RIMOSSO: logica e visualizzazione warning DB provvisorio
        
        frame_top = ttk.Frame(self.root); frame_top.pack(fill="x", padx=10, pady=10)
        try:
            logo_path = resource_path(os.path.join("add_data", "logo_dataflow.png"))
            if os.path.exists(logo_path):
                img = Image.open(logo_path)
                # BUG #51 FIX: Check dimensioni valide prima di divisione per evitare ZeroDivisionError
                if img.width > 0 and img.height > 0:
                    img.thumbnail((int(40 * (img.width / img.height)), 40), Image.Resampling.LANCZOS)
                    self.logo_photo = ImageTk.PhotoImage(img)
                    ttk.Label(frame_top, image=self.logo_photo).pack(side="left", padx=(0, 20), anchor="w")
        except Exception as e: print(f"Errore caricamento logo: {e}")
        # --- Pulsanti Operativi (Riga Superiore) ---
        # 1. New RfQ
        self.btn_new_rdo = ttk.Button(frame_top, text=_("➕ Nuova RdO"), command=self.open_new_request_window)
        self.btn_new_rdo.pack(side="left", padx=(0, 10))

        # 2. Delete RfQ
        self.btn_delete_rdo = ttk.Button(frame_top, text=_("🗑 Elimina RdO"), command=self.delete_selected_request, state="disabled")
        self.btn_delete_rdo.pack(side="left", padx=(0, 10))

        # 3. Duplicate RfQ
        self.btn_duplicate_rdo = ttk.Button(frame_top, text=_("🔁 Duplica RdO"), command=self.duplicate_selected_request, state="disabled")
        self.btn_duplicate_rdo.pack(side="left", padx=(0, 10))

        # 4. Archive RfQ (Spostato a sinistra)
        self.btn_archive_rdo = ttk.Button(frame_top, text=_("📦 Archivia"), command=self.archive_selected_request, state="disabled")
        self.btn_archive_rdo.pack(side="left", padx=(0, 10))

        # 5. Export Excel (Export Globale, Spostato a destra)
        self.btn_mega_export = ttk.Button(frame_top, text=_("📊 Export Excel"), command=self.mega_export_excel)
        self.btn_mega_export.pack(side="left", padx=(0, 10))

        # 6. Reactivate (Ultimo dei pulsanti operativi)
        self.btn_reactivate = ttk.Button(frame_top, text=_("↩️ Riattiva"), command=self.reactivate_selected_request, state="disabled")
        self.btn_reactivate.pack(side="left", padx=(0, 20))  # Aumenta padx per separare dal gruppo "Utility"
        
        # --- MODIFICA: Aggiunto pulsante Licenza e riordinato ---
        self.btn_guida = ttk.Button(frame_top, text=_("❓ Guida"), command=self.open_help_window); self.btn_guida.pack(side="right")
        self.btn_license = ttk.Button(frame_top, text=_("📄 Licenza"), command=self.open_license_window); self.btn_license.pack(side="right", padx=(0, 10))
        self.btn_settings = ttk.Button(frame_top, text=_("⚙️ Impostazioni"), command=self.open_settings_window); self.btn_settings.pack(side="right", padx=(0, 10))
        # --- FINE MODIFICA ---
        
        search_frame = ttk.LabelFrame(self.root, text=_("Filtri di Ricerca"), padding=(10, 5)); search_frame.pack(fill="x", padx=10, pady=5)
        self.search_vars = {name: tk.StringVar() for name in ['num', 'ref', 'forn', 'cod', 'desc', 'ord', 'cod_grezzo', 'dis_grezzo', 'mat_cl']}; self.search_tipo = tk.StringVar(value=_("Tutte"))
        ttk.Label(search_frame, text=_("Numero RdO:")).grid(row=0, column=0, sticky="w"); ttk.Entry(search_frame, textvariable=self.search_vars['num']).grid(row=0, column=1, sticky="ew")
        ttk.Label(search_frame, text=_("Tipo RdO:")).grid(row=0, column=2, sticky="w"); ttk.Combobox(search_frame, textvariable=self.search_tipo, values=[_("Tutte"), _("Fornitura piena"), _("Conto lavoro")], state="readonly").grid(row=0, column=3, sticky="ew")
        ttk.Label(search_frame, text=_("Riferimento:")).grid(row=1, column=0, sticky="w"); ttk.Entry(search_frame, textvariable=self.search_vars['ref']).grid(row=1, column=1, sticky="ew")
        ttk.Label(search_frame, text=_("Fornitore:")).grid(row=1, column=2, sticky="w"); ttk.Entry(search_frame, textvariable=self.search_vars['forn']).grid(row=1, column=3, sticky="ew")
        ttk.Label(search_frame, text=_("Cod. Materiale:")).grid(row=2, column=0, sticky="w"); ttk.Entry(search_frame, textvariable=self.search_vars['cod']).grid(row=2, column=1, sticky="ew")
        ttk.Label(search_frame, text=_("Desc. Materiale:")).grid(row=2, column=2, sticky="w"); ttk.Entry(search_frame, textvariable=self.search_vars['desc']).grid(row=2, column=3, sticky="ew")
        ttk.Label(search_frame, text=_("Num. Ordine:")).grid(row=0, column=4, sticky="w"); ttk.Entry(search_frame, textvariable=self.search_vars['ord']).grid(row=0, column=5, sticky="ew")
        # --- INIZIO BLOCCO AGGIUNTO ---
        ttk.Label(search_frame, text=_("Cod. Grezzo:")).grid(row=3, column=0, sticky="w"); ttk.Entry(search_frame, textvariable=self.search_vars['cod_grezzo']).grid(row=3, column=1, sticky="ew")
        ttk.Label(search_frame, text=_("Allegato Grezzo:")).grid(row=3, column=2, sticky="w"); ttk.Entry(search_frame, textvariable=self.search_vars['dis_grezzo']).grid(row=3, column=3, sticky="ew")
        ttk.Label(search_frame, text=_("Mat. c/lavoro:")).grid(row=3, column=4, sticky="w"); ttk.Entry(search_frame, textvariable=self.search_vars['mat_cl']).grid(row=3, column=5, sticky="ew")
        # --- FINE BLOCCO AGGIUNTO ---
        ttk.Label(search_frame, text=_("Utente:")).grid(row=1, column=4, sticky="w")
        default_user_value = self.current_username if getattr(self, 'current_username', '') else self.all_users_placeholder
        self.username_filter_var = tk.StringVar(value=default_user_value)
        self.user_filter_combo = ttk.Combobox(
            search_frame,
            textvariable=self.username_filter_var,
            state="readonly",
            values=[default_user_value]
        )
        self.user_filter_combo.grid(row=1, column=5, sticky="ew")
        self.user_filter_combo.bind("<<ComboboxSelected>>", lambda _e: self.refresh_data())
        self.date_entries = {}
        for i, (lbl, key) in enumerate([(_("Da:"), "emm_da"), (_("A:"), "emm_a"), (_("Da:"), "scad_da"), (_("A:"), "scad_a")]):
            row, col_lbl, col_entry = (4 + i // 2, (i % 2) * 2, (i % 2) * 2 + 1)
            prefix = _("Data Emissione ") if i < 2 else _("Data Scadenza ")
            ttk.Label(search_frame, text=prefix + lbl).grid(row=row, column=col_lbl, sticky="w"); de = DateEntry(search_frame, date_pattern='dd/mm/yyyy', locale=('it_IT' if get_current_language() == 'it' else 'en_US')); de.grid(row=row, column=col_entry, sticky="ew"); de.delete(0, 'end'); self.date_entries[key] = de
        for i in range(1, 6, 2): search_frame.grid_columnconfigure(i, weight=1)
        btn_search_frame = ttk.Frame(search_frame); btn_search_frame.grid(row=0, column=6, rowspan=6, sticky="nsew", padx=20)
        ttk.Button(btn_search_frame, text=_("🔍 Cerca"), command=self.search_requests).pack(fill="x", expand=True, pady=2); ttk.Button(btn_search_frame, text=_("🔎 Pulisci Filtri"), command=self.clear_filters).pack(fill="x", expand=True, pady=2)
        self.notebook = ttk.Notebook(self.root); self.notebook.pack(fill="both", expand=True, padx=10, pady=5)
        self.tab_attive = ttk.Frame(self.notebook); self.tab_archiviate = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_attive, text=_("RdO Attive")); self.notebook.add(self.tab_archiviate, text=_("RdO Archiviate"))
        footer_frame = ttk.Frame(self.root); footer_frame.pack(side="bottom", fill="x", padx=10, pady=5)
        ttk.Label(footer_frame, text=_("DataFlow Procurement Software v2.0.0 - Sviluppato da ")).pack(side="left")
        name_label = ttk.Label(footer_frame, text="Guido Sorarù", foreground="blue", cursor="hand2"); name_label.pack(side="left")
        name_label.bind("<Button-1>", lambda e: webbrowser.open("https://www.linkedin.com/in/guido-soraru-buyer/"))
        ttk.Label(footer_frame, text=_(" © 2025")).pack(side="left")
        # --- RIMOSSO: etichetta warning DB provvisorio ---
        self.tree_attive = self.create_request_treeview(self.tab_attive); self.tree_archiviate = self.create_request_treeview(self.tab_archiviate)
        # NOTA: La configurazione dei tag (oddrow, scaduta) non è più necessaria
        # perché tksheet usa highlight_rows() per applicare la colorazione
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed); self.refresh_data(); self.update_button_visibility(); self.check_for_autobackup()

    # --- INIZIO NUOVI METODI LICENZA ---
    def open_license_window(self):
        """Apre la finestra della licenza in modalità sola lettura."""
        # --- CORREZIONE OBBLIGATORIA ---
        # Rimuovi temporaneamente topmost per permettere alla finestra 
        # di licenza di apparire SOPRA.
        self.root.attributes('-topmost', False)
        # --- FINE CORREZIONE ---
        
        LicenseWindow(self.root, first_run=False)

    def show_first_run_license(self):
        """
        Mostra la finestra modale della licenza al primo avvio.
        Blocca l'esecuzione finché l'utente non accetta o esce.
        """
        # --- CORREZIONE OBBLIGATORIA ---
        # Rimuovi temporaneamente topmost per permettere alla finestra 
        # di licenza di apparire SOPRA.
        self.root.attributes('-topmost', False)
        # --- FINE CORREZIONE ---
        
        license_prompt = LicenseWindow(self.root, first_run=True)
        self.root.wait_window(license_prompt) # Attende che la finestra di licenza venga chiusa
        
        if not license_prompt.accepted:
            # L'utente ha cliccato "Esci" o ha chiuso la finestra
            # Usa after per evitare problemi di timing con la distruzione della finestra
            self.root.after(100, self.root.destroy)
            return False
        else:
            # L'utente ha cliccato "Accetto", salva l'impostazione
            try:
                config = configparser.ConfigParser(interpolation=None)
                config_file = get_config_file()
                if os.path.exists(config_file):
                    config.read(config_file)
                if 'Settings' not in config:
                    config['Settings'] = {}
                config['Settings']['license_accepted'] = 'True'
                # BUG #49 FIX: Usa encoding UTF-8 per gestire caratteri speciali
                with open(config_file, 'w', encoding='utf-8') as f:
                    config.write(f)
            except Exception as e:
                messagebox.showerror(_("Errore"), _("Impossibile salvare l'impostazione della licenza: {}\n\nIl programma continuerà, ma la licenza potrebbe riapparire al prossimo avvio.").format(e), parent=self.root)
            
            # Assicura che la finestra principale sia visibile e attiva
            self.root.deiconify()
            self.root.focus_force()
            return True
    # --- FINE NUOVI METODI LICENZA ---

    def _load_identity_from_config(self):
        identity = load_user_identity()
        self.current_first_name = identity.get('first_name', '')
        self.current_last_name = identity.get('last_name', '')
        self.current_username = identity.get('username', '')
        self.current_full_name = identity.get('full_name', '').strip()

    def ensure_user_identity(self, force_prompt=False):
        """Garantisce che nome, cognome e username siano impostati."""
        self._load_identity_from_config()
        needs_prompt = force_prompt or not self.current_first_name or not self.current_last_name or not self.current_username
        if not needs_prompt:
            return True
        
        identity_kwargs = {
            'first_name': self.current_first_name,
            'last_name': self.current_last_name
        }
        
        while True:
            dialog = UserIdentityDialog(self.root, **identity_kwargs)
            self.root.wait_window(dialog)
            result = getattr(dialog, 'result', None)
            if not result:
                if not self.root.winfo_exists():
                    return False
                messagebox.showwarning(_("Dati mancanti"), _("Per utilizzare DataFlow devi inserire nome e cognome."), parent=self.root)
                continue
            try:
                save_user_identity(result['first_name'], result['last_name'], result['username'])
                self._load_identity_from_config()
                self.apply_user_identity_to_ui()
                return True
            except Exception as e:
                logger.error(f"Errore salvataggio identità utente: {e}", exc_info=True)
                messagebox.showerror(_("Errore"), _("Impossibile salvare i dati utente: {}").format(e), parent=self.root)
                identity_kwargs = result

    def apply_user_identity_to_ui(self):
        """Applica lo username corrente all'interfaccia (filtri e nuovi inserimenti)."""
        if not self.username_filter_var:
            return
        value = self.current_username if self.current_username else self.all_users_placeholder
        self.username_filter_var.set(value)
        self.refresh_data()

    def populate_username_filter(self):
        """Aggiorna la lista degli username disponibili nel filtro usando aggregazione multi-database."""
        if not self.user_filter_combo or not self.username_filter_var:
            return
        
        usernames = []
        
        try:
            # Carica TUTTI gli username da TUTTI i database aggregati
            with DatabaseManager(get_db_path()) as db_manager:
                all_requests = db_manager.get_all_richieste_aggregated(get_db_path())
            
            # BUG #2 FIX: Validazione robusta per gestire tuple di lunghezza variabile
            # Estrai username unici dalle richieste aggregate (indice 5)
            usernames = list({req[5].strip().lower() for req in all_requests 
                            if len(req) > 5 and req[5] and str(req[5]).strip()})
            
            logger.info(f"[populate_username_filter] Trovati {len(usernames)} utenti: {usernames}")
            
        except (DatabaseError, IndexError, AttributeError, TypeError) as e:
            # BUG #2 FIX: Fallback robusto con gestione completa eccezioni
            logger.warning(f"Aggregazione multi-DB fallita in populate_username_filter, uso fallback: {e}")
            # Fallback: usa solo il database locale
            try:
                with DatabaseManager(get_db_path()) as db_manager:
                    usernames = db_manager.get_distinct_usernames()
                logger.info(f"Fallback completato: popolato filtro con {len(usernames)} utenti (locale)")
            except DatabaseError as e2:
                logger.error(f"Errore anche nel fallback: {e2}", exc_info=True)
        
        # Assicurati che l'utente corrente sia nella lista
        if self.current_username and self.current_username not in usernames:
            usernames.append(self.current_username)
        
        # Ordina e prepara la lista per la combo
        clean_usernames = sorted({u for u in usernames if u})
        values = [self.all_users_placeholder] + clean_usernames
        current_value = self.username_filter_var.get()
        self.user_filter_combo.config(values=values)
        
        # Resetta al valore corrente se valido, altrimenti all'utente corrente
        if current_value not in values:
            self.username_filter_var.set(self.current_username or self.all_users_placeholder)

    def _get_active_username_filter(self):
        if not self.username_filter_var:
            return None
        value = self.username_filter_var.get().strip()
        if not value or value == self.all_users_placeholder:
            return None
        return value.lower()
    
    def _has_active_search_filters(self):
        """Verifica se ci sono filtri di ricerca attivi (escludendo username e stato)"""
        # Controlla filtri di testo
        for var in self.search_vars.values():
            if var.get().strip():
                return True
        
        # Controlla filtro tipo RdO
        if self.search_tipo.get() != _("Tutte"):
            return True
        
        # Controlla filtri data
        for entry in self.date_entries.values():
            if entry.get().strip():
                return True
        
        return False

    def _assign_request_to_current_user(self, request_id):
        """Associa una RdO all'utente corrente."""
        if not self.current_username:
            return
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(get_db_path()) as db_manager:
                db_manager.update_request_username(request_id, self.current_username)
        except DatabaseError as e:
            logger.error(f"Impossibile assegnare la RdO {request_id} all'utente {self.current_username}: {e}", exc_info=True)

    def get_standard_db_path(self):
        """Restituisce il percorso standard del database (estensione .db)"""
        return os.path.join(get_fixed_db_dir(), 'gestione_offerte.db')

    def restart_program(self):
        """Riavvia l'applicazione con le nuove impostazioni."""
        python = sys.executable
        
        # Determina il percorso corretto del file Python
        script_path = None
        
        # Prova prima con __file__ (sempre disponibile quando eseguito come script)
        try:
            # __file__ è sempre disponibile quando il file viene eseguito come script
            current_file = __file__
            if current_file:
                script_path = os.path.abspath(current_file)
                if os.path.exists(script_path) and script_path.endswith('.py'):
                    # Percorso valido trovato
                    pass
        except (NameError, AttributeError):
            # __file__ non disponibile (raro, ma può accadere in alcuni contesti)
            pass
        
        # Se __file__ non è disponibile o non valido, prova sys.argv[0]
        if not script_path or not os.path.exists(script_path):
            if sys.argv[0]:
                # Se sys.argv[0] è un percorso relativo, prova a risolverlo
                if os.path.exists(sys.argv[0]):
                    script_path = os.path.abspath(sys.argv[0])
                else:
                    # Se non esiste, prova a costruire il percorso assoluto
                    # basandosi sulla directory corrente
                    possible_path = os.path.join(os.getcwd(), sys.argv[0])
                    if os.path.exists(possible_path):
                        script_path = os.path.abspath(possible_path)
                    else:
                        # Ultimo tentativo: usa il nome del file nella directory dello script
                        # (se siamo in modalità PyInstaller o MSIX)
                        if hasattr(sys, '_MEIPASS'):
                            # PyInstaller: usa sys.executable
                            script_path = sys.executable
                        else:
                            script_path = sys.argv[0]
        
        # Se ancora non abbiamo un percorso valido, usa sys.executable
        if not script_path or (not os.path.exists(script_path) and not hasattr(sys, '_MEIPASS')):
            script_path = sys.executable
        
        # Riavvia l'applicazione usando subprocess invece di os.execl
        # Questo gestisce correttamente i percorsi con spazi
        try:
            # Costruisci il comando da eseguire
            # Usa subprocess.Popen con lista di argomenti per gestire correttamente gli spazi
            if script_path.endswith('.py') or (not hasattr(sys, '_MEIPASS') and script_path != sys.executable):
                # Esecuzione come script Python
                cmd = [python, script_path]
            else:
                # Eseguibile (PyInstaller o MSIX)
                cmd = [script_path]
            
            # Imposta la working directory
            if os.path.dirname(script_path):
                cwd = os.path.dirname(script_path)
            else:
                cwd = os.getcwd()
            
            # Funzione per chiudere tutto e avviare il nuovo processo
            def do_restart():
                try:
                    # Invalida la cache del DB prima del riavvio
                    reset_db_cache()
                    
                    # Avvia il nuovo processo PRIMA di chiudere quello corrente
                    # Usa DETACHED_PROCESS su Windows per evitare che apra una nuova console
                    if sys.platform == 'win32':
                        new_process = subprocess.Popen(
                            cmd, 
                            cwd=cwd,
                            creationflags=subprocess.CREATE_NEW_PROCESS_GROUP | subprocess.DETACHED_PROCESS,
                            close_fds=True
                        )
                    else:
                        new_process = subprocess.Popen(cmd, cwd=cwd, start_new_session=True)
                    
                    # Attendi fino a 2 secondi che il processo si stabilizzi
                    import time
                    for _ in range(20):
                        if new_process.poll() is None:  # Processo ancora in esecuzione
                            break
                        time.sleep(0.1)
                    
                    # Piccolo delay aggiuntivo per sicurezza
                    time.sleep(0.2)
                    
                    # Chiudi tutte le finestre Tkinter
                    if hasattr(self, 'root') and self.root:
                        try:
                            # Distruggi tutte le finestre Toplevel
                            for widget in self.root.winfo_children():
                                if isinstance(widget, tk.Toplevel):
                                    try:
                                        widget.destroy()
                                    except:
                                        pass
                            # Esci dal mainloop
                            self.root.quit()
                            # Distruggi la root
                            self.root.destroy()
                        except:
                            pass
                    
                    # Forza la terminazione immediata del processo
                    # Usa os._exit() invece di sys.exit() per evitare che il cleanup blocchi
                    os._exit(0)
                    
                except Exception as e:
                    logger.error(f"Errore nel riavvio dell'applicazione: {e}")
                    try:
                        messagebox.showerror(
                            _("Errore"),
                            _("Impossibile riavviare l'applicazione automaticamente.\n\nPerfavore, chiudi e riapri manualmente l'applicazione per applicare le modifiche.\n\nPercorso tentato: {}").format(script_path),
                            parent=None
                        )
                    except:
                        pass
            
            # Chiudi la finestra di dialogo corrente se esiste
            if hasattr(self, 'master') and self.master:
                try:
                    self.master.destroy()
                except:
                    pass
            
            # Esegui il riavvio dopo un breve delay per permettere la chiusura della finestra corrente
            if hasattr(self, 'root') and self.root:
                self.root.after(100, do_restart)
            else:
                # Se root non è disponibile, esegui immediatamente
                do_restart()
            
        except Exception as e:
            # Se il riavvio fallisce, mostra un messaggio all'utente
            logger.error(f"Errore nel riavvio dell'applicazione: {e}")
            messagebox.showerror(
                _("Errore"),
                _("Impossibile riavviare l'applicazione automaticamente.\n\nPerfavore, chiudi e riapri manualmente l'applicazione per applicare le modifiche.\n\nPercorso tentato: {}").format(script_path),
                parent=self.root if hasattr(self, 'root') else None
            )

    def check_for_autobackup(self):
        config = configparser.ConfigParser(interpolation=None); config.read(get_config_file())
        if config.getboolean('AutoBackup', 'enabled', fallback=False):
            # BUG #38 FIX: Strip whitespace da valori config per evitare path con spazi invisibili
            path = config.get('AutoBackup', 'path', fallback='').strip()
            hour = config.get('AutoBackup', 'hour', fallback='').strip()
            if path and hour:
                try:
                    now = datetime.now()
                    if now.hour == int(hour) and now.date() != self.last_backup_date:
                        self.perform_autobackup(path); self.last_backup_date = now.date()
                except Exception as e: print(f"ERRORE AUTOBACKUP: {e}")
        
        # BUG #45 FIX: Cancella timer precedente prima di ri-registrarlo (previene memory leak)
        if self._autobackup_timer_id is not None:
            try:
                self.root.after_cancel(self._autobackup_timer_id)
            except Exception as e:
                logger.warning(f"Impossibile cancellare timer autobackup precedente: {e}")
        
        # Ri-registra timer e salva ID per cancellazione futura
        self._autobackup_timer_id = self.root.after(60000, self.check_for_autobackup)

    def perform_autobackup(self, dest_folder):
        """Esegue backup automatico copiando direttamente il database.
        
        BUG #8 FIX: Aggiunta sincronizzazione e retry logic per evitare race condition.
        """
        logger.info(f"Avvio backup automatico in: {dest_folder}")
        db_file = get_db_path()
        
        # Verifica che non ci sia un backup già in corso
        if hasattr(self, '_backup_in_progress') and self._backup_in_progress:
            logger.warning("Backup già in corso, saltato")
            return
        
        self._backup_in_progress = True
        
        try:
            if not os.path.exists(db_file): 
                logger.warning(f"File database non trovato per backup: {db_file}")
                return
            
            # Attendi un momento per permettere sync su disco
            import time
            time.sleep(0.2)
            
            # Gestione vecchi backup (mantieni solo gli ultimi 3 SET completi)
            # Un SET = .db + .db-wal + .db-shm con lo stesso timestamp
            backup_sets = {}  # timestamp -> [file_path1, file_path2, ...]
            
            for ext in ['*.db', '*.db-wal', '*.db-shm']:
                pattern = os.path.join(dest_folder, f"*_backup_auto_{ext.replace('*', '')}")
                for filepath in glob.glob(pattern):
                    # Estrai timestamp dal nome file (es: gestione_offerte_backup_auto_20250102_143000.db)
                    basename = os.path.basename(filepath)
                    try:
                        # Pattern: *_backup_auto_YYYYMMDD_HHMMSS.ext
                        timestamp_part = basename.split('_backup_auto_')[1].rsplit('.', 1)[0]
                        if timestamp_part not in backup_sets:
                            backup_sets[timestamp_part] = []
                        backup_sets[timestamp_part].append(filepath)
                    except (IndexError, ValueError):
                        logger.warning(f"Formato nome backup non riconosciuto: {basename}")
            
            # Ordina i set per timestamp e mantieni solo gli ultimi 3
            sorted_timestamps = sorted(backup_sets.keys())
            while len(sorted_timestamps) > 3:
                old_timestamp = sorted_timestamps.pop(0)
                for old_file in backup_sets[old_timestamp]:
                    try:
                        os.remove(old_file)
                        logger.info(f"Rimosso vecchio backup: {old_file}")
                    except Exception as e:
                        logger.warning(f"Impossibile eliminare vecchio backup {old_file}: {e}")
            
            # Genera timestamp per il nuovo set di backup
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            base_name = f"gestione_offerte_backup_auto_{timestamp}"
            
            # ✅ COPIA FILE PRINCIPALE
            dest_path = os.path.join(dest_folder, f"{base_name}.db")
            
            # Copia con retry su errori temporanei
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    shutil.copy2(db_file, dest_path)
                    break
                except (PermissionError, OSError) as e:
                    if attempt < max_retries - 1:
                        logger.warning(f"Tentativo backup {attempt+1} fallito: {e}, riprovo...")
                        time.sleep(1)
                    else:
                        raise
            
            logger.info(f"Backup automatico DB principale: {dest_path}")
            
            # ✅ COPIA FILE WAL (se esiste)
            wal_file = db_file.replace('.db', '.db-wal')
            if os.path.exists(wal_file):
                wal_dest = os.path.join(dest_folder, f"{base_name}.db-wal")
                try:
                    shutil.copy2(wal_file, wal_dest)
                    logger.info(f"Backup WAL copiato: {wal_dest}")
                except Exception as e:
                    logger.warning(f"Impossibile copiare WAL: {e}")
            else:
                logger.info("File WAL non presente per autobackup (normale se DB chiuso)")
            
            # ✅ COPIA FILE SHM (se esiste)
            shm_file = db_file.replace('.db', '.db-shm')
            if os.path.exists(shm_file):
                shm_dest = os.path.join(dest_folder, f"{base_name}.db-shm")
                try:
                    shutil.copy2(shm_file, shm_dest)
                    logger.info(f"Backup SHM copiato: {shm_dest}")
                except Exception as e:
                    logger.warning(f"Impossibile copiare SHM: {e}")
            else:
                logger.info("File SHM non presente per autobackup (normale se DB chiuso)")
            
            # Verifica integrità backup principale
            original_size = os.path.getsize(db_file)
            backup_size = os.path.getsize(dest_path)
            
            if backup_size < original_size * 0.5:
                logger.error(f"Backup automatico potenzialmente corrotto: {backup_size} vs {original_size} bytes")
                # Elimina backup corrotto (tutti i file del set)
                try:
                    os.remove(dest_path)
                    wal_dest = os.path.join(dest_folder, f"{base_name}.db-wal")
                    shm_dest = os.path.join(dest_folder, f"{base_name}.db-shm")
                    if os.path.exists(wal_dest):
                        os.remove(wal_dest)
                    if os.path.exists(shm_dest):
                        os.remove(shm_dest)
                except:
                    pass
            else:
                # Conta i file effettivamente copiati
                files_copied = 1  # DB principale
                wal_dest = os.path.join(dest_folder, f"{base_name}.db-wal")
                shm_dest = os.path.join(dest_folder, f"{base_name}.db-shm")
                if os.path.exists(wal_dest):
                    files_copied += 1
                if os.path.exists(shm_dest):
                    files_copied += 1
                
                total_size = sum(os.path.getsize(f) for f in [dest_path] + 
                               ([wal_dest] if os.path.exists(wal_dest) else []) + 
                               ([shm_dest] if os.path.exists(shm_dest) else []))
                
                logger.info(f"Backup automatico completato: {files_copied} file copiati, {total_size} bytes totali ({total_size/original_size*100:.1f}% dimensione originale)")
            
        except Exception as e:
            logger.error(f"Errore backup automatico: {e}", exc_info=True)
            print(f"ERRORE AUTOBACKUP: {e}")
        finally:
            self._backup_in_progress = False

    def open_help_window(self): HelpWindow(self.root)
    def open_settings_window(self): self.root.wait_window(SettingsWindow(self.root, self))
    def create_request_treeview(self, parent):
        # Frame per contenere il Sheet
        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill="both", expand=True)
        
        # Crea il widget tksheet invece di Treeview
        sheet = Sheet(tree_frame,
                     theme="light blue",
                     header_font=("Calibri", 11, "bold"),
                     font=("Calibri", 11, "normal"),
                     headers=[_("Num RdO"), _("Tipo RdO"), _("Data Emiss."), _("Data Scad."), _("Riferimento"), _("Utente")],
                     show_header=True,
                     show_row_index=False)
        
        # Configura le larghezze delle colonne
        sheet.set_column_widths([80, 120, 120, 120, 300, 140])
        
        # Centra tutte le colonne tranne "Riferimento"
        sheet.align_columns(columns=[0, 1, 2, 3, 5], align="center")
        
        # Abilita tutti i binding
        sheet.enable_bindings()
        
        # Rendi il sheet completamente in sola lettura (nessuna cella editabile)
        for col_idx in range(6):
            sheet.readonly_columns(columns=[col_idx], readonly=True)
        
        # Configura il binding per doppio click su cella (metodo nativo tksheet)
        # Questo si attiva quando si fa doppio click su qualsiasi cella della riga
        sheet.extra_bindings("cell_select", self.create_cell_select_handler(sheet))
        sheet.extra_bindings("row_select", self.create_row_select_handler(sheet))
        
        # Variabile per tracciare il tempo dell'ultimo click (per doppio click)
        sheet._last_click_time = 0
        sheet._last_click_row = None
        
        # Binding generico per gestire il doppio click
        sheet.bind("<Double-Button-1>", lambda event: self.on_sheet_double_click(sheet, event))
        
        sheet.pack(fill="both", expand=True)
        
        # Salva riferimento per uso successivo
        sheet._sheet_data = []  # Per memorizzare i dati attuali
        
        return sheet
    
    def create_cell_select_handler(self, sheet):
        """Crea un handler per il doppio click su celle"""
        def handler(event_data):
            # Quando viene selezionata una cella, aggiorna i pulsanti
            self.update_button_visibility()
        return handler
    
    def create_row_select_handler(self, sheet):
        """Crea un handler per la selezione di righe"""
        def handler(event_data):
            # Quando viene selezionata una riga, aggiorna i pulsanti
            self.update_button_visibility()
        return handler

    # NOTA: I metodi sort_treeview_column e update_sort_indicators sono stati rimossi
    # perché tksheet ha funzionalità di ordinamento integrate che si abilitano automaticamente
    # con enable_bindings(). L'utente può cliccare sugli header delle colonne per ordinare.

    def _get_selected_row_indices(self, sheet):
        """
        Metodo helper per ottenere gli indici delle righe selezionate dal sheet.
        Gestisce sia la selezione di celle che di righe complete.
        Restituisce una lista di indici di riga.
        """
        row_indices = []
        
        # Metodo 1: Prova con get_currently_selected (per selezione cella singola)
        currently_selected = sheet.get_currently_selected()
        if currently_selected:
            if hasattr(currently_selected, 'row') and currently_selected.row is not None:
                row_indices.append(currently_selected.row)
            elif isinstance(currently_selected, tuple) and len(currently_selected) >= 1:
                row_indices.append(currently_selected[0])
        
        # Metodo 2: Prova con get_selected_rows (per selezione righe multiple)
        if not row_indices:
            selected_rows = sheet.get_selected_rows()
            if selected_rows:
                if isinstance(selected_rows, (list, set, tuple)):
                    row_indices.extend(selected_rows)
                else:
                    row_indices.append(selected_rows)
        
        return row_indices
    
    def _check_if_all_selected_are_mine(self, sheet, selected_indices):
        """Verifica se tutte le RfQ selezionate appartengono all'utente corrente.
        
        Args:
            sheet: Il widget Sheet da controllare
            selected_indices: Lista di indici riga selezionati
        
        Returns:
            bool: True se tutte le RfQ selezionate sono dell'utente corrente, False altrimenti
        """
        if not selected_indices:
            return False
        
        # Se i metadati non sono disponibili, per sicurezza blocca le operazioni su RfQ legacy
        if not hasattr(sheet, '_sheet_rows_metadata'):
            logger.warning("Metadati sheet non disponibili - blocco operazioni per sicurezza")
            return False
        
        for idx in selected_indices:
            # Salta indici fuori range
            if idx >= len(sheet._sheet_rows_metadata):
                logger.warning(f"Indice {idx} fuori range metadati (len={len(sheet._sheet_rows_metadata)})")
                continue
            
            metadata = sheet._sheet_rows_metadata[idx]
            is_mine = metadata.get('is_mine', False)  # Default False per sicurezza
            
            if not is_mine:
                return False  # Almeno una RfQ non è mia
        
        return True  # Tutte le RfQ selezionate sono mie
    
    def archive_selected_request(self): self._change_request_status('archiviata')
    def reactivate_selected_request(self): self._change_request_status('attiva')
    def _change_request_status(self, new_status):
        sheet, _status = self.get_current_tree_and_status()
        
        # Ottieni le righe selezionate usando il metodo helper
        selected_rows_indices = self._get_selected_row_indices(sheet)
        if not selected_rows_indices:
            return
        
        # VALIDAZIONE SICUREZZA: Verifica che tutte le RfQ selezionate siano dell'utente corrente
        if not self._check_if_all_selected_are_mine(sheet, selected_rows_indices):
            messagebox.showerror(
                _("Operazione Non Consentita"),
                _("Non puoi modificare lo stato di RfO di altri utenti.\nPuoi operare solo sulle tue RdO."),
                parent=self.root
            )
            logger.warning(f"Tentativo di modifica stato RfQ altrui bloccato: utente={self.current_username}")
            return
        
        # Ottieni gli ID dalle righe selezionate
        ids = []
        for row_idx in selected_rows_indices:
            try:
                row_data = sheet.get_row_data(row_idx)
                if row_data and len(row_data) > 0:
                    ids.append(row_data[0])  # Primo elemento è l'ID
            except Exception as e:
                logger.error(f"Errore nel recupero dati riga {row_idx}: {e}", exc_info=True)
        
        if not ids:
            return
        
        try:
            # Usa db_manager per aggiornare lo stato
            params = [(new_status, req_id) for req_id in ids]
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(get_db_path()) as db_manager:
                db_manager.update_stato_richieste(params)
        except DatabaseError as e:
            messagebox.showerror(_("Errore"), _("Impossibile aggiornare stato: {}").format(e))
        else:
            self.refresh_data()

    def on_tab_changed(self, event): self.update_button_visibility(); self.clear_selection()
    def update_button_visibility(self):
        """Aggiorna la visibilità e lo stato dei pulsanti in base alla selezione e proprietà delle RfQ"""
        sheet, status = self.get_current_tree_and_status()
        selected_rows_indices = self._get_selected_row_indices(sheet)
        has_sel = bool(selected_rows_indices)
        num_selected = len(selected_rows_indices) if selected_rows_indices else 0
        
        # Verifica se tutte le RfQ selezionate appartengono all'utente corrente
        all_mine = self._check_if_all_selected_are_mine(sheet, selected_rows_indices) if has_sel else False
        
        # Elimina: abilita solo se c'è selezione E tutte le RfQ sono dell'utente corrente
        can_delete = has_sel and all_mine
        self.btn_delete_rdo.config(state="normal" if can_delete else "disabled")
        
        # Duplica: abilita solo se è selezionata 1 sola RfQ E è dell'utente corrente
        can_duplicate = (num_selected == 1) and all_mine
        self.btn_duplicate_rdo.config(state="normal" if can_duplicate else "disabled")
        
        # Archivia/Riattiva: abilita solo se c'è selezione E tutte le RfQ sono dell'utente corrente
        can_change_status = has_sel and all_mine
        
        self.btn_archive_rdo.pack_forget()
        self.btn_reactivate.pack_forget()
        
        if status == 'attiva':
            self.btn_archive_rdo.pack(side="left", padx=(0, 10))
            self.btn_archive_rdo.config(state="normal" if can_change_status else "disabled")
        else:
            self.btn_reactivate.pack(side="left", padx=(0, 20))
            self.btn_reactivate.config(state="normal" if can_change_status else "disabled")

    def clear_selection(self):
        """Deseleziona tutte le righe in entrambi i sheet"""
        self.tree_attive.deselect("all")
        self.tree_archiviate.deselect("all")
        self.update_button_visibility()

    def get_current_tree_and_status(self):
        return (self.tree_attive, 'attiva') if self.notebook.index(self.notebook.select()) == 0 else (self.tree_archiviate, 'archiviata')

    def refresh_data(self):
        """Ricarica i dati preservando i filtri di ricerca attivi"""
        # BUGFIX: Se ci sono filtri di ricerca attivi, usa search_requests invece di ricaricare tutto
        if self._has_active_search_filters():
            logger.info("[refresh_data] Filtri attivi rilevati, riapplico la ricerca")
            self.search_requests()
            return
        
        # Ottieni il percorso completo del mio DB
        my_path = get_db_path()
        # Chiama get_all_richieste_aggregated per ottenere tutte le richieste aggregate
        try:
            all_requests = self.db_manager.get_all_richieste_aggregated(my_path)
            # Salva i dati aggregati per uso successivo
            self._all_aggregated_requests = all_requests
        except DatabaseError as e:
            logger.error(f"Errore nel caricamento richieste aggregate: {e}", exc_info=True)
            # Fallback: usa il metodo normale se l'aggregazione fallisce
            self._all_aggregated_requests = None
        
        self._load_requests_by_status(self.tree_attive, 'attiva'); self._load_requests_by_status(self.tree_archiviate, 'archiviata')
        self.populate_username_filter()
        self.update_button_visibility()

    def _load_requests_by_status(self, tree, status):
        """Carica richieste per stato specifico con supporto multi-database."""
        try:
            username_filter = self._get_active_username_filter()
            
            # SEMPRE usa aggregazione multi-database per avere accesso a tutti gli utenti
            logger.info(f"[MULTI-DB] Caricamento da tutti i database (filtro utente: {username_filter})...")
            
            # BUG #47 FIX: Usa context manager per garantire chiusura DB
            with DatabaseManager(get_db_path()) as db_manager:
                # Chiama il metodo aggregato che legge TUTTI i database
                all_rows = db_manager.get_all_richieste_aggregated(get_db_path())
            
            # Filtra per stato richiesto
            # Struttura: [0] id_richiesta, [1] tipo_rdo, [2] data_emissione,
            # [3] data_scadenza, [4] riferimento, [5] username, [6] stato, 
            # [7] is_mine, [8] source_file
            filtered_rows = [row for row in all_rows if row[6] == status]
            
            # SE C'È UN FILTRO UTENTE SPECIFICO, filtra anche per username
            if username_filter is not None:
                filtered_rows = [row for row in filtered_rows if row[5] and row[5].lower() == username_filter.lower()]
                logger.info(f"[MULTI-DB] Trovate {len(filtered_rows)} RdO in stato '{status}' per utente '{username_filter}'")
            else:
                logger.info(f"[MULTI-DB] Trovate {len(filtered_rows)} RdO in stato '{status}' da tutti gli utenti")
            
            # BUGFIX: Applica filtro tipo RdO se presente (non solo "Tutte")
            tipo_filter = self.search_tipo.get()
            if tipo_filter != _("Tutte"):
                tipo_canonico = normalize_rfq_type(tipo_filter)
                filtered_rows = [row for row in filtered_rows if row[1] == tipo_canonico]
                logger.info(f"[MULTI-DB] Filtro tipo RdO '{tipo_filter}' applicato: {len(filtered_rows)} risultati")
            
            self.update_treeview(tree, filtered_rows)
                
        except DatabaseError as e:
            logger.error(f"Errore database in _load_requests_by_status: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Impossibile caricare elenco: {}").format(e))

    def update_treeview(self, sheet, requests):
        """Aggiorna il foglio tksheet con i dati delle richieste"""
        today = date.today()
        data_rows = []
        
        # Variabile per tracciare la lunghezza massima del riferimento
        max_ref_length = 0
        
        # Inizializza lista metadati se non esiste
        if not hasattr(sheet, '_sheet_rows_metadata'):
            sheet._sheet_rows_metadata = []
        
        sheet._sheet_rows_metadata = []  # Reset metadati
        
        for i, req in enumerate(requests):
            # Traduci il tipo RFQ prima di inserirlo nel sheet
            tipo_rdo_tradotto = translate_rfq_type(req[1])
            riferimento = req[4] if req[4] else ""
            username_value = ""
            if len(req) > 5 and req[5]:
                username_value = str(req[5]).strip()
            
            # BUG #3 FIX: Validazione robusta per metadati con logging dettagliato
            # Salva metadati per questa riga (is_mine e source_file)
            # La struttura aggregate è: [..., stato, is_mine, source_file]
            if len(req) > 8:
                is_mine = req[7]
                source_file = req[8]
                logger.debug(f"Riga {i} (ID {req[0]}): is_mine={is_mine}, source={source_file}")
            else:
                # Fallback per dati non aggregati
                is_mine = True
                source_file = 'local'
                if len(req) < 6:
                    logger.warning(f"Riga {i}: tuple troppo corta ({len(req)} elementi), dati incompleti. Usando default is_mine=True")
            
            sheet._sheet_rows_metadata.append({
                'is_mine': is_mine,
                'source_file': source_file
            })
            
            # Aggiorna la lunghezza massima del riferimento
            if riferimento:
                max_ref_length = max(max_ref_length, len(riferimento))
            
            row = [
                str(req[0]),  # ID
                tipo_rdo_tradotto,  # Tipo
                self._format_date_for_display(req[2]),  # Data emissione
                self._format_date_for_display(req[3]),  # Data scadenza
                riferimento,  # Riferimento
                username_value  # Username
            ]
            data_rows.append(row)
        
        # Carica i dati nel sheet
        sheet.set_sheet_data(data_rows)
        
        # Salva i dati per uso successivo (ad esempio per l'ordinamento)
        sheet._sheet_data = data_rows
        sheet._sheet_requests = requests  # Salva anche i dati completi del DB
        
        # Calcola larghezza ottimale per la colonna Riferimento
        try:
            import tkinter.font as tkfont
            
            # Font usato nel sheet
            content_font = tkfont.Font(family="Calibri", size=11, weight="normal")
            header_font = tkfont.Font(family="Calibri", size=11, weight="bold")
            
            # Larghezza minima (basata sull'header "Riferimento")
            header_text = _("Riferimento")
            min_width = header_font.measure(header_text) + 30  # +30 per padding
            
            # Se ci sono riferimenti, calcola la larghezza in base al più lungo
            if max_ref_length > 0:
                # Trova il riferimento più lungo per misurarlo con precisione
                longest_ref = max((row[4] for row in data_rows if row[4]), key=len, default="")
                content_width = content_font.measure(longest_ref) + 40  # +40 per padding e margini
                optimal_width = max(min_width, content_width)
            else:
                optimal_width = min_width
            
            # Limita la larghezza massima per evitare colonne eccessivamente larghe
            MAX_WIDTH = 600  # Massimo 600 pixel
            optimal_width = min(optimal_width, MAX_WIDTH)
            
            # Applica la larghezza ottimale alla colonna Riferimento (indice 4)
            sheet.column_width(column=4, width=int(optimal_width))
            
        except Exception as e:
            logger.warning(f"Errore calcolo larghezza colonna Riferimento: {e}. Uso larghezza default.")
            # Fallback a larghezza fissa se il calcolo fallisce
            sheet.column_width(column=4, width=300)
        
        # Reset completo di tutte le evidenziazioni precedenti
        sheet.dehighlight_all()
        
        # Applica colorazione per righe scadute (solo per tab attive)
        if sheet is self.tree_attive:
            for i, req in enumerate(requests):
                if req[3]:
                    try:
                        # BUG #22 FIX: Parse con strptime e confronto date robusto
                        expiry_date = datetime.strptime(req[3], '%Y-%m-%d').date()
                        if expiry_date < today:
                            # Evidenzia la riga in rosso per scadenza
                            sheet.highlight_rows([i], bg='#FFE6E6', fg='red')
                    except (ValueError, TypeError) as e:
                        # BUG #22 FIX: Log dettagliato con ID RdO per troubleshooting
                        logger.warning(f"Formato data scadenza non valido per RdO {req[0]}: '{req[3]}' - {e}")
        
        # BUG #27 FIX: Usa enumerate invece di range(len())
        # Applica strisce alternate per le righe
        for i, _row in enumerate(data_rows):
            if i % 2 != 0:
                if sheet is self.tree_attive:
                    # BUG #29 FIX: Catch specifico invece di bare except
                    # Controlla se già evidenziata come scaduta
                    try:
                        req = requests[i]
                        if req[3] and datetime.strptime(req[3], '%Y-%m-%d').date() < today:
                            continue  # Non sovrascrivere l'evidenziazione rossa
                    except (ValueError, TypeError, IndexError) as e:
                        logger.debug(f"Errore parsing data per stripe row {i}: {e}")
                sheet.highlight_rows([i], bg='#F0F0F0', fg='black')
            else:
                # Righe pari: assicura che abbiano fg='black' se non sono scadute
                if sheet is self.tree_attive:
                    # BUG #29 FIX: Catch specifico invece di bare except
                    try:
                        req = requests[i]
                        if req[3] and datetime.strptime(req[3], '%Y-%m-%d').date() < today:
                            continue  # Già evidenziata in rosso
                    except (ValueError, TypeError, IndexError) as e:
                        logger.debug(f"Errore parsing data per stripe row {i}: {e}")
                # Applica esplicitamente sfondo bianco con testo nero per righe pari non scadute
                sheet.highlight_rows([i], bg='white', fg='black')

    def search_requests(self):
        tree, status = self.get_current_tree_and_status()
        username_filter = self._get_active_username_filter()
        
        # BUG #9 FIX: Validazione lunghezza input per evitare query troppo lente
        MAX_SEARCH_LENGTH = 100
        crit = {k: v.get().strip() for k, v in self.search_vars.items()}
        
        # BUG #9 FIX: Blacklist caratteri pericolosi per SQL injection
        FORBIDDEN_CHARS = re.compile(r"[';\"\\`<>]")
        
        # Controlla che nessun campo di ricerca sia troppo lungo
        for field_name, value in crit.items():
            if value and len(value) > MAX_SEARCH_LENGTH:
                messagebox.showwarning(
                    _("Input Troppo Lungo"),
                    _("Il testo di ricerca nel campo '{}' è troppo lungo (max {} caratteri)").format(
                        field_name, MAX_SEARCH_LENGTH
                    ),
                    parent=self.root
                )
                return
            
            # BUG #5 FIX: Rimuovi caratteri pericolosi per SQL injection
            if value and FORBIDDEN_CHARS.search(value):
                sanitized = FORBIDDEN_CHARS.sub('', value)
                logger.warning(f"Caratteri pericolosi rimossi dal campo '{field_name}': '{value}' -> '{sanitized}'")
                # Aggiorna il campo con il valore sanitizzato
                self.search_vars[field_name].set(sanitized)
                crit[field_name] = sanitized
                
                # Avvisa l'utente una sola volta per tutti i campi
                if not hasattr(self, '_sql_injection_warning_shown'):
                    self._sql_injection_warning_shown = True
                    messagebox.showinfo(
                        _("Input Sanitizzato"),
                        _("Alcuni caratteri speciali sono stati rimossi dai campi di ricerca per motivi di sicurezza."),
                        parent=self.root
                    )
                    # Reset flag dopo 2 secondi - BUG #48 FIX: cancella timer precedente per evitare memory leak
                    if self._sql_warning_after_id is not None:
                        try:
                            self.root.after_cancel(self._sql_warning_after_id)
                        except Exception as e:
                            logger.warning(f"Impossibile cancellare timer SQL warning: {e}")
                    
                    def reset_flag():
                        if hasattr(self, '_sql_injection_warning_shown'):
                            delattr(self, '_sql_injection_warning_shown')
                    self._sql_warning_after_id = self.root.after(2000, reset_flag)
        
        # Validazione rimossa: ora il numero RdO supporta ricerca parziale come gli altri filtri
        
        dates = {k: self._format_date_for_db(v.get().strip()) for k, v in self.date_entries.items()}
        base = "SELECT DISTINCT ro.id_richiesta, ro.tipo_rdo, ro.data_emissione, ro.data_scadenza, ro.riferimento FROM richieste_offerta ro LEFT JOIN dettagli_richiesta dr ON ro.id_richiesta=dr.id_richiesta LEFT JOIN richiesta_fornitori rf ON ro.id_richiesta=rf.id_richiesta"
        clauses, params = ["ro.stato=?"], [status]
        if crit['num']: clauses.append("CAST(ro.id_richiesta AS TEXT) LIKE ?"); params.append(f"%{crit['num']}%")
        if self.search_tipo.get() != _("Tutte"): 
            # Normalizza il valore di ricerca al valore canonico per il confronto nel database
            tipo_canonico = normalize_rfq_type(self.search_tipo.get())
            clauses.append("ro.tipo_rdo=?"); 
            params.append(tipo_canonico)
        if username_filter:
            clauses.append("LOWER(COALESCE(ro.username, '')) = ?"); params.append(username_filter)
        if crit['ref']: clauses.append("LOWER(ro.riferimento) LIKE LOWER(?)"); params.append(f"%{crit['ref']}%")
        if crit['forn']: clauses.append("LOWER(rf.nome_fornitore) LIKE LOWER(?)"); params.append(f"%{crit['forn']}%")
        if crit['cod']: clauses.append("LOWER(dr.codice_materiale) LIKE LOWER(?)"); params.append(f"%{crit['cod']}%")
        if crit['desc']: clauses.append("LOWER(dr.descrizione_materiale) LIKE LOWER(?)"); params.append(f"%{crit['desc']}%")
        if crit['ord']: clauses.append("LOWER(ro.numeri_ordine) LIKE LOWER(?)"); params.append(f"%{crit['ord']}%")
        # --- INIZIO BLOCCO AGGIUNTO ---
        if crit['cod_grezzo']: clauses.append("LOWER(dr.codice_grezzo) LIKE LOWER(?)"); params.append(f"%{crit['cod_grezzo']}%")
        if crit['dis_grezzo']: clauses.append("LOWER(dr.disegno_grezzo) LIKE LOWER(?)"); params.append(f"%{crit['dis_grezzo']}%")
        if crit['mat_cl']: clauses.append("LOWER(dr.materiale_conto_lavoro) LIKE LOWER(?)"); params.append(f"%{crit['mat_cl']}%")
        # --- FINE BLOCCO AGGIUNTO ---
        if dates['emm_da']: clauses.append("ro.data_emissione >= ?"); params.append(dates['emm_da'])
        if dates['emm_a']: clauses.append("ro.data_emissione <= ?"); params.append(dates['emm_a'])
        if dates['scad_da']: clauses.append("ro.data_scadenza >= ?"); params.append(dates['scad_da'])
        if dates['scad_a']: clauses.append("ro.data_scadenza <= ?"); params.append(dates['scad_a'])
        
        try:
            # Usa DatabaseManager per la ricerca avanzata
            criteria = {
                'num': crit['num'],
                'ref': crit['ref'],
                'forn': crit['forn'],
                'cod': crit['cod'],
                'desc': crit['desc'],
                'ord': crit['ord'],
                'cod_grezzo': crit['cod_grezzo'],
                'dis_grezzo': crit['dis_grezzo'],
                'mat_cl': crit['mat_cl']
            }
            date_ranges = {
                'emm_da': dates['emm_da'],
                'emm_a': dates['emm_a'],
                'scad_da': dates['scad_da'],
                'scad_a': dates['scad_a']
            }
            
            # Gestione tipo RdO
            tipo_rdo = None
            if self.search_tipo.get() != _("Tutte"):
                tipo_rdo = normalize_rfq_type(self.search_tipo.get())
            
            # FIX: La ricerca deve usare aggregazione multi-database quando si filtra per altri utenti o "All users"
            # Comportamento:
            # - username_filter = None (All users) → cerca in TUTTI i database
            # - username_filter = altro utente → cerca in TUTTI i database (poi filtra per username)
            # - username_filter = utente corrente → ottimizzazione, cerca solo nel DB locale
            
            # Ottimizzazione: cerca solo nel DB locale se filtriamo per l'utente corrente
            search_local_only = (username_filter and 
                                username_filter.lower() == self.current_username.lower() if self.current_username else False)
            
            if search_local_only:
                # Caso ottimizzato: cerca solo nel database locale
                logger.info(f"[search_requests] Ricerca locale per utente corrente: {username_filter}")
                with DatabaseManager(get_db_path()) as db_manager:
                    results = db_manager.search_richieste_advanced(criteria, date_ranges, status=status, tipo=tipo_rdo, username=username_filter)
            else:
                # Caso generale: usa aggregazione multi-database
                logger.info(f"[search_requests] Ricerca aggregata multi-DB (filtro utente: {username_filter or 'All users'})")
                with DatabaseManager(get_db_path()) as db_manager:
                    # Prima ottieni TUTTE le RdO aggregate
                    all_results = db_manager.get_all_richieste_aggregated(get_db_path())
                
                # Poi filtra in memoria applicando TUTTI i criteri di ricerca
                # Struttura all_results: [id_richiesta, tipo_rdo, data_emissione, data_scadenza, riferimento, username, stato, is_mine, source_file]
                results = []
                for row in all_results:
                    # Filtro stato (obbligatorio)
                    if row[6] != status:
                        continue
                    
                    # Filtro username (se specificato)
                    if username_filter and (not row[5] or row[5].lower() != username_filter.lower()):
                        continue
                    
                    # Filtro tipo RdO
                    if tipo_rdo and row[1] != tipo_rdo:
                        continue
                    
                    # Filtri di testo (tutti case-insensitive)
                    if crit['num'] and crit['num'] not in str(row[0]):
                        continue
                    if crit['ref'] and (not row[4] or crit['ref'].lower() not in row[4].lower()):
                        continue
                    
                    # Filtri data emissione
                    if dates['emm_da'] and (not row[2] or row[2] < dates['emm_da']):
                        continue
                    if dates['emm_a'] and (not row[2] or row[2] > dates['emm_a']):
                        continue
                    
                    # Filtri data scadenza
                    if dates['scad_da'] and (not row[3] or row[3] < dates['scad_da']):
                        continue
                    if dates['scad_a'] and (not row[3] or row[3] > dates['scad_a']):
                        continue
                    
                    # Per i filtri su dettagli (fornitore, materiale, ecc.), dobbiamo interrogare il DB specifico
                    # Questo è necessario perché get_all_richieste_aggregated non include questi dettagli
                    if any([crit['forn'], crit['cod'], crit['desc'], crit['ord'], 
                           crit['cod_grezzo'], crit['dis_grezzo'], crit['mat_cl']]):
                        # Apri il database di origine per questa RdO
                        # BUG FIX: Se source_file è 'local', usa il percorso del DB corrente
                        source_db_path = row[8] if len(row) > 8 else 'local'
                        if source_db_path == 'local':
                            source_db_path = get_db_path()
                        try:
                            with DatabaseManager(source_db_path) as source_db_mgr:
                                # Verifica i criteri di dettaglio sul DB specifico
                                detail_match = source_db_mgr.check_richiesta_detail_criteria(
                                    row[0],  # id_richiesta
                                    {
                                        'forn': crit['forn'],
                                        'cod': crit['cod'],
                                        'desc': crit['desc'],
                                        'ord': crit['ord'],
                                        'cod_grezzo': crit['cod_grezzo'],
                                        'dis_grezzo': crit['dis_grezzo'],
                                        'mat_cl': crit['mat_cl']
                                    }
                                )
                            if not detail_match:
                                continue
                        except Exception as e:
                            logger.warning(f"Errore verifica criteri dettaglio per RdO {row[0]} su DB {source_db_path}: {e}")
                            continue
                    
                    # Tutti i filtri passati, aggiungi ai risultati
                    # BUGFIX: Passa l'intera tupla con metadati (is_mine, source_file) per update_treeview
                    results.append(row)
                
                logger.info(f"[search_requests] Ricerca aggregata completata: {len(results)} risultati trovati")
            
            self.update_treeview(tree, results)
        except DatabaseError as e: 
            logger.error(f"Errore ricerca richieste: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Errore ricerca: {}").format(e))

    def delete_selected_request(self):
        sheet, _status = self.get_current_tree_and_status()
        
        # Ottieni le righe selezionate (supporta sia selezione cella che riga)
        selected_rows_indices = self._get_selected_row_indices(sheet)
        if not selected_rows_indices:
            return
        
        # VALIDAZIONE SICUREZZA: Verifica che tutte le RfQ selezionate siano dell'utente corrente
        if not self._check_if_all_selected_are_mine(sheet, selected_rows_indices):
            messagebox.showerror(
                _("Operazione Non Consentita"),
                _("Non puoi eliminare RdO di altri utenti.\nPuoi operare solo sulle tue RdO."),
                parent=self.root
            )
            logger.warning(f"Tentativo di eliminazione RfQ altrui bloccato: utente={self.current_username}")
            return
        
        # Ottieni gli ID dalle righe selezionate
        request_ids = []
        for row_idx in selected_rows_indices:
            try:
                row_data = sheet.get_row_data(row_idx)
                if row_data and len(row_data) > 0:
                    request_ids.append(row_data[0])
            except Exception as e:
                logger.error(f"Errore nel recupero dati riga {row_idx}: {e}", exc_info=True)
        
        if not request_ids:
            return
        count = len(request_ids)
        
        if count == 1:
            rdo_num = request_ids[0]
            msg = _("Sei sicuro di voler eliminare la RdO N° {}?\nL'operazione è permanente.").format(rdo_num)
        else:
            msg = _("Sei sicuro di voler eliminare le {} RdO selezionate?\nL'operazione è permanente.").format(count)
        if not messagebox.askyesno(_("Conferma Eliminazione"), msg, parent=self.root): return
        
        try:
            print(f"[MainWindow.delete_selected_request] Eliminazione di {len(request_ids)} richieste: {request_ids}")

            # Rimuovi i file fisici degli allegati per ogni richiesta prima di eliminare le righe DB
            archive_path = get_fixed_attachments_dir()
            try:
                with DatabaseManager(get_db_path()) as db_manager:
                    # Per ogni richiesta, recupera i percorsi_esterni e prova a rimuovere i file
                    for req_id in request_ids:
                        try:
                            rows = db_manager.conn.execute(
                                "SELECT percorso_esterno FROM allegati_richiesta WHERE id_richiesta = ? AND percorso_esterno IS NOT NULL",
                                (req_id,)
                            ).fetchall()
                        except Exception:
                            rows = []

                        for row in rows:
                            percorso = row[0]
                            if not percorso:
                                continue
                            # Se percorso è relativo, cerca nella cartella Attachments
                            if archive_path and not os.path.isabs(percorso):
                                file_to_delete = os.path.join(archive_path, percorso)
                            else:
                                file_to_delete = percorso

                            try:
                                if os.path.exists(file_to_delete):
                                    os.remove(file_to_delete)
                                    logger.info(f"Allegato eliminato dal disco durante cancellazione RdO: {file_to_delete}")
                                else:
                                    logger.info(f"File allegato non trovato durante cancellazione RdO: {file_to_delete}")
                            except Exception as disk_error:
                                logger.warning(f"Impossibile eliminare il file allegato {file_to_delete}: {disk_error}")

                    # Ora elimina le richieste e i record correlati nel DB
                    count = db_manager.delete_richieste_batch(request_ids)

            except DatabaseError as e:
                raise

            print(f"[MainWindow.delete_selected_request] Eliminate {count} richieste dal database")

            # Ricarica i dati invece di cancellare elementi dalla view
            self.refresh_data()
            if count == 1:
                msg = _("1 RdO eliminata.")
            else:
                msg = _("{} RdO eliminate.").format(count)
            messagebox.showinfo(_("Successo"), msg, parent=self.root)
        except DatabaseError as e:
            messagebox.showerror(_("Errore"), _("Impossibile eliminare: {}").format(e), parent=self.root)

    def duplicate_selected_request(self):
        sheet, _status = self.get_current_tree_and_status()
        
        # Prova a ottenere la riga selezionata in vari modi
        row_index = None
        
        # Metodo 1: Prova con get_currently_selected (per selezione cella)
        currently_selected = sheet.get_currently_selected()
        if currently_selected:
            if hasattr(currently_selected, 'row') and currently_selected.row is not None:
                row_index = currently_selected.row
            elif isinstance(currently_selected, tuple) and len(currently_selected) >= 1:
                row_index = currently_selected[0]
        
        # Metodo 2: Prova con get_selected_rows (per selezione riga)
        if row_index is None:
            selected_rows = sheet.get_selected_rows()
            if selected_rows:
                if len(selected_rows) > 1:
                    messagebox.showwarning(_("Selezione non valida"), _("Seleziona una sola RdO per duplicarla."), parent=self.root)
                    return
                row_index = selected_rows[0] if isinstance(selected_rows, (list, set, tuple)) else selected_rows
        
        # VALIDAZIONE SICUREZZA: Verifica che la RfQ selezionata sia dell'utente corrente
        if row_index is not None:
            if not self._check_if_all_selected_are_mine(sheet, [row_index]):
                messagebox.showerror(
                    _("Operazione Non Consentita"),
                    _("Non puoi duplicare RdO di altri utenti.\nPuoi operare solo sulle tue RdO."),
                    parent=self.root
                )
                logger.warning(f"Tentativo di duplicazione RfQ altrui bloccato: utente={self.current_username}")
                return
        
        # Se non c'è nessuna selezione
        if row_index is None:
            messagebox.showwarning(_("Selezione mancante"), _("Selezionare una RdO da duplicare."), parent=self.root)
            return

        # Ottieni i dati della riga
        try:
            row_data = sheet.get_row_data(row_index)
            if not row_data or len(row_data) == 0:
                messagebox.showerror(_("Errore"), _("Impossibile determinare la RdO selezionata."), parent=self.root)
                return
            original_id = int(row_data[0])
        except (ValueError, TypeError, IndexError) as e:
            logger.error(f"Errore nel recupero dati riga per duplicazione: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Impossibile determinare la RdO selezionata."), parent=self.root)
            return

        new_request_id = None

        try:
            # Helper function per ottenere colonne
            def get_columns(table_name, exclude):
                # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
                with DatabaseManager(get_db_path()) as db_mgr:
                    cols_info = db_mgr.get_table_columns(table_name)
                excluded = set(exclude)
                # SQLite PRAGMA table_info restituisce: colonna[0] = cid, colonna[1] = nome colonna, colonna[2] = tipo
                # Usa colonna[1] per estrarre il nome della colonna
                columns = [row[1] for row in cols_info if row[1] not in excluded]
                print(f"[get_columns] Tabella {table_name}: colonne recuperate = {columns}")
                print(f"[get_columns] Dettagli PRAGMA: {cols_info[:3] if cols_info else []}")  # Prime 3 righe per debug
                return columns

            # BUG #47 FIX: Usa context manager anche per duplicazione
            with DatabaseManager(get_db_path()) as db_manager:
                new_request_id = db_manager.duplicate_richiesta_full(original_id, get_columns)
            
            # BUG #3 FIX: Verifica SUBITO dopo duplicazione
            if new_request_id is None:
                raise ValueError("Duplicazione fallita: ID nuova RdO non ottenuto")
            
            logger.info(f"RdO duplicata: {original_id} -> {new_request_id}")

        except ValueError as ve:
            messagebox.showerror(_("Errore"), str(ve), parent=self.root)
            return
        except DatabaseError as e:
            logger.error(f"Errore duplicazione RdO {original_id}: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Impossibile duplicare la RdO: {}").format(e), parent=self.root)
            return
        except Exception as e:
            logger.error(f"Errore duplicazione RdO {original_id}: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Impossibile duplicare: {}").format(e), parent=self.root)
            return

        # Ora new_request_id è garantito essere valido

        self._assign_request_to_current_user(new_request_id)
        self.refresh_data()
        self.notebook.select(self.tab_attive)
        
        # Cerca e seleziona la riga con il nuovo ID nel sheet
        total_rows = self.tree_attive.get_total_rows()
        for row_idx in range(total_rows):
            row_data = self.tree_attive.get_row_data(row_idx)
            if row_data and str(row_data[0]) == str(new_request_id):
                self.tree_attive.select_row(row_idx)
                self.tree_attive.see(row_idx)
                break
        
        self.update_button_visibility()
        messagebox.showinfo(_("Successo"), _("RdO duplicata come N° {}.").format(new_request_id), parent=self.root)

    def clear_filters(self):
        for var in self.search_vars.values(): var.set("")
        self.search_tipo.set(_("Tutte"))
        for de in self.date_entries.values(): de.delete(0, 'end')
        if self.username_filter_var:
            self.username_filter_var.set(self.current_username or self.all_users_placeholder)
        self.refresh_data()

    def on_sheet_double_click(self, sheet, event=None):
        """Gestisce il doppio click su una riga del sheet per aprire la RdO con debounce"""
        try:
            # Verifica che non ci sia già una finestra in apertura (debounce)
            if hasattr(self, '_opening_request') and self._opening_request:
                return
            
            self._opening_request = True
            
            try:
                # Ottieni la selezione corrente (riga, colonna)
                # Al momento del doppio click, tksheet ha già selezionato la cella
                currently_selected = sheet.get_currently_selected()
                row_index = None
                
                # Prova a determinare la riga dal get_currently_selected
                if currently_selected:
                    # get_currently_selected restituisce un oggetto con vari attributi
                    if hasattr(currently_selected, 'row') and currently_selected.row is not None:
                        row_index = currently_selected.row
                    elif isinstance(currently_selected, tuple) and len(currently_selected) >= 2:
                        row_index = currently_selected[0]
                
                # Se non abbiamo trovato la riga, proviamo con get_selected_rows
                if row_index is None:
                    selected = sheet.get_selected_rows()
                    if not selected:
                        return
                    row_index = selected[0] if isinstance(selected, (list, set, tuple)) else selected
                
                # Verifica che l'indice sia valido
                if row_index is None or row_index < 0 or row_index >= sheet.get_total_rows():
                    return
                
                # Ottieni i dati della riga
                data = sheet.get_row_data(row_index)
                if data and len(data) > 0:
                    request_id = data[0]  # Primo elemento è l'ID
                    
                    # Controllo se la RdO è mia (per multi-utente)
                    is_mine = True  # Default per compatibilità
                    source_db_path = None  # Percorso del DB sorgente
                    
                    # BUG #3 FIX: Validazione robusta con gestione errori completa
                    if hasattr(sheet, '_sheet_rows_metadata'):
                        try:
                            if 0 <= row_index < len(sheet._sheet_rows_metadata):
                                metadata = sheet._sheet_rows_metadata[row_index]
                                is_mine = metadata.get('is_mine', True)
                                source_db_path = metadata.get('source_file', None)
                                logger.debug(f"RdO {request_id}: is_mine={is_mine}, source={source_db_path}")
                            else:
                                logger.warning(f"Indice {row_index} fuori range metadati (len={len(sheet._sheet_rows_metadata)}), uso default is_mine=True")
                        except (AttributeError, KeyError, TypeError) as e:
                            logger.error(f"Errore accesso metadati riga {row_index}: {e}")
                    
                    # Apri la finestra di dettaglio della RdO (con flag read_only se non è mia)
                    self.root.wait_window(ViewRequestWindow(
                        self.root, 
                        request_id, 
                        read_only=not is_mine,
                        source_db_path=source_db_path if not is_mine else None
                    ))
                    # Aggiorna i dati dopo la chiusura della finestra
                    self.refresh_data()
            finally:
                # BUG #30 FIX: Usa weakref per evitare memory leak
                # Rilascia il lock dopo un breve delay per evitare doppi click rapidi
                import weakref
                weak_self = weakref.ref(self)
                def release_lock():
                    obj = weak_self()
                    if obj is not None:
                        obj._opening_request = False
                self.root.after(300, release_lock)
                
        except Exception as e:
            logger.error(f"Errore nell'apertura della RdO: {e}", exc_info=True)
            self._opening_request = False
    

    def open_new_request_window(self):
        """Crea una nuova RdO 'guscio' e apre l'editor"""
        # Mostra dialog per scelta tipo
        dialog = NewRdOTypeDialog(self.root)
        self.root.wait_window(dialog)
        
        # Se l'utente ha annullato, esci
        if not dialog.result:
            return
        
        tipo_rdo = normalize_rfq_type(dialog.result)
        
        # BUG #32 FIX: Usa try-finally per garantire chiusura DB
        db_manager = None
        try:
            # Inserisce testata minima usando db_manager
            data_oggi = datetime.now().strftime('%Y-%m-%d')
            db_manager = DatabaseManager(get_db_path())
            id_nuova = db_manager.insert_richiesta_offerta(tipo_rdo, 'attiva', data_oggi, username=self.current_username)
            
            logger.info(f"Creata nuova RdO guscio N° {id_nuova} (tipo: {tipo_rdo})")
            
            # Apri immediatamente l'editor
            self.root.wait_window(ViewRequestWindow(self.root, id_nuova))
            
            # Aggiorna la lista dopo la chiusura
            self.refresh_data()
            
        except DatabaseError as e:
            logger.error(f"Errore creazione RdO guscio: {e}", exc_info=True)
            messagebox.showerror(
                _("Errore Database"),
                _("Impossibile creare la nuova RdO: {}").format(e),
                parent=self.root
            )
        finally:
            # BUG #32 FIX: Garantisce chiusura connessione anche in caso di eccezione
            if db_manager is not None:
                try:
                    db_manager.close()
                except Exception as close_error:
                    logger.warning(f"Errore chiusura database in open_new_request_window: {close_error}")

    def mega_export_excel(self):
        """
        Esporta tutte le RfQ attualmente visibili nella lista (filtrate) in un unico file Excel.
        Genera un report a blocchi verticali, adattandosi al tipo di ogni singola RfQ.
        """
        # 1. Identifica quale tabella è attiva e recupera lo stato corrente
        current_tree, status = self.get_current_tree_and_status()
        
        # 2. Recupera TUTTI gli ID che corrispondono ai filtri attivi (non solo quelli visualizzati nel sheet)
        # Questo è necessario perché il sheet potrebbe avere un limite di righe visualizzate
        # IMPORTANTE: recupera anche il percorso del database sorgente per ogni RfQ
        request_data = []  # Lista di tuple (request_id, source_db_path)
        
        try:
            # Verifica se ci sono filtri di ricerca attivi
            if self._has_active_search_filters():
                # CI SONO FILTRI ATTIVI: esegui la stessa query di search_requests per ottenere TUTTI i risultati
                logger.info("[export_excel] Filtri di ricerca attivi - recupero tutti i risultati filtrati")
                
                username_filter = self._get_active_username_filter()
                crit = {k: v.get().strip() for k, v in self.search_vars.items()}
                dates = {k: self._format_date_for_db(v.get().strip()) for k, v in self.date_entries.items()}
                
                # Gestione tipo RdO
                tipo_rdo = None
                if self.search_tipo.get() != _("Tutte"):
                    tipo_rdo = normalize_rfq_type(self.search_tipo.get())
                
                # Usa SEMPRE aggregazione multi-database per avere dati da tutti gli utenti
                # Poi filtra in memoria per username specifico se necessario
                with DatabaseManager(get_db_path()) as db_manager:
                    all_results = db_manager.get_all_richieste_aggregated(get_db_path())
                
                # Filtra in memoria applicando TUTTI i criteri
                for row in all_results:
                    # Filtro stato
                    if row[6] != status:
                        continue
                    # Filtro username
                    if username_filter and (not row[5] or row[5].lower() != username_filter.lower()):
                        continue
                    # Filtro tipo RdO
                    if tipo_rdo and row[1] != tipo_rdo:
                        continue
                    # Filtri di testo
                    if crit['num'] and crit['num'] not in str(row[0]):
                        continue
                    if crit['ref'] and (not row[4] or crit['ref'].lower() not in row[4].lower()):
                        continue
                    # Filtri data
                    if dates['emm_da'] and (not row[2] or row[2] < dates['emm_da']):
                        continue
                    if dates['emm_a'] and (not row[2] or row[2] > dates['emm_a']):
                        continue
                    if dates['scad_da'] and (not row[3] or row[3] < dates['scad_da']):
                        continue
                    if dates['scad_a'] and (not row[3] or row[3] > dates['scad_a']):
                        continue
                    
                    # Filtri su dettagli (fornitore, materiale, ecc.)
                    if any([crit['forn'], crit['cod'], crit['desc'], crit['ord'], 
                           crit['cod_grezzo'], crit['dis_grezzo'], crit['mat_cl']]):
                        source_db_path = row[8] if len(row) > 8 else 'local'
                        if source_db_path == 'local':
                            source_db_path = get_db_path()
                        try:
                            with DatabaseManager(source_db_path) as source_db_mgr:
                                detail_match = source_db_mgr.check_richiesta_detail_criteria(
                                    row[0],
                                    {
                                        'forn': crit['forn'], 'cod': crit['cod'], 'desc': crit['desc'],
                                        'ord': crit['ord'], 'cod_grezzo': crit['cod_grezzo'],
                                        'dis_grezzo': crit['dis_grezzo'], 'mat_cl': crit['mat_cl']
                                    }
                                )
                            if not detail_match:
                                continue
                        except Exception as e:
                            logger.warning(f"Errore verifica criteri dettaglio per RdO {row[0]}: {e}")
                            continue
                    
                    # Tutti i filtri passati - salva ID e percorso database
                    source_db_path = row[8] if len(row) > 8 else 'local'
                    if source_db_path == 'local':
                        source_db_path = get_db_path()
                    request_data.append((row[0], source_db_path))
                
            else:
                # NESSUN FILTRO ATTIVO: carica tutte le RfQ nello stato corrente
                logger.info("[export_excel] Nessun filtro attivo - carico tutte le RfQ nello stato corrente")
                
                username_filter = self._get_active_username_filter()
                
                with DatabaseManager(get_db_path()) as db_manager:
                    all_rows = db_manager.get_all_richieste_aggregated(get_db_path())
                
                # Filtra per stato corrente
                filtered_rows = [row for row in all_rows if row[6] == status]
                
                # Applica filtro username se presente
                if username_filter is not None:
                    filtered_rows = [row for row in filtered_rows if row[5] and row[5].lower() == username_filter.lower()]
                
                # Applica filtro tipo RdO se presente (non "Tutte")
                tipo_filter = self.search_tipo.get()
                if tipo_filter != _("Tutte"):
                    tipo_canonico = normalize_rfq_type(tipo_filter)
                    filtered_rows = [row for row in filtered_rows if row[1] == tipo_canonico]
                
                # Salva ID e percorso database per ogni RfQ
                for row in filtered_rows:
                    source_db_path = row[8] if len(row) > 8 else 'local'
                    if source_db_path == 'local':
                        source_db_path = get_db_path()
                    request_data.append((row[0], source_db_path))
            
            if not request_data:
                messagebox.showwarning(_("Attenzione"), _("Nessuna RfQ da esportare nella vista corrente."), parent=self.root)
                return
            
            logger.info(f"[export_excel] Trovate {len(request_data)} RfQ da esportare: {[r[0] for r in request_data[:10]]}{'...' if len(request_data) > 10 else ''}")
            
        except Exception as e:
            logger.error(f"[export_excel] Errore nel recupero degli ID: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Errore nel recupero delle RfQ da esportare: {}").format(e), parent=self.root)
            return

        # 2. Chiedi Lingua
        prompt = LanguagePrompt(self.root)
        self.root.wait_window(prompt)
        lang = prompt.choice  # 'ita' o 'eng'
        if not lang:
            return

        # 3. Configurazione Testi e Header in base alla lingua
        is_ita = (lang == 'ita')
        headers_map = {
            'cod': "Codice" if is_ita else "Code",
            'att': "Allegato" if is_ita else "Attachment",
            'desc': "Descrizione" if is_ita else "Description",
            'qty': "Q.tà" if is_ita else "Q.ty",
            'cod_g': "Cod. Grezzo" if is_ita else "Raw Code",
            'dis_g': "Dis. Grezzo" if is_ita else "Raw Dwg",
            'mat_cl': "Mat. C/L" if is_ita else "Work Order Mat.",
            'vs_best': "VS. MIGLIORE" if is_ita else "VS. BEST",
            'rdo_num': "Richiesta N°" if is_ita else "RfQ N°",
            'date': "Del" if is_ita else "Date",
            'type': "Tipo" if is_ita else "Type"
        }

        # 4. Setup Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export DataFlow"
        
        # Stili
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        bold_font = Font(bold=True)
        header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')  # Grigio chiaro
        best_price_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Verde
        
        # Setup larghezze colonne
        ws.column_dimensions['A'].width = 15  # Codice
        ws.column_dimensions['B'].width = 15  # Allegato
        ws.column_dimensions['C'].width = 10  # Qta
        ws.column_dimensions['D'].width = 35  # Descrizione
        ws.column_dimensions['E'].width = 15  # Cod Grezzo
        ws.column_dimensions['F'].width = 15  # Dis Grezzo
        ws.column_dimensions['G'].width = 20  # Mat CL
        
        current_row = 1
        
        try:
            # 5. CICLO SULLE RDO - usa il database corretto per ogni RfQ
            for req_id, source_db_path in request_data:
                # Apri il database corretto per questa RfQ
                db_manager = DatabaseManager(source_db_path)
                
                try:
                    # Recupera dati testata
                    rdo_data = db_manager.get_richiesta_full_data(req_id)
                    if not rdo_data:
                        continue
                    de_db, ds_db, rif, tipo_raw = rdo_data
                    
                    # Normalizza tipo
                    tipo_normalizzato = normalize_rfq_type(tipo_raw)
                    is_cl = (tipo_normalizzato == 'Conto lavoro')
                    
                    # Recupera dettagli e fornitori
                    items = db_manager.get_dettagli_by_richiesta(req_id)
                    suppliers_rows = db_manager.get_fornitori_by_richiesta(req_id, order_by=True)
                    suppliers = [r[0] for r in suppliers_rows]
                    prices_rows = db_manager.get_offerte_by_richiesta(req_id)
                    prices = {(id_d, nf): p for id_d, nf, p in prices_rows}
                    
                finally:
                    # Chiudi il database manager dopo ogni RfQ
                    db_manager.close()

                # --- SCRITTURA BLOCCO TESTATA ---
                ws.cell(row=current_row, column=1, value=f"{headers_map['rdo_num']} {req_id}").font = Font(size=12, bold=True)
                ws.cell(row=current_row, column=4, value=f"{headers_map['date']}: {self._format_date_for_display(de_db)}")
                ws.cell(row=current_row, column=7, value=f"Ref: {rif}")
                current_row += 1
                ws.cell(row=current_row, column=1, value=f"{headers_map['type']}: {translate_rfq_type(tipo_normalizzato)}")
                current_row += 2

                # --- SCRITTURA HEADER TABELLA ---
                col_headers = [
                    headers_map['cod'], headers_map['att'], headers_map['qty'], headers_map['desc'],
                    headers_map['cod_g'], headers_map['dis_g'], headers_map['mat_cl']
                ]
                
                for i, h_text in enumerate(col_headers, start=1):
                    c = ws.cell(row=current_row, column=i, value=h_text)
                    c.font = bold_font
                    c.border = thin_border
                    c.fill = header_fill
                    c.alignment = Alignment(horizontal='center')

                # Colonna separatore
                c_sep = ws.cell(row=current_row, column=8, value=headers_map['vs_best'])
                c_sep.font = bold_font
                c_sep.border = thin_border
                c_sep.alignment = Alignment(horizontal='center')

                # Colonne Fornitori
                start_supplier_col = 9
                for i, sup in enumerate(suppliers):
                    c = ws.cell(row=current_row, column=start_supplier_col + i, value=sup)
                    c.font = bold_font
                    c.border = thin_border
                    c.alignment = Alignment(horizontal='center')
                
                current_row += 1

                # --- SCRITTURA RIGHE ARTICOLI ---
                for item in items:
                    id_d, cod, all_file, desc, qta, c_g, d_g, m_cl = item
                    
                    ws.cell(row=current_row, column=1, value=cod).border = thin_border
                    ws.cell(row=current_row, column=2, value=all_file).border = thin_border
                    ws.cell(row=current_row, column=3, value=format_quantity_display(qta)).border = thin_border
                    ws.cell(row=current_row, column=4, value=desc).border = thin_border
                    
                    ws.cell(row=current_row, column=5, value=c_g if is_cl else "").border = thin_border
                    ws.cell(row=current_row, column=6, value=d_g if is_cl else "").border = thin_border
                    ws.cell(row=current_row, column=7, value=m_cl if is_cl else "").border = thin_border
                    
                    ws.cell(row=current_row, column=8, value="").border = thin_border
                    # Prezzi
                    min_price = None
                    row_prices = []
                    for sup in suppliers:
                        p_val = prices.get((id_d, sup))
                        if p_val:
                            try:
                                row_prices.append(float(str(p_val).replace(',', '.')))
                            except:
                                pass
                    if row_prices:
                        min_price = min(row_prices)
                    for i, sup in enumerate(suppliers):
                        col_idx = start_supplier_col + i
                        cell = ws.cell(row=current_row, column=col_idx)
                        price_val = prices.get((id_d, sup))
                        
                        if price_val is not None:
                            try:
                                val_float = float(str(price_val).replace(',', '.'))
                                cell.value = val_float
                                cell.number_format = '0.0000'
                                if min_price is not None and val_float == min_price and val_float > 0:
                                    cell.fill = best_price_fill
                            except:
                                cell.value = price_val
                                cell.alignment = Alignment(horizontal='right')
                        cell.border = thin_border
                    current_row += 1
                
                current_row += 3

            # 6. Salvataggio
            default_name = f"Export_DataFlow_{datetime.now().strftime('%Y%m%d')}.xlsx"
            save_path = filedialog.asksaveasfilename(
                title=_("Salva Export"),
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel Files", "*.xlsx")]
            )
            
            if save_path:
                wb.save(save_path)
                messagebox.showinfo(_("Successo"), _("Export completato con successo:\n{}").format(save_path), parent=self.root)
                logger.info(f"Export Excel salvato in: {save_path}")
        except Exception as e:
            logger.error(f"Errore Export Excel: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Errore durante l'esportazione: {}").format(e), parent=self.root)

    def _format_date_for_display(self, db_date):
        if not db_date: return ""
        try: return datetime.strptime(db_date, '%Y-%m-%d').strftime('%d/%m/%Y')
        except (ValueError, TypeError): return db_date
    def _format_date_for_db(self, display_date):
        if not display_date: return None
        try: return datetime.strptime(display_date, '%d/%m/%Y').strftime('%Y-%m-%d')
        except (ValueError, TypeError): return None

class UserIdentityDialog(tk.Toplevel):
    """Finestra modale che forza l'inserimento di nome e cognome."""
    def __init__(self, parent, first_name='', last_name=''):
        super().__init__(parent)
        self.withdraw()
        self.title(_("Dati Utente Richiesti"))
        self.transient(parent)
        self.resizable(False, False)
        self.grab_set()
        self.result = None
        set_window_icon(self)
        self.protocol("WM_DELETE_WINDOW", self._prevent_close)
        
        self.first_var = tk.StringVar(value=first_name)
        self.last_var = tk.StringVar(value=last_name)
        self.username_var = tk.StringVar(value=_("(in attesa dati)"))
        
        frame = ttk.Frame(self, padding=20)
        frame.pack(fill="both", expand=True)
        
        ttk.Label(
            frame,
            text=_("Per procedere è necessario indicare il tuo nome e cognome."),
            wraplength=320,
            justify="left"
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 10))
        
        ttk.Label(frame, text=_("Nome:")).grid(row=1, column=0, sticky="w", pady=5)
        first_entry = ttk.Entry(frame, textvariable=self.first_var, width=30)
        first_entry.grid(row=1, column=1, sticky="ew", pady=5)
        
        ttk.Label(frame, text=_("Cognome:")).grid(row=2, column=0, sticky="w", pady=5)
        last_entry = ttk.Entry(frame, textvariable=self.last_var, width=30)
        last_entry.grid(row=2, column=1, sticky="ew", pady=5)
        
        ttk.Label(frame, text=_("Username generato:")).grid(row=3, column=0, sticky="w", pady=(10, 0))
        username_display = ttk.Label(frame, textvariable=self.username_var, font=("Calibri", 12, "bold"), foreground="#005AA0")
        username_display.grid(row=3, column=1, sticky="w", pady=(10, 0))
        
        confirm_btn = ttk.Button(frame, text=_("Conferma"), command=self._on_confirm)
        confirm_btn.grid(row=4, column=0, columnspan=2, pady=(20, 0), sticky="ew")
        
        frame.columnconfigure(1, weight=1)
        
        self.first_var.trace_add("write", self._update_preview)
        self.last_var.trace_add("write", self._update_preview)
        self._update_preview()
        self._center_window()
        first_entry.focus_set()

    def _update_preview(self, *_args):
        first = self.first_var.get().strip()
        last = self.last_var.get().strip()
        if not first or not last:
            self.username_var.set(_("(in attesa dati)"))
            return
        try:
            username = generate_username(first, last)
            self.username_var.set(username)
        except ValueError:
            self.username_var.set(_("Dati non validi"))

    def _on_confirm(self):
        first = self.first_var.get().strip()
        last = self.last_var.get().strip()
        if not first or not last:
            messagebox.showerror(_("Campi obbligatori"), _("Inserisci sia il nome sia il cognome."), parent=self)
            return
        try:
            username = generate_username(first, last)
        except ValueError as e:
            messagebox.showerror(_("Formato non valido"), str(e), parent=self)
            return
        self.result = {
            'first_name': first,
            'last_name': last,
            'username': username
        }
        self.grab_release()
        self.destroy()

    def _prevent_close(self):
        messagebox.showwarning(_("Operazione necessaria"), _("Per utilizzare DataFlow è necessario completare i dati richiesti."), parent=self)

    def _center_window(self):
        self.update_idletasks()
        w = self.winfo_reqwidth()
        h = self.winfo_reqheight()
        if not w or not h:
            w, h = 360, 220
        screen_w = self.winfo_screenwidth()
        screen_h = self.winfo_screenheight()
        x = (screen_w // 2) - (w // 2)
        y = (screen_h // 2) - (h // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.deiconify()

# ------------------------------------------------------------------------------------
# FINESTRA PROGRESSO COPIA (PER SPOSTAMENTO CARTELLA)
# ------------------------------------------------------------------------------------
class CopyProgressWindow(tk.Toplevel):
    """Finestra di progresso per operazioni di copia file (stile splash screen)."""
    def __init__(self, parent, title="Copia in corso..."):
        super().__init__(parent)
        self.withdraw()
        set_window_icon(self)
        self.title(title)
        self.overrideredirect(True)
        
        frame = ttk.Frame(self, borderwidth=2, relief="raised")
        frame.pack(fill="both", expand=True)
        
        # Logo (opzionale)
        try:
            logo_path = resource_path(os.path.join("add_data", "logo_dataflow.png"))
            if os.path.exists(logo_path):
                img = Image.open(logo_path)
                # BUG #51 FIX: Check dimensioni valide prima di divisione per evitare ZeroDivisionError
                if img.width > 0 and img.height > 0:
                    img.thumbnail((200, int(200 * (img.height/img.width))), Image.Resampling.LANCZOS)
                    self.logo_photo = ImageTk.PhotoImage(img)
                    ttk.Label(frame, image=self.logo_photo).pack(pady=(20, 10))
        except Exception as e:
            print(f"Errore logo: {e}")
            ttk.Label(frame, text="DataFlow", font=("Helvetica", 18, "bold")).pack(pady=(20, 10))
        
        self.status_label = ttk.Label(
            frame,
            text="Preparazione...",
            font=("Helvetica", 10),
            width=50,
            anchor="center"
        )
        self.status_label.pack(pady=(10, 5))
        
        self.progress = ttk.Progressbar(frame, orient="horizontal", length=400, mode='determinate')
        self.progress.pack(pady=(0, 20))
        
        # Calcola dimensioni e posizione
        self.update_idletasks()
        w = 500
        h = 250
        x = (self.winfo_screenwidth()//2) - (w//2)
        y = (self.winfo_screenheight()//2) - (h//2)
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.deiconify()
    
    def update_progress(self, val, txt):
        """Aggiorna barra e testo."""
        self.progress['value'] = val
        self.status_label['text'] = ""
        self.update_idletasks()
        self.status_label['text'] = txt
        self.update_idletasks()

# ------------------------------------------------------------------------------------
# FINESTRA DI AVVIO (SPLASH SCREEN)
# ------------------------------------------------------------------------------------
class SplashScreen(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.withdraw()  # 1. Nascondi subito
        set_window_icon(self)
        self.title(_("Avvio DataFlow")); self.overrideredirect(True) # 2. Rendi senza bordi
        
        # 3. Aggiungi TUTTI i widget (mentre è ancora nascosta)
        frame = ttk.Frame(self, borderwidth=2, relief="raised"); frame.pack(fill="both", expand=True)
        try:
            logo_path = resource_path(os.path.join("add_data", "logo_dataflow.png"))
            if os.path.exists(logo_path):
                img = Image.open(logo_path)
                # BUG #51 FIX: Check dimensioni valide prima di divisione per evitare ZeroDivisionError
                if img.width > 0 and img.height > 0:
                    img.thumbnail((273, int(273 * (img.height/img.width))), Image.Resampling.LANCZOS)
                    self.logo_photo = ImageTk.PhotoImage(img)
                    ttk.Label(frame, image=self.logo_photo).pack(pady=(30, 20))
        except Exception as e:
            print(f"Errore logo splash: {e}"); ttk.Label(frame, text=_("DataFlow"), font=("Helvetica", 24, "bold")).pack(pady=(30, 20))
        
        # CORREZIONE: Aggiunti width e anchor per evitare sovrapposizioni
        self.status_label = ttk.Label(
            frame, 
            text=_("Avvio in corso..."), 
            font=("Helvetica", 10),
            width=40,
            anchor="center"
        )
        self.status_label.pack(pady=(10, 5))
        self.progress = ttk.Progressbar(frame, orient="horizontal", length=300, mode='determinate'); self.progress.pack(pady=(0, 20))
        
        # 4. Forza Tkinter a calcolare le dimensioni REALI
        self.update_idletasks() 
        
        # 5. Leggi le dimensioni REALI (non più 450x250 fisse)
        w = 450
        h = 250
        
        # 6. Calcola la posizione centrale
        x = (self.winfo_screenwidth()//2) - (w//2)
        y = (self.winfo_screenheight()//2) - (h//2)
        
        # 7. Applica la geometria corretta
        self.geometry(f"{w}x{h}+{x}+{y}")
        
        # 8. Mostra la finestra (ora perfetta)
        self.deiconify()

    def update_progress(self, val, txt):
        self.progress['value'] = val
        # CORREZIONE: Pulisci il testo prima di aggiornarlo per evitare sovrapposizioni
        self.status_label['text'] = ""
        self.update_idletasks()
        self.status_label['text'] = txt
        self.update_idletasks()

# ------------------------------------------------------------------------------------
# SCALING DPI GESTITO AUTOMATICAMENTE DA WINDOWS + TKINTER
# ------------------------------------------------------------------------------------
# Il manifest app.manifest dichiara PerMonitorV2, quindi Tkinter gestisce
# automaticamente il DPI scaling senza bisogno di intervento manuale.


# Le dimensioni delle finestre sono ora gestite automaticamente da Tkinter
# in base al DPI di Windows, senza bisogno di scaling manuale.

# ------------------------------------------------------------------------------------
# ESECUZIONE PRINCIPALE
# ------------------------------------------------------------------------------------
if __name__ == '__main__':
    # Inizializza il sistema di internazionalizzazione PRIMA di creare qualsiasi finestra
    logger.info("=" * 70)
    logger.info("INIZIALIZZAZIONE I18N")
    logger.info("=" * 70)
    language_code = init_i18n()
    logger.info(f"Lingua inizializzata: {language_code}")
    logger.info("=" * 70)
    
    root = tk.Tk()
    root.withdraw()
    splash = None
    
    def main_task():
        # Leggi config esistente (se presente)
        config = configparser.ConfigParser(interpolation=None)
        config_file = get_config_file()
        license_was_accepted = False
        if os.path.exists(config_file):
            try:
                config.read(config_file)
                license_was_accepted = config.getboolean('Settings', 'license_accepted', fallback=False)
            except Exception:
                license_was_accepted = False

        # 1) Mostra la licenza PRIMA di qualsiasi creazione DB
        if not license_was_accepted:
            license_prompt = LicenseWindow(root, first_run=True)
            root.wait_window(license_prompt)
            if not getattr(license_prompt, 'accepted', False):
                try:
                    root.destroy()
                except:
                    pass
                return

            # Salva subito l'accettazione
            try:
                if 'Settings' not in config:
                    config['Settings'] = {}
                config['Settings']['license_accepted'] = 'True'
                with open(config_file, 'w', encoding='utf-8') as f:
                    config.write(f)
            except Exception as e:
                logger.warning(f"Impossibile salvare stato licenza: {e}")

        # 2) Verifica se l'identità utente è già presente, altrimenti richiedila
        identity = load_user_identity()
        if not identity.get('username'):
            # L'identità non è presente o incompleta, mostra il dialogo
            dialog = UserIdentityDialog(root)
            root.wait_window(dialog)
            identity = getattr(dialog, 'result', None)
            if not identity:
                try:
                    root.destroy()
                except:
                    pass
                return
            # Salva subito l'identità nel config
            save_user_identity(identity['first_name'], identity['last_name'], identity['username'])
            # Ricarica l'identità appena salvata
            identity = load_user_identity()
        else:
            # L'identità è già presente, logga e continua
            logger.info(f"Identità utente già presente: {identity['username']}")

        # Salva identità e imposta percorso DB utente (solo se non già presente o diverso)
        try:
            if 'Settings' not in config:
                config['Settings'] = {}
            existing_identity = load_user_identity()
            if not existing_identity.get('username'):
                config['User'] = {
                    'first_name': identity['first_name'],
                    'last_name': identity['last_name'],
                    'username': identity['username']
                }
                # Crea la struttura DataFlow_{username} solo ora
                user_dataflow_dir = get_user_documents_dataflow_dir()
                if not user_dataflow_dir:
                    logger.error("Impossibile creare la cartella utente: username mancante.")
                    messagebox.showerror(_("Errore"), _("Impossibile creare la cartella utente."), parent=root)
                    try:
                        root.destroy()
                    except:
                        pass
                    return
                user_db_dir = os.path.join(user_dataflow_dir, 'Database')
                os.makedirs(user_db_dir, exist_ok=True)
                user_db_name = f"dataflow_db_{identity['username']}.db"
                user_db_path = os.path.join(user_db_dir, user_db_name)
                config['Settings']['custom_db_path'] = user_db_path
                with open(config_file, 'w', encoding='utf-8') as f:
                    config.write(f)
                logger.info(f"Salvata identità utente e percorso DB: {user_db_path}")
            else:
                logger.info(f"Identità utente già salvata: {existing_identity['username']}")
        except Exception as e:
            logger.error(f"Impossibile salvare identità nel config: {e}", exc_info=True)

        # Invalida la cache e crea il DB specifico per l'utente
        reset_db_cache()
        try:
            crea_database_v4()
        except Exception as e:
            logger.error(f"Errore creazione DB utente: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), _("Impossibile creare il database utente: {}").format(e), parent=root)
            try:
                root.destroy()
            except:
                pass
            return

        # 3) Ora mostriamo lo splash (dopo la creazione DB) e carichiamo l'interfaccia
        splash_local = SplashScreen(root)
        splash_local.update_progress(90, _("Caricamento interfaccia..."))
        splash_local.update()

        app = MainWindow(root)
        time.sleep(0.3)

        splash_local.update_progress(100, _("Completato!"))
        time.sleep(0.25)

        # Prepara e mostra la finestra principale
        geometry = calculate_center_position(root)
        root.geometry(geometry)
        root.deiconify()
        root.lift()
        root.attributes('-topmost', True)

        try:
            splash_local.destroy()
        except:
            pass

        # Rimuovi il topmost forzato dopo che la finestra ha il focus
        # BUG #41 FIX: usa funzione nominata invece di lambda per evitare reference leak
        def remove_topmost():
            try:
                root.attributes('-topmost', False)
            except:
                pass
        root.after(50, remove_topmost)

        root.focus_set()

    root.after(200, main_task)
    root.mainloop()

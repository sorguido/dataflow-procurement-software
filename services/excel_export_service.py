"""Excel export services for RFQ, VSM, and Derisking dashboard tabs.

This module intentionally keeps UI prompts/dialog behavior aligned with the
previous in-method implementation in dataflow.py while moving workbook logic
out of the kernel file.
"""

from __future__ import annotations

import logging
from copy import copy

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from tkinter import filedialog

from database_manager import DatabaseManager
from services.app_paths import get_db_path
from ui.dialogs.common_dialogs import LanguagePrompt, SimpleMessageDialog
from utils.format_utils import (
    format_quantity_display,
    get_currency_excel_number_format,
)
from utils.export_filename import (
    build_excel_export_filename,
    normalize_export_lang,
)
from utils.i18n_utils import (
    normalize_rfq_type,
    tr,
    translate_derisking_status,
    translate_vsm_action,
)


logger = logging.getLogger(__name__)


def export_rfq_requests_excel(*, parent, request_data, format_date_for_display):
    """Export RFQ request bundles to a single Excel file.

    Args:
        parent: Tk parent window used for dialogs.
        request_data: list[(request_id, source_db_path)]
        format_date_for_display: callback(db_date) -> display_date
    """
    prompt = LanguagePrompt(parent)
    parent.wait_window(prompt)
    lang = prompt.choice
    if not lang:
        return

    is_ita = lang == "ita"
    headers_map = {
        "cod": "Codice" if is_ita else "Code",
        "att": "Allegato" if is_ita else "Attachment",
        "desc": "Descrizione" if is_ita else "Description",
        "qty": "Q.tà" if is_ita else "Q.ty",
        "cod_g": "Cod. Grezzo" if is_ita else "Raw Code",
        "dis_g": "Dis. Grezzo" if is_ita else "Raw Dwg",
        "mat_cl": "Mat. C/L" if is_ita else "Work Order Mat.",
        "vs_best": "VS. MIGLIORE" if is_ita else "YOUR BEST",
        "rdo_num": "Richiesta N°" if is_ita else "RfQ N°",
        "date": "Del" if is_ita else "Date",
        "type": "Tipo" if is_ita else "Type",
    }
    rfq_type_en = {
        "Fornitura piena": "Full Supply",
        "Conto lavoro": "Work Order",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export DataFlow"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    best_price_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    col_rules = {
        1: {"min_width": 15, "max_width": 18},
        2: {"min_width": 15, "max_width": 28, "wrap_threshold": 34},
        3: {"min_width": 10, "max_width": 12},
        4: {"min_width": 35, "max_width": 52, "wrap_threshold": 46},
        5: {"min_width": 15, "max_width": 24, "wrap_threshold": 32},
        6: {"min_width": 15, "max_width": 24, "wrap_threshold": 32},
        7: {"min_width": 20, "max_width": 30, "wrap_threshold": 36},
        8: {"min_width": 16, "max_width": 24, "wrap_threshold": 24},
    }
    supplier_min_width = 11
    supplier_max_width = 24
    supplier_wrap_threshold = 22
    text_padding = 2

    def text_len(value):
        if value is None:
            return 0
        text = str(value).strip()
        if not text:
            return 0
        return max(len(line) for line in text.splitlines())

    def enable_wrap_text(cell):
        if cell.alignment and cell.alignment.wrap_text:
            return
        if cell.alignment:
            new_alignment = copy(cell.alignment)
            new_alignment.wrap_text = True
            cell.alignment = new_alignment
        else:
            cell.alignment = Alignment(wrap_text=True)

    col_text_max = {
        1: len(headers_map["cod"]),
        2: len(headers_map["att"]),
        3: len(headers_map["qty"]),
        4: len(headers_map["desc"]),
        5: len(headers_map["cod_g"]),
        6: len(headers_map["dis_g"]),
        7: len(headers_map["mat_cl"]),
        8: len(headers_map["vs_best"]),
    }
    supplier_col_text_max = {}

    current_row = 1

    try:
        for req_id, source_db_path in request_data:
            db_manager = DatabaseManager(source_db_path)
            try:
                rdo_data = db_manager.get_richiesta_full_data(req_id)
                if not rdo_data:
                    continue
                de_db, _ds_db, rif, tipo_raw = rdo_data

                tipo_normalizzato = normalize_rfq_type(tipo_raw)
                is_cl = tipo_normalizzato == "Conto lavoro"

                items = db_manager.get_dettagli_by_richiesta(req_id)
                suppliers_rows = db_manager.get_fornitori_by_richiesta(req_id, order_by=True)
                suppliers = [r[0] for r in suppliers_rows]
                prices_rows = db_manager.get_offerte_by_richiesta(req_id)
                prices = {(id_d, nf): p for id_d, nf, p in prices_rows}
            finally:
                db_manager.close()

            ws.cell(row=current_row, column=1, value=f"{headers_map['rdo_num']} {req_id}").font = Font(size=12, bold=True)
            ws.cell(row=current_row, column=4, value=f"{headers_map['date']}: {format_date_for_display(de_db)}")
            ws.cell(row=current_row, column=7, value=f"Ref: {rif}")
            current_row += 1
            ws.cell(
                row=current_row,
                column=1,
                value=f"{headers_map['type']}: {tipo_normalizzato if is_ita else rfq_type_en.get(tipo_normalizzato, tipo_normalizzato)}",
            )
            current_row += 2

            col_headers = [
                headers_map["cod"], headers_map["att"], headers_map["qty"], headers_map["desc"],
                headers_map["cod_g"], headers_map["dis_g"], headers_map["mat_cl"],
            ]

            for i, h_text in enumerate(col_headers, start=1):
                c = ws.cell(row=current_row, column=i, value=h_text)
                c.font = bold_font
                c.border = thin_border
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center")
                col_text_max[i] = max(col_text_max.get(i, 0), text_len(h_text))

            c_sep = ws.cell(row=current_row, column=8, value=headers_map["vs_best"])
            c_sep.font = bold_font
            c_sep.border = thin_border
            c_sep.alignment = Alignment(horizontal="center")
            col_text_max[8] = max(col_text_max.get(8, 0), text_len(headers_map["vs_best"]))

            start_supplier_col = 9
            for i, sup in enumerate(suppliers):
                supplier_col = start_supplier_col + i
                c = ws.cell(row=current_row, column=supplier_col, value=sup)
                c.font = bold_font
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")
                sup_len = text_len(sup)
                supplier_col_text_max[supplier_col] = max(supplier_col_text_max.get(supplier_col, 0), sup_len)
                if sup_len > supplier_wrap_threshold:
                    enable_wrap_text(c)

            current_row += 1

            for item in items:
                id_d, cod, all_file, desc, qta, c_g, d_g, m_cl = item

                c_cod = ws.cell(row=current_row, column=1, value=cod)
                c_cod.border = thin_border
                col_text_max[1] = max(col_text_max.get(1, 0), text_len(cod))

                c_att = ws.cell(row=current_row, column=2, value=all_file)
                c_att.border = thin_border
                att_len = text_len(all_file)
                col_text_max[2] = max(col_text_max.get(2, 0), att_len)
                if att_len > col_rules[2]["wrap_threshold"]:
                    enable_wrap_text(c_att)

                qty_text = format_quantity_display(qta)
                c_qty = ws.cell(row=current_row, column=3, value=qty_text)
                c_qty.border = thin_border
                col_text_max[3] = max(col_text_max.get(3, 0), text_len(qty_text))

                c_desc = ws.cell(row=current_row, column=4, value=desc)
                c_desc.border = thin_border
                desc_len = text_len(desc)
                col_text_max[4] = max(col_text_max.get(4, 0), desc_len)
                if desc_len > col_rules[4]["wrap_threshold"]:
                    enable_wrap_text(c_desc)

                val_cg = c_g if is_cl else ""
                c_cg = ws.cell(row=current_row, column=5, value=val_cg)
                c_cg.border = thin_border
                cg_len = text_len(val_cg)
                col_text_max[5] = max(col_text_max.get(5, 0), cg_len)
                if cg_len > col_rules[5]["wrap_threshold"]:
                    enable_wrap_text(c_cg)

                val_dg = d_g if is_cl else ""
                c_dg = ws.cell(row=current_row, column=6, value=val_dg)
                c_dg.border = thin_border
                dg_len = text_len(val_dg)
                col_text_max[6] = max(col_text_max.get(6, 0), dg_len)
                if dg_len > col_rules[6]["wrap_threshold"]:
                    enable_wrap_text(c_dg)

                val_mcl = m_cl if is_cl else ""
                c_mcl = ws.cell(row=current_row, column=7, value=val_mcl)
                c_mcl.border = thin_border
                mcl_len = text_len(val_mcl)
                col_text_max[7] = max(col_text_max.get(7, 0), mcl_len)
                if mcl_len > col_rules[7]["wrap_threshold"]:
                    enable_wrap_text(c_mcl)

                ws.cell(row=current_row, column=8, value="").border = thin_border

                min_price = None
                row_prices = []
                for sup in suppliers:
                    p_val = prices.get((id_d, sup))
                    if p_val:
                        try:
                            row_prices.append(float(str(p_val).replace(",", ".")))
                        except Exception:
                            pass
                if row_prices:
                    min_price = min(row_prices)

                for i, sup in enumerate(suppliers):
                    col_idx = start_supplier_col + i
                    cell = ws.cell(row=current_row, column=col_idx)
                    price_val = prices.get((id_d, sup))

                    if price_val is not None:
                        try:
                            val_float = float(str(price_val).replace(",", "."))
                            cell.value = val_float
                            cell.number_format = "0.0000"
                            if min_price is not None and val_float == min_price and val_float > 0:
                                cell.fill = best_price_fill
                        except Exception:
                            cell.value = price_val
                            cell.alignment = Alignment(horizontal="right")
                        supplier_col_text_max[col_idx] = max(
                            supplier_col_text_max.get(col_idx, 0),
                            text_len(cell.value),
                        )
                    cell.border = thin_border
                current_row += 1

            current_row += 3

        for col_idx, rule in col_rules.items():
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            measured = col_text_max.get(col_idx, 0) + text_padding
            min_w = rule["min_width"]
            max_w = max(rule["max_width"], min_w)
            ws.column_dimensions[col_letter].width = min(max(measured, min_w), max_w)

        for col_idx, measured_len in supplier_col_text_max.items():
            if col_idx < 9:
                continue
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            measured = measured_len + text_padding
            ws.column_dimensions[col_letter].width = min(max(measured, supplier_min_width), supplier_max_width)

        max_used_row = max(1, current_row - 1)
        max_used_col = max(1, ws.max_column)
        for row_cells in ws.iter_rows(min_row=1, max_row=max_used_row, min_col=1, max_col=max_used_col):
            for cell in row_cells:
                if cell.value is None:
                    continue
                alignment = cell.alignment
                if alignment:
                    if alignment.vertical == "center":
                        continue
                    new_alignment = copy(alignment)
                    new_alignment.vertical = "center"
                    cell.alignment = new_alignment
                else:
                    cell.alignment = Alignment(vertical="center")

        default_name = build_excel_export_filename(
            "DataFlow",
            "Dashboard",
            "RFQ",
            "Export",
            normalize_export_lang(lang),
        )
        save_path = filedialog.asksaveasfilename(
            title=tr("Save Export"),
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if save_path:
            wb.save(save_path)
            SimpleMessageDialog(parent, tr("Success"), tr("Export completed successfully:\n{}").format(save_path), "info")
            logger.info("Export Excel salvato in: %s", save_path)
    except Exception as e:
        logger.error("Errore Export Excel: %s", e, exc_info=True)
        SimpleMessageDialog(parent, tr("Error"), tr("Error during export: {}").format(e), "error")


def export_vsm_events_excel(*, parent, status, sheet_col_widths, events):
    """Export pre-filtered VSM events to Excel."""
    status_to_event_type = {
        "vsm_saving": "Saving",
        "vsm_cost_avoidance": "Cost Avoidance",
    }
    event_type = status_to_event_type.get(status, status)

    prompt = LanguagePrompt(parent)
    parent.wait_window(prompt)
    lang = prompt.choice
    if not lang:
        return
    is_ita = lang == "ita"

    if not events:
        SimpleMessageDialog(parent, tr("Warning"), tr("No data to export in the current view."), "warning")
        return

    use_dual = event_type in ("Saving", "Cost Avoidance")
    if is_ita:
        if event_type == "Saving":
            headers = [
                "Data", "Tipo", "Azione", "Descrizione",
                "Saving Teorico", "Saving Effettivo", "Realizzo %", "Variance %", "Ripetitivo", "Utente",
            ]
        else:
            headers = [
                "Data", "Tipo", "Azione", "Descrizione",
                "CA Teorico", "CA Effettivo", "Realizzo %", "Variance %", "Ripetitivo", "Utente",
            ]
    else:
        if event_type == "Saving":
            headers = [
                "Date", "Type", "Action", "Description",
                "Theoretical Savings", "Actual Savings", "Realization %", "Variance %", "Repetitive", "User",
            ]
        else:
            headers = [
                "Date", "Type", "Action", "Description",
                "CA Theoretical", "CA Actual", "Realization %", "Variance %", "Repetitive", "User",
            ]

    data_rows = []
    for event in events:
        valore_teorico = event.calculate_theoretical_value() or 0.0
        date_str = event.event_date.strftime("%d/%m/%Y") if event.event_date else ""
        desc = (event.description or event.reference or "")[:50]
        action_str = translate_vsm_action(event.action, language_code=lang)

        if use_dual:
            valore_effettivo = event.calculate_effective_value() or 0.0
            if event.driver == "Pagamenti" and event.giorni_pagamento_attuali is not None and event.giorni_pagamento_negoziati is not None:
                delta = event.giorni_pagamento_negoziati - event.giorni_pagamento_attuali
                variance_pct = round((delta / 30.0) * event.effective_payments_rate_pct, 2)
            elif event_type == "Cost Avoidance":
                baseline = event.importo_richiesto_iniziale or 0.0
                variance_pct = round((baseline - (event.importo_negoziato or 0.0)) / baseline * 100, 1) if baseline != 0.0 else 0.0
            else:
                baseline = event.importo_bdg or 0.0
                variance_pct = round((baseline - (event.importo_negoziato or 0.0)) / baseline * 100, 1) if baseline != 0.0 else 0.0
            row = [
                date_str,
                event.event_type,
                action_str,
                desc,
                round(valore_teorico, 2),
                round(valore_effettivo, 2),
                round(event.percent_realizzo, 1),
                variance_pct,
                "✓" if event.opex_ripetitivo else "",
                event.username,
            ]
            data_rows.append(row)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = event_type[:31]

    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font
        cell.border = thin_border
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    monetary_cols = {5, 6} if use_dual else {6}
    pct_cols = {7, 8} if use_dual else {7}
    rep_col = 9 if use_dual else 8
    money_fmt = get_currency_excel_number_format()

    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in monetary_cols:
                cell.number_format = money_fmt
            elif col_idx in pct_cols:
                cell.number_format = "0.0"
            elif col_idx == rep_col:
                cell.alignment = Alignment(horizontal="center")

    if sheet_col_widths:
        for i, px_width in enumerate(sheet_col_widths):
            col_letter = ws.cell(row=1, column=i + 1).column_letter
            ws.column_dimensions[col_letter].width = max(10, px_width / 7)

    if event_type == "Saving":
        section = "VSM_Saving"
    elif event_type == "Cost Avoidance":
        section = "VSM_CostAvoidance"
    else:
        section = f"VSM_{event_type}"
    default_name = build_excel_export_filename(
        "DataFlow",
        "Dashboard",
        section,
        "Export",
        normalize_export_lang(lang),
    )
    try:
        save_path = filedialog.asksaveasfilename(
            title=tr("Save Export"),
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if save_path:
            wb.save(save_path)
            SimpleMessageDialog(parent, tr("Success"), tr("Export completed successfully:\n{}").format(save_path), "info")
            logger.info("Export VSM Excel salvato in: %s", save_path)
    except Exception as e:
        logger.error("Errore Export VSM Excel: %s", e, exc_info=True)
        SimpleMessageDialog(parent, tr("Error"), tr("Error during export: {}").format(e), "error")


def export_derisking_suppliers_excel(*, parent, suppliers):
    """Export supplier list for Derisking tab to Excel."""
    prompt = LanguagePrompt(parent)
    parent.wait_window(prompt)
    lang = prompt.choice
    if not lang:
        return
    is_ita = lang == "ita"

    if not suppliers:
        SimpleMessageDialog(parent, tr("Warning"), tr("No data to export in the current view."), "warning")
        return

    if is_ita:
        headers = ["Fornitore", "Categoria", "Stato", "Contatto", "E-mail", "Telefono", "Web", "Note", "User"]
    else:
        headers = ["Supplier", "Category", "Status", "Contact", "E-mail", "Phone", "Web", "Notes", "User"]

    data_rows = []
    for s in suppliers:
        status_display = translate_derisking_status(
            s.supplier_status,
            language_code="ita" if is_ita else "eng",
        ) if s.supplier_status else ""
        data_rows.append([
            s.supplier_name,
            s.category,
            status_display,
            s.contact_name,
            s.email,
            s.phone,
            s.website,
            s.notes,
            s.username,
        ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Derisking"

    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font
        cell.border = thin_border
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    col_widths = [30, 20, 18, 25, 30, 18, 30, 40, 15]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    default_name = build_excel_export_filename(
        "DataFlow",
        "Dashboard",
        "VSM_Derisking",
        "Export",
        normalize_export_lang(lang),
    )
    try:
        save_path = filedialog.asksaveasfilename(
            title=tr("Save Export"),
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if save_path:
            wb.save(save_path)
            SimpleMessageDialog(parent, tr("Success"), tr("Export completed successfully:\n{}").format(save_path), "info")
            logger.info("Export Derisking Excel salvato in: %s", save_path)
    except Exception as e:
        logger.error("Errore Export Derisking Excel: %s", e, exc_info=True)
        SimpleMessageDialog(parent, tr("Error"), tr("Error during export: {}").format(e), "error")


def load_derisking_suppliers_for_export(*, username_filter):
    """Load suppliers for Derisking export from persistence layer."""
    from services.supplier_persistence import get_all_suppliers

    with DatabaseManager(get_db_path()) as db_manager:
        return get_all_suppliers(db_manager, username=username_filter)

# -*- coding: utf-8 -*-
"""
KPI Excel Export — generazione workbook openpyxl per KPI Analysis.

Nessuna UI, nessuna logica KPI, nessun accesso diretto al DB.
Riceve i dati già calcolati e genera un openpyxl.Workbook.

Funzione pubblica:
    build_kpi_workbook(...) → openpyxl.Workbook
"""

import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from utils.format_utils import get_currency_excel_number_format

logger = logging.getLogger('DataFlow.KpiExcelExport')

# ---------------------------------------------------------------------------
# Stili — stessa convenzione di mega_export_excel / sqdc_analysis_window
# ---------------------------------------------------------------------------

_THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)
_BOLD        = Font(bold=True)
_BOLD_11     = Font(bold=True, size=11)
_BOLD_13     = Font(bold=True, size=13)
_ITALIC_9    = Font(italic=True, size=9, color="555555")
_FILL_HEADER = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
_FILL_TITLE  = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
_CENTER      = Alignment(horizontal='center', vertical='center')
_LEFT        = Alignment(horizontal='left',   vertical='center')


def _hdr(ws, row, col, text):
    """Scrive una cella header (bold, fill grigio, bordo, centrata)."""
    c = ws.cell(row=row, column=col, value=text)
    c.font      = _BOLD
    c.fill      = _FILL_HEADER
    c.border    = _THIN
    c.alignment = _CENTER
    return c


def _dat(ws, row, col, value, fmt=None, align=None):
    """Scrive una cella dati (bordo, allineamento opzionale, formato numerico opzionale)."""
    c = ws.cell(row=row, column=col, value=value)
    c.border    = _THIN
    c.alignment = align or _LEFT
    if fmt:
        c.number_format = fmt
    return c


def _w(ws, letter, width):
    ws.column_dimensions[letter].width = width


# ---------------------------------------------------------------------------
# Helpers bilingua e numerici
# ---------------------------------------------------------------------------

def _t(is_ita, ita, eng):
    return ita if is_ita else eng


# Mapping canonical status (sempre IT nel DB) → label EN per export KPI.
# Mantenere allineato a:
#   • models/potential_supplier.SUPPLIER_STATUS_*  (costanti canoniche)
#   • dataflow._export_derisking_excel._STATUS_EXPORT_EN  (stesso mapping, percorso export standalone)
_STATUS_DERISKING_EN = {
    "Nuovo":          "New",
    "In valutazione": "Under Evaluation",
    "Qualificato":    "Qualified",
    "Scartato":       "Rejected",
}


def _i(v):
    """Valore intero, 0 su errore."""
    try:
        return int(v or 0)
    except (TypeError, ValueError):
        return 0


def _f(v):
    """Valore float, 0.0 su errore."""
    try:
        return float(v or 0.0)
    except (TypeError, ValueError):
        return 0.0


# Formato percentuale: es. 5.23
_FMT_PCT   = '0.00'


# ---------------------------------------------------------------------------
# API pubblica
# ---------------------------------------------------------------------------

def build_kpi_workbook(
    rfq_data:       dict,
    saving_data:    dict,
    ca_data:        dict,
    derisking_data: dict,
    filter_label:   str,
    scope:          str,
    current_section: str,
    lang:           str,
    currency_code:  str = "NONE",
) -> openpyxl.Workbook:
    """
    Genera il workbook Excel KPI Analysis.

    Args:
        rfq_data:         dict da get_rfq_kpi()
        saving_data:      dict da get_saving_kpi()
        ca_data:          dict da get_cost_avoidance_kpi()
        derisking_data:   dict da get_derisking_kpi()
        filter_label:     es. "Anno: 2025" | "Periodo: 12M" | "Tutti i dati"
        scope:            'current' | 'all'
        current_section:  'RFQ' | 'Saving' | 'Cost Avoidance' | 'Derisking'
        lang:             'ita' | 'eng'

    Returns:
        openpyxl.Workbook (non ancora salvato su disco)
    """
    is_ita = (lang == 'ita')
    money_fmt = get_currency_excel_number_format(currency_code)

    sections = (
        [current_section] if scope == 'current'
        else ['RFQ', 'Saving', 'Cost Avoidance', 'Derisking']
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # Rimuove foglio "Sheet" di default

    _build_summary(
        wb, rfq_data, saving_data, ca_data, derisking_data,
        filter_label, scope, current_section, sections, is_ita, money_fmt
    )

    if 'RFQ' in sections:
        _build_rfq(wb, rfq_data, filter_label, is_ita)
    if 'Saving' in sections:
        _build_saving(wb, saving_data, filter_label, is_ita, money_fmt)
    if 'Cost Avoidance' in sections:
        _build_ca(wb, ca_data, filter_label, is_ita, money_fmt)
    if 'Derisking' in sections:
        _build_derisking(wb, derisking_data, filter_label, is_ita)

    return wb


# ---------------------------------------------------------------------------
# Foglio SUMMARY / RIEPILOGO
# ---------------------------------------------------------------------------

def _build_summary(wb, rfq, sav, ca, der,
                   filter_label, scope, current_section, sections, is_ita, money_fmt):
    sheet_name = _t(is_ita, "Riepilogo", "Summary")
    ws = wb.create_sheet(sheet_name)
    _w(ws, 'A', 34)
    _w(ws, 'B', 30)
    _w(ws, 'C', 26)

    # --- Blocco metadata ---
    ws['A1'] = _t(is_ita,
                  "Export KPI Analysis — DataFlow",
                  "KPI Analysis Export — DataFlow")
    ws['A1'].font = _BOLD_13

    now_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws['A3'] = _t(is_ita, "Data export:", "Export date:")
    ws['A3'].font = _BOLD
    ws['B3'] = now_str

    ws['A4'] = _t(is_ita, "Filtro applicato:", "Applied filter:")
    ws['A4'].font = _BOLD
    ws['B4'] = filter_label

    scope_str = _t(is_ita,
                   "Sezione corrente" if scope == 'current' else "Tutte le sezioni",
                   "Current section"  if scope == 'current' else "All sections")
    if scope == 'current':
        scope_str += f" ({current_section})"
    ws['A5'] = _t(is_ita, "Sezione esportata:", "Exported scope:")
    ws['A5'].font = _BOLD
    ws['B5'] = scope_str

    # --- Tabella KPI sintetica ---
    r = 7
    for col, hdr in enumerate(
        [_t(is_ita, "Sezione", "Section"),
         "KPI",
         _t(is_ita, "Valore", "Value")],
        start=1,
    ):
        _hdr(ws, r, col, hdr)
    r += 1

    all_rows = (
        _rows_rfq(rfq, is_ita)
        + _rows_saving(sav, is_ita, money_fmt)
        + _rows_ca(ca, is_ita, money_fmt)
        + _rows_derisking(der, is_ita)
    )
    for sec_key, kpi_name, value, fmt in all_rows:
        if sec_key not in sections:
            continue
        _dat(ws, r, 1, sec_key)
        _dat(ws, r, 2, kpi_name)
        _dat(ws, r, 3, value, fmt=fmt, align=_CENTER)
        r += 1


# ---------------------------------------------------------------------------
# Righe dati per Summary (e riuso nei fogli sezione)
# ---------------------------------------------------------------------------

def _rows_rfq(data, is_ita):
    if not data:
        return []
    return [
        ('RFQ', _t(is_ita, "RFQ Attive",           "RFQ Active"),       _i(data.get('rfq_active')),       None),
        ('RFQ', _t(is_ita, "RFQ Archiviate",        "RFQ Archived"),     _i(data.get('rfq_archived')),     None),
        ('RFQ', _t(is_ita, "RFQ Totali",            "RFQ Total"),        _i(data.get('rfq_total')),        None),
        ('RFQ', _t(is_ita, "RFQ Non Scadute",       "RFQ Not Expired"),  _i(data.get('rfq_not_expired')),  None),
        ('RFQ', _t(is_ita, "RFQ Scadute",           "RFQ Expired"),      _i(data.get('rfq_expired')),      None),
        ('RFQ', _t(is_ita, "Conto Lavoro",          "Work Order"),       _i(data.get('work_order')),       None),
        ('RFQ', _t(is_ita, "Fornitura Piena",       "Full Supply"),      _i(data.get('full_supply')),      None),
    ]


def _rows_saving(data, is_ita, money_fmt):
    if not data:
        return []
    rows = [
        ('Saving', _t(is_ita, "Saving Teorico",              "Theoretical Saving"),        _f(data.get('theoretical_saving')),   money_fmt),
        ('Saving', _t(is_ita, "Saving Effettivo",            "Actual Saving"),             _f(data.get('actual_saving')),         money_fmt),
        ('Saving', _t(is_ita, "Media % Saving Teorico",      "Avg Theoretical Saving %"),  _f(data.get('average_saving_pct')),   _FMT_PCT),
        ('Saving', _t(is_ita, "Best Saving %",               "Best Saving %"),             _f(data.get('best_saving_pct')),      _FMT_PCT),
        ('Saving', _t(is_ita, "Worst Saving %",              "Worst Saving %"),            _f(data.get('worst_saving_pct')),     _FMT_PCT),
        ('Saving', _t(is_ita, "Mediana Saving %",            "Median Saving %"),           _f(data.get('median_saving_pct')),    _FMT_PCT),
        ('Saving', _t(is_ita, "Impatto Ricorrente",          "Recurring Impact"),          _f(data.get('recurring_impact')),     money_fmt),
        ('Saving', _t(is_ita, "Impatto Non Ricorrente",      "Non-Recurring Impact"),      _f(data.get('non_recurring_impact')), money_fmt),
    ]
    co = data.get('carry_over_to_next_year')
    if co is not None:
        rows.append((
            'Saving',
            _t(is_ita, "Carry-over Anno Successivo", "Carry-over to Next Year"),
            _f(co),
            money_fmt,
        ))
    return rows


def _rows_ca(data, is_ita, money_fmt):
    if not data:
        return []
    rows = [
        ('Cost Avoidance', _t(is_ita, "CA Teorico",             "Theoretical CA"),        _f(data.get('theoretical_cost_avoidance')),  money_fmt),
        ('Cost Avoidance', _t(is_ita, "CA Effettivo",           "Actual CA"),             _f(data.get('actual_cost_avoidance')),        money_fmt),
        ('Cost Avoidance', _t(is_ita, "Media % CA Teorico",     "Avg Theoretical CA %"),  _f(data.get('average_pct')),                  _FMT_PCT),
        ('Cost Avoidance', _t(is_ita, "Best %",                 "Best %"),                _f(data.get('best_pct')),                     _FMT_PCT),
        ('Cost Avoidance', _t(is_ita, "Worst %",                "Worst %"),               _f(data.get('worst_pct')),                    _FMT_PCT),
        ('Cost Avoidance', _t(is_ita, "Mediana %",              "Median %"),              _f(data.get('median_pct')),                   _FMT_PCT),
        ('Cost Avoidance', _t(is_ita, "Ricorrente",             "Recurring"),             _f(data.get('recurring')),                    money_fmt),
        ('Cost Avoidance', _t(is_ita, "Non Ricorrente",         "Non-Recurring"),         _f(data.get('non_recurring')),               money_fmt),
    ]
    co = data.get('carry_over_to_next_year')
    if co is not None:
        rows.append((
            'Cost Avoidance',
            _t(is_ita, "Carry-over Anno Successivo", "Carry-over to Next Year"),
            _f(co),
            money_fmt,
        ))
    return rows


def _rows_derisking(data, is_ita):
    if not data:
        return []
    rows = [
        ('Derisking',
         _t(is_ita, "Totale Fornitori Potenziali", "Total Potential Suppliers"),
         _i(data.get('total_suppliers')),
         None),
        ('Derisking',
         _t(is_ita, "Categorie Uniche", "Unique Categories"),
         _i(data.get('unique_categories')),
         None),
    ]
    for stato, count in data.get('status_counts', {}).items():
        rows.append(('Derisking',
                     stato if is_ita else (_STATUS_DERISKING_EN.get(stato) or stato),
                     _i(count), None))
    return rows


# ---------------------------------------------------------------------------
# Foglio RFQ
# ---------------------------------------------------------------------------

def _build_rfq(wb, data, filter_label, is_ita):
    ws = wb.create_sheet("RFQ")
    _w(ws, 'A', 36)
    _w(ws, 'B', 20)
    _sheet_meta(ws, "RFQ", filter_label, is_ita)
    r = 4
    _hdr(ws, r, 1, "KPI")
    _hdr(ws, r, 2, _t(is_ita, "Valore", "Value"))
    r += 1
    for _, kpi_name, value, fmt in _rows_rfq(data, is_ita):
        _dat(ws, r, 1, kpi_name)
        _dat(ws, r, 2, value, fmt=fmt, align=_CENTER)
        r += 1


# ---------------------------------------------------------------------------
# Foglio SAVING
# ---------------------------------------------------------------------------

def _build_saving(wb, data, filter_label, is_ita, money_fmt):
    ws = wb.create_sheet("Saving")
    _w(ws, 'A', 38)
    _w(ws, 'B', 22)
    _sheet_meta(ws, _t(is_ita, "Saving KPI", "Saving KPI"), filter_label, is_ita)
    r = 4
    _hdr(ws, r, 1, "KPI")
    _hdr(ws, r, 2, _t(is_ita, "Valore", "Value"))
    r += 1
    for _, kpi_name, value, fmt in _rows_saving(data, is_ita, money_fmt):
        _dat(ws, r, 1, kpi_name)
        _dat(ws, r, 2, value, fmt=fmt, align=_CENTER)
        r += 1


# ---------------------------------------------------------------------------
# Foglio COST AVOIDANCE
# ---------------------------------------------------------------------------

def _build_ca(wb, data, filter_label, is_ita, money_fmt):
    ws = wb.create_sheet("Cost Avoidance")
    _w(ws, 'A', 38)
    _w(ws, 'B', 22)
    _sheet_meta(ws, _t(is_ita, "Cost Avoidance KPI", "Cost Avoidance KPI"), filter_label, is_ita)
    r = 4
    _hdr(ws, r, 1, "KPI")
    _hdr(ws, r, 2, _t(is_ita, "Valore", "Value"))
    r += 1
    for _, kpi_name, value, fmt in _rows_ca(data, is_ita, money_fmt):
        _dat(ws, r, 1, kpi_name)
        _dat(ws, r, 2, value, fmt=fmt, align=_CENTER)
        r += 1


# ---------------------------------------------------------------------------
# Foglio DERISKING
# ---------------------------------------------------------------------------

def _build_derisking(wb, data, filter_label, is_ita):
    ws = wb.create_sheet("Derisking")
    _w(ws, 'A', 42)
    _w(ws, 'B', 26)
    _sheet_meta(ws, _t(is_ita, "Derisking KPI", "Derisking KPI"), filter_label, is_ita)
    r = 4
    _hdr(ws, r, 1, "KPI")
    _hdr(ws, r, 2, _t(is_ita, "Valore", "Value"))
    r += 1
    for _, kpi_name, value, fmt in _rows_derisking(data, is_ita):
        _dat(ws, r, 1, kpi_name)
        _dat(ws, r, 2, value, fmt=fmt, align=_CENTER)
        r += 1


# ---------------------------------------------------------------------------
# Helper: intestazione comune ai fogli sezione
# ---------------------------------------------------------------------------

def _sheet_meta(ws, section_title, filter_label, is_ita):
    """Scrive titolo sezione e riga filtro nelle righe 1–2."""
    ws['A1'] = section_title
    ws['A1'].font = _BOLD_11
    ws['A2'] = _t(is_ita, "Filtro: ", "Filter: ") + filter_label
    ws['A2'].font = _ITALIC_9
    ws['A3'] = ""

"""
SQDCAnalysisWindow - Finestra per l'analisi SQDC dei fornitori.
Estratta da dataflow.py per compatibilità con PyInstaller.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import logging
import os
from tksheet import Sheet
from datetime import datetime
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

from database_manager import DatabaseManager, DatabaseError
from services.app_paths import get_db_path, get_fixed_attachments_dir
from utils.window_utils import center_window
from utils.resource_utils import set_window_icon, resource_path
from utils.i18n_utils import _, get_current_language
from utils.format_utils import parse_float_from_comma_string
from utils.validation_utils import sanitize_filename

logger = logging.getLogger(__name__)


class SQDCAnalysisWindow(tk.Toplevel):
    def __init__(self, parent, request_id, existing_data=None):
        super().__init__(parent)
        self.withdraw()
        set_window_icon(self)
        self.request_id = request_id
        self.db_path = get_db_path()
        self.title(_("Analisi SQDC - RdO N° {}").format(self.request_id))
        self.transient(parent)
        self.grab_set()
        
        # Inizializza variabili
        self.weights = {'safety': tk.StringVar(value='25'), 'quality': tk.StringVar(value='25'),
                       'delivery': tk.StringVar(value='25'), 'cost': tk.StringVar(value='25')}
        self.scores = {}  # {'supplier': {'safety': score, 'quality': score, ...}}
        self.suppliers = []
        self.automatic_cost = False  # Se True, Cost è calcolato automaticamente
        self.missing_price_suppliers = []  # Lista fornitori con prezzi mancanti/incompleti
        
        # Carica fornitori
        self.load_suppliers()
        
        # Pulsanti (sempre in fondo)
        button_frame = ttk.Frame(self)
        button_frame.pack(side="bottom", fill="x", padx=10, pady=10)
        ttk.Button(button_frame, text=_("📊 Esporta Excel"), command=self.export_to_excel).pack(side="left", padx=5)
        ttk.Button(button_frame, text=_("💾 Salva SQDC"), command=self.save_sqdc).pack(side="left", padx=5)
        ttk.Button(button_frame, text=_("❌ Chiudi"), command=self.destroy).pack(side="right", padx=5)
        
        # UI principale (espandibile)
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(side="top", fill="both", expand=True)
        
        # Notepad con tab
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill="both", expand=True, pady=10)
        
        # Tab 1: Pesi
        tab_weights = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab_weights, text=_("Pesi (%)"))
        self.create_weights_tab(tab_weights)
        
        # Tab 2: Voti
        tab_scores = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab_scores, text=_("Voti (1-10)"))
        self.create_scores_tab(tab_scores)
        
        # Binding per validazione quando si passa al tab Voti
        def on_tab_changed(event):
            selected = self.notebook.index(self.notebook.select())
            if selected == 1:  # Tab Voti (indice 1, 0 è Pesi)
                if not self.validate_weights_only():
                    self.notebook.select(0)
        
        self.notebook.bind("<<NotebookTabChanged>>", on_tab_changed)
        
        # Inizializza dati esistenti se forniti - DOPO aver creato l'UI
        if existing_data:
            self.load_from_existing_data(existing_data)
        
        # Imposta dimensione finestra
        width = 950
        height = 680
        
        # Centra la finestra
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        # Mostra la finestra
        self.deiconify()
    
    def load_suppliers(self):
        """Carica la lista dei fornitori per questa RdO"""
        try:
            with DatabaseManager(getattr(self, 'db_path', get_db_path())) as db_manager:
                rows = db_manager.get_fornitori_by_richiesta(self.request_id, order_by=True)
            self.suppliers = [row[0] for row in rows]
            
            # Inizializza scores se non ancora fatto
            for supplier in self.suppliers:
                if supplier not in self.scores:
                    self.scores[supplier] = {'safety': tk.StringVar(value=''), 'quality': tk.StringVar(value=''),
                                           'delivery': tk.StringVar(value=''), 'cost': tk.StringVar(value='')}
        except DatabaseError as e:
            logger.error(f"Errore database in load_suppliers: {e}", exc_info=True)
            messagebox.showerror(_("Errore Database"), _("Impossibile caricare i fornitori: {}").format(e), parent=self)
    
    def create_weights_tab(self, parent):
        """Crea il tab per l'inserimento dei pesi percentuali"""
        info_label = ttk.Label(parent, 
                              text=_("Assegna i pesi percentuali a ciascun criterio (la somma deve essere 100%):"),
                              font=(None, 10))
        info_label.pack(pady=(0, 10))
        
        descs_frame = ttk.LabelFrame(parent, text=_("Descrizioni Criteri"), padding="10")
        descs_frame.pack(fill="x", pady=(0, 5))
        
        descriptions = [
            (_("Safety"), _("aderenza del fornitore agli standard di sicurezza, alle conformità normativa relativa al prodotto/servizio e al rischio finanziario e geopolitico.")),
            (_("Quality"), _("capacità del fornitore di rispettare integralmente le specifiche tecniche concordate.")),
            (_("Delivery"), _("capacità di rispettare il tempo di consegna offerto e la flessibilità a fronte di eventuali variazioni richieste.")),
            (_("Cost"), _("competitività relativa al prezzo totale offerto, considerando i termini di pagamento e i costi accessori (es. trasporto, installazione)."))
        ]
        
        for i, (criterion, desc) in enumerate(descriptions):
            row_frame = ttk.Frame(descs_frame)
            row_frame.pack(fill="x", pady=2)
            ttk.Label(row_frame, text=criterion + ":", font=(None, 10, 'bold'), width=12, anchor='w').pack(side="left", padx=(0, 10))
            ttk.Label(row_frame, text=desc, font=(None, 9), foreground="gray", wraplength=500).pack(side="left", fill="x")
        
        weights_frame = ttk.LabelFrame(parent, text=_("Pesi Percentuali"), padding="10")
        weights_frame.pack(fill="x", pady=(5, 10))
        
        for criterion in ['safety', 'quality', 'delivery', 'cost']:
            row = ttk.Frame(weights_frame)
            row.pack(fill="x", pady=5)
            label_text = {'safety': _("Safety"), 'quality': _("Quality"), 
                         'delivery': _("Delivery"), 'cost': _("Cost")}[criterion]
            ttk.Label(row, text=label_text + ":", width=15, anchor='w').pack(side="left")
            entry = ttk.Entry(row, textvariable=self.weights[criterion], width=10)
            entry.pack(side="left", padx=5)
            ttk.Label(row, text="%").pack(side="left")
    
    def create_scores_tab(self, parent):
        """Crea il tab per l'inserimento dei voti usando tksheet"""
        info_label = ttk.Label(parent, 
                              text=_("Assegna un voto da 1 a 10 per ciascun criterio e fornitore (doppio click per modificare):"),
                              font=(None, 10))
        info_label.pack(pady=(0, 10))
        
        # Pulsante per calcolare automaticamente i voti Cost
        calc_cost_frame = ttk.Frame(parent)
        calc_cost_frame.pack(fill="x", pady=(0, 10))
        ttk.Button(calc_cost_frame, text=_("🔄 Calcola Cost Automaticamente"), 
                  command=self.auto_calculate_cost).pack()
        
        # Frame per avviso prezzi mancanti (inizialmente nascosto)
        self.price_warning_frame = ttk.Frame(parent)
        self.price_warning_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        self.price_warning_label = tk.Label(
            self.price_warning_frame,
            text="",
            bg="red",
            fg="white",
            font=("Calibri", 11, "bold"),
            wraplength=800,
            justify="center",
            padx=10,
            pady=10
        )
        self.price_warning_label.pack(fill="x")
        
        # Nascondi inizialmente il frame
        self.price_warning_frame.pack_forget()
        
        # Frame per la tabella
        sheet_frame = ttk.Frame(parent)
        sheet_frame.pack(fill="both", expand=True)
        
        # Crea il widget tksheet
        self.sheet_scores = Sheet(sheet_frame,
                                 theme="light blue",
                                 header_font=("Calibri", 11, "bold"),
                                 font=("Calibri", 11, "normal"))
        self.sheet_scores.enable_bindings()
        self.sheet_scores.pack(fill="both", expand=True)
        
        # Imposta la funzione di validazione per l'editing delle celle
        self.sheet_scores.edit_validation(self.validate_sqdc_cell_input)
        
        # Popola il sheet con i fornitori
        self.refresh_scores_sheet()
        
        # Binding per catturare le modifiche e aggiornare i totali
        self.sheet_scores.extra_bindings([
            ("end_edit_cell", self.on_sqdc_cell_edit)
        ])
    
    def refresh_scores_sheet(self):
        """Popola o aggiorna il sheet con i dati dei fornitori"""
        if not hasattr(self, 'sheet_scores'):
            return
        
        # Definisci gli headers
        headers = [_("Fornitore"), _("Safety"), _("Quality"), _("Delivery"), _("Cost"), _("TOTALE")]
        
        # Costruisci le righe di dati e trova i vincitori
        data_rows = []
        max_total = -1
        winner_rows = []
        all_scores_complete = True
        
        for idx, supplier in enumerate(self.suppliers):
            # Recupera i valori correnti
            safety_val = self.scores[supplier]['safety'].get() or ''
            quality_val = self.scores[supplier]['quality'].get() or ''
            delivery_val = self.scores[supplier]['delivery'].get() or ''
            cost_val = self.scores[supplier]['cost'].get() or ''
            
            # Verifica se tutti i voti sono completi per questo fornitore
            if not (safety_val and quality_val and delivery_val and cost_val):
                all_scores_complete = False
            
            # Calcola il totale
            try:
                w_safety = float(self.weights['safety'].get() or 0) / 100
                w_quality = float(self.weights['quality'].get() or 0) / 100
                w_delivery = float(self.weights['delivery'].get() or 0) / 100
                w_cost = float(self.weights['cost'].get() or 0) / 100
                
                score_safety = float(safety_val or 0)
                score_quality = float(quality_val or 0)
                score_delivery = float(delivery_val or 0)
                score_cost = float(cost_val or 0)
                
                total = (score_safety * w_safety + score_quality * w_quality + 
                        score_delivery * w_delivery + score_cost * w_cost)
                
                # Forza sempre due decimali nel formato con virgola come separatore
                total_str = f"{total:.2f}".replace('.', ',')
                
                # Trova i vincitori (considera parità con tolleranza 0.01)
                if total > max_total + 0.01:
                    max_total = total
                    winner_rows = [idx]
                elif abs(total - max_total) <= 0.01:
                    winner_rows.append(idx)
                
            except (ValueError, TypeError):
                total_str = "0,00"
            
            # Aggiungi la riga
            data_rows.append([supplier, safety_val, quality_val, delivery_val, cost_val, total_str])
        
        # Carica i dati nel sheet
        self.sheet_scores.set_sheet_data(data_rows)
        self.sheet_scores.headers(headers)
        
        # Configura larghezze colonne
        self.sheet_scores.column_width(column=0, width=200)  # Fornitore
        for col_idx in range(1, 6):  # Safety, Quality, Delivery, Cost, TOTALE
            self.sheet_scores.column_width(column=col_idx, width=100)
        
        # Rendi le colonne Fornitore (0) e TOTALE (5) in sola lettura
        self.sheet_scores.readonly_columns(columns=[0, 5], readonly=True)
        
        # Prima rimuovi TUTTE le evidenziazioni dalla colonna TOTALE
        for row_idx, _supplier in enumerate(self.suppliers):
            self.sheet_scores.dehighlight_cells(row=row_idx, column=5)
        
        # Poi evidenzia in verde le celle TOTALE di TUTTI i vincitori se tutti i voti sono completi
        if winner_rows and all_scores_complete:
            for winner_row_idx in winner_rows:
                self.sheet_scores.highlight_cells(row=winner_row_idx, column=5, bg="#90EE90", fg="black")
    
    def validate_sqdc_cell_input(self, event):
        """Valida l'input delle celle SQDC durante l'editing - SOLO interi 1-10"""
        col_idx = event.get('column')
        raw_value = event.get('text', event.get('value', ''))
        new_value = raw_value.strip() if isinstance(raw_value, str) else str(raw_value)
        
        # Le colonne Fornitore (0) e TOTALE (5) sono già readonly
        if col_idx == 0 or col_idx == 5:
            return None
        
        # Se il campo è vuoto, permetti
        if new_value == '':
            return ''
        
        # Controlla che sia un numero intero tra 1 e 10
        try:
            score = int(new_value)
            if 1 <= score <= 10:
                return str(score)
            else:
                messagebox.showwarning(_("Valore Non Valido"), 
                                      _("I voti devono essere tra 1 e 10."),
                                      parent=self)
                return None
        except ValueError:
            messagebox.showwarning(_("Valore Non Valido"), 
                                  _("I voti devono essere numeri interi da 1 a 10."),
                                  parent=self)
            return None
    
    def on_sqdc_cell_edit(self, event):
        """Gestisce le modifiche alle celle del sheet SQDC - aggiorna i totali"""
        if not hasattr(self, 'sheet_scores'):
            return
        
        row_idx = event.row
        col_idx = event.column
        new_value = str(event.value).strip() if event.value is not None else ''
        
        if row_idx is None or col_idx is None:
            return
        
        # Non permettere editing sulla colonna Fornitore (0) o TOTALE (5)
        if col_idx == 0 or col_idx == 5:
            return
        
        # Mappa indice colonna a criterio
        col_to_criterion = {1: 'safety', 2: 'quality', 3: 'delivery', 4: 'cost'}
        if col_idx not in col_to_criterion:
            return
        
        # Verifica che row_idx sia valido
        if row_idx < 0 or row_idx >= len(self.suppliers):
            return
        
        # Ottieni il fornitore dalla riga
        supplier = self.suppliers[row_idx]
        criterion = col_to_criterion[col_idx]
        
        # Aggiorna il valore nella variabile StringVar
        self.scores[supplier][criterion].set(new_value)
        
        # Ricarica completamente il sheet per ricalcolare totali e evidenziare il vincitore
        self.refresh_scores_sheet()
    
    def update_price_warning(self):
        """Aggiorna l'avviso per i prezzi mancanti nella scheda Scores"""
        if not hasattr(self, 'price_warning_frame') or not hasattr(self, 'price_warning_label'):
            return
        
        if self.missing_price_suppliers:
            suppliers_list = ", ".join(self.missing_price_suppliers)
            
            if len(self.missing_price_suppliers) == 1:
                warning_text = _("⚠️ ATTENZIONE: Impossibile calcolare il prezzo automaticamente per il fornitore {} per mancanza di prezzi o quantità nella tabella di RdO. Il voto Cost è stato impostato a 0.").format(suppliers_list)
            else:
                warning_text = _("⚠️ ATTENZIONE: Impossibile calcolare il prezzo automaticamente per i fornitori {} per mancanza di prezzi o quantità nella tabella di RdO. I voti Cost sono stati impostati a 0.").format(suppliers_list)
            
            self.price_warning_label.config(text=warning_text)
            self.price_warning_frame.pack(fill="x", padx=10, pady=(0, 10))
            logger.info(f"SQDC: Avviso prezzi mancanti mostrato per: {suppliers_list}")
        else:
            self.price_warning_frame.pack_forget()
            logger.info("SQDC: Avviso prezzi mancanti nascosto")
    
    def auto_calculate_cost(self):
        """Calcola automaticamente i voti Cost basati sui prezzi"""
        # Prima verifica la somma dei pesi
        try:
            total_weight = (float(self.weights['safety'].get() or 0) + 
                           float(self.weights['quality'].get() or 0) + 
                           float(self.weights['delivery'].get() or 0) + 
                           float(self.weights['cost'].get() or 0))
        except (ValueError, TypeError):
            messagebox.showerror(_("Errore Pesi"),
                               _("I pesi devono essere numeri validi."),
                               parent=self)
            return
        
        # Verifica che la somma sia 100%
        if abs(total_weight - 100) > 0.01:
            messagebox.showerror(_("Errore Pesi"),
                               _("La somma dei pesi deve essere 100%. Attualmente: {:.1f}%").format(total_weight),
                               parent=self)
            return
        
        # Carica prezzi da database
        supplier_prices = {}
        self.missing_price_suppliers = []
        
        try:
            logger.info(f"SQDC auto_calculate_cost: Processing {len(self.suppliers)} suppliers")
            
            db_manager = DatabaseManager(get_db_path())
            total_items = db_manager.get_dettagli_count_by_richiesta(self.request_id)
            logger.info(f"SQDC auto_calculate_cost: Total items in RdO: {total_items}")
            
            for supplier in self.suppliers:
                results = db_manager.get_prezzo_quantita_by_fornitore(self.request_id, supplier)
                
                if len(results) < total_items:
                    logger.warning(f"SQDC auto_calculate_cost: Supplier {supplier} ha solo {len(results)} prezzi su {total_items} articoli - Cost impostato a 0")
                    self.scores[supplier]['cost'].set('0')
                    self.missing_price_suppliers.append(supplier)
                    continue
                
                total_price = 0
                valid_prices = 0
                has_invalid_price = False
                
                for price_val, qty in results:
                    price_str_raw = str(price_val).strip()
                    if not price_str_raw:
                        has_invalid_price = True
                        logger.warning(f"SQDC auto_calculate_cost: Supplier {supplier} ha un prezzo vuoto")
                        break
                    
                    price_str = price_str_raw.upper()
                    if price_str in ('X', 'ND'):
                        has_invalid_price = True
                        logger.warning(f"SQDC auto_calculate_cost: Supplier {supplier} ha prezzo non numerico: {price_str}")
                        break
                    
                    try:
                        price_float = parse_float_from_comma_string(str(price_val))
                        qty_float = parse_float_from_comma_string(str(qty))
                        total_price += price_float * qty_float
                        valid_prices += 1
                    except (ValueError, TypeError) as e:
                        logger.warning(f"SQDC auto_calculate_cost: Invalid price/qty for supplier {supplier}: {e}")
                        has_invalid_price = True
                        break
                
                if has_invalid_price or valid_prices < total_items:
                    logger.warning(f"SQDC auto_calculate_cost: Supplier {supplier} ha prezzi non validi - Cost impostato a 0")
                    self.scores[supplier]['cost'].set('0')
                    self.missing_price_suppliers.append(supplier)
                else:
                    supplier_prices[supplier] = total_price
                    logger.info(f"SQDC auto_calculate_cost: Supplier {supplier} total price: {total_price}")
            
            db_manager.close()
            
            # Se almeno un fornitore ha prezzi validi, calcola i voti
            if supplier_prices:
                min_price = min(supplier_prices.values())

                for supplier, price in supplier_prices.items():
                    try:
                        if min_price == 0:
                            score = 10 if price == 0 else 1
                        else:
                            price_ratio = price / min_price
                            score = 10 / price_ratio
                        
                        score = int(score + 0.5)
                        score = max(1, min(10, score))
                        
                        self.scores[supplier]['cost'].set(str(score))
                        
                    except (ZeroDivisionError, ValueError, TypeError) as e:
                        logger.error(f"Errore calcolo score per {supplier}: price={price}, min={min_price}, error={e}")
                        self.scores[supplier]['cost'].set('1')

                logger.info("SQDC auto_calculate_cost: Calculation completed successfully")

            else:
                logger.warning("SQDC auto_calculate_cost: Nessun fornitore ha prezzi completi")

        except DatabaseError as e:
            logger.error(f"SQDC auto_calculate_cost: Database error: {e}", exc_info=True)
            messagebox.showerror(_("Errore Database"),
                               _("Errore nel recupero dei prezzi dal database: {}").format(e),
                               parent=self)
        finally:
            self.refresh_scores_sheet()
            self.update_price_warning()
            
            # Cambia al tab Scores
            if hasattr(self, 'notebook'):
                try:
                    for idx in range(self.notebook.index('end')):
                        tab_text = self.notebook.tab(idx, 'text')
                        if tab_text and ('scores' in tab_text.lower() or 'voti' in tab_text.lower()):
                            self.notebook.select(idx)
                            logger.info(f"SQDC: Switched to Scores tab (index {idx})")
                            break
                except Exception as e:
                    logger.error(f"SQDC: Failed to switch to Scores tab: {e}", exc_info=True)
    
    def validate_weights_only(self):
        """Valida solo i pesi senza controllare i voti"""
        total_weight = 0
        for criterion in ['safety', 'quality', 'delivery', 'cost']:
            try:
                weight = float(self.weights[criterion].get() or 0)
                if weight < 0 or weight > 100:
                    messagebox.showerror(_("Errore Pesi"), 
                                       _("I pesi devono essere tra 0 e 100."), 
                                       parent=self)
                    return False
                total_weight += weight
            except ValueError:
                messagebox.showerror(_("Errore Pesi"), 
                                   _("I pesi devono essere numeri validi."), 
                                   parent=self)
                return False
        
        if abs(total_weight - 100) > 0.01:
            messagebox.showerror(_("Errore Pesi"), 
                               _("La somma dei pesi deve essere 100%. Attualmente: {:.1f}%").format(total_weight), 
                               parent=self)
            return False
        
        return True
    
    def validate_inputs(self):
        """Valida che i pesi sommino a 100% e i voti siano 1-10"""
        if not self.validate_weights_only():
            return False
        
        # Valida voti (solo interi 1-10) - TUTTI devono essere valorizzati
        for supplier in self.suppliers:
            for criterion in ['safety', 'quality', 'delivery', 'cost']:
                value = self.scores[supplier][criterion].get()
                if not value:
                    messagebox.showerror(_("Errore Voti"), 
                                       _("Devi compilare tutti i voti."), 
                                       parent=self)
                    return False
                try:
                    score = int(value)
                    if score < 1 or score > 10:
                        messagebox.showerror(_("Errore Voti"), 
                                           _("I voti devono essere tra 1 e 10.\nFornitore: {}\nCriterio: {}").format(supplier, criterion), 
                                           parent=self)
                        return False
                except ValueError:
                    messagebox.showerror(_("Errore Voti"), 
                                       _("I voti devono essere numeri interi da 1 a 10.\nFornitore: {}\nCriterio: {}").format(supplier, criterion), 
                                       parent=self)
                    return False
        
        return True
    
    def load_from_existing_data(self, data):
        """Carica dati da un'analisi SQDC esistente"""
        if 'weights' in data:
            for criterion in ['safety', 'quality', 'delivery', 'cost']:
                if criterion in data['weights']:
                    self.weights[criterion].set(str(data['weights'][criterion]))
        
        if 'scores' in data:
            for supplier, scores_dict in data['scores'].items():
                if supplier in self.scores:
                    for criterion in ['safety', 'quality', 'delivery', 'cost']:
                        if criterion in scores_dict:
                            self.scores[supplier][criterion].set(str(scores_dict[criterion]))
        
        if 'automatic_cost' in data:
            self.automatic_cost = data['automatic_cost']
        
        # Aggiorna il Sheet dopo caricamento dati
        if hasattr(self, 'sheet_scores'):
            self.refresh_scores_sheet()
    
    def export_to_excel(self):
        """Esporta l'analisi SQDC in un file Excel"""
        if not self.validate_inputs():
            return
        
        language = get_current_language()
        
        if language == 'it':
            template_name = "template_sqdc.xlsx"
            default_name = f"SQDC_Analisi_RdO_{self.request_id}.xlsx"
        else:
            template_name = "template_sqdc_eng.xlsx"
            default_name = f"SQDC_Analysis_RfQ_{self.request_id}.xlsx"
        
        template_path = resource_path(os.path.join("add_data", template_name))
        
        if not os.path.exists(template_path):
            messagebox.showerror(_("Errore"), 
                               _("File modello non trovato!\nAssicurarsi che '{}' esista nella cartella 'add_data'.").format(template_name), 
                               parent=self)
            return
        
        wb = None
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # Popola dati
            ws['B1'] = self.request_id
            ws['B2'] = datetime.now().strftime('%d/%m/%Y')
            
            ws['B5'] = float(self.weights['safety'].get() or 0)
            ws['B6'] = float(self.weights['quality'].get() or 0)
            ws['B7'] = float(self.weights['delivery'].get() or 0)
            ws['B8'] = float(self.weights['cost'].get() or 0)
            
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            
            start_row = 17
            max_total = -1
            winner_row = None
            
            def get_total(supplier):
                try:
                    w_safety = float(self.weights['safety'].get() or 0) / 100
                    w_quality = float(self.weights['quality'].get() or 0) / 100
                    w_delivery = float(self.weights['delivery'].get() or 0) / 100
                    w_cost = float(self.weights['cost'].get() or 0) / 100
                    
                    score_safety = float(self.scores[supplier]['safety'].get() or 0)
                    score_quality = float(self.scores[supplier]['quality'].get() or 0)
                    score_delivery = float(self.scores[supplier]['delivery'].get() or 0)
                    score_cost = float(self.scores[supplier]['cost'].get() or 0)
                    
                    return (score_safety * w_safety + score_quality * w_quality + 
                           score_delivery * w_delivery + score_cost * w_cost)
                except (ValueError, TypeError):
                    return 0.0
            
            suppliers_sorted = sorted(self.suppliers, key=get_total, reverse=True)
            
            for i, supplier in enumerate(suppliers_sorted):
                row = start_row + i
                ws.cell(row=row, column=1, value=supplier)
                
                try:
                    ws.cell(row=row, column=2, value=float(self.scores[supplier]['safety'].get() or 0))
                    ws.cell(row=row, column=3, value=float(self.scores[supplier]['quality'].get() or 0))
                    ws.cell(row=row, column=4, value=float(self.scores[supplier]['delivery'].get() or 0))
                    ws.cell(row=row, column=5, value=float(self.scores[supplier]['cost'].get() or 0))
                except (ValueError, TypeError) as e:
                    logger.warning(f"Errore conversione punteggio per {supplier}: {e}. Uso 0.")
                    ws.cell(row=row, column=2, value=0)
                    ws.cell(row=row, column=3, value=0)
                    ws.cell(row=row, column=4, value=0)
                    ws.cell(row=row, column=5, value=0)
                
                total = get_total(supplier)
                total_cell = ws.cell(row=row, column=6)
                total_cell.value = total
                total_cell.number_format = '0.00'
                
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')
                
                if total > max_total:
                    max_total = total
                    winner_row = row
            
            if winner_row:
                for col in range(1, 7):
                    ws.cell(row=winner_row, column=col).fill = green_fill
            
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[(_("File Excel"), "*.xlsx")],
                title=_("Salva Analisi SQDC"),
                initialfile=default_name,
                parent=self
            )
            
            if filepath:
                wb.save(filepath)
                logger.info(f"Analisi SQDC esportata: {filepath}")
                messagebox.showinfo(_("Successo"), 
                                  _("Analisi SQDC esportata con successo:\n{}").format(filepath), 
                                  parent=self)
        
        except Exception as e:
            logger.error(f"Errore esportazione SQDC: {e}", exc_info=True)
            messagebox.showerror(_("Errore Esportazione"), 
                               _("Impossibile esportare l'analisi: {}").format(e), 
                               parent=self)
        finally:
            if wb is not None:
                try:
                    wb.close()
                    logger.debug(f"Workbook SQDC chiuso: {template_path}")
                except Exception as close_error:
                    logger.warning(f"Errore chiusura workbook SQDC: {close_error}")
    
    def save_sqdc(self):
        """Salva l'analisi SQDC come Documento Interno"""
        if not self.validate_inputs():
            return
        
        language = get_current_language()
        
        sqdc_filename = f"SQDC_Analysis_RfQ_{self.request_id}.xlsx"
        
        if language == 'it':
            template_name = "template_sqdc.xlsx"
        else:
            template_name = "template_sqdc_eng.xlsx"
        
        template_path = resource_path(os.path.join("add_data", template_name))
        
        wb = None
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            ws['B1'] = self.request_id
            ws['B2'] = datetime.now().strftime('%d/%m/%Y')
            
            ws['B5'] = float(self.weights['safety'].get() or 0)
            ws['B6'] = float(self.weights['quality'].get() or 0)
            ws['B7'] = float(self.weights['delivery'].get() or 0)
            ws['B8'] = float(self.weights['cost'].get() or 0)
            
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            
            start_row = 17
            max_total = -1
            winner_row = None
            
            def get_total(supplier):
                try:
                    w_safety = float(self.weights['safety'].get() or 0) / 100
                    w_quality = float(self.weights['quality'].get() or 0) / 100
                    w_delivery = float(self.weights['delivery'].get() or 0) / 100
                    w_cost = float(self.weights['cost'].get() or 0) / 100
                    
                    score_safety = float(self.scores[supplier]['safety'].get() or 0)
                    score_quality = float(self.scores[supplier]['quality'].get() or 0)
                    score_delivery = float(self.scores[supplier]['delivery'].get() or 0)
                    score_cost = float(self.scores[supplier]['cost'].get() or 0)
                    
                    return (score_safety * w_safety + score_quality * w_quality + 
                           score_delivery * w_delivery + score_cost * w_cost)
                except (ValueError, TypeError):
                    return 0.0
            
            suppliers_sorted = sorted(self.suppliers, key=get_total, reverse=True)
            
            for i, supplier in enumerate(suppliers_sorted):
                row = start_row + i
                ws.cell(row=row, column=1, value=supplier)
                
                try:
                    ws.cell(row=row, column=2, value=float(self.scores[supplier]['safety'].get() or 0))
                    ws.cell(row=row, column=3, value=float(self.scores[supplier]['quality'].get() or 0))
                    ws.cell(row=row, column=4, value=float(self.scores[supplier]['delivery'].get() or 0))
                    ws.cell(row=row, column=5, value=float(self.scores[supplier]['cost'].get() or 0))
                except (ValueError, TypeError) as e:
                    logger.warning(f"Errore conversione punteggio per {supplier}: {e}. Uso 0.")
                    ws.cell(row=row, column=2, value=0)
                    ws.cell(row=row, column=3, value=0)
                    ws.cell(row=row, column=4, value=0)
                    ws.cell(row=row, column=5, value=0)
                
                total = get_total(supplier)
                total_cell = ws.cell(row=row, column=6)
                total_cell.value = total
                total_cell.number_format = '0.00'
                
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')
                
                if total > max_total:
                    max_total = total
                    winner_row = row
            
            if winner_row:
                for col in range(1, 7):
                    ws.cell(row=winner_row, column=col).fill = green_fill
            
            archive_path = get_fixed_attachments_dir()
            if not archive_path:
                messagebox.showerror(_("Errore"), _("Percorso allegati non disponibile."), parent=self)
                return
            
            try:
                db_manager_temp = DatabaseManager(get_db_path())
                try:
                    next_id = db_manager_temp.get_max_allegato_id() + 1
                finally:
                    try:
                        db_manager_temp.close()
                    except Exception:
                        pass
                
                file_ext = ".xlsx"
                new_filename = f"RfQ{self.request_id}_SQDC_ID{next_id}{file_ext}"
                dest_path = os.path.join(archive_path, new_filename)
                
                wb.save(dest_path)
                
                with DatabaseManager(get_db_path()) as db_manager:
                    db_manager.insert_or_update_allegato_sqdc(self.request_id, sqdc_filename, new_filename)
                
                logger.info(f"Analisi SQDC salvata come Documento Interno: {new_filename} -> {dest_path}")
                messagebox.showinfo(_("Successo"), 
                                  _("Analisi SQDC salvata correttamente nei Documenti Interni."), 
                                  parent=self)
                self.destroy()
                
            except DatabaseError as e:
                logger.error(f"Errore database in save_sqdc: {e}", exc_info=True)
                messagebox.showerror(_("Errore Database"), 
                                   _("Impossibile salvare l'analisi: {}").format(e), 
                                   parent=self)
        
        except Exception as e:
            logger.error(f"Errore nella creazione file SQDC: {e}", exc_info=True)
            messagebox.showerror(_("Errore"), 
                               _("Impossibile creare il file: {}").format(e), 
                               parent=self)
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception as close_error:
                    logger.warning(f"Errore chiusura workbook SQDC: {close_error}")

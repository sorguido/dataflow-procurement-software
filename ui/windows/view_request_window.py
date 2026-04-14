"""
ViewRequestWindow - Finestra di gestione dettagli RdO.
Estratta da dataflow.py per migliore organizzazione del codice.
"""

import tkinter as tk
from tkinter import ttk, filedialog
from tksheet import Sheet
import os
from datetime import datetime, date
from copy import copy
import openpyxl
import time
import logging
import tempfile

# Import da database
from database_manager import DatabaseManager, DatabaseError

# Import da utils
from utils.resource_utils import resource_path, set_window_icon
from utils.window_utils import center_window, calculate_optimal_window_size
from utils.i18n_utils import (
    tr,
    get_current_language,
    get_qty_column_text,
    normalize_rfq_type,
    translate_rfq_type
)
from utils.format_utils import parse_float_from_comma_string, format_quantity_display
from utils.validation_utils import format_date_for_db, format_price_display

# Import da services
from services.app_paths import get_db_path, get_fixed_attachments_dir

# Import finestre (estratte per compatibilità PyInstaller)
from ui.windows.edit_suppliers_window import EditSuppliersWindow
from ui.windows.edit_reference_window import EditReferenceWindow
from ui.windows.notes_window import NotesWindow
from ui.windows.purchase_order_window import PurchaseOrderWindow
from ui.windows.attachment_window import AttachmentWindow
from ui.windows.sqdc_analysis_window import SQDCAnalysisWindow
from ui.dialogs.common_dialogs import (
    LanguagePrompt,
    show_error,
    show_warning,
    show_info,
    show_confirm,
    show_ok_cancel
)

logger = logging.getLogger(__name__)


class ViewRequestWindow(tk.Toplevel):
    # Export RFQ Excel: tuning conservativo per colonne testuali del template
    _EXPORT_TEXT_MIN_WIDTH = 12
    _EXPORT_TEXT_MAX_WIDTH = 45
    _EXPORT_WRAP_THRESHOLD = 40

    def __init__(self, parent, request_id, read_only=False, source_db_path=None):
        super().__init__(parent)
        self.withdraw()
        set_window_icon(self)
        self.request_id = request_id
        self.selected_detail_id = None
        self.selected_supplier_name = None
        self.read_only = read_only  # Flag per modalità sola lettura
        
        # Determina quale database usare
        # Se source_db_path è fornito (RdO di altro utente), usa quello
        # Altrimenti usa il database corrente
        if source_db_path and os.path.exists(source_db_path):
            self.db_path = source_db_path
            print(f"[ViewRequestWindow] Usando DB remoto: {source_db_path}")
        else:
            self.db_path = get_db_path()
            print(f"[ViewRequestWindow] Usando DB locale: {self.db_path}")
        
        # Recupera il tipo di RdO e lo username dal database per il titolo
        tipo_rdo = "Fornitura piena"  # Valore di default
        username = None
        try:
            # BUG #47 FIX: Usa context manager per garantire chiusura DB anche su eccezione
            with DatabaseManager(self.db_path) as db_manager:
                result = db_manager.get_tipo_rdo(request_id)
                if result and result[0]:
                    tipo_rdo = result[0]
                username = db_manager.get_username_by_richiesta(request_id)
        except DatabaseError as e:
            logger.error(f"Errore database nel recupero dati per titolo: {e}", exc_info=True)
        
        # Traduci il tipo di RdO e imposta il titolo con User
        tipo_rdo_tradotto = translate_rfq_type(tipo_rdo)
        if username:
            title_base = tr("Control Panel - User: {} - Request N° {} - {}").format(username, request_id, tipo_rdo_tradotto)
        else:
            # Fallback se username non disponibile
            title_base = tr("Control Panel - Request N° {} - {}").format(request_id, tipo_rdo_tradotto)
        if self.read_only:
            title_base += tr(" [READ-ONLY]")
        self.title(title_base)
        
        # Rendi la finestra ridimensionabile e massimizzabile
        self.resizable(True, True)
        
        # Gestisci correttamente la chiusura della finestra (NON chiudere l'app principale)
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Frame pulsanti articoli (sempre in fondo)
        frame_article_buttons = ttk.Frame(self)
        frame_article_buttons.pack(side="bottom", fill="x", padx=10, pady=10)
        self.btn_add_article = ttk.Button(frame_article_buttons, text=tr("➕ Add Item"), command=self.add_new_article_row)
        self.btn_add_article.pack(side="left", padx=5)
        self.btn_remove_article = ttk.Button(frame_article_buttons, text=tr("🗑 Remove Selected Item"), command=self.remove_selected_article)
        self.btn_remove_article.pack(side="left", padx=5)
        self.btn_import_excel = ttk.Button(frame_article_buttons, text=tr("📊 Import from Excel"), command=self.import_from_excel)
        self.btn_import_excel.pack(side="left", padx=5)
        
        # Frame comandi (in alto)
        frame_comandi = ttk.Frame(self)
        frame_comandi.pack(side="top", fill="x", padx=10, pady=5)
        ttk.Button(frame_comandi, text=tr("📄 Manage Supplier Offers"), command=lambda: self.open_attachment_window("Offerta Fornitore")).pack(side="left")
        ttk.Button(frame_comandi, text=tr("📁 Manage Internal Documents"), command=lambda: self.open_attachment_window("Documento Interno")).pack(side="left", padx=10)
        # --- MODIFICA: Pulsante dinamico Fornitori ---
        self.btn_suppliers = ttk.Button(frame_comandi, text="...", command=self.open_edit_suppliers_window)
        self.btn_suppliers.pack(side="left")
        # --- FINE MODIFICA ---
        # --- MODIFICA RICHIESTA: Aggiunta pulsante Note ---
        self.btn_notes = ttk.Button(frame_comandi, text="...", command=self.open_notes_window)
        self.btn_notes.pack(side="left", padx=10)
        # --- FINE MODIFICA ---
        export_label = tr("Export")
        self.btn_export = ttk.Menubutton(
            frame_comandi,
            text=f"📊 {export_label}"
        )
        self.btn_export.pack(side="left", padx=10)
        self.export_menu = tk.Menu(self.btn_export, tearoff=0)
        self.export_menu.add_command(label=tr("📗 Excel"), command=self.export_to_excel)
        self.export_menu.add_command(label=tr("📄 PDF"), command=self.open_rfq_pdf_export_dialog)
        self.btn_export.config(menu=self.export_menu)
        # --- MODIFICA RICHIESTA: Aggiunta pulsante SQDC ---
        self.btn_sqdc = ttk.Button(frame_comandi, text="...", command=self.open_sqdc_analysis)
        self.btn_sqdc.pack(side="left", padx=10)
        # --- FINE MODIFICA ---
        
        # Frame dettagli richiesta con layout a griglia
        details_frame = ttk.LabelFrame(self, text=tr("Request Details"), padding="10")
        details_frame.pack(fill="x", padx=10, pady=5)
        
        s = ttk.Style(); s.configure("Clickable.TLabel", foreground="black", font=('Calibri', 9, 'underline'))
        
        # --- LAYOUT MODIFICATO CON GRID ---
        # Riga 0: Data Emissione + Data Scadenza + Pulsante Add PO
        ttk.Label(details_frame, text=tr("Issue Date: ")).grid(row=0, column=0, sticky="w", padx=(0,5), pady=5)
        
        # Import DateEntry qui per evitare circular dependency
        from tkcalendar import DateEntry
        
        self.entry_data_emissione = DateEntry(details_frame, width=12, date_pattern='dd/mm/yyyy', locale=('it_IT' if get_current_language() == 'it' else 'en_US'))
        self.entry_data_emissione.grid(row=0, column=1, sticky="w", padx=(0,20), pady=5)
        self.entry_data_emissione.bind('<<DateEntrySelected>>', self._on_date_changed)
        self.entry_data_emissione.bind('<FocusOut>', self._on_date_changed)  # Salva quando l'utente esce dal campo
        self.entry_data_emissione.bind('<Return>', self._on_date_changed)  # Salva quando l'utente preme Invio
        
        ttk.Label(details_frame, text=tr("Expiry Date: ")).grid(row=0, column=2, sticky="w", padx=(0,5), pady=5)
        self.entry_data_scadenza = DateEntry(details_frame, width=12, date_pattern='dd/mm/yyyy', locale=('it_IT' if get_current_language() == 'it' else 'en_US'))
        self.entry_data_scadenza.grid(row=0, column=3, sticky="w", padx=(0,20), pady=5)
        self.entry_data_scadenza.bind('<<DateEntrySelected>>', self._on_date_changed)
        self.entry_data_scadenza.bind('<FocusOut>', self._on_date_changed)  # Salva quando l'utente esce dal campo
        self.entry_data_scadenza.bind('<Return>', self._on_date_changed)  # Salva quando l'utente preme Invio
        
        # Pulsante Add PO allineato a destra nella stessa riga
        po_btn_text = tr("📋 Add PO")
        self.btn_po = ttk.Button(details_frame, text=po_btn_text, command=self.open_po_window)
        self.btn_po.grid(row=0, column=4, sticky="e", padx=(20,0), pady=5)
        
        # Riga 1: Riferimento
        ttk.Label(details_frame, text=tr("Reference: ")).grid(row=1, column=0, sticky="w", padx=(0,5), pady=5)
        self.lbl_riferimento = ttk.Label(details_frame, text="...", style="Clickable.TLabel", cursor="hand2")
        self.lbl_riferimento.grid(row=1, column=1, columnspan=3, sticky="w", pady=5)
        self.lbl_riferimento.bind("<Button-1>", self._on_reference_click)
        
        # Configura le colonne per espandere il pulsante Add PO a destra
        details_frame.grid_columnconfigure(4, weight=1)
        # --- FINE LAYOUT MODIFICATO ---
        
        # Frame griglia (espandibile, tra i pulsanti sopra e sotto)
        frame_grid = ttk.LabelFrame(self, text=tr("Price Table: Materials / Suppliers"))
        frame_grid.pack(side="top", fill="both", expand=True, padx=10, pady=5)
        
        # Crea il widget tksheet
        self.sheet = Sheet(frame_grid, 
                          theme="light blue",
                          header_font=("Calibri", 11, "bold"),
                          font=("Calibri", 11, "normal"))
        self.sheet.enable_bindings()
        self.sheet.pack(fill="both", expand=True)
        
        self.load_rdo_details(); self.build_grid()
        self.check_suppliers_status_and_update_button() # MODIFICA: Controlla lo stato dei fornitori all'avvio
        self.check_note_status_and_update_button() # MODIFICA RICHIESTA: Controlla lo stato della nota all'avvio
        self.check_sqdc_status_and_update_button() # MODIFICA: Controlla lo stato dell'analisi SQDC all'avvio
        
        # Disabilita pulsanti di modifica se in modalità read-only
        if self.read_only:
            self._disable_edit_controls()
            
            # Mostra un messaggio informativo sulla modalità sola lettura
            info_frame = ttk.Frame(self, style='Warning.TFrame')
            info_frame.pack(side='bottom', fill='x', padx=10, pady=(0, 5))
            
            info_label = ttk.Label(
                info_frame,
                text=tr("⚠️ READ-ONLY MODE: You are viewing an RfQ from another user. You cannot modify the data."),
                foreground='#d63031',
                font=('Calibri', 10, 'bold'),
                anchor='center'
            )
            info_label.pack(pady=5)
        
        # 🆕 MODIFICA: Calcola e imposta la larghezza ottimale della finestra in base ai fornitori
        num_suppliers = len(self.suppliers) if hasattr(self, 'suppliers') else 0
        is_cl = self.is_conto_lavoro if hasattr(self, 'is_conto_lavoro') else False
        optimal_geometry = calculate_optimal_window_size(self, num_suppliers, is_cl)
        self.geometry(optimal_geometry)

        try:
            self.state('zoomed')
        except Exception:
            self.attributes('-zoomed', True)

        self.deiconify()

    def _disable_edit_controls(self):
        """Disabilita tutti i controlli di modifica quando in modalità read-only."""
        # Disabilita pulsanti articoli
        if hasattr(self, 'btn_add_article'):
            self.btn_add_article.config(state='disabled')
        if hasattr(self, 'btn_remove_article'):
            self.btn_remove_article.config(state='disabled')
        if hasattr(self, 'btn_import_excel'):
            self.btn_import_excel.config(state='disabled')
        
        # Disabilita pulsanti gestione
        if hasattr(self, 'btn_suppliers'):
            self.btn_suppliers.config(state='disabled')
        if hasattr(self, 'btn_notes'):
            self.btn_notes.config(state='disabled')
        if hasattr(self, 'btn_sqdc'):
            self.btn_sqdc.config(state='disabled')
        if hasattr(self, 'btn_po'):
            self.btn_po.config(state='disabled')
        
        # NON disabilitare i pulsanti "Gestisci Offerte/Documenti" e "Export"
        # perché sono operazioni di sola lettura/visualizzazione
        
        # Disabilita date
        if hasattr(self, 'entry_data_emissione'):
            self.entry_data_emissione.config(state='disabled')
        if hasattr(self, 'entry_data_scadenza'):
            self.entry_data_scadenza.config(state='disabled')
        
        # Disabilita riferimento
        if hasattr(self, 'entry_riferimento'):
            self.entry_riferimento.config(state='disabled')
        
        # Disabilita click su label riferimento
        if hasattr(self, 'lbl_riferimento'):
            self.lbl_riferimento.unbind("<Button-1>")
            self.lbl_riferimento.config(cursor="")
        
        # Rendi la griglia read-only
        if hasattr(self, 'sheet'):
            self.sheet.disable_bindings()
            self.sheet.enable_bindings(
                "single_select",
                "row_select",
                "column_width_resize",
                "double_click_column_resize",
                "arrowkeys",
                "right_click_popup_menu",
                "copy"
            )

    def on_closing(self):
        """Gestisce la chiusura della finestra rilasciando tutte le risorse per prevenire memory leak."""
        # BUG #7 FIX: Cleanup completo per prevenire memory leak
        try:
            # Rilascia binding eventi per prevenire memory leak
            if hasattr(self, 'entry_data_emissione'):
                try:
                    self.entry_data_emissione.unbind('<<DateEntrySelected>>')
                    self.entry_data_emissione.unbind('<FocusOut>')
                    self.entry_data_emissione.unbind('<Return>')
                except Exception as e:
                    logger.debug(f"Errore unbind entry_data_emissione: {e}")
            
            if hasattr(self, 'entry_data_scadenza'):
                try:
                    self.entry_data_scadenza.unbind('<<DateEntrySelected>>')
                    self.entry_data_scadenza.unbind('<FocusOut>')
                    self.entry_data_scadenza.unbind('<Return>')
                except Exception as e:
                    logger.debug(f"Errore unbind entry_data_scadenza: {e}")
            
            # Rilascia binding sheet per prevenire memory leak
            if hasattr(self, 'sheet'):
                try:
                    self.sheet.unbind("<Double-Button-1>")
                    # Pulisci i dati memorizzati
                    if hasattr(self.sheet, '_sheet_data'):
                        del self.sheet._sheet_data
                    if hasattr(self.sheet, '_last_click_time'):
                        del self.sheet._last_click_time
                    if hasattr(self.sheet, '_last_click_row'):
                        del self.sheet._last_click_row
                except Exception as e:
                    logger.debug(f"Errore rilascio binding sheet: {e}")
            
            # Pulisci riferimenti circolari che potrebbero impedire garbage collection
            if hasattr(self, 'materials'):
                self.materials = None
            if hasattr(self, 'suppliers'):
                self.suppliers = None
            if hasattr(self, 'prices'):
                self.prices = None
                
        except Exception as e:
            logger.error(f"Errore in on_closing cleanup: {e}", exc_info=True)
        finally:
            self.destroy()

    def _get_db_manager(self):
        """Helper per ottenere il DatabaseManager corretto (locale o remoto).
        
        Se self.read_only è True (database di altro utente), apre in modalità sola lettura
        per permettere accesso concorrente senza lock.
        """
        return DatabaseManager(self.db_path, read_only=self.read_only)

    def open_notes_window(self):
        """Apre la finestra delle note"""
        win = NotesWindow(self, self.request_id)
        setattr(win, 'db_path', getattr(self, 'db_path', None))
        setattr(win, 'read_only', getattr(self, 'read_only', False))
        self.wait_window(win)
        self.check_note_status_and_update_button()
        
    def check_note_status_and_update_button(self):
        """Controlla se esiste una nota e aggiorna il testo del pulsante."""
        db_manager = None
        try:
            db_manager = self._get_db_manager()
            result = db_manager.get_note_formattate(self.request_id)
            
            if result and result[0] and result[0] != "()":
                self.btn_notes.config(text="📝 " + tr("View note"))
            else:
                self.btn_notes.config(text="📝 " + tr("Add note"))
        except DatabaseError as e:
            logger.error(f"Errore database in check_note_status_and_update_button: {e}", exc_info=True)
            self.btn_notes.config(text="⚠️ " + tr("Note Error"), state="disabled")
        finally:
            if db_manager:
                try:
                    db_manager.close()
                except Exception as close_err:
                    logger.warning(f"Errore chiusura DatabaseManager in check_note_status_and_update_button: {close_err}")
    
    def check_suppliers_status_and_update_button(self):
        """Controlla se esistono fornitori associati e aggiorna il testo del pulsante."""
        try:
            with DatabaseManager(self.db_path, read_only=self.read_only) as db_manager:
                result = db_manager.get_fornitori_count(self.request_id)
                
                if result and result > 0:
                    self.btn_suppliers.config(text="✏️ " + tr("Modify Suppliers"))
                else:
                    self.btn_suppliers.config(text="➕ " + tr("Add Suppliers"))
        except DatabaseError as e:
            logger.error(f"Errore database in check_suppliers_status_and_update_button: {e}", exc_info=True)
            self.btn_suppliers.config(text="⚠️ " + tr("Suppliers Error"), state="disabled")
    
    def check_sqdc_status_and_update_button(self):
        """Controlla se esiste un'analisi SQDC salvata e aggiorna il testo del pulsante."""
        try:
            sqdc_filename = f"SQDC_Analysis_RfQ_{self.request_id}.xlsx"
            
            with DatabaseManager(self.db_path, read_only=self.read_only) as db_manager_sqdc:
                existing = db_manager_sqdc.get_allegato_id_by_filename(self.request_id, sqdc_filename, 'Documento Interno')
            
            if existing:
                self.btn_sqdc.config(text=tr("📈 Open SQDC Analysis"))
            else:
                self.btn_sqdc.config(text=tr("📊 Create SQDC Analysis"))
                
        except DatabaseError as e:
            logger.error(f"Errore database in check_sqdc_status_and_update_button: {e}", exc_info=True)
            self.btn_sqdc.config(text=tr("SQDC Error"), state="disabled")
    
    def open_sqdc_analysis(self):
        """Apre la finestra di analisi SQDC (nuova o esistente)"""
        
        # Carica dati esistenti se presenti
        existing_data = None
        
        try:
            sqdc_display_name = f"SQDC_Analysis_RfQ_{self.request_id}.xlsx"
            
            db_manager_sqdc = DatabaseManager(self.db_path, read_only=self.read_only)
            try:
                allegato = db_manager_sqdc.get_allegato_id_by_filename(
                    self.request_id, 
                    sqdc_display_name, 
                    'Documento Interno'
                )
            finally:
                try:
                    db_manager_sqdc.close()
                except Exception:
                    pass
            
            if allegato:
                allegato_id = allegato[0]
                
                db_manager_file = DatabaseManager(self.db_path, read_only=self.read_only)
                try:
                    result = db_manager_file.get_allegato_file_data(allegato_id)
                finally:
                    try:
                        db_manager_file.close()
                    except Exception:
                        pass
                
                if result:
                    nome_file, dati_file, percorso_esterno = result
                    temp_blob_path = None
                    
                    try:
                        if percorso_esterno:
                            base_path = get_fixed_attachments_dir()
                            if not base_path:
                                raise ValueError("Percorso allegati non configurato")
                            
                            excel_path = os.path.join(base_path, percorso_esterno)
                            
                            if not os.path.exists(excel_path):
                                raise FileNotFoundError(f"File SQDC non trovato: {excel_path}")
                            
                            logger.info(f"Caricamento SQDC da file fisico: {excel_path}")
                        elif dati_file:
                            logger.warning("SQDC trovato come BLOB - conversione consigliata")
                            fd, temp_blob_path = tempfile.mkstemp(suffix='.xlsx')
                            os.close(fd)
                            with open(temp_blob_path, 'wb') as f:
                                f.write(dati_file)
                            excel_path = temp_blob_path
                        else:
                            raise ValueError("Nessun dato allegato disponibile")
                        
                        wb = openpyxl.load_workbook(excel_path, data_only=True)
                        ws = wb.active
                        
                        weights = {
                            'safety': ws['B5'].value or 25,
                            'quality': ws['B6'].value or 25,
                            'delivery': ws['B7'].value or 25,
                            'cost': ws['B8'].value or 25
                        }
                        
                        scores = {}
                        start_row = 17
                        row = start_row
                        
                        while row < 100:
                            supplier_cell = ws.cell(row=row, column=1)
                            if not supplier_cell.value:
                                break
                            
                            supplier = supplier_cell.value
                            scores[supplier] = {
                                'safety': ws.cell(row=row, column=2).value or 0,
                                'quality': ws.cell(row=row, column=3).value or 0,
                                'delivery': ws.cell(row=row, column=4).value or 0,
                                'cost': ws.cell(row=row, column=5).value or 0
                            }
                            row += 1
                        
                        wb.close()
                        
                        existing_data = {
                            'weights': weights,
                            'scores': scores,
                            'automatic_cost': False
                        }
                        
                        logger.info(f"Dati SQDC caricati per RdO {self.request_id}: {len(scores)} fornitori")
                        
                    except Exception as parse_error:
                        logger.error(f"Errore parsing file SQDC: {parse_error}", exc_info=True)
                        show_warning(
                            tr("Warning"),
                            tr("SQDC file found but unable to load data.\nA new empty analysis will be opened.\n\nError: {}").format(parse_error),
                            parent=self
                        )
                        existing_data = None
                    finally:
                        if temp_blob_path and os.path.exists(temp_blob_path):
                            try:
                                os.remove(temp_blob_path)
                                logger.debug(f"File temporaneo BLOB rimosso: {temp_blob_path}")
                            except Exception as e:
                                logger.warning(f"Impossibile rimuovere file temporaneo {temp_blob_path}: {e}")
        
        except (DatabaseError, Exception) as e:
            logger.error(f"Errore nel caricamento dati SQDC: {e}", exc_info=True)
        
        win = SQDCAnalysisWindow(self, self.request_id, existing_data)
        setattr(win, 'db_path', getattr(self, 'db_path', None))
        setattr(win, 'read_only', getattr(self, 'read_only', False))
        self.wait_window(win)
        self.check_sqdc_status_and_update_button()
    
    def _format_date_for_display(self, db_date):
        if not db_date: return tr("N/A")
        try: return datetime.strptime(db_date, '%Y-%m-%d').strftime('%d/%m/%Y')
        except (ValueError, TypeError): return db_date

    def load_rdo_details(self):
        try:
            with self._get_db_manager() as db_manager:
                result = db_manager.get_richiesta_basic_data(self.request_id)
            r, de, ds = result if result else (tr("N/A"), None, None)
            self.lbl_riferimento.config(text=r if r else tr("N/A"))
            
            try:
                if de: self.entry_data_emissione.set_date(datetime.strptime(de, '%Y-%m-%d'))
                else: self.entry_data_emissione.delete(0, 'end')
            except (ValueError, TypeError): self.entry_data_emissione.delete(0, 'end')
                
            try:
                if ds: self.entry_data_scadenza.set_date(datetime.strptime(ds, '%Y-%m-%d'))
                else: self.entry_data_scadenza.delete(0, 'end')
            except (ValueError, TypeError): self.entry_data_scadenza.delete(0, 'end')
                
        except DatabaseError as e:
            logger.error(f"Errore database in load_rdo_details: {e}", exc_info=True)
            show_error(tr("Error"), tr("Unable to load details: {}").format(e), parent=self)

    def open_edit_reference_window(self):
        """Apre la finestra di modifica riferimento."""
        if self.read_only:
            show_warning(
                tr("Operation Not Allowed"),
                tr("You cannot edit the reference of other users' RfQs."),
                parent=self
            )
            return
        
        win = EditReferenceWindow(self, self.request_id)
        setattr(win, 'db_path', getattr(self, 'db_path', None))
        setattr(win, 'read_only', getattr(self, 'read_only', False))
        self.wait_window(win)
        self.load_rdo_details()
    
    def open_po_window(self):
        """Apre la finestra di gestione numeri ordine di acquisto."""
        if self.read_only:
            show_warning(
                tr("Operation Not Allowed"),
                tr("You cannot edit purchase order numbers of other users' RfQs."),
                parent=self
            )
            return
        
        win = PurchaseOrderWindow(self, self.request_id)
        setattr(win, 'db_path', getattr(self, 'db_path', None))
        setattr(win, 'read_only', getattr(self, 'read_only', False))
        self.wait_window(win)

    def _on_date_changed(self, event=None):
        """Handler per eventi di cambio data."""
        self.auto_save_dates()
    
    def _on_reference_click(self, event=None):
        """Handler per click su etichetta riferimento."""
        self.open_edit_reference_window()

    def auto_save_dates(self):
        """Salva automaticamente le date quando vengono modificate dai calendari."""
        if hasattr(self, '_date_save_pending') and self._date_save_pending:
            return
        
        self._date_save_pending = True
        
        try:
            new_date_em = format_date_for_db(self.entry_data_emissione.get())
            new_date_sc = format_date_for_db(self.entry_data_scadenza.get())
            
            with self._get_db_manager() as db_manager:
                db_manager.update_date_richiesta(self.request_id, new_date_em, new_date_sc)
            
            if hasattr(self.master, 'refresh_data'):
                self.master.refresh_data()
                
        except Exception as e:
            logger.error(f"Errore database in auto_save_dates: {e}", exc_info=True)
            show_error(tr("Error"), tr("Unable to save dates: {}").format(e), parent=self)
        finally:
            if hasattr(self, 'after'):
                import weakref
                weak_self = weakref.ref(self)
                def release_lock():
                    obj = weak_self()
                    if obj is not None:
                        obj._date_save_pending = False
                self.after(300, release_lock)
            else:
                self._date_save_pending = False

    def save_dates(self):
        """Salva la data di emissione e scadenza modificate."""
        try:
            new_date_em = format_date_for_db(self.entry_data_emissione.get())
            new_date_sc = format_date_for_db(self.entry_data_scadenza.get())
        except Exception as e:
            show_error(tr("Date Format Error"), tr("Invalid dates: {}").format(e), parent=self)
            return

        try:
            with self._get_db_manager() as db_manager:
                db_manager.update_date_richiesta(self.request_id, new_date_em, new_date_sc)
            
            show_info(tr("Success"), tr("Dates updated."), parent=self)
            
            if hasattr(self.master, 'refresh_data'):
                self.master.refresh_data()
                
        except DatabaseError as e:
            logger.error(f"Errore database in save_dates: {e}", exc_info=True)
            show_error(tr("Database Error"), tr("Unable to save dates: {}").format(e), parent=self)

    def export_to_excel(self):
        """Esporta i dati della RdO in Excel"""
        logger.info(f"Esportazione Excel per RdO {self.request_id}")
        
        prompt = LanguagePrompt(self)
        self.wait_window(prompt)
        chosen_language = prompt.choice
        
        if not chosen_language:
            # BUG FIX: Riporta focus a ViewRequestWindow in caso di annullamento
            self.lift()
            self.focus_force()
            return
            
        try:
            with self._get_db_manager() as db_manager:
                rdo_type_result = db_manager.get_tipo_rdo(self.request_id)
            if not rdo_type_result:
                show_error(tr("Error"), tr("RfQ type not found."), parent=self)
                return
            tipo_normalizzato = normalize_rfq_type(rdo_type_result[0])
            is_cl = tipo_normalizzato == 'Conto lavoro'
        except Exception as e:
            show_error(tr("Database Error"), tr("Unable to determine the RfQ type: {}").format(e), parent=self)
            return

        template_name = ""
        texts = {}

        if is_cl:
            if chosen_language == 'ita':
                template_name = "template_rdo_cl.xlsx"
            else:
                template_name = "template_rdo_eng_cl.xlsx"
        else:
            if chosen_language == 'ita':
                template_name = "template_rdo.xlsx"
            else:
                template_name = "template_rdo_eng.xlsx"

        if chosen_language == 'ita':
            texts = {
                "save_title": "Salva Riepilogo",
                "initial_file": f"Riepilogo_RdO_{self.request_id}.xlsx",
                "vs_best": "VS. MIGLIORE"
            }
        else:
            texts = {
                "save_title": "Save Summary",
                "initial_file": f"Summary_RfQ_{self.request_id}.xlsx",
                "vs_best": "BEST DELIVERY"
            }
            
        template_path = resource_path(os.path.join("add_data", template_name))

        if not os.path.exists(template_path):
            show_error(tr("Error"), tr("Template file not found!\nMake sure '{}' exists in the 'add_data' folder.").format(template_name), parent=self)
            return
        
        wb = None
        try:
            with self._get_db_manager() as db_manager:
                rdo_det = db_manager.get_richiesta_full_data(self.request_id)
                if not rdo_det:
                    show_error(tr("Error"), tr("RfQ details not found."), parent=self)
                    return
                de_db, ds_db, rif, tipo = rdo_det
                suppliers_rows = db_manager.get_fornitori_by_richiesta(self.request_id, order_by=True)
                suppliers = [r[0] for r in suppliers_rows]
                items = db_manager.get_dettagli_by_richiesta(self.request_id)
                prices_rows = db_manager.get_offerte_by_richiesta(self.request_id)
            prices = {(id_d, nf): p for id_d, nf, p in prices_rows}
            
            wb = openpyxl.load_workbook(template_path); ws = wb.active
            
            from openpyxl.styles import Border, Side, Font, Alignment
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            bold = Font(bold=True); center = Alignment(horizontal='center', vertical='center'); left_align = Alignment(horizontal='left', vertical='center'); price_fmt = '0.0000'
            
            try:
                ws['B1'] = datetime.strptime(de_db, '%Y-%m-%d').strftime('%d/%m/%Y') if de_db else tr("N/A")
            except (ValueError, TypeError) as e:
                logger.error(f"Formato data emissione non valido per RdO {self.request_id}: '{de_db}' - {e}")
                ws['B1'] = tr("Invalid date")
            
            try:
                ws['B2'] = datetime.strptime(ds_db, '%Y-%m-%d').strftime('%d/%m/%Y') if ds_db else tr("N/A")
            except (ValueError, TypeError) as e:
                logger.error(f"Formato data scadenza non valido per RdO {self.request_id}: '{ds_db}' - {e}")
                ws['B2'] = tr("Invalid date")
            
            ws['C1'] = self.request_id
            for i, s_name in enumerate(suppliers): ws.cell(row=3, column=14+i, value=s_name).font=bold; ws.cell(row=3, column=14+i).alignment=center; ws.cell(row=3, column=14+i).border=border
            for i, item in enumerate(items):
                id_d, cod, all, desc, qta, c_g, d_g, m_cl = item; row = 4+i
                ws.cell(row=row, column=1, value=cod); ws.cell(row=row, column=2, value=all).alignment=left_align
                ws.cell(row=row, column=3, value=format_quantity_display(qta)).alignment=center; ws.cell(row=row, column=5, value=desc)
                if is_cl: ws.cell(row=row, column=6, value=c_g); ws.cell(row=row, column=7, value=d_g); ws.cell(row=row, column=8, value=m_cl).alignment=left_align
                
                ws.cell(row=row, column=10, value=texts["vs_best"]).alignment=center
                
                ws.cell(row=row, column=12, value=rif)
                for col in range(1, 13): ws.cell(row=row, column=col).border = border
                for j, s_name in enumerate(suppliers):
                    p_cell = ws.cell(row=row, column=14+j)
                    price_val = prices.get((id_d, s_name))
                    if price_val is not None:
                        try:
                            p_cell.value = float(str(price_val).replace(',', '.'))
                            p_cell.number_format = price_fmt
                        except (ValueError, TypeError):
                            p_cell.value = price_val
                            p_cell.number_format = '@'
                    p_cell.border = border

            self._apply_export_text_layout_tuning(
                ws=ws,
                suppliers_count=len(suppliers),
                item_count=len(items),
                is_cl=is_cl
            )

            # BUG FIX: Assicura che ViewRequestWindow sia in primo piano prima di aprire il file dialog
            # Questo previene che asksaveasfilename si apra dietro altre finestre
            self.lift()
            self.focus_force()
            self.update_idletasks()  # Forza aggiornamento UI prima del dialog
            
            filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[(tr("Excel file"), "*.xlsx")], title=texts["save_title"], initialfile=texts["initial_file"])
            
            if filepath: 
                wb.save(filepath)
                logger.info(f"Excel esportato: {filepath}")
                show_info(tr("Success"), tr("File saved in:\n{}").format(filepath), parent=self)
            else:
                # BUG FIX: Se annullato, riporta focus a ViewRequestWindow
                self.lift()
                self.focus_force()
        except Exception as e: 
            logger.error(f"Errore esportazione Excel: {e}", exc_info=True)
            show_error(tr("Export Error"), tr("Error: {}").format(e), parent=self)
        finally:
            if wb is not None:
                try:
                    wb.close()
                    logger.debug(f"Workbook Excel chiuso: {template_path}")
                except Exception as close_error:
                    logger.warning(f"Errore chiusura workbook Excel: {close_error}")

    def _apply_export_text_layout_tuning(self, ws, suppliers_count, item_count, is_cl):
        """Migliora la leggibilita' del template RFQ su colonne testuali mirate."""
        row_start = 4
        row_end = row_start + max(0, item_count) - 1

        column_rules = {
            2: {"min_width": 15, "max_width": 28, "wrap_threshold": 35},  # Allegato
            5: {"min_width": 35, "max_width": 52, "wrap_threshold": 55},  # Descrizione
            10: {"min_width": 18, "max_width": 24, "wrap_threshold": 24},  # Delivery required / VS. MIGLIORE
            12: {"min_width": 16, "max_width": 36, "wrap_threshold": 42},  # Riferimento
        }
        if is_cl:
            column_rules[6] = {"min_width": 15, "max_width": 24, "wrap_threshold": 30}  # Cod Grezzo
            column_rules[7] = {"min_width": 15, "max_width": 26, "wrap_threshold": 32}  # Dis Grezzo
            column_rules[8] = {"min_width": 20, "max_width": 30, "wrap_threshold": 36}  # Mat CL

        self._tune_export_text_columns(
            ws=ws,
            row_start=row_start,
            row_end=row_end,
            column_rules=column_rules
        )

        if suppliers_count > 0:
            supplier_rules = {
                14 + i: {"min_width": 12, "max_width": 28, "wrap_threshold": 24}
                for i in range(suppliers_count)
            }
            self._tune_export_text_columns(
                ws=ws,
                row_start=3,
                row_end=3,
                column_rules=supplier_rules
            )

    def _tune_export_text_columns(self, ws, row_start, row_end, column_rules):
        """Applica stima larghezza + wrap su range/colonne specifiche, senza autofit globale."""
        if row_end < row_start:
            return

        for col_idx, rule in column_rules.items():
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            col_dim = ws.column_dimensions[col_letter]
            current_width = col_dim.width
            if current_width is None:
                current_width = rule.get("min_width", self._EXPORT_TEXT_MIN_WIDTH)

            max_text_len = 0
            wrap_threshold = rule.get("wrap_threshold", self._EXPORT_WRAP_THRESHOLD)

            for row_idx in range(row_start, row_end + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                text_len = self._estimate_export_text_len(cell.value)
                max_text_len = max(max_text_len, text_len)

                if text_len > wrap_threshold:
                    self._set_wrap_text_preserving_alignment(cell)

            estimated_width = max_text_len + 2
            min_width = max(rule.get("min_width", self._EXPORT_TEXT_MIN_WIDTH), current_width)
            max_width = max(rule.get("max_width", self._EXPORT_TEXT_MAX_WIDTH), min_width)
            bounded_width = min(max(estimated_width, min_width), max_width)

            if bounded_width > current_width:
                col_dim.width = bounded_width

    @staticmethod
    def _estimate_export_text_len(value):
        if value is None:
            return 0
        text = str(value).strip()
        if not text:
            return 0
        return max(len(line) for line in text.splitlines())

    @staticmethod
    def _set_wrap_text_preserving_alignment(cell):
        """Abilita wrap_text senza perdere gli altri attributi di allineamento."""
        if cell.coordinate in cell.parent.merged_cells:
            return

        current_alignment = cell.alignment
        if current_alignment and current_alignment.wrap_text:
            return

        if current_alignment:
            new_alignment = copy(current_alignment)
            new_alignment.wrap_text = True
            cell.alignment = new_alignment
            return

        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    def open_rfq_pdf_export_dialog(self):
        """Apre il dialog di export RFQ PDF."""
        try:
            from ui.dialogs.rfq_pdf_export_dialog import RfqPdfExportDialog
        except Exception as e:
            logger.error(f"Errore import dialog RFQ PDF: {e}", exc_info=True)
            show_error(
                tr("Error"),
                tr("RFQ PDF feature not available: {}").format(e),
                parent=self
            )
            return

        dialog = RfqPdfExportDialog(
            parent=self,
            request_id=self.request_id,
            db_path=self.db_path,
            read_only=self.read_only,
        )
        self.wait_window(dialog)
        self.lift()
        self.focus_force()

    def open_edit_suppliers_window(self):
        """Apre la finestra di modifica fornitori"""
        print(f"[ViewRequestWindow] Prima di aprire EditSuppliersWindow - Richiesta {self.request_id}")
        win = EditSuppliersWindow(self, self.request_id)
        setattr(win, 'db_path', getattr(self, 'db_path', None))
        setattr(win, 'read_only', getattr(self, 'read_only', False))
        print(f"[ViewRequestWindow] Attendendo chiusura EditSuppliersWindow...")
        self.wait_window(win)
        print(f"[ViewRequestWindow] EditSuppliersWindow CHIUSA - Richiesta {self.request_id}")
        
        print(f"[ViewRequestWindow] Attesa 300ms per consolidamento WAL prima del refresh...")
        self.after(300, self._delayed_refresh_after_suppliers)
    
    def _delayed_refresh_after_suppliers(self):
        """Refresh ritardato dopo modifica fornitori per evitare race condition."""
        print(f"[ViewRequestWindow] === INIZIO REFRESH RITARDATO ===")
        print(f"[ViewRequestWindow] Chiamando refresh_grid() per richiesta {self.request_id}...")
        try:
            self.refresh_grid()
            print(f"[ViewRequestWindow] refresh_grid() COMPLETATO con successo")
        except Exception as e:
            print(f"[ViewRequestWindow] ERRORE in refresh_grid(): {e}")
            import traceback
            traceback.print_exc()
        
        print(f"[ViewRequestWindow] Aggiornando pulsante fornitori...")
        try:
            self.check_suppliers_status_and_update_button()
            print(f"[ViewRequestWindow] Pulsante fornitori aggiornato")
        except Exception as e:
            print(f"[ViewRequestWindow] ERRORE in check_suppliers_status_and_update_button(): {e}")
        
        print(f"[ViewRequestWindow] === FINE REFRESH RITARDATO ===")
    
    def refresh_grid(self):
        """Aggiorna la griglia ricostruendola"""
        self.build_grid()
    
    def build_grid(self):
        """Costruisce la griglia prezzi usando tksheet"""
        try:
            print(f"[ViewRequestWindow.build_grid] === INIZIO BUILD_GRID ===")
            print(f"[ViewRequestWindow.build_grid] Richiesta: {self.request_id}, DB path: {self.db_path}")
            
            self.update_idletasks()
            
            print(f"[ViewRequestWindow.build_grid] Apertura DatabaseManager...")
            with DatabaseManager(self.db_path, read_only=self.read_only) as db_manager:
                print(f"[ViewRequestWindow.build_grid] DatabaseManager aperto, recupero tipo RdO...")
                result = db_manager.get_tipo_rdo(self.request_id)
                tipo_normalizzato = normalize_rfq_type(result[0] if result and result[0] else "Fornitura piena")
                self.is_conto_lavoro = tipo_normalizzato == "Conto lavoro"
                print(f"[ViewRequestWindow.build_grid] Tipo RdO: {tipo_normalizzato}")
                
                print(f"[ViewRequestWindow.build_grid] Recupero fornitori...")
                suppliers_rows = db_manager.get_fornitori_by_richiesta(self.request_id, order_by=True)
                suppliers = [r[0] for r in suppliers_rows]
                print(f"[ViewRequestWindow.build_grid] Fornitori recuperati: {len(suppliers)} - {suppliers}")
                
                print(f"[ViewRequestWindow.build_grid] Recupero materiali...")
                materials = db_manager.get_dettagli_by_richiesta(self.request_id)
                print(f"[ViewRequestWindow.build_grid] Recupero prezzi...")
                prices_rows = db_manager.get_offerte_by_richiesta(self.request_id)
                print(f"[ViewRequestWindow.build_grid] Chiusura DatabaseManager...")
            
            print(f"[ViewRequestWindow.build_grid] DatabaseManager chiuso, elaborazione dati...")
            prices = {(id_d, nf): p for id_d, nf, p in prices_rows}
            print(f"[ViewRequestWindow.build_grid] Materiali: {len(materials)}, Prezzi: {len(prices)}")
            if materials:
                print(f"[ViewRequestWindow.build_grid] Primi 3 materiali: {materials[:3]}")
        except DatabaseError as e:
            logger.error(f"Errore database in build_grid: {e}", exc_info=True)
            show_error(tr("Database Error"), tr("Unable to load grid: {}").format(e), parent=self)
            return
        
        self.suppliers = suppliers
        self.materials = materials
        self.prices = prices
        
        self.sheet.set_sheet_data([])
        self.sheet.headers([])
        self.sheet.set_all_cell_sizes_to_text()
        
        base_headers = [
            tr("Code"),
            tr("Attachment"),
            tr("Description"),
            get_qty_column_text()
        ]
        
        if self.is_conto_lavoro:
            cl_headers = [tr("Raw Code"), tr("Raw Attachment"), tr("Material for Processing")]
            headers = base_headers + cl_headers + suppliers
            num_article_cols = 7
        else:
            headers = base_headers + suppliers
            num_article_cols = 4
        
        data_rows = []
        for i, (id_d, cod, all, desc, qta, c_g, d_g, m_cl) in enumerate(materials, start=1):
            row = [
                cod or "",
                all or "",
                desc or "",
                format_quantity_display(qta) or ""
            ]
            
            if self.is_conto_lavoro:
                row.extend([c_g or "", d_g or "", m_cl or ""])
            
            for s in suppliers:
                price = prices.get((id_d, s))
                display_val = ""
                if price is not None:
                    try:
                        price_str = str(price)
                        price_str_upper = price_str.upper()
                        if price_str_upper in ('X', 'ND'):
                            display_val = price_str_upper
                        else:
                            price_float = parse_float_from_comma_string(price_str)
                            display_val = format_price_display(price_float)
                    except (ValueError, TypeError):
                        display_val = str(price)
                row.append(display_val)
            
            data_rows.append(row)
        
        self.sheet.headers(headers)
        self.sheet.set_sheet_data(data_rows)
        
        self.sheet.update_idletasks()
        print(f"[ViewRequestWindow.build_grid] Griglia aggiornata con {len(data_rows)} righe di dati")
        
        try:
            import tkinter.font as tkfont
            header_font = tkfont.Font(family="Calibri", size=10, weight="bold")
            content_font = tkfont.Font(family="Calibri", size=10, weight="normal")
            
            PADDING_PX = 20
            
            article_min_widths = {
                0: 80,
                1: 80,
                2: 200,
                3: 60
            }
            
            if self.is_conto_lavoro:
                article_min_widths.update({
                    4: 100,
                    5: 100,
                    6: 150
                })
            
            DEFAULT_MIN_WIDTH = 100
            
            for col_idx in range(len(headers)):
                header_text = headers[col_idx]
                max_width = header_font.measure(header_text)
                
                if col_idx < num_article_cols:
                    for row in data_rows:
                        if col_idx < len(row):
                            cell_value = str(row[col_idx])
                            cell_width = content_font.measure(cell_value)
                            max_width = max(max_width, cell_width)
                    
                    min_width = article_min_widths.get(col_idx, 80)
                else:
                    min_width = DEFAULT_MIN_WIDTH
                
                column_width = max(int(max_width + PADDING_PX), min_width)
                self.sheet.column_width(column=col_idx, width=column_width)
                
        except Exception as e:
            logger.warning(f"Errore calcolo larghezza colonne: {e}. Uso larghezze default.")
            widths = [120, 120, 250, 60]
            if self.is_conto_lavoro:
                widths.extend([100, 100, 150])
            for col_idx in range(len(widths)):
                self.sheet.column_width(column=col_idx, width=widths[col_idx])
            for col_idx in range(len(widths), len(headers)):
                self.sheet.column_width(column=col_idx, width=120)
        
        align_cols = [3]
        self.sheet.align_columns(columns=align_cols, align="center")
        
        price_columns = list(range(num_article_cols, len(headers)))
        if price_columns:
            self.sheet.align_columns(columns=price_columns, align="right")
        
        for row_idx in range(len(data_rows)):
            if row_idx % 2 != 0:
                self.sheet.highlight_rows(rows=[row_idx], bg="#F0F0F0", fg="black", highlight_index=False)
        
        self.sheet.enable_bindings(
            "single_select",
            "drag_select",
            "column_select",
            "row_select",
            "column_width_resize",
            "double_click_column_resize",
            "row_width_resize",
            "column_height_resize",
            "arrowkeys",
            "row_height_resize",
            "double_click_row_resize",
            "right_click_popup_menu",
            "rc_select",
            "copy",
            "cut",
            "paste",
            "delete",
            "undo",
            "edit_cell"
        )
        
        self.sheet.extra_bindings("end_edit_cell", self.on_sheet_edit_cell)

    def on_sheet_edit_cell(self, event):
        """Gestisce la modifica di una cella nella tabella tksheet"""
        try:
            row = event.row
            col = event.column
            new_value = event.value
            
            if row is None or col is None or row >= len(self.materials):
                return
            
            if new_value is None:
                new_value = ""
            
            num_article_cols = 7 if self.is_conto_lavoro else 4
            
            if col < num_article_cols:
                success = self.save_article_field(row, col, str(new_value))
                
                if not success:
                    if row < len(self.materials):
                        old_material = self.materials[row]
                        if old_material:
                            material_index_map = {
                                0: 1,
                                1: 2,
                                2: 3,
                                3: 4,
                                4: 5,
                                5: 6,
                                6: 7
                            }
                            if col in material_index_map:
                                old_value = old_material[material_index_map[col]]
                                self.sheet.set_cell_data(row, col, str(old_value) if old_value else "")
                    return
                
                if col == 3:
                    updated_material = self.materials[row]
                    if updated_material and len(updated_material) > 4:
                        saved_qty = updated_material[4]
                        if saved_qty is not None:
                            self.sheet.set_cell_data(row, col, str(saved_qty))
            
            elif col >= num_article_cols:
                detail_id = self.materials[row][0]
                
                supplier_idx = col - num_article_cols
                if supplier_idx >= len(self.suppliers):
                    return
                
                supplier_name = self.suppliers[supplier_idx]
                
                formatted_value = self.save_price_in_db_no_refresh(detail_id, supplier_name, str(new_value))
                
                if formatted_value is not None:
                    self.sheet.set_cell_data(row, col, formatted_value)
                else:
                    self.sheet.set_cell_data(row, col, "")
            
        except Exception as e:
            logger.error(f"Errore in on_sheet_edit_cell: {e}", exc_info=True)

    def save_article_field(self, row_idx, col_idx, new_value):
        """Salva una modifica a un campo articolo nel database."""
        if row_idx >= len(self.materials):
            return False
        
        detail_id = self.materials[row_idx][0]
        
        field_map = {
            0: 'codice_materiale',
            1: 'disegno',
            2: 'descrizione_materiale',
            3: 'quantita',
            4: 'codice_grezzo',
            5: 'disegno_grezzo',
            6: 'materiale_conto_lavoro'
        }
        
        if col_idx not in field_map:
            return False
        
        field_name = field_map[col_idx]
        
        if col_idx == 3:
            if new_value and new_value.strip():
                if '.' in new_value and ',' not in new_value:
                    show_warning(
                        tr("Decimal Separator"),
                        tr("You used the dot (.) as decimal separator.\n\nThis program uses the COMMA (,) as decimal separator.\n\nCorrect example: 12,5 instead of 12.5"),
                        parent=self
                    )
                    return False
                
                try:
                    qty_float = parse_float_from_comma_string(new_value)
                    qty_float = round(qty_float, 4)
                    if qty_float == int(qty_float):
                        new_value = str(int(qty_float))
                    else:
                        new_value = str(qty_float).replace('.', ',').rstrip('0').rstrip(',')
                except (ValueError, TypeError):
                    pass
        
        try:
            db_manager = DatabaseManager(get_db_path())
            db_manager.update_dettaglio_field(detail_id, field_name, new_value)
            
            updated_row = db_manager.get_dettaglio_row_by_id(detail_id)
            if updated_row:
                self.materials[row_idx] = updated_row
            
            db_manager.close()
            return True
                
        except DatabaseError as e:
            logger.error(f"Errore database in save_article_field: {e}", exc_info=True)
            show_error(tr("Database Error"), 
                               tr("Unable to save changes: {}").format(e), 
                               parent=self)
            return False

    def save_price_in_db_no_refresh(self, detail_id, supplier_name, price_str):
        """Salva un valore nel DB senza aggiornare la griglia."""
        try:
            if not price_str:
                with DatabaseManager(get_db_path()) as db_manager:
                    db_manager.delete_offerta_by_dettaglio_fornitore(detail_id, supplier_name)
                return ""
            else:
                value_to_save = price_str
                
                price_str_upper = price_str.upper()
                
                if price_str_upper != 'X' and price_str_upper != 'ND':
                    try:
                        price_float = parse_float_from_comma_string(price_str)
                        value_to_save = format_price_display(price_float)
                    except ValueError:
                        show_error(tr("Format Error"), tr("The price must be a valid number (e.g. 123,45), 'X' or 'ND'.\nUse comma as decimal separator."), parent=self)
                        return None

                if price_str_upper == 'X' or price_str_upper == 'ND':
                    value_to_save = price_str_upper

                with DatabaseManager(get_db_path()) as db_manager:
                    db_manager.insert_or_replace_offerta(detail_id, supplier_name, value_to_save)
                
                return value_to_save
                
        except DatabaseError as e:
            logger.error(f"Errore database in save_price_in_db_no_refresh: {e}", exc_info=True)
            show_error(tr("Database Error"), tr("Unable to save price: {}").format(e), parent=self)
            return None

    def save_price_in_db(self, detail_id, supplier_name, price_str):
        """Salva un valore nel DB e aggiorna la griglia."""
        self.save_price_in_db_no_refresh(detail_id, supplier_name, price_str)
        self.refresh_grid()
    
    def add_new_article_row(self):
        """Aggiunge una nuova riga articolo vuota"""
        try:
            with DatabaseManager(get_db_path()) as db_manager:
                db_manager.insert_dettaglio_richiesta(self.request_id)
            
            self.refresh_grid()
            
            total_rows = self.sheet.get_total_rows()
            if total_rows > 0:
                self.sheet.see(row=total_rows-1, column=0, keep_yscroll=False, keep_xscroll=False, 
                              bottom_right_corner=False, check_cell_visibility=True)
                self.sheet.select_row(total_rows-1)
                
        except DatabaseError as e:
            logger.error(f"Errore database in add_new_article_row: {e}", exc_info=True)
            show_error(tr("Database Error"), 
                               tr("Unable to add the item: {}").format(e), 
                               parent=self)
    
    def remove_selected_article(self):
        """Rimuove gli articoli selezionati"""
        selected = self.sheet.get_selected_rows()
        print(f"[ViewRequestWindow.remove_selected_article] Righe selezionate: {selected}")
        
        if not selected:
            show_warning(tr("Warning"), 
                                  tr("Select at least one item to remove."), 
                                  parent=self)
            return
        
        if not hasattr(self, 'materials') or not self.materials:
            show_warning(tr("Warning"), 
                                  tr("No items available for deletion."), 
                                  parent=self)
            return
        
        print(f"[ViewRequestWindow.remove_selected_article] Numero materiali disponibili: {len(self.materials)}")
        
        if not show_confirm(tr("Delete Confirmation"), 
                                   tr("Are you sure you want to delete {} selected item(s)?\nAll associated prices will also be deleted.").format(len(selected)), 
                                   parent=self):
            return
        
        try:
            ids_to_delete = []
            invalid_indices = []
            
            for row_idx in selected:
                print(f"[ViewRequestWindow.remove_selected_article] Elaborazione riga {row_idx} (range: 0-{len(self.materials)-1})")
                if not isinstance(row_idx, int):
                    logger.error(f"remove_selected_article: row_idx non è int: {type(row_idx)} = {row_idx}")
                    invalid_indices.append(str(row_idx))
                    continue
                
                if 0 <= row_idx < len(self.materials):
                    detail_id = self.materials[row_idx][0]
                    ids_to_delete.append(detail_id)
                    print(f"[ViewRequestWindow.remove_selected_article] Aggiunto id_dettaglio {detail_id} alla lista di eliminazione")
                else:
                    print(f"[ViewRequestWindow.remove_selected_article] WARNING: Indice {row_idx} fuori range (0-{len(self.materials)-1})")
                    invalid_indices.append(str(row_idx))
            
            if invalid_indices:
                logger.warning(f"remove_selected_article: {len(invalid_indices)} indici invalidi: {invalid_indices}")
                show_warning(
                    tr("Warning"),
                    tr("Some selected indices are invalid and will be ignored: {}").format(", ".join(invalid_indices)),
                    parent=self
                )
            
            if not ids_to_delete:
                show_warning(tr("Warning"), 
                                      tr("No valid items selected for deletion."), 
                                      parent=self)
                return
            
            print(f"[ViewRequestWindow.remove_selected_article] Eliminazione di {len(ids_to_delete)} articoli: {ids_to_delete}")
            
            with DatabaseManager(get_db_path()) as db_manager:
                count = db_manager.delete_dettagli_batch(ids_to_delete)
            
            print(f"[ViewRequestWindow.remove_selected_article] Eliminati {count} articoli dal database")
            
            self.refresh_grid()
            
            show_info(tr("Success"), 
                               tr("{} item(s) successfully deleted.").format(count), 
                               parent=self)
            
        except DatabaseError as e:
            logger.error(f"Errore database in remove_selected_article: {e}", exc_info=True)
            show_error(tr("Database Error"), 
                               tr("Unable to delete the item: {}").format(e), 
                               parent=self)

    def import_from_excel(self):
        """Importa articoli da un file Excel"""
        try:
            with DatabaseManager(get_db_path()) as db_manager:
                result = db_manager.get_tipo_rdo(self.request_id)
            if not result:
                show_error(tr("Error"), tr("RfQ not found."), parent=self)
                return
            tipo_rdo = result[0]
        except DatabaseError as e:
            logger.error(f"Errore database in import_from_excel: {e}", exc_info=True)
            show_error(tr("Database Error"), tr("Unable to determine the RfQ type: {}").format(e), parent=self)
            return
        
        is_cl = (tipo_rdo == "Conto lavoro")
        
        msg = (tr("Make sure the Excel file has the following structure:\n\n")
               + tr("TYPE '{}' (4 columns):\n").format(tr("Full Supply"))
               + tr("A: Code, B: Attachment, C: Description, D: Quantity\n\n")
               + tr("TYPE '{}' (7 columns):\n").format(tr("Work Order"))
               + tr("A-D as above, E: Raw Code, F: Raw Attachment, G: Work Order Material"))
        if not show_ok_cancel(tr("Excel Import Instructions"), msg, parent=self):
            return

        filepath = filedialog.askopenfilename(
            title=tr("Select Excel file"), 
            filetypes=[(tr("Excel file"), "*.xlsx"), (tr("All files"), "*.*")],
            parent=self
        )
        if not filepath:
            return

        expected_cols = 7 if is_cl else 4
        workbook = None
        try:
            workbook = openpyxl.load_workbook(filepath, read_only=True)
            sheet = workbook.active
            
            if sheet.max_column < expected_cols:
                raise ValueError(tr("The Excel file must have at least {} columns for an RfQ '{}'.").format(
                    expected_cols, 
                    tr("Work Order") if is_cl else tr("Full Supply")
                ))
            
            items_to_add = []
            for row in sheet.iter_rows(min_row=1):
                cod = row[0].value
                allegato = row[1].value
                desc = row[2].value
                qta = row[3].value
                
                if cod is None or qta is None:
                    continue
                
                if is_cl:
                    cod_grezzo = str(row[4].value or "") if len(row) > 4 else ""
                    dis_grezzo = str(row[5].value or "") if len(row) > 5 else ""
                    mat_cl = str(row[6].value or "") if len(row) > 6 else ""
                    items_to_add.append((str(cod), str(allegato or ""), str(desc or ""), str(qta), cod_grezzo, dis_grezzo, mat_cl))
                else:
                    items_to_add.append((str(cod), str(allegato or ""), str(desc or ""), str(qta), "", "", ""))
            
            if not items_to_add:
                show_warning(tr("Warning"), tr("No valid items found in the Excel file."), parent=self)
                return
            
            db_manager = DatabaseManager(get_db_path())
            count = db_manager.import_dettagli_from_list(self.request_id, items_to_add)
            
            db_manager.close()
            
            print(f"[ViewRequestWindow.import_from_excel] Chiamata refresh_grid() dopo importazione di {count} articoli")
            self.refresh_grid()
            
            show_info(tr("Import Completed"), 
                               tr("{} items imported.").format(count), 
                               parent=self)
            
        except ValueError as e:
            show_error(tr("File Format Error"), str(e), parent=self)
        except Exception as e:
            logger.error(f"Errore in import_from_excel: {e}", exc_info=True)
            show_error(tr("Import Error"), 
                               tr("Unable to read the Excel file.\n{}").format(e), 
                               parent=self)
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                    logger.debug(f"Workbook Excel chiuso: {filepath}")
                except Exception as close_error:
                    logger.warning(f"Errore chiusura workbook Excel: {close_error}")

    def open_attachment_window(self, attachment_type):
        """Apre la finestra gestione allegati"""
        AttachmentWindow(self, self.request_id, attachment_type, read_only=self.read_only, source_db_path=self.db_path)
